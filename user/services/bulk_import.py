"""Spreadsheet import for master pages, reusing what the admin already has.

24 ``ModelResource`` classes already exist across the apps' ``admin.py`` files,
each with its own column names and foreign-key lookups. They work — but only
inside the Django admin, which is superuser-only, so nobody who actually
maintains a master can use them.

This exposes the same resources on the master pages themselves, gated by the
Web-Access matrix rather than by ``is_superuser``. Writing a second importer
would mean two sets of column rules to keep in step, and this codebase has
already paid that price elsewhere.

Two safeguards worth stating:

* An import is a bulk **add**, so it needs the tab's ``add`` right — viewing a
  master is not permission to rewrite it.
* Every import is dry-run first. The row errors come back before anything is
  written, so a bad column or a missing foreign key is a message rather than a
  half-finished master.
"""
from importlib import import_module

#: tab code -> "app.ResourceClassName". Only masters whose resource is a
#: straightforward column mapping; transaction resources are left out because
#: importing a voucher is not a spreadsheet problem.
IMPORTABLE = {
    "items": "inventory.admin.ItemResource",
    "customer": "sales.admin.CustomerWebImportResource",
    "sales_price_master": "sales.admin.SalesPriceMasterResource",
    "farmer_group": "broiler.admin.FarmerGroupResource",
    "branch_template": "broiler.admin.BranchResource",
    "supervisor_template": "broiler.admin.SupervisorResource",
    "broiler_line": "broiler.admin.BroilerLineResource",
    "branch_farm": "broiler.admin.BroilerFarmResource",
    "broiler_disease": "broiler.admin.BroilerDiseaseResource",
    "coa": "account.admin.ChartOfAccountResource",
    "bank_cash": "account.admin.BankCashMasterResource",
    "employee_list": "hr.admin.EmployeeResource",
    "employee_leave_details": "hr.admin.EmployeeLeaveResource",
    "employee_attendance": "hr.admin.AttendanceResource",
}

#: Masters with no resource yet — Item Category, Office/Warehouse and Supplier
#: among them. They need a ModelResource in their app's admin.py before they can
#: be listed above; a registry entry pointing at a class that does not exist
#: would render an Import button that fails when clicked.


def resource_for(tab_code):
    """The resource class registered for a tab, or None.

    Resolved lazily: the admin modules import a great deal, and a master page
    should not pay for that unless someone actually opens the import dialog.
    """
    path = IMPORTABLE.get(tab_code)
    if not path:
        return None
    module_path, _, name = path.rpartition(".")
    try:
        return getattr(import_module(module_path), name)
    except (ImportError, AttributeError):
        return None


def template_columns(tab_code):
    """Header row for the blank template someone downloads before filling it in.

    Taken from the resource itself, so the template and the parser can never
    disagree about what the columns are. The internal database id is left
    out: every import here is add-only, and on a row that does not exist yet
    there is no right answer to put there.
    """
    resource = resource_for(tab_code)
    if resource is None:
        return []
    return [c for c in resource().get_export_headers() if c != "id"]


def _choice_options(model, field_name):
    """Values from a plain Django ``choices=`` field, if the column is one."""
    try:
        field = model._meta.get_field(field_name)
    except Exception:
        return None
    choices = getattr(field, "choices", None)
    return [str(value) for value, _label in choices] if choices else None


def _picklist_options(app_label, model_name, field_name):
    """Live values from Picklist Master, for a field bound to one.

    A binding can be edited without a deploy, so this is read at request time
    rather than baked into the resource — the same values the web form's own
    dropdown for this field would show.
    """
    try:
        from picklist.services import get_field_config
    except ImportError:
        return None
    config = get_field_config(app_label, model_name, field_name)
    if config["mode"] != "PICKLIST":
        return None
    values = [value for value, _label in config["options"]]
    return values or None


def _fk_options(field):
    """Live values for a ForeignKeyWidget column, by the natural key it matches on."""
    from import_export.widgets import ForeignKeyWidget

    widget = getattr(field, "widget", None)
    if not isinstance(widget, ForeignKeyWidget):
        return None
    lookup = widget.field
    values = list(
        widget.model._default_manager
        .exclude(**{lookup: None}).exclude(**{lookup: ""})
        .order_by(lookup).values_list(lookup, flat=True).distinct()
    )
    return [str(v) for v in values] or None


def dropdown_options(tab_code):
    """``{column_name: [valid values]}`` for every column with a bounded
    value set — a Picklist-bound choice, a plain model ``choices`` field, or
    a foreign-key column matched by its natural key.

    Used to turn a column in the blank template into a real Excel drop-down,
    so filling it in is a pick rather than a guess at the exact spelling a
    foreign-key lookup or a choice field needs.
    """
    resource = resource_for(tab_code)
    if resource is None:
        return {}
    instance = resource()
    model = instance._meta.model
    app_label, model_name = model._meta.app_label, model.__name__
    out = {}
    for field in instance.get_export_fields():
        name = field.column_name
        attribute = field.attribute or name
        values = (_picklist_options(app_label, model_name, attribute)
                  or _fk_options(field)
                  or _choice_options(model, attribute))
        if values:
            out[name] = values
    return out


def build_template_workbook(tab_code):
    """A blank .xlsx template for a master's import: the resource's own
    columns as the header row, with an Excel drop-down list on every column
    ``dropdown_options`` can bound — so a "list" column in the sheet is an
    actual pick-list, the same as the field is on the master's own web form.
    """
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    columns = template_columns(tab_code)
    options = dropdown_options(tab_code)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Import"
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Data validation reads its list from a cell range, not an inline value —
    # so the option values live on a second, hidden sheet rather than being
    # typed into the formula (which also caps out on anything but a short list).
    if options:
        lists_sheet = workbook.create_sheet("Lists")
        lists_sheet.sheet_state = "hidden"
        list_col = 1
        for column_name, values in options.items():
            if column_name not in columns:
                continue
            letter = get_column_letter(list_col)
            for row_index, value in enumerate(values, start=1):
                lists_sheet.cell(row=row_index, column=list_col, value=value)
            validation = DataValidation(
                type="list",
                formula1=f"Lists!${letter}$1:${letter}${len(values)}",
                allow_blank=True,
            )
            validation.error = "Choose a value from the list."
            validation.errorTitle = "Invalid entry"
            sheet.add_data_validation(validation)
            target = get_column_letter(columns.index(column_name) + 1)
            validation.add(f"{target}2:{target}1000")
            list_col += 1

    return workbook


def run_import(tab_code, file_obj, filename, commit=False):
    """Parse and validate a spreadsheet; write it only when ``commit``.

    Returns ``(result, errors)`` where errors is a list of
    ``(row number, message)`` — row numbers as the user sees them in the sheet,
    counting the header, because "row 7 is wrong" is only useful if it matches
    what is on their screen.
    """
    import tablib

    resource = resource_for(tab_code)
    if resource is None:
        return None, [(0, "This page does not support import.")]

    suffix = (filename or "").rsplit(".", 1)[-1].lower()
    if suffix not in ("csv", "xlsx", "xls", "tsv"):
        return None, [(0, "Upload a .csv or .xlsx file.")]

    raw = file_obj.read()
    try:
        if suffix == "csv":
            dataset = tablib.Dataset().load(raw.decode("utf-8-sig"), format="csv")
        elif suffix == "tsv":
            dataset = tablib.Dataset().load(raw.decode("utf-8-sig"), format="tsv")
        else:
            dataset = tablib.Dataset().load(raw, format="xlsx")
    except Exception as exc:
        return None, [(0, f"Could not read the file: {exc}")]

    instance = resource()
    result = instance.import_data(dataset, dry_run=not commit,
                                  raise_errors=False, collect_failed_rows=True)

    id_fields = list(getattr(resource._meta, "import_id_fields", None) or [])
    errors = []
    for number, row in enumerate(result.rows, start=2):   # +1 header, +1 to 1-index
        for error in row.errors:
            errors.append((number, str(error.error)))
        for _field, messages in (row.validation_error.message_dict.items()
                                 if row.validation_error else []):
            errors.append((number, "; ".join(messages)))
        if row.is_skip():
            # A resource that matches on a natural key (e.g. Customer.mobile)
            # skips a row that already exists rather than raising the DB's
            # unique-constraint error — name the key so "skipped" reads as
            # "already on file" rather than an unexplained no-op.
            on = f" ({', '.join(id_fields)})" if id_fields else ""
            errors.append((number, f"Duplicate{on} — a matching record already exists; skipped."))
    for error in result.base_errors:
        errors.append((0, str(error.error)))
    return result, errors
