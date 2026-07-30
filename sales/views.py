#pylint: disable=no-member

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_POST
from django.http import Http404, JsonResponse
from django.utils import timezone

import json

from inventory.models import Item, ItemCategory
from account.models import ChartOfAccount
from hatchery_master.models import STATES_AND_TERRITORIES
from picklist.services import validate_value
from sales.models import (Customer, CustomerCreditNote, CustomerDebitNote, CustomerGroup,
                          CustomerShippingAddress, SalesPriceMaster)

# Used only by the billing/shipping address modals (state field itself is
# picklist-bound, see picklist.bindable_fields.BINDABLE_FIELDS).
states_and_union_territories = STATES_AND_TERRITORIES

@login_required
def customer(request):
    return render(request, "customer.html", {
        "customers": Customer.objects.select_related("customer_group").all()
    })


def _customer_form_context(customer=None):
    return {
        "customer": customer,
        "next_code": Customer.next_code() if not customer else None,
        "states_and_union_territories": states_and_union_territories,
        "to_pay_to_receive_choices": Customer.ToPayToReceive.choices,
        "customer_groups": CustomerGroup.objects.order_by("description", "code"),
        "today": timezone.localdate().isoformat(),
    }


def _apply_posted_customer_fields(instance, request):
    instance.name = request.POST.get("name", "").strip()
    instance.address = request.POST.get("address", "").strip()
    instance.mobile = request.POST.get("mobile", "").strip()
    instance.mobile_2 = request.POST.get("mobile_2", "").strip()
    instance.customer_group_id = request.POST.get("customer_group") or None
    instance.email = request.POST.get("email", "").strip()
    instance.pan_tin = request.POST.get("pan_tin", "").strip()
    instance.aadhar = request.POST.get("aadhar", "").strip()
    instance.contact_type = request.POST.get("contact_type") or Customer.ContactType.BOTH
    instance.party_category = request.POST.get("party_category") or None
    instance.gstin = request.POST.get("gstin", "").strip()
    instance.state = request.POST.get("state", "").strip()
    instance.opening_balance = request.POST.get("opening_balance") or None
    instance.to_pay_to_receive = request.POST.get("to_pay_to_receive") or None
    instance.as_on_date = request.POST.get("as_on_date") or None
    instance.note = request.POST.get("note", "").strip()
    instance.credit_period = request.POST.get("credit_period") or None
    instance.credit_limit = request.POST.get("credit_limit") or 0
    instance.country = request.POST.get("country", "").strip()
    instance.currency = request.POST.get("currency", "").strip()
    instance.account_no = request.POST.get("account_no", "").strip()
    instance.ifsc_code = request.POST.get("ifsc_code", "").strip()
    instance.bank_details = request.POST.get("bank_details", "").strip()
    instance.terms = request.POST.get("terms", "").strip()
    instance.agreement_start_date = request.POST.get("agreement_start_date") or None
    instance.agreement_months = request.POST.get("agreement_months") or None
    if request.FILES.get("agreement_copy"):
        instance.agreement_copy = request.FILES["agreement_copy"]
    if request.FILES.get("other_documents"):
        instance.other_documents = request.FILES["other_documents"]
    for field in ("state", "contact_type", "party_category"):
        validate_value("sales", "Customer", field, getattr(instance, field))


@login_required(login_url="login")
def create_customer(request):
    """Add a new customer master record."""
    if request.method == "POST":
        instance = Customer()
        try:
            _apply_posted_customer_fields(instance, request)
            instance.full_clean()
            instance.save()
            _create_posted_shipping_addresses(instance, request)
            messages.success(request, "Customer added successfully.")
            return redirect("customer")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "customer_form.html", _customer_form_context())


def _create_posted_shipping_addresses(instance, request):
    try:
        addresses = json.loads(request.POST.get("shipping_addresses_json") or "[]")
    except json.JSONDecodeError:
        addresses = []
    if not addresses and instance.address:
        addresses = [{"label": instance.address[:100], "address": instance.address, "is_default": True}]
    default_assigned = False
    for entry in addresses:
        label = (entry.get("label") or "").strip()
        address_text = (entry.get("address") or "").strip()
        if not label or not address_text:
            continue
        is_default = bool(entry.get("is_default")) and not default_assigned
        default_assigned = default_assigned or is_default
        CustomerShippingAddress.objects.create(
            customer=instance, label=label, address=address_text,
            contact_person=(entry.get("contact_person") or "").strip(),
            mobile=(entry.get("mobile") or "").strip(),
            is_default=is_default,
        )

def _sync_default_shipping_address(instance, previous_address):
    """Keep the default shipping address in step with the billing address.

    On create the default shipping address is seeded from billing, so the two
    start out identical. Editing billing therefore has to move it along too —
    but only while it is still a mirror of the old billing text. Once someone
    has given the shipping address its own content, it is left alone.
    """
    new_address = (instance.address or "").strip()
    if not new_address or new_address == (previous_address or "").strip():
        return
    default = instance.shipping_addresses.filter(is_default=True).first()
    if default is None:
        CustomerShippingAddress.objects.create(
            customer=instance, label=new_address[:100], address=new_address, is_default=True)
        return
    if (default.address or "").strip() == (previous_address or "").strip():
        default.label = new_address[:100]
        default.address = new_address
        default.save(update_fields=["label", "address"])


@login_required(login_url="login")
def edit_customer(request, id):
    """Edit an existing customer master record."""
    instance = get_object_or_404(Customer, id=id)

    if request.method == "POST":
        previous_address = instance.address
        try:
            _apply_posted_customer_fields(instance, request)
            instance.full_clean()
            instance.save()
            _sync_default_shipping_address(instance, previous_address)
            messages.success(request, "Customer updated successfully.")
            return redirect("customer")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "customer_form.html", _customer_form_context(instance))


@login_required(login_url="login")
@require_POST
def delete_customer(request, id):
    """Delete a customer master record."""
    instance = get_object_or_404(Customer, id=id)
    instance.delete()
    messages.success(request, "Customer deleted successfully.")
    return redirect("customer")


@login_required
def customer_groups(request):
    return render(request, "customer_groups.html", {
        "coa_accounts": ChartOfAccount.objects.filter(status="Active").order_by("code"),
    })


@login_required
def sales_price(request):
    return render(request, "sales_price_master.html")


@method_decorator(login_required, name="dispatch")
class CustomerShippingAddressAPI(View):
    """Customer Master addresses, also used by transaction forms."""
    def get(self, request, customer_id, id=None):
        customer = Customer.objects.get(id=customer_id)
        addresses = customer.shipping_addresses.all()
        if id:
            addresses = addresses.filter(id=id)
        result = [{
            "id": address.id, "label": address.label, "address": address.address,
            "contact_person": address.contact_person, "mobile": address.mobile,
            "is_default": address.is_default,
        } for address in addresses]
        if id:
            if not result:
                return JsonResponse({"error": "Shipping address not found"}, status=404)
            return JsonResponse(result[0])
        return JsonResponse(result, safe=False)

    def post(self, request, customer_id):
        try:
            data = json.loads(request.body)
            customer = Customer.objects.get(id=customer_id)
            if not data.get("label") or not data.get("address"):
                return JsonResponse({"error": "Address label and address are required"}, status=400)
            if data.get("is_default"):
                customer.shipping_addresses.update(is_default=False)
            address = CustomerShippingAddress.objects.create(
                customer=customer, label=data["label"], address=data["address"],
                contact_person=data.get("contact_person", ""), mobile=data.get("mobile", ""),
                is_default=bool(data.get("is_default")),
            )
            return JsonResponse({"id": address.id, "message": "Shipping address saved"}, status=201)
        except Customer.DoesNotExist:
            return JsonResponse({"error": "Customer not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def put(self, request, customer_id, id):
        try:
            data = json.loads(request.body)
            address = CustomerShippingAddress.objects.get(id=id, customer_id=customer_id)
            if data.get("is_default"):
                CustomerShippingAddress.objects.filter(customer_id=customer_id).exclude(id=id).update(is_default=False)
            for field in ("label", "address", "contact_person", "mobile"):
                if field in data:
                    setattr(address, field, data[field])
            address.is_default = bool(data.get("is_default", address.is_default))
            address.full_clean(); address.save()
            return JsonResponse({"message": "Shipping address updated"})
        except CustomerShippingAddress.DoesNotExist:
            return JsonResponse({"error": "Shipping address not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def delete(self, request, customer_id, id):
        try:
            CustomerShippingAddress.objects.get(id=id, customer_id=customer_id).delete()
            return JsonResponse({"message": "Shipping address deleted"})
        except CustomerShippingAddress.DoesNotExist:
            return JsonResponse({"error": "Shipping address not found"}, status=404)


@method_decorator(login_required, name="dispatch")
class CustomerGroupAPI(View):

    @staticmethod
    def _serialize(group):
        return {
            "id": group.id,
            "code": group.code,
            "description": group.description,
            "currency": group.currency,
            "control_account": group.control_account_id,
            "control_account_display": str(group.control_account) if group.control_account else "",
            "advance_account": group.advance_account_id,
            "advance_account_display": str(group.advance_account) if group.advance_account else "",
        }

    def get(self, request, id=None):
        """
        Handle GET requests to retrieve either a list of customer groups or a specific customer group.
        """
        if id:
            try:
                customer_group = CustomerGroup.objects.select_related("control_account", "advance_account").get(id=id)
                return JsonResponse(self._serialize(customer_group))
            except CustomerGroup.DoesNotExist:
                raise Http404("Customer group not found")
        else:
            customer_groups = [
                self._serialize(group)
                for group in CustomerGroup.objects.select_related("control_account", "advance_account")
            ]
            return JsonResponse(customer_groups, safe=False)

    def post(self, request):
        """
        Handle POST requests to create a new customer group.
        """
        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

        required_fields = ["code", "description", "currency", "control_account", "advance_account"]
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({"error": f"{field} field is required"}, status=400)

        CustomerGroup.objects.create(
            code=data["code"],
            description=data["description"],
            currency=data["currency"],
            control_account_id=data["control_account"],
            advance_account_id=data["advance_account"]
        )
        return JsonResponse({"message": "Customer group created successfully"}, status=201)

    def put(self, request, id):
        """
        Handle PUT requests to update an existing customer group.
        """
        try:
            customer_group = CustomerGroup.objects.get(id=id)
        except CustomerGroup.DoesNotExist:
            return JsonResponse({"error": "Customer group not found"}, status=404)

        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)

        for field in ["code", "description", "currency"]:
            setattr(customer_group, field, data.get(field, getattr(customer_group, field)))
        if "control_account" in data:
            customer_group.control_account_id = data["control_account"] or None
        if "advance_account" in data:
            customer_group.advance_account_id = data["advance_account"] or None

        customer_group.save()
        return JsonResponse({"message": "Customer group updated successfully"})

    def delete(self, request, id):
        """
        Handle DELETE requests to delete an existing customer group.
        """
        try:
            customer_group = CustomerGroup.objects.get(id=id)
        except CustomerGroup.DoesNotExist:
            return JsonResponse({"error": "Customer group not found"}, status=404)

        customer_group.delete()
        return JsonResponse({"message": "Customer group deleted successfully"}, status=204)
    

@method_decorator(login_required, name="dispatch")
class SalesPriceMasterAPI(View):
    def get(self, request, id=None):
        if id:
            try:
                sales_price = SalesPriceMaster.objects.get(id=id)
                sales_price_data = {
                    "id": sales_price.id,
                    "item_category": sales_price.item_category.id,
                    "item": sales_price.item.id,
                    "price": sales_price.price,
                    "date": sales_price.date,
                }
                return JsonResponse(sales_price_data)
            except SalesPriceMaster.DoesNotExist:
                raise Http404("Sales price not found")
        else:
            sales_prices = list(SalesPriceMaster.objects.values(
                "id", "item_category", "item", "price", "date"
            ))
            return JsonResponse(sales_prices, safe=False)

    def post(self, request):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        try:
            item_category = ItemCategory.objects.filter(id=data["item_category"]).first()
            item = Item.objects.filter(id=data["item"]).first()

            sales_price = SalesPriceMaster.objects.create(
                item_category=item_category,
                item=item,
                price=data["price"],
                date=data["date"],
            )

            return JsonResponse({"message": "Sales price created", "id": sales_price.id}, status=201)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    def put(self, request, id):
        try:
            sales_price = SalesPriceMaster.objects.get(id=id)
        except SalesPriceMaster.DoesNotExist:
            return JsonResponse({"error": "Sales price not found"}, status=404)

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        sales_price.price = data.get("price", sales_price.price)
        sales_price.save()
        return JsonResponse({"message": "Sales price updated"})

    def delete(self, request, id):
        try:
            sales_price = SalesPriceMaster.objects.get(id=id)
        except SalesPriceMaster.DoesNotExist:
            return JsonResponse({"error": "Sales price not found"}, status=404)

        sales_price.delete()
        return JsonResponse({"message": "Sales price deleted"}, status=204)    


# ---------------------------------------------------------------------------
# Sales Invoice (Sales > Transactions)
# ---------------------------------------------------------------------------
from decimal import Decimal, InvalidOperation
from sales.models import SalesInvoice, SalesInvoiceItem
from inventory.models import Warehouse


def _si_num(v):
    try:
        return Decimal(str(v)) if v not in (None, "") else Decimal("0")
    except Exception:
        return Decimal("0")


def _sales_invoice_list_dict(inv):
    return {
        "id": inv.id, "date": inv.date.isoformat(), "invoice_no": inv.invoice_no,
        "transaction_type": inv.transaction_type,
        "customer_name": inv.customer.name if inv.customer_id else "",
        "reference_no": inv.reference_no,
        "total_items": inv.total_items(),
        "amount": str(inv.net_amount),
        "is_active": inv.is_active,
    }


def _sales_invoice_item_dict(row):
    return {
        "item": row.item_id, "item_code": row.item.item_code,
        "item_name": row.item.description, "batch_no": row.batch_no,
        "hsn_sac": row.hsn_sac, "uom": row.uom,
        "quantity": str(row.quantity), "free_qty": str(row.free_qty), "rate": str(row.rate),
        "discount_percent": str(row.discount_percent),
        "taxable_amount": str(row.taxable_amount),
        "gst_percent": str(row.gst_percent), "gst_amount": str(row.gst_amount),
        "amount": str(row.amount),
    }


def _sales_invoice_form_context(inv=None):
    from account.models import TermsConditions, BankCashMaster, CompanyProfile
    items = list(Item.objects.order_by("item_code").values(
        "id", "item_code", "description", "hsn_code", "storage_uom", "standard_cost_per_unit"))
    return {
        "invoice": inv,
        "next_no": SalesInvoice._next_no() if not inv else None,
        "customers": Customer.objects.order_by("name"),
        "branches": Warehouse.objects.order_by("name"),
        "org_centres": __import__("account.models", fromlist=["OrganizationCentre"]).OrganizationCentre.objects.order_by("name"),
        "items_json": json.dumps(items, default=str),
        "transaction_types": SalesInvoice.TRANSACTION_TYPE_CHOICES,
        "today": timezone.localdate().isoformat(),
        "existing_items_json": json.dumps(
            [_sales_invoice_item_dict(r) for r in inv.items.select_related("item")]) if inv else "[]",
        "terms_list": list(TermsConditions.objects.order_by("type").values("id", "type", "party_type", "condition")),
        "banks": list(BankCashMaster.objects.order_by("name").values("id", "code", "name", "micr", "address", "contact_person", "is_cash")),
        "company": CompanyProfile.get_solo(),
        "states_and_union_territories": states_and_union_territories,
        "terms_json": json.dumps({t.id: t.condition for t in TermsConditions.objects.all()}),
        "banks_json": json.dumps({b["id"]: b for b in BankCashMaster.objects.values("id", "code", "name", "micr", "address", "contact_person")}, default=str),
    }


def _apply_posted_invoice(instance, request):
    d = request.POST
    instance.transaction_type = d.get("transaction_type") or "Sales Invoice"
    instance.date = d.get("date") or timezone.localdate()
    instance.customer_id = d.get("customer") or None
    instance.billing_address = d.get("billing_address") or ""
    instance.shipping_address = d.get("shipping_address") or ""
    instance.gstin = d.get("gstin") or ""
    instance.reference_no = d.get("reference_no") or ""
    instance.reference_date = d.get("reference_date") or None
    instance.transportation = d.get("transportation") or ""
    instance.vehicle_no = d.get("vehicle_no") or ""
    instance.place_of_supply = d.get("place_of_supply") or ""
    instance.eway_bill_no = d.get("eway_bill_no") or ""
    instance.branch_id = d.get("branch") or None
    instance.organization_centre_id = d.get("organization_centre") or None
    instance.sales_person = d.get("sales_person") or ""
    instance.payment_terms = d.get("payment_terms") or ""
    instance.due_date = d.get("due_date") or None
    instance.remarks = d.get("remarks") or ""
    instance.terms_conditions_id = d.get("terms_conditions") or None
    instance.bank_account_id = d.get("bank_account") or None
    instance.print_bank_details = bool(d.get("print_bank_details"))
    instance.other_charges_amount = _si_num(d.get("other_charges_amount"))
    instance.round_off = _si_num(d.get("round_off"))


def _save_invoice_items(instance, request):
    try:
        rows = json.loads(request.POST.get("items_json") or "[]")
    except json.JSONDecodeError:
        rows = []
    instance.items.all().delete()
    for r in rows:
        if not r.get("item"):
            continue
        qty = _si_num(r.get("quantity"))
        free = _si_num(r.get("free_qty"))
        rate = _si_num(r.get("rate"))
        disc = _si_num(r.get("discount_percent"))
        gross = qty * rate
        taxable = gross - (gross * disc / Decimal("100"))
        gst_pct = _si_num(r.get("gst_percent"))
        gst_amt = taxable * gst_pct / Decimal("100")
        SalesInvoiceItem.objects.create(
            invoice=instance, item_id=r["item"], batch_no=r.get("batch_no") or "",
            hsn_sac=r.get("hsn_sac") or "", uom=r.get("uom") or "",
            quantity=qty, free_qty=free, rate=rate, discount_percent=disc,
            taxable_amount=taxable.quantize(Decimal("0.01")),
            gst_percent=gst_pct, gst_amount=gst_amt.quantize(Decimal("0.01")),
            amount=(taxable + gst_amt).quantize(Decimal("0.01")),
        )
    instance.net_amount = instance.compute_net_amount()
    # "remarks" is included so an auto-generated description picks up the total
    # that only became known once the line items were saved.
    instance.save(update_fields=["net_amount", "remarks"])


@login_required(login_url="login")
def sales_invoice_list(request):
    return render(request, "sales_invoice_list.html")


@login_required(login_url="login")
def sales_invoice_api_list(request):
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    qs = SalesInvoice.objects.select_related("customer")
    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    return JsonResponse([_sales_invoice_list_dict(i) for i in qs.order_by("-date", "-id")], safe=False)


@login_required(login_url="login")
def create_sales_invoice(request):
    if request.method == "POST":
        instance = SalesInvoice(created_by=request.user if request.user.is_authenticated else None)
        try:
            _apply_posted_invoice(instance, request)
            if not instance.customer_id:
                raise ValidationError("Select a customer.")
            instance.full_clean(exclude=["invoice_no"])
            with transaction.atomic():
                instance.save()
                _save_invoice_items(instance, request)
                if not instance.items.exists():
                    raise ValidationError("Add at least one item.")
            messages.success(request, "Sales Invoice created successfully.")
            return redirect("sales_invoice_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))
    return render(request, "sales_invoice_form.html", _sales_invoice_form_context())


@login_required(login_url="login")
def edit_sales_invoice(request, id):
    instance = get_object_or_404(SalesInvoice, id=id)
    if request.method == "POST":
        try:
            _apply_posted_invoice(instance, request)
            if not instance.customer_id:
                raise ValidationError("Select a customer.")
            instance.full_clean(exclude=["invoice_no"])
            with transaction.atomic():
                instance.save()
                _save_invoice_items(instance, request)
                if not instance.items.exists():
                    raise ValidationError("Add at least one item.")
            messages.success(request, "Sales Invoice updated successfully.")
            return redirect("sales_invoice_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))
    return render(request, "sales_invoice_form.html", _sales_invoice_form_context(instance))


@login_required(login_url="login")
@require_POST
def delete_sales_invoice(request, id):
    get_object_or_404(SalesInvoice, id=id).delete()
    messages.success(request, "Sales Invoice deleted successfully.")
    return redirect("sales_invoice_list")


# ---------------------------------------------------------------------------
# Sales > Reports > Customer Ledger (Customer History Report)
# ---------------------------------------------------------------------------
# A customer's running account: sales (Bird Sales + Sales Invoices) raise what
# the customer owes us (Dr), receipts lower it (Cr), carried forward from an
# opening/previous balance. Mirrors purchase.views.supplier_ledger_report but
# from the receivables side. Bird-sale rows carry poultry columns (Birds /
# Avg.Weight); Sales-Invoice rows carry one detail row per item.

def _cl_overdue_days(as_of, due_date):
    """Days a debit is past due as of the report end (0 when not overdue)."""
    if not as_of or not due_date:
        return ""
    diff = (as_of - due_date).days
    return diff if diff > 0 else 0


@login_required(login_url="login")
def customer_ledger_report(request):
    from decimal import Decimal
    from collections import OrderedDict
    from datetime import date as _date, timedelta
    from django.utils.dateparse import parse_date
    from account.models import CompanyProfile
    from broiler.models import BirdSale, BirdSaleReceipt
    from hatchery.models import ChickSale, ChickSaleReceipt
    from sales.models import SalesReceipt

    q2 = Decimal("0.01")
    customer_id = (request.GET.get("customer") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    export = (request.GET.get("export") or "").strip().lower()
    fd = parse_date(from_date) if from_date else None
    td = parse_date(to_date) if to_date else None
    as_of = td or _date.today()

    customer = Customer.objects.filter(id=customer_id).first() if customer_id.isdigit() else None
    groups = []
    totals = {"debit": Decimal("0"), "credit": Decimal("0")}
    prev_balance = Decimal("0")
    running = Decimal("0")
    sale_count = receipt_count = 0
    sales_total = receipts_total = Decimal("0")

    if customer:
        # Opening balance — a receivable (customer owes us) is Dr.
        opening = _si_num(customer.opening_balance)
        if str(customer.to_pay_to_receive or "").lower().startswith("pay"):
            opening = -opening  # we owe the customer (advance) → Cr

        credit_period = int(customer.credit_period or 0)

        bird_sales = list(BirdSale.objects.filter(sale_type="customer", customer=customer)
                          .select_related("farm").order_by("date", "id"))
        invoices = list(SalesInvoice.objects.filter(customer=customer, is_active=True)
                        .select_related("branch").prefetch_related("items__item").order_by("date", "id"))
        receipts = list(BirdSaleReceipt.objects.filter(sale_type="customer", customer=customer)
                        .select_related("receipt_account").order_by("date", "id"))
        chick_sales = list(ChickSale.objects.filter(customer=customer)
                           .select_related("warehouse").prefetch_related("items__item").order_by("date", "id"))
        chick_receipts = list(ChickSaleReceipt.objects.filter(customer=customer)
                              .select_related("receipt_account").order_by("date", "id"))
        sales_receipts = list(SalesReceipt.objects.filter(customer=customer)
                              .select_related("receipt_account").order_by("date", "id"))
        # A debit note raises what the customer owes, a credit note reduces it.
        debit_notes = list(CustomerDebitNote.objects.filter(customer=customer)
                           .select_related("account", "sector").order_by("date", "id"))
        credit_notes = list(CustomerCreditNote.objects.filter(customer=customer)
                            .select_related("account", "sector").order_by("date", "id"))

        # previous balance = opening + sales - receipts strictly before window
        prev_balance = opening
        for bs in bird_sales:
            if fd and bs.date and bs.date < fd:
                prev_balance += _si_num(bs.amount)
        for inv in invoices:
            if fd and inv.date and inv.date < fd:
                prev_balance += _si_num(inv.net_amount)
        for cs in chick_sales:
            if fd and cs.date and cs.date < fd:
                prev_balance += _si_num(cs.final_amount)
        for rc in receipts:
            if fd and rc.date and rc.date < fd:
                prev_balance -= _si_num(rc.amount)
        for crc in chick_receipts:
            if fd and crc.date and crc.date < fd:
                prev_balance -= _si_num(crc.amount)
        for sr in sales_receipts:
            if fd and sr.date and sr.date < fd:
                prev_balance -= _si_num(sr.amount)
        for dn in debit_notes:
            if fd and dn.date and dn.date < fd:
                prev_balance += _si_num(dn.amount)
        for cn in credit_notes:
            if fd and cn.date and cn.date < fd:
                prev_balance -= _si_num(cn.amount)

        # in-window events, ordered by (date, kind)
        events = []
        for bs in bird_sales:
            if bs.date and ((fd and bs.date < fd) or (td and bs.date > td)):
                continue
            events.append((bs.date, 0, "BS", bs))
        for inv in invoices:
            if inv.date and ((fd and inv.date < fd) or (td and inv.date > td)):
                continue
            events.append((inv.date, 1, "INV", inv))
        for cs in chick_sales:
            if cs.date and ((fd and cs.date < fd) or (td and cs.date > td)):
                continue
            events.append((cs.date, 1, "CS", cs))
        for rc in receipts:
            if rc.date and ((fd and rc.date < fd) or (td and rc.date > td)):
                continue
            events.append((rc.date, 2, "RC", rc))
        for crc in chick_receipts:
            if crc.date and ((fd and crc.date < fd) or (td and crc.date > td)):
                continue
            events.append((crc.date, 2, "CRC", crc))
        for sr in sales_receipts:
            if sr.date and ((fd and sr.date < fd) or (td and sr.date > td)):
                continue
            events.append((sr.date, 2, "SR", sr))
        # Notes sort with the sales (order 1): they adjust a sale, so they read
        # naturally next to it rather than after the day's receipts.
        for dn in debit_notes:
            if dn.date and ((fd and dn.date < fd) or (td and dn.date > td)):
                continue
            events.append((dn.date, 1, "CDN", dn))
        for cn in credit_notes:
            if cn.date and ((fd and cn.date < fd) or (td and cn.date > td)):
                continue
            events.append((cn.date, 1, "CCN", cn))
        events.sort(key=lambda e: (e[0] or parse_date("1900-01-01"), e[1]))

        month_groups = OrderedDict()

        def _grp(d):
            key = d.strftime("%B %Y") if d else "Undated"
            if key not in month_groups:
                month_groups[key] = {"label": key, "rows": [], "debit": Decimal("0"),
                                     "credit": Decimal("0"), "closing": Decimal("0"), "count": 0}
            return month_groups[key]

        running = prev_balance
        for d, _o, kind, obj in events:
            grp = _grp(d)
            grp["count"] += 1
            if kind == "BS":
                sale_count += 1
                amt = _si_num(obj.amount)
                running += amt
                grp["debit"] += amt
                totals["debit"] += amt
                sales_total += amt
                grp["rows"].append({
                    "date": d, "trnum": obj.sale_no, "doc_no": obj.doc_no or "",
                    "type": "Bird Sale", "type_slug": "bird-sale",
                    "item": "Broiler Birds",
                    "birds": obj.birds or "", "quantity": _si_num(obj.net_weight),
                    "avg_weight": _si_num(obj.avg_weight), "free": "", "rate": _si_num(obj.rate),
                    "amount": amt.quantize(q2), "debit": amt.quantize(q2), "credit": "",
                    "balance": abs(running).quantize(q2), "cr_dr": "Dr" if running >= 0 else "Cr",
                    "sector": obj.farm.farm_name if obj.farm_id else "",
                    "vehicle": obj.vehicle or "", "remarks": obj.remarks or "",
                    "overdue": _cl_overdue_days(as_of, d + timedelta(days=credit_period)) if d else "",
                })
            elif kind == "INV":
                sale_count += 1
                amt = _si_num(obj.net_amount)
                running += amt
                grp["debit"] += amt
                totals["debit"] += amt
                sales_total += amt
                due = obj.due_date or (d + timedelta(days=credit_period) if d else None)
                overdue = _cl_overdue_days(as_of, due)
                items = list(obj.items.all()) or [None]
                for i, it in enumerate(items):
                    first = i == 0
                    grp["rows"].append({
                        "date": d if first else None,
                        "trnum": obj.invoice_no if first else "",
                        "doc_no": (obj.reference_no or "") if first else "",
                        "type": "Sales Invoice" if first else "",
                        "type_slug": "sales-invoice" if first else "",
                        "item": it.item.description if it and it.item_id else "",
                        "birds": "", "quantity": _si_num(it.quantity) if it else "",
                        "avg_weight": "", "free": _si_num(it.free_qty) if it else "",
                        "rate": _si_num(it.rate) if it else "",
                        "amount": _si_num(it.amount) if it else "",
                        "debit": amt.quantize(q2) if first else "", "credit": "",
                        "balance": abs(running).quantize(q2) if first else "",
                        "cr_dr": ("Dr" if running >= 0 else "Cr") if first else "",
                        "sector": obj.branch.name if obj.branch_id else "",
                        "vehicle": (obj.vehicle_no or "") if first else "",
                        "remarks": (obj.remarks or "") if first else "",
                        "overdue": overdue if first else "",
                    })
            elif kind == "CS":  # hatchery chick sale — debit
                sale_count += 1
                amt = _si_num(obj.final_amount)
                running += amt
                grp["debit"] += amt
                totals["debit"] += amt
                sales_total += amt
                overdue = _cl_overdue_days(as_of, d + timedelta(days=credit_period)) if d else ""
                items = list(obj.items.all()) or [None]
                for i, it in enumerate(items):
                    first = i == 0
                    grp["rows"].append({
                        "date": d if first else None,
                        "trnum": obj.bill_no if first else "",
                        "doc_no": "",
                        "type": "Chick Sale" if first else "",
                        "type_slug": "chick-sale" if first else "",
                        "item": it.item.description if it and it.item_id else "Chicks",
                        "birds": "", "quantity": _si_num(it.net_qty) if it else "",
                        "avg_weight": "", "free": _si_num(it.free_qty) if it else "",
                        "rate": _si_num(it.sale_rate) if it else "",
                        "amount": _si_num(it.amount) if it else "",
                        "debit": amt.quantize(q2) if first else "", "credit": "",
                        "balance": abs(running).quantize(q2) if first else "",
                        "cr_dr": ("Dr" if running >= 0 else "Cr") if first else "",
                        "sector": obj.warehouse.name if obj.warehouse_id else "",
                        "vehicle": (obj.vehicle or "") if first else "",
                        "remarks": (obj.remarks or "") if first else "",
                        "overdue": overdue if first else "",
                    })
            elif kind in ("CDN", "CCN"):
                # Customer debit/credit note. Counted in neither the sale nor
                # the receipt tally - it adjusts a sale rather than being one.
                amt = _si_num(obj.amount)
                is_debit = kind == "CDN"
                if is_debit:
                    running += amt
                    grp["debit"] += amt
                    totals["debit"] += amt
                else:
                    running -= amt
                    grp["credit"] += amt
                    totals["credit"] += amt
                grp["rows"].append({
                    "date": d, "trnum": obj.note_no, "doc_no": obj.against_bill or "",
                    "type": "Debit Note" if is_debit else "Credit Note",
                    "type_slug": "note",
                    "item": obj.account.description if obj.account_id else "",
                    "birds": "", "quantity": "", "avg_weight": "", "free": "", "rate": "",
                    "amount": amt.quantize(q2),
                    "debit": amt.quantize(q2) if is_debit else "",
                    "credit": "" if is_debit else amt.quantize(q2),
                    "balance": abs(running).quantize(q2),
                    "cr_dr": "Dr" if running >= 0 else "Cr",
                    # Sector = the office/branch, matching the column's meaning
                    # elsewhere in the ledger and on the Journal screen.
                    "sector": obj.sector.name if obj.sector_id else "",
                    "vehicle": "", "remarks": obj.remarks or "", "overdue": "",
                })
            else:  # RC / CRC / SR — receipt (bird / chick / sales)
                receipt_count += 1
                amt = _si_num(obj.amount)
                running -= amt
                grp["credit"] += amt
                totals["credit"] += amt
                receipts_total += amt
                acct = obj.receipt_account.description if obj.receipt_account_id else ""
                type_label = {"CRC": "Chick Receipt", "SR": "Sales Receipt"}.get(kind, "Bird Receipt")
                grp["rows"].append({
                    "date": d, "trnum": obj.receipt_no, "doc_no": obj.reference_no or "",
                    "type": type_label, "type_slug": "receipt",
                    "item": obj.get_mode_display(),
                    "birds": "", "quantity": "", "avg_weight": "", "free": "", "rate": "", "amount": "",
                    "debit": "", "credit": amt.quantize(q2),
                    "balance": abs(running).quantize(q2), "cr_dr": "Dr" if running >= 0 else "Cr",
                    # Sector shows the cash/bank account the receipt was received into.
                    "sector": acct, "vehicle": "", "remarks": obj.remarks or "", "overdue": "",
                })
            grp["closing"] = running

        for g in month_groups.values():
            groups.append({
                "label": g["label"], "count": g["count"],
                "debit": g["debit"].quantize(q2), "credit": g["credit"].quantize(q2),
                "closing": abs(g["closing"]).quantize(q2),
                "cr_dr": "Dr" if g["closing"] >= 0 else "Cr",
                "rows": g["rows"],
            })

    kpi = {
        "opening": abs(prev_balance).quantize(q2), "opening_cr_dr": "Dr" if prev_balance >= 0 else "Cr",
        "sales": sales_total.quantize(q2), "sale_count": sale_count,
        "receipts": receipts_total.quantize(q2), "receipt_count": receipt_count,
        "closing": abs(running).quantize(q2), "closing_cr_dr": "Dr" if running >= 0 else "Cr",
    }
    company = CompanyProfile.get_solo()
    ctx = {
        "customers": Customer.objects.order_by("name"),
        "customer": customer, "customer_id": customer_id,
        "from_date": from_date, "to_date": to_date,
        "groups": groups,
        "totals": {"debit": totals["debit"].quantize(q2), "credit": totals["credit"].quantize(q2)},
        "prev_balance": abs(prev_balance).quantize(q2), "prev_cr_dr": "Dr" if prev_balance >= 0 else "Cr",
        "closing": abs(running).quantize(q2), "closing_cr_dr": "Dr" if running >= 0 else "Cr",
        "kpi": kpi, "company": company,
    }
    if export == "excel" and customer:
        return _customer_ledger_excel(company, customer, from_date, to_date, ctx["prev_balance"],
                                      ctx["prev_cr_dr"], groups, ctx["totals"], ctx["closing"], ctx["closing_cr_dr"])
    return render(request, "customer_ledger_report.html", ctx)


def _customer_balance_row(cust, fd, td, ref_date):
    """One Customer Balance row: opening (signed receivable before the window),
    period Birds/Weight/Amount (Bird Sales + Sales Invoices) and Receipt (Bird
    Sale Receipts), the between-days movement, the closing split into Debit
    (customer owes) / Credit (advance), credit-limit position, and days since
    the last receipt."""
    from decimal import Decimal
    from broiler.models import BirdSale, BirdSaleReceipt
    from hatchery.models import ChickSale, ChickSaleReceipt
    from sales.models import SalesReceipt

    q2 = Decimal("0.01")
    signed = _si_num(cust.opening_balance)
    if str(cust.to_pay_to_receive or "").lower().startswith("pay"):
        signed = -signed  # we owe the customer (advance) → Cr

    bird_sales = list(BirdSale.objects.filter(sale_type="customer", customer=cust))
    invoices = list(SalesInvoice.objects.filter(customer=cust, is_active=True))
    receipts = list(BirdSaleReceipt.objects.filter(sale_type="customer", customer=cust))
    chick_sales = list(ChickSale.objects.filter(customer=cust).prefetch_related("items"))
    chick_receipts = list(ChickSaleReceipt.objects.filter(customer=cust))
    sales_receipts = list(SalesReceipt.objects.filter(customer=cust))
    debit_notes = list(CustomerDebitNote.objects.filter(customer=cust))
    credit_notes = list(CustomerCreditNote.objects.filter(customer=cust))

    def before(d):
        return d and fd and d < fd

    def within(d):
        return d and (not fd or d >= fd) and (not td or d <= td)

    opening = signed
    amount = Decimal("0")   # receivable-raising in period (sales)
    receipt = Decimal("0")  # receivable-lowering in period (receipts)
    birds = 0
    weight = Decimal("0")
    chicks = Decimal("0")
    for bs in bird_sales:
        v = _si_num(bs.amount)
        opening += v if before(bs.date) else 0
        if within(bs.date):
            amount += v
            birds += int(bs.birds or 0)
            weight += _si_num(bs.net_weight)
    for inv in invoices:
        v = _si_num(inv.net_amount)
        opening += v if before(inv.date) else 0
        amount += v if within(inv.date) else 0
    for cs in chick_sales:
        v = _si_num(cs.final_amount)
        opening += v if before(cs.date) else 0
        if within(cs.date):
            amount += v
            chicks += _si_num(cs.total_net_qty())
    for rc in receipts:
        v = _si_num(rc.amount)
        opening -= v if before(rc.date) else 0
        receipt += v if within(rc.date) else 0
    for crc in chick_receipts:
        v = _si_num(crc.amount)
        opening -= v if before(crc.date) else 0
        receipt += v if within(crc.date) else 0
    for sr in sales_receipts:
        v = _si_num(sr.amount)
        opening -= v if before(sr.date) else 0
        receipt += v if within(sr.date) else 0
    # A debit note behaves like a sale (raises the receivable), a credit note
    # like a receipt (reduces it), so they join those two period buckets.
    for dn in debit_notes:
        v = _si_num(dn.amount)
        opening += v if before(dn.date) else 0
        amount += v if within(dn.date) else 0
    for cn in credit_notes:
        v = _si_num(cn.amount)
        opening -= v if before(cn.date) else 0
        receipt += v if within(cn.date) else 0

    bw = amount - receipt
    closing = opening + bw

    credit_limit = _si_num(cust.credit_limit)
    limit_exceeded = closing - credit_limit
    available = credit_limit - closing

    rcpt_dates = [rc.date for rc in receipts if rc.date and (not td or rc.date <= td)] \
        + [crc.date for crc in chick_receipts if crc.date and (not td or crc.date <= td)] \
        + [sr.date for sr in sales_receipts if sr.date and (not td or sr.date <= td)]
    if rcpt_dates:
        gap = (ref_date - max(rcpt_dates)).days
    else:
        txn_dates = ([bs.date for bs in bird_sales if bs.date] + [inv.date for inv in invoices if inv.date]
                     + [cs.date for cs in chick_sales if cs.date])
        base = min(txn_dates) if txn_dates else cust.as_on_date
        gap = (ref_date - base).days if base else 0

    return {
        "id": cust.id, "code": cust.code or "", "name": cust.name,
        "group": (cust.customer_group.description or cust.customer_group.code
                  if cust.customer_group_id else "") or "Sundry Debtors",
        "opening": opening.quantize(q2),
        "birds": birds, "weight": weight.quantize(q2), "chicks": chicks.quantize(q2),
        "amount": amount.quantize(q2), "receipt": receipt.quantize(q2), "bw": bw.quantize(q2),
        "debit": closing.quantize(q2) if closing >= 0 else Decimal("0.00"),
        "credit": (-closing).quantize(q2) if closing < 0 else Decimal("0.00"),
        "credit_limit": credit_limit.quantize(q2),
        "limit_exceeded": limit_exceeded.quantize(q2) if limit_exceeded > 0 else Decimal("0.00"),
        "available": available.quantize(q2) if available > 0 else Decimal("0.00"),
        "gap": max(gap, 0),
        "has_activity": bool(opening or amount or receipt or closing),
    }


@login_required(login_url="login")
def customer_balance_report(request):
    """Sales > Reports > Customer Balance — every customer's receivable position:
    opening, this period's Birds/Weight/Amount/Receipt movement, closing Debit
    (customer owes) / Credit (advance), credit-limit position and last-receipt gap."""
    from decimal import Decimal
    from django.utils.dateparse import parse_date
    from account.models import CompanyProfile

    q2 = Decimal("0.01")
    group = (request.GET.get("customer_group") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    fd = parse_date(from_date) if from_date else None
    td = parse_date(to_date) if to_date else None
    ref_date = td or timezone.localdate()

    customers = Customer.objects.select_related("customer_group").order_by("name")
    if group.isdigit():
        customers = customers.filter(customer_group_id=group)

    rows = [_customer_balance_row(c, fd, td, ref_date) for c in customers]

    tkeys = ["opening", "amount", "receipt", "bw", "debit", "credit",
             "credit_limit", "limit_exceeded", "available"]
    totals = {k: sum((r[k] for r in rows), Decimal("0")).quantize(q2) for k in tkeys}
    totals["birds"] = sum((r["birds"] for r in rows), 0)
    totals["weight"] = sum((r["weight"] for r in rows), Decimal("0")).quantize(q2)
    totals["chicks"] = sum((r["chicks"] for r in rows), Decimal("0")).quantize(q2)

    groups = [{"id": g.id, "name": g.description or g.code or f"Group {g.id}"}
              for g in CustomerGroup.objects.order_by("description")]

    return render(request, "customer_balance_report.html", {
        "rows": rows, "totals": totals,
        "customer_groups": groups, "group": group,
        "from_date": from_date, "to_date": to_date,
        "company": CompanyProfile.get_solo(),
    })


def _customer_ledger_excel(company, customer, from_date, to_date, prev_balance, prev_cr_dr,
                           groups, totals, closing, closing_cr_dr):
    """Stream the Customer Ledger as an .xlsx workbook (openpyxl)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"
    head_fill = PatternFill("solid", fgColor="1B3A6B")
    head_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    headers = ["Date", "Transaction No.", "Doc No.", "Type", "Item", "Birds", "Quantity",
               "Avg.Weight", "Free", "Rate", "Amount", "Debit", "Credit", "Balance", "Sector",
               "Vehicle", "Remarks", "Over Due By Days"]

    ws.append([company.name if company else ""])
    ws.append([f"Customer History Report — {customer.name}"])
    ws.append([f"Period: {from_date or '—'} to {to_date or '—'}"])
    ws.append([])
    ws.append(headers)
    for c in ws[ws.max_row]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center")

    prev = ["", "", "", "", "", "", "", "", "", "", "", "", "", f"{prev_balance} {prev_cr_dr}",
            "Previous Balance", "", "", ""]
    ws.append(prev)
    for g in groups:
        ws.append([g["label"], f"{g['count']} transaction(s)", "", "", "", "", "", "", "", "", "",
                   float(g["debit"]), float(g["credit"]), f"{g['closing']} {g['cr_dr']}", "", "", "", ""])
        for c in ws[ws.max_row]:
            c.font = bold
        for r in g["rows"]:
            ws.append([
                r["date"].strftime("%d.%m.%Y") if r["date"] else "", r["trnum"], r["doc_no"],
                r["type"], r["item"], r["birds"], r["quantity"], r["avg_weight"], r["free"], r["rate"],
                r["amount"], r["debit"], r["credit"],
                f"{r['balance']} {r['cr_dr']}" if r["balance"] != "" else "",
                r["sector"], r["vehicle"], r["remarks"], r["overdue"],
            ])
    ws.append(["Grand Total", "", "", "", "", "", "", "", "", "", "", float(totals["debit"]),
               float(totals["credit"]), f"{closing} {closing_cr_dr}", "", "", "", ""])
    for c in ws[ws.max_row]:
        c.font = bold

    widths = [11, 18, 14, 14, 22, 8, 11, 10, 9, 9, 14, 14, 14, 16, 18, 14, 20, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="customer_ledger_{customer.code or customer.id}.xlsx"'
    wb.save(resp)
    return resp


# ---------------------------------------------------------------------------
# Sales Receipt (Sales > Transactions) — customer payments against sales,
# mirroring the broiler Bird Receipt / hatchery Chick Receipt (customer-only).
# ---------------------------------------------------------------------------
from sales.models import SalesReceipt


def _sales_receipt_to_dict(row):
    return {
        "id": row.id, "receipt_no": row.receipt_no, "date": row.date.isoformat(),
        "location": row.location_id, "location_name": row.location.name if row.location_id else "",
        "customer": row.customer_id, "customer_name": row.customer.name if row.customer_id else "",
        "mode": row.mode,
        "receipt_account": row.receipt_account_id,
        "receipt_account_name": (f"{row.receipt_account.code} - {row.receipt_account.description}"
                                 if row.receipt_account_id else ""),
        "amount": str(row.amount), "reference_no": row.reference_no, "remarks": row.remarks,
    }


def _apply_sales_receipt(instance, data):
    if data.get("date"):
        instance.date = timezone.datetime.fromisoformat(data["date"]).date()
    instance.location_id = data.get("location") or None
    instance.customer_id = data.get("customer") or None
    instance.mode = data.get("mode") or "Cash"
    instance.receipt_account_id = data.get("receipt_account") or None
    instance.amount = _si_num(data.get("amount"))
    instance.reference_no = data.get("reference_no") or ""
    instance.remarks = data.get("remarks") or ""


@login_required(login_url="login")
def sales_receipt_list(request):
    return render(request, "sales_receipt_list.html")


@login_required(login_url="login")
def sales_receipt_form(request, id=None):
    from account.services.bank_cash import bank_cash_accounts, active_payment_modes, payment_mode_map
    import json as _json
    return render(request, "sales_receipt_form.html", {
        "instance": SalesReceipt.objects.filter(id=id).first() if id else None,
        "locations": Warehouse.objects.order_by("name"),
        "customers": Customer.objects.order_by("name"),
        "accounts": bank_cash_accounts(),   # receipt into a Bank/Cash master account
        "payment_modes": active_payment_modes("receipt"),
        "payment_mode_map_json": _json.dumps(payment_mode_map("receipt")),
        "today": timezone.localdate().isoformat(),
    })


@method_decorator(login_required, name="dispatch")
class SalesReceiptAPI(View):
    def get(self, request, id=None):
        try:
            if id:
                row = SalesReceipt.objects.select_related("customer", "location", "receipt_account").get(id=id)
                return JsonResponse(_sales_receipt_to_dict(row))
            qs = SalesReceipt.objects.select_related("customer", "location", "receipt_account")
            from_date = (request.GET.get("from_date") or "").strip()
            to_date = (request.GET.get("to_date") or "").strip()
            if from_date:
                qs = qs.filter(date__gte=from_date)
            if to_date:
                qs = qs.filter(date__lte=to_date)
            return JsonResponse([_sales_receipt_to_dict(r) for r in qs.order_by("-date", "-id")], safe=False)
        except SalesReceipt.DoesNotExist:
            raise Http404("Receipt not found")
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def post(self, request):
        try:
            data = json.loads(request.body or "{}")
            rows = data.get("rows") or []
            created = []
            with transaction.atomic():
                for row in rows:
                    if not row.get("customer"):
                        continue
                    instance = SalesReceipt(created_by=request.user if request.user.is_authenticated else None)
                    _apply_sales_receipt(instance, row)
                    instance.full_clean(exclude=["receipt_no"])
                    instance.save()
                    created.append(instance.id)
            if not created:
                return JsonResponse({"error": "Add at least one receipt row with a Customer selected"}, status=400)
            return JsonResponse({"message": "Receipt(s) created", "ids": created}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def put(self, request, id):
        try:
            instance = SalesReceipt.objects.get(id=id)
            data = json.loads(request.body or "{}")
            _apply_sales_receipt(instance, data)
            instance.full_clean(exclude=["receipt_no"])
            instance.save()
            return JsonResponse({"message": "Receipt updated"})
        except SalesReceipt.DoesNotExist:
            raise Http404("Receipt not found")
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def delete(self, request, id):
        try:
            SalesReceipt.objects.get(id=id).delete()
            return JsonResponse({"message": "Receipt deleted"})
        except SalesReceipt.DoesNotExist:
            raise Http404("Receipt not found")
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)


def _customer_current_balance(customer_id, exclude_sales_receipt_id=None,
                              exclude_bird_receipt_id=None, exclude_chick_receipt_id=None):
    """A customer's total outstanding balance across every module — the same
    figure the Customer Ledger shows as its closing (opening + all Bird/Chick/
    Invoice sales − all Bird/Chick/Sales receipts). Positive = customer owes us.
    The per-module exclude_* ids drop the receipt currently being edited so its
    own amount isn't double-counted in the balance shown on its edit form."""
    from broiler.models import BirdSale, BirdSaleReceipt
    from hatchery.models import ChickSale, ChickSaleReceipt
    from django.db.models import Sum

    cust = Customer.objects.filter(id=customer_id).first() if customer_id else None
    if not cust:
        return Decimal("0")
    bal = _si_num(cust.opening_balance)
    if str(cust.to_pay_to_receive or "").lower().startswith("pay"):
        bal = -bal
    def _sum(qs, field):
        return _si_num(qs.aggregate(t=Sum(field))["t"])
    bal += _sum(BirdSale.objects.filter(sale_type="customer", customer=cust), "amount")
    bal += _sum(SalesInvoice.objects.filter(customer=cust, is_active=True), "net_amount")
    bal += _sum(ChickSale.objects.filter(customer=cust), "final_amount")

    bird_rc = BirdSaleReceipt.objects.filter(sale_type="customer", customer=cust)
    if exclude_bird_receipt_id:
        bird_rc = bird_rc.exclude(id=exclude_bird_receipt_id)
    bal -= _sum(bird_rc, "amount")
    chick_rc = ChickSaleReceipt.objects.filter(customer=cust)
    if exclude_chick_receipt_id:
        chick_rc = chick_rc.exclude(id=exclude_chick_receipt_id)
    bal -= _sum(chick_rc, "amount")
    sr = SalesReceipt.objects.filter(customer=cust)
    if exclude_sales_receipt_id:
        sr = sr.exclude(id=exclude_sales_receipt_id)
    bal -= _sum(sr, "amount")
    return bal


@login_required(login_url="login")
def sales_receipt_balance_lookup(request):
    """Customer's full outstanding balance (across all modules), matching the
    Customer Ledger closing — not just the sales-invoice portion."""
    balance = _customer_current_balance(request.GET.get("customer"),
                                        exclude_sales_receipt_id=request.GET.get("exclude_id"))
    return JsonResponse({"balance": str(balance)})

@login_required
def customer_ledger_balance(request):
    """Outstanding ledger balance of one customer, for forms that raise a
    document against them (Bird Sale, Delivery Challan).

    Reuses the Customer Balance report's own row builder rather than adding a
    second calculation, so the figure shown while raising a document cannot
    drift from the report the same customer is judged by.
    """
    customer_id = (request.GET.get("customer") or "").strip()
    customer = (Customer.objects.select_related("customer_group")
                .filter(id=customer_id).first()) if customer_id.isdigit() else None
    if not customer:
        return JsonResponse({"balance": "", "label": "", "available": ""})

    # No window: opening plus every movement to date is the current balance.
    row = _customer_balance_row(customer, None, None, timezone.localdate())
    if row["debit"] > 0:
        label = "%s Dr" % row["debit"]          # customer owes us
    elif row["credit"] > 0:
        label = "%s Cr" % row["credit"]         # customer is in advance
    else:
        label = "0.00"
    return JsonResponse({
        "balance": str(row["debit"] - row["credit"]),
        "label": label,
        "credit_limit": str(row["credit_limit"]),
        "available": str(row["available"]),
        "limit_exceeded": str(row["limit_exceeded"]),
    })


# ---------------------------------------------------- customer credit/debit notes
#
# Deliberately parallel to purchase's supplier Debit/Credit Notes: same fields,
# same helper shape, same auto-number series style. A Debit Note raises the
# receivable, a Credit Note reduces it; both feed the Customer Ledger and
# Customer Balance.

def _customer_note_list_dict(n):
    return {
        "id": n.id, "date": n.date.isoformat(), "note_no": n.note_no,
        "customer_name": n.customer.name if n.customer_id else "",
        "against_bill": n.against_bill,
        "account_name": n.account.description if n.account_id else "",
        "sector_name": n.sector.name if n.sector_id else "",
        "amount": str(n.amount), "remarks": n.remarks,
    }


def _customer_note_row_dict(n):
    """One row for the entry grid, so editing reopens exactly what was saved."""
    return {
        "id": n.id, "date": n.date.isoformat() if n.date else "",
        "note_no": n.note_no, "customer": n.customer_id or "",
        "against_bill": n.against_bill or "", "account": n.account_id or "",
        "amount": str(n.amount or 0), "sector": n.sector_id or "",
        "remarks": n.remarks or "",
    }


def _customer_note_form_context(model, instance=None):
    from account.models import ChartOfAccount
    from inventory.models import Warehouse
    return {
        "note": instance,
        "next_no": model._next_no() if not instance else None,
        "note_kind": ("Customer Debit Note" if model is CustomerDebitNote
                      else "Customer Credit Note"),
        "customers": Customer.objects.order_by("name"),
        "accounts": ChartOfAccount.objects.order_by("code"),
        "sectors": Warehouse.objects.order_by("name"),
        "today": timezone.localdate().isoformat(),
        "existing_rows_json": json.dumps(
            [_customer_note_row_dict(instance)] if instance else []),
    }


def _apply_customer_note_row(instance, row):
    """Copy one posted grid row onto a note, validating as we go."""
    instance.date = row.get("date") or timezone.localdate()
    instance.customer_id = row.get("customer") or None
    instance.against_bill = (row.get("against_bill") or "").strip()
    instance.account_id = row.get("account") or None
    instance.sector_id = row.get("sector") or None
    instance.remarks = (row.get("remarks") or "").strip()
    try:
        instance.amount = Decimal(str(row.get("amount") or 0))
    except (InvalidOperation, TypeError, ValueError):
        raise ValidationError("Enter a valid amount.")
    if not instance.customer_id:
        raise ValidationError("Select a customer on every row.")
    if instance.amount <= 0:
        raise ValidationError("Enter an amount greater than zero on every row.")


def _save_customer_notes(request, model, kind, template, redirect_name, instance=None):
    """Grid entry: one note per row, all saved together or not at all.

    The Add screen takes as many rows as wanted; the Edit screen reopens the one
    note being edited as a single row.
    """
    if request.method == "POST":
        try:
            rows = json.loads(request.POST.get("rows_json") or "[]")
        except json.JSONDecodeError:
            rows = []
        rows = [r for r in rows if any(str(v).strip() for v in r.values())]
        try:
            if not rows:
                raise ValidationError("Add at least one row.")
            with transaction.atomic():
                saved = 0
                for row in rows:
                    note = (get_object_or_404(model, id=row["id"])
                            if row.get("id") else model())
                    _apply_customer_note_row(note, row)
                    note.full_clean(exclude=["note_no"])
                    note.save()
                    saved += 1
            messages.success(
                request,
                "%s %s successfully." % (
                    kind if saved == 1 else "%d %ss" % (saved, kind),
                    "updated" if instance else "added"))
            return redirect(redirect_name)
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))
    return render(request, template, _customer_note_form_context(model, instance))


def _customer_note_api(request, model):
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    customer_id = (request.GET.get("customer") or "").strip()
    qs = model.objects.select_related("customer", "account", "sector")
    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    if customer_id.isdigit():
        qs = qs.filter(customer_id=customer_id)
    return JsonResponse([_customer_note_list_dict(n) for n in qs.order_by("-date", "-id")],
                        safe=False)


# --- Customer Debit Note ---
@login_required(login_url="login")
def customer_debit_note_list(request):
    return render(request, "customer_debit_note_list.html")


@login_required(login_url="login")
def create_customer_debit_note(request):
    return _save_customer_notes(request, CustomerDebitNote,
                                "Customer Debit Note", "customer_debit_note_form.html",
                                "customer_debit_note_list")


@login_required(login_url="login")
def edit_customer_debit_note(request, id):
    return _save_customer_notes(request, CustomerDebitNote,
                                "Customer Debit Note", "customer_debit_note_form.html", "customer_debit_note_list",
                                instance=get_object_or_404(CustomerDebitNote, id=id))


@login_required(login_url="login")
@require_POST
def delete_customer_debit_note(request, id):
    get_object_or_404(CustomerDebitNote, id=id).delete()
    messages.success(request, "Customer Debit Note deleted successfully.")
    return redirect("customer_debit_note_list")


@login_required
def customer_debit_note_api_list(request):
    return _customer_note_api(request, CustomerDebitNote)


# --- Customer Credit Note ---
@login_required(login_url="login")
def customer_credit_note_list(request):
    return render(request, "customer_credit_note_list.html")


@login_required(login_url="login")
def create_customer_credit_note(request):
    return _save_customer_notes(request, CustomerCreditNote,
                                "Customer Credit Note", "customer_credit_note_form.html",
                                "customer_credit_note_list")


@login_required(login_url="login")
def edit_customer_credit_note(request, id):
    return _save_customer_notes(request, CustomerCreditNote,
                                "Customer Credit Note", "customer_credit_note_form.html", "customer_credit_note_list",
                                instance=get_object_or_404(CustomerCreditNote, id=id))


@login_required(login_url="login")
@require_POST
def delete_customer_credit_note(request, id):
    get_object_or_404(CustomerCreditNote, id=id).delete()
    messages.success(request, "Customer Credit Note deleted successfully.")
    return redirect("customer_credit_note_list")


@login_required
def customer_credit_note_api_list(request):
    return _customer_note_api(request, CustomerCreditNote)
