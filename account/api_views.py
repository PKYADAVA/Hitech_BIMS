"""REST APIs for the chart-of-accounts engine.

All endpoints require login; write endpoints record an AccountAuditLog row
with the acting user and client IP. Company scoping comes from the
``company`` query/body parameter, defaulting to the primary company (pk=1),
so the API is multi-company ready while staying transparent for the current
single-company deployment.

Routes (see account/urls.py):
    GET  /api/chart-of-accounts/                list (paginated, filterable)
    POST /api/chart-of-accounts/                create (code auto-generated)
    GET  /api/chart-of-accounts/tree/           nested tree (lazy: ?parent=<id>)
    GET  /api/chart-of-accounts/search/?q=      flat search
    GET/PUT/DELETE /api/chart-of-accounts/<id>/ detail / update / soft delete
    GET  /api/chart-of-accounts/templates/      available COA templates
    POST /api/chart-of-accounts/generate/       run the COA generator
    POST /api/chart-of-accounts/import/         CSV/Excel import (?dry_run=1)
    GET  /api/chart-of-accounts/export/         CSV/Excel export
    GET/POST /api/chart-of-accounts/opening-balance/  read / bulk set
"""
import csv
import io
import json
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import Http404, HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views import View

import openpyxl

from account.models import (
    AccountAuditLog,
    AccountGroup,
    AccountType,
    ChartOfAccount,
    CoATemplate,
    CompanyProfile,
)
from account.services import AccountCodeGenerator, CoAGeneratorService

MAX_PAGE_SIZE = 500


# ------------------------------------------------------------------- helpers

def _client_ip(request):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    return forwarded.split(",")[0].strip() if forwarded else request.META.get("REMOTE_ADDR")


def _audit(request, action, account=None, company=None, old=None, new=None, reason=""):
    AccountAuditLog.objects.create(
        company=company or (account.company if account else None),
        account=account,
        action=action,
        old_values=old or {},
        new_values=new or {},
        reason=reason,
        ip_address=_client_ip(request),
        user=request.user if request.user.is_authenticated else None,
    )


def _company(request):
    company_id = (
        request.GET.get("company")
        or getattr(request, "parsed_body", {}).get("company")
        or 1
    )
    company = CompanyProfile.objects.filter(pk=company_id).first()
    if company is None:
        company = CompanyProfile.get_solo()
    return company


def _parse_body(request):
    try:
        request.parsed_body = json.loads(request.body or b"{}")
    except json.JSONDecodeError:
        request.parsed_body = {}
    return request.parsed_body


def _serialize(account, children_count=None):
    return {
        "id": account.id,
        "company": account.company_id,
        "parent": account.parent_id,
        "code": account.code,
        "description": account.description,
        "account_type": account.account_type.name if account.account_type_id else account.type,
        "account_group": account.account_group.name if account.account_group_id else None,
        "type": account.type,
        "control_type": account.control_type,
        "schedule": account.schedule_id,
        "schedule__name": account.schedule.name if account.schedule_id else None,
        "level": account.level,
        "path": account.path,
        "currency": account.currency,
        "opening_balance": str(account.opening_balance),
        "opening_type": account.opening_type,
        "is_group": account.is_group,
        "is_postable": account.is_postable,
        "allow_manual_entry": account.allow_manual_entry,
        "system_generated": account.system_generated,
        "system_role": account.system_role,
        "is_locked": account.is_locked,
        "status": account.status,
        "children_count": children_count,
        "created_at": account.created_at.strftime("%d-%b-%Y %H:%M") if account.created_at else None,
    }


def _snapshot(account):
    """Plain-value dict used for audit old/new comparison."""
    return {
        "code": account.code,
        "description": account.description,
        "parent": account.parent_id,
        "account_type": account.account_type_id,
        "account_group": account.account_group_id,
        "status": account.status,
        "is_group": account.is_group,
        "is_postable": account.is_postable,
        "allow_manual_entry": account.allow_manual_entry,
        "opening_balance": str(account.opening_balance),
        "opening_type": account.opening_type,
        "schedule": account.schedule_id,
        "control_type": account.control_type,
    }


def _base_queryset(company):
    return (
        ChartOfAccount.objects.filter(company=company)
        .select_related("account_type", "account_group", "schedule")
    )


def _get_account(company, id):
    account = _base_queryset(company).filter(id=id).first()
    if account is None:
        raise Http404("Chart of Account not found")
    return account


# ------------------------------------------------------------------ list/CRUD

@method_decorator(login_required, name="dispatch")
class CoAListCreateAPI(View):
    def get(self, request):
        company = _company(request)
        qs = _base_queryset(company)

        q = request.GET.get("q", "").strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(code__icontains=q) | Q(description__icontains=q))
        for param, field in [
            ("type", "type"),
            ("status", "status"),
            ("account_type", "account_type__code"),
            ("group", "account_group__name"),
            ("parent", "parent_id"),
            ("system_role", "system_role"),
        ]:
            value = request.GET.get(param)
            if value:
                qs = qs.filter(**{field: value})
        if request.GET.get("postable") in ("1", "true"):
            qs = qs.filter(is_postable=True)

        try:
            page = max(int(request.GET.get("page", 1)), 1)
            page_size = min(int(request.GET.get("page_size", 50)), MAX_PAGE_SIZE)
        except ValueError:
            return JsonResponse({"error": "Invalid pagination parameters."}, status=400)

        total = qs.count()
        offset = (page - 1) * page_size
        results = [_serialize(a) for a in qs.order_by("code")[offset:offset + page_size]]
        return JsonResponse({
            "count": total,
            "page": page,
            "page_size": page_size,
            "results": results,
        })

    def post(self, request):
        data = _parse_body(request)
        company = _company(request)
        try:
            with transaction.atomic():
                account = self._create(request, company, data)
        except (KeyError, ValueError) as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        except Exception as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        return JsonResponse(_serialize(account), status=201)

    def _create(self, request, company, data):
        parent = None
        if data.get("parent"):
            parent = _get_account(company, data["parent"])
            if not parent.is_group:
                raise ValueError("Parent must be a group account.")
        account_type = None
        if data.get("account_type"):
            account_type = AccountType.objects.filter(code=data["account_type"]).first()
            if account_type is None:
                raise ValueError(f"Unknown account type '{data['account_type']}'")
        if account_type is None and parent is not None:
            account_type = parent.account_type
        if account_type is None:
            raise ValueError("account_type is required for root accounts.")

        is_group = bool(data.get("is_group"))
        codegen = AccountCodeGenerator(company)
        account = ChartOfAccount(
            company=company,
            parent=parent,
            code=codegen.next_code(parent=parent, account_type=account_type, is_group=is_group),
            description=data["description"],
            account_type=account_type,
            account_group=AccountGroup.objects.filter(name=data.get("account_group", "")).first()
            or (parent.account_group if parent else None),
            control_type=data.get("control_type"),
            schedule_id=data.get("schedule") or None,
            is_group=is_group,
            is_postable=not is_group,
            allow_manual_entry=data.get("allow_manual_entry", True),
            status=data.get("status", "Active"),
            created_by=request.user,
        )
        account.full_clean(exclude=["type", "code"])
        account.save()
        _audit(request, "create", account=account, new=_snapshot(account))
        return account


@method_decorator(login_required, name="dispatch")
class CoADetailAPI(View):
    def get(self, request, id):
        company = _company(request)
        account = _get_account(company, id)
        return JsonResponse(_serialize(account, children_count=account.children.count()))

    def put(self, request, id):
        data = _parse_body(request)
        company = _company(request)
        account = _get_account(company, id)
        if account.is_locked:
            return JsonResponse({"error": "This account is locked."}, status=400)

        old = _snapshot(account)
        editable = ["description", "control_type", "status", "allow_manual_entry"]
        try:
            with transaction.atomic():
                for field in editable:
                    if field in data:
                        setattr(account, field, data[field])
                if "account_group" in data:
                    account.account_group = AccountGroup.objects.filter(name=data["account_group"]).first()
                if "schedule" in data:
                    account.schedule_id = data["schedule"] or None
                if "parent" in data and data["parent"] != account.parent_id:
                    if account.system_generated:
                        return JsonResponse({"error": "System accounts cannot be moved."}, status=400)
                    account.parent = _get_account(company, data["parent"]) if data["parent"] else None
                if "is_group" in data and bool(data["is_group"]) != account.is_group:
                    error = self._convert_group_flag(account, bool(data["is_group"]))
                    if error:
                        return JsonResponse({"error": error}, status=400)
                account.modified_by = request.user
                account.full_clean(exclude=["type", "code"])
                account.save()
                self._repath_descendants(account)
        except Exception as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        _audit(request, "update", account=account, old=old, new=_snapshot(account),
               reason=data.get("reason", ""))
        return JsonResponse(_serialize(account))

    @staticmethod
    def _convert_group_flag(account, make_group):
        """Convert a postable leaf to a group (or back). Returns an error string
        or None. Once a journal engine exists this must also check for posted
        transactions."""
        if make_group:
            if account.opening_balance:
                return (
                    f"{account.code} holds an opening balance of {account.opening_balance}. "
                    "Move it to a sub-account (or zero it) before converting to a group."
                )
            account.is_group = True
            account.is_postable = False
        else:
            if account.children.exists():
                return "Cannot convert a group that has sub-accounts back to a postable account."
            account.is_group = False
            account.is_postable = True
        return None

    def _repath_descendants(self, account):
        """Recompute level/path down the subtree after a move or rename."""
        for child in account.children.all():
            child.save()
            self._repath_descendants(child)

    def delete(self, request, id):
        company = _company(request)
        account = _get_account(company, id)
        if account.is_locked:
            return JsonResponse({"error": "This account is locked."}, status=400)
        if account.system_role:
            return JsonResponse({"error": "System anchor accounts cannot be deleted."}, status=400)
        try:
            old = _snapshot(account)
            account.soft_delete(user=request.user)
        except Exception as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        _audit(request, "delete", account=account, old=old)
        return JsonResponse({"message": "Account deleted (recoverable)."})


# ------------------------------------------------------------------ tree/search

@method_decorator(login_required, name="dispatch")
class CoATreeAPI(View):
    """Nested tree. With ?parent=<id> returns only that node's children
    (lazy loading); otherwise the whole tree in one response built from a
    single query."""

    def get(self, request):
        company = _company(request)
        parent_id = request.GET.get("parent")
        if parent_id:
            children = _base_queryset(company).filter(parent_id=parent_id).order_by("code")
            return JsonResponse(
                [_serialize(c, children_count=c.children.count()) for c in children],
                safe=False,
            )

        accounts = list(_base_queryset(company).order_by("code"))
        nodes = {a.id: {**_serialize(a), "children": []} for a in accounts}
        roots = []
        for account in accounts:
            node = nodes[account.id]
            parent_node = nodes.get(account.parent_id)
            if parent_node is not None:
                parent_node["children"].append(node)
            else:
                roots.append(node)
        return JsonResponse(roots, safe=False)


@method_decorator(login_required, name="dispatch")
class CoASearchAPI(View):
    def get(self, request):
        company = _company(request)
        q = request.GET.get("q", "").strip()
        if len(q) < 2:
            return JsonResponse([], safe=False)
        from django.db.models import Q
        matches = (
            _base_queryset(company)
            .filter(Q(code__icontains=q) | Q(description__icontains=q))
            .order_by("code")[:50]
        )
        return JsonResponse([_serialize(a) for a in matches], safe=False)


# ------------------------------------------------------------------ generation

@method_decorator(login_required, name="dispatch")
class CoATemplatesAPI(View):
    def get(self, request):
        from django.db.models import Count
        templates = (
            CoATemplate.objects.filter(status="Active")
            .annotate(account_count=Count("accounts"))
            .order_by("industry")
        )
        return JsonResponse([
            {
                "id": t.id,
                "template_name": t.template_name,
                "industry": t.industry,
                "country": t.country,
                "currency": t.currency,
                "description": t.description,
                "account_count": t.account_count,
            }
            for t in templates
        ], safe=False)


@method_decorator(login_required, name="dispatch")
class CoAGenerateAPI(View):
    def post(self, request):
        data = _parse_body(request)
        company = _company(request)
        template = CoATemplate.objects.filter(pk=data.get("template")).first()
        if template is None:
            return JsonResponse({"error": "Select a valid COA template."}, status=400)
        options = {
            key: bool(data.get(key, True))
            for key in ("with_tax", "with_inventory", "with_cash", "with_banks", "with_fixed_assets")
        }
        service = CoAGeneratorService(company, template, user=request.user, options=options)
        try:
            log = service.generate()
        except Exception as exc:
            return JsonResponse({"error": f"Generation failed and was rolled back: {exc}"}, status=400)
        _audit(request, "generate", company=company, new={"template": template.template_name, **log.summary})
        return JsonResponse({
            "message": "Chart of accounts generated.",
            "log_id": log.id,
            "summary": log.summary,
        }, status=201)


# --------------------------------------------------------------- import/export

IMPORT_COLUMNS = ("code", "name", "type", "group", "parent_code", "is_group", "opening_balance", "opening_type")


@method_decorator(login_required, name="dispatch")
class CoAImportAPI(View):
    """Validate-then-import. ?dry_run=1 only validates and returns row errors;
    a real run is all-or-nothing inside one transaction."""

    def post(self, request):
        from account.views import _parse_csv_rows, _parse_excel_rows

        company = _company(request)
        upload = request.FILES.get("file")
        if not upload:
            return JsonResponse({"error": "No file uploaded."}, status=400)
        filename = upload.name.lower()
        try:
            if filename.endswith(".csv"):
                rows = _parse_csv_rows(upload)
            elif filename.endswith((".xlsx", ".xls")):
                rows = _parse_excel_rows(upload)
            else:
                return JsonResponse({"error": "Unsupported file type. Upload .csv or .xlsx."}, status=400)
        except Exception as exc:
            return JsonResponse({"error": f"Could not read file: {exc}"}, status=400)

        parsed, errors = self._validate(company, rows)
        dry_run = request.GET.get("dry_run") in ("1", "true")
        if errors or dry_run:
            return JsonResponse({
                "valid": not errors,
                "rows": len(parsed),
                "errors": errors,
                "imported": 0,
            }, status=200 if (dry_run and not errors) else (400 if errors else 200))

        try:
            with transaction.atomic():
                imported = self._import(request, company, parsed)
        except Exception as exc:
            return JsonResponse({"error": f"Import failed and was rolled back: {exc}"}, status=400)
        _audit(request, "import", company=company, new={"rows": imported, "file": upload.name})
        return JsonResponse({"message": f"{imported} accounts imported.", "imported": imported})

    def _validate(self, company, rows):
        types = {t.name.lower(): t for t in AccountType.objects.all()}
        types.update({t.code.lower(): t for t in AccountType.objects.all()})
        groups = {g.name.lower(): g for g in AccountGroup.objects.all()}
        existing_codes = set(
            ChartOfAccount.all_objects.filter(company=company).values_list("code", flat=True)
        )
        parsed, errors, seen_codes = [], [], set()
        for row_num, row in rows:
            name = str(row.get("name") or row.get("description") or "").strip()
            code = str(row.get("code") or "").strip()
            type_key = str(row.get("type") or "").strip().lower()
            group_key = str(row.get("group") or "").strip().lower()
            parent_code = str(row.get("parent_code") or row.get("parent") or "").strip()
            is_group = str(row.get("is_group") or "").strip().lower() in ("1", "true", "yes", "y")
            balance_raw = str(row.get("opening_balance") or "0").strip() or "0"
            opening_type = str(row.get("opening_type") or "Debit").strip().title()

            if not name:
                errors.append({"row": row_num, "error": "Name is required."})
                continue
            account_type = types.get(type_key)
            if account_type is None:
                errors.append({"row": row_num, "error": f"Unknown account type '{row.get('type')}'."})
                continue
            if code and (code in existing_codes or code in seen_codes):
                errors.append({"row": row_num, "error": f"Code '{code}' already exists."})
                continue
            if parent_code and parent_code not in existing_codes and parent_code not in seen_codes:
                errors.append({"row": row_num, "error": f"Parent code '{parent_code}' not found (parents must appear before children)."})
                continue
            try:
                balance = Decimal(balance_raw)
            except InvalidOperation:
                errors.append({"row": row_num, "error": f"Invalid opening balance '{balance_raw}'."})
                continue
            if opening_type not in ("Debit", "Credit"):
                errors.append({"row": row_num, "error": f"Opening type must be Debit or Credit, got '{opening_type}'."})
                continue
            if code:
                seen_codes.add(code)
            parsed.append({
                "row": row_num,
                "code": code,
                "name": name,
                "account_type": account_type,
                "group": groups.get(group_key),
                "parent_code": parent_code,
                "is_group": is_group,
                "opening_balance": balance,
                "opening_type": opening_type,
            })
        return parsed, errors

    def _import(self, request, company, parsed):
        codegen = AccountCodeGenerator(company)
        by_code = {}
        imported = 0
        for row in parsed:
            parent = None
            if row["parent_code"]:
                parent = by_code.get(row["parent_code"]) or ChartOfAccount.objects.filter(
                    company=company, code=row["parent_code"]
                ).first()
                if parent is None or not parent.is_group:
                    raise ValueError(f"Row {row['row']}: parent '{row['parent_code']}' is missing or not a group.")
            account = ChartOfAccount(
                company=company,
                parent=parent,
                code=row["code"] or codegen.next_code(
                    parent=parent, account_type=row["account_type"], is_group=row["is_group"]
                ),
                description=row["name"],
                account_type=row["account_type"],
                account_group=row["group"] or (parent.account_group if parent else None),
                is_group=row["is_group"],
                is_postable=not row["is_group"],
                opening_balance=row["opening_balance"] if not row["is_group"] else 0,
                opening_type=row["opening_type"],
                created_by=request.user,
            )
            account.save()
            if account.code:
                by_code[account.code] = account
            imported += 1
        return imported


@method_decorator(login_required, name="dispatch")
class CoAImportTemplateAPI(View):
    """Downloadable sample file showing the import columns and example rows."""

    HEADER = ["code", "name", "type", "group", "parent_code", "is_group", "opening_balance", "opening_type"]
    SAMPLE_ROWS = [
        ["611000", "Marketing Expenses", "Expense", "Administrative Expense", "610000", "Yes", "", ""],
        ["611001", "Google Ads", "Expense", "Administrative Expense", "611000", "No", "0", ""],
        ["611002", "Print Media", "Expense", "Administrative Expense", "611000", "No", "2500", "Debit"],
        ["", "Shop Counter Cash", "Asset", "Current Assets", "111000", "No", "10000", "Debit"],
        ["", "Festival Advance", "Liability", "Current Liability", "210000", "No", "0", "Credit"],
    ]

    def _instructions(self):
        types = ", ".join(AccountType.objects.values_list("name", flat=True))
        groups = ", ".join(AccountGroup.objects.filter(is_active=True).values_list("name", flat=True))
        return [
            ["Column", "Required", "What to enter"],
            ["code", "No", "Leave blank to auto-generate the next code. Fill only to force a specific code (must be unique)."],
            ["name", "Yes", "Account name, e.g. 'Google Ads'."],
            ["type", "Yes", f"One of: {types}."],
            ["group", "No", f"One of: {groups}. Blank = inherited from the parent."],
            ["parent_code", "No", "Code of the parent GROUP account. Blank = root account. Parents must appear on an earlier row or already exist (use Export to see your current codes)."],
            ["is_group", "No", "Yes/No. 'Yes' makes a group that holds sub-accounts (cannot be posted to). Blank = No."],
            ["opening_balance", "No", "Amount for postable accounts only. Blank = 0."],
            ["opening_type", "No", "Debit or Credit. Blank = Debit."],
            [],
            ["Note", "", "The example rows are illustrative - replace parent_code values with codes from your own chart. The file is fully validated before anything is saved."],
        ]

    def get(self, request):
        fmt = request.GET.get("format", "csv").lower()
        if fmt == "xlsx":
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Accounts"
            sheet.append(self.HEADER)
            for row in self.SAMPLE_ROWS:
                sheet.append(row)
            info = workbook.create_sheet("Instructions")
            for row in self._instructions():
                info.append(row)
            info.column_dimensions["C"].width = 110
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="coa_import_template.xlsx"'
            return response

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(self.HEADER)
        writer.writerows(self.SAMPLE_ROWS)
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="coa_import_template.csv"'
        return response


@method_decorator(login_required, name="dispatch")
class CoAExportAPI(View):
    HEADER = ["Code", "Name", "Type", "Group", "Parent Code", "Level", "Is Group",
              "Postable", "Opening Balance", "Opening Type", "Status"]

    def get(self, request):
        company = _company(request)
        fmt = request.GET.get("format", "csv").lower()
        accounts = (
            _base_queryset(company)
            .select_related("parent")
            .order_by("path")
        )
        rows = [
            [
                a.code,
                a.description,
                a.account_type.name if a.account_type_id else a.type,
                a.account_group.name if a.account_group_id else "",
                a.parent.code if a.parent_id else "",
                a.level,
                "Yes" if a.is_group else "No",
                "Yes" if a.is_postable else "No",
                str(a.opening_balance),
                a.opening_type,
                a.status,
            ]
            for a in accounts
        ]
        if fmt == "xlsx":
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Chart of Accounts"
            sheet.append(self.HEADER)
            for row in rows:
                sheet.append(row)
            buffer = io.BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="chart_of_accounts.xlsx"'
            return response

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(self.HEADER)
        writer.writerows(rows)
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="chart_of_accounts.csv"'
        return response


# ------------------------------------------------------------ opening balance

@method_decorator(login_required, name="dispatch")
class CoAOpeningBalanceAPI(View):
    """GET lists postable accounts with balances; POST applies a batch of
    {account, opening_balance, opening_type} rows and rebalances the
    Opening Balance Equity anchor so total debits equal total credits."""

    def get(self, request):
        company = _company(request)
        accounts = _base_queryset(company).filter(is_postable=True).order_by("code")
        totals = self._totals(accounts)
        return JsonResponse({
            "results": [_serialize(a) for a in accounts],
            "total_debit": str(totals[0]),
            "total_credit": str(totals[1]),
        })

    def post(self, request):
        data = _parse_body(request)
        company = _company(request)
        entries = data.get("entries", [])
        if not isinstance(entries, list) or not entries:
            return JsonResponse({"error": "Provide a non-empty 'entries' list."}, status=400)
        try:
            with transaction.atomic():
                updated = self._apply(request, company, entries)
                balancer = self._rebalance(company)
        except Exception as exc:
            return JsonResponse({"error": str(exc)}, status=400)

        # Mirror the balances into the system Opening voucher so ledgers and
        # the trial balance see them. Failure here (e.g. no open financial
        # year yet) must not lose the saved balances.
        opening_voucher = None
        voucher_warning = ""
        try:
            from account.services import journal
            opening_voucher = journal.sync_opening_voucher(company, user=request.user)
        except Exception as exc:
            voucher_warning = f"Balances saved, but the opening voucher was not updated: {exc}"
        return JsonResponse({
            "message": f"{updated} opening balances saved.",
            "balancing_account": balancer and balancer.code,
            "balancing_amount": balancer and str(balancer.opening_balance),
            "balancing_side": balancer and balancer.opening_type,
            "opening_voucher": opening_voucher and opening_voucher.voucher_no,
            "warning": voucher_warning,
        })

    def _apply(self, request, company, entries):
        updated = 0
        for entry in entries:
            account = _get_account(company, entry.get("account"))
            if not account.is_postable:
                raise ValueError(f"{account.code} is a group account and cannot hold an opening balance.")
            try:
                balance = Decimal(str(entry.get("opening_balance", "0")))
            except InvalidOperation:
                raise ValueError(f"Invalid opening balance for {account.code}.")
            side = str(entry.get("opening_type", account.opening_type)).title()
            if side not in ("Debit", "Credit"):
                raise ValueError(f"Opening type for {account.code} must be Debit or Credit.")
            old = _snapshot(account)
            account.opening_balance = balance
            account.opening_type = side
            account.modified_by = request.user
            account.save()
            _audit(request, "opening_balance", account=account, old=old, new=_snapshot(account))
            updated += 1
        return updated

    def _rebalance(self, company):
        balancer = ChartOfAccount.objects.filter(
            company=company, system_role="OPENING_BALANCE_EQUITY"
        ).first()
        if balancer is None:
            return None
        accounts = ChartOfAccount.objects.filter(company=company, is_postable=True).exclude(pk=balancer.pk)
        debit, credit = self._totals(accounts)
        difference = debit - credit
        balancer.opening_balance = abs(difference)
        balancer.opening_type = "Credit" if difference > 0 else "Debit"
        balancer.save()
        return balancer

    @staticmethod
    def _totals(accounts):
        debit = credit = Decimal("0")
        for account in accounts:
            if account.opening_type == "Debit":
                debit += account.opening_balance
            else:
                credit += account.opening_balance
        return debit, credit


# ------------------------------------------------------------------ audit trail

@method_decorator(login_required, name="dispatch")
class CoAAuditLogAPI(View):
    def get(self, request):
        company = _company(request)
        qs = AccountAuditLog.objects.filter(company=company).select_related("user", "account")
        account_id = request.GET.get("account")
        if account_id:
            qs = qs.filter(account_id=account_id)
        try:
            page = max(int(request.GET.get("page", 1)), 1)
            page_size = min(int(request.GET.get("page_size", 50)), MAX_PAGE_SIZE)
        except ValueError:
            return JsonResponse({"error": "Invalid pagination parameters."}, status=400)
        offset = (page - 1) * page_size
        return JsonResponse({
            "count": qs.count(),
            "results": [
                {
                    "id": log.id,
                    "action": log.action,
                    "account": log.account and f"{log.account.code} - {log.account.description}",
                    "old_values": log.old_values,
                    "new_values": log.new_values,
                    "reason": log.reason,
                    "user": log.user and log.user.get_username(),
                    "ip_address": log.ip_address,
                    "timestamp": log.timestamp.strftime("%d-%b-%Y %H:%M:%S"),
                }
                for log in qs[offset:offset + page_size]
            ],
        })
