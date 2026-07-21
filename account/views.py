import csv
import io

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, Http404
from django.views import View
from django.utils.decorators import method_decorator
import json
import openpyxl
from django.http import HttpResponse

from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from django.db.models.deletion import ProtectedError

from account.models import BankCashMaster, ChartOfAccount, CompanyProfile, FinancialYear, OrganizationCentre, TermsConditions
from hatchery_master.models import STATES_AND_TERRITORIES
from inventory.models import Mapping, Sector, Warehouse

# Create your views here.
@login_required
def coa(request):
    from account.models import AccountGroup, AccountType
    context = {
        'account_types': AccountType.objects.all(),
        'account_groups': AccountGroup.objects.filter(is_active=True),
    }

    return render(request, "coa.html", context)


@login_required
def ledger_report(request):
    return render(request, "ledger_report.html")


@login_required
def profit_loss_report(request):
    return render(request, "profit_loss_report.html")


@login_required
def balance_sheet_report(request):
    return render(request, "balance_sheet_report.html")


@login_required
def vouchers(request):
    from account.models import Voucher
    from inventory.models import Warehouse
    return render(request, "journal.html", {
        "voucher_types": Voucher.TYPE_CHOICES,
        "sectors": Warehouse.objects.all().order_by("name"),
    })


def _parse_csv_rows(upload):
    text = io.TextIOWrapper(upload.file, encoding="utf-8-sig")
    reader = csv.DictReader(text)
    reader.fieldnames = [(f or "").strip().lower() for f in (reader.fieldnames or [])]
    return list(enumerate(reader, start=2))


def _parse_excel_rows(upload):
    workbook = openpyxl.load_workbook(upload, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(h or "").strip().lower() for h in next(rows_iter)]
    rows = []
    for row_num, values in enumerate(rows_iter, start=2):
        if not any(values):
            continue
        rows.append((row_num, dict(zip(headers, values))))
    return rows


@login_required
def fin_year(request):
    return render(request, "fin_year.html")


def _coa_head_label(is_cash):
    """The COA anchor group a Bank/Cash row's auto-created ledger files
    under — 'Bank Accounts' for a bank, 'Cash & Cash Equivalents' for cash.
    Shown read-only on the Bank/Cash Master so the mapping is visible before
    and after saving; falls back gracefully if the company's COA hasn't
    been generated yet."""
    role = 'CASH' if is_cash else 'BANK_ACCOUNTS'
    anchor = ChartOfAccount.objects.filter(system_role=role).first()
    return f"{anchor.code} - {anchor.description}" if anchor else "Not yet generated"


@login_required
def bank_cash(request):
    return render(request, "bank_cash.html", {
        "sectors": Warehouse.objects.all().order_by("name"),
        "bank_coa_head": _coa_head_label(is_cash=False),
        "cash_coa_head": _coa_head_label(is_cash=True),
    })


def _resolve_sectors(raw_ids):
    """(warehouses, error) for a list of office ids. Empty is valid and
    means 'All Offices' — a bank/cash record isn't pinned to one place."""
    ids = raw_ids or []
    warehouses = list(Warehouse.objects.filter(id__in=ids))
    if len(warehouses) != len(set(ids)):
        return None, "Invalid office ID"
    return warehouses, None


def _bank_cash_dict(row):
    offices = list(row.sectors.all())
    all_offices = not offices
    return {
        "id": row.id, "code": row.code, "is_cash": row.is_cash, "name": row.name,
        "sectors": [o.id for o in offices],
        "sector_names": "All Offices" if all_offices else ", ".join(o.name for o in offices),
        "micr": row.micr, "address": row.address, "email": row.email,
        "phone": row.phone, "fax": row.fax, "contact_person": row.contact_person,
        "coa_head": _coa_head_label(row.is_cash),
    }


@method_decorator(login_required, name="dispatch")
class BankCashMasterAPI(View):
    def get(self, request, id=None):
        if id:
            try:
                return JsonResponse(_bank_cash_dict(BankCashMaster.objects.prefetch_related("sectors").get(id=id)))
            except BankCashMaster.DoesNotExist:
                raise Http404("Record not found")
        rows = BankCashMaster.objects.prefetch_related("sectors").order_by("name")
        return JsonResponse([_bank_cash_dict(r) for r in rows], safe=False)

    def post(self, request):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        if not data.get("name"):
            return JsonResponse({"error": "Name is required"}, status=400)

        offices, error = _resolve_sectors(data.get("sectors"))
        if error:
            return JsonResponse({"error": error}, status=400)

        try:
            row = BankCashMaster.objects.create(
                is_cash=bool(data.get("is_cash")), name=data["name"],
                micr=data.get("micr") or None,
                address=data.get("address") or None, email=data.get("email") or None,
                phone=data.get("phone") or None, fax=data.get("fax") or None,
                contact_person=data.get("contact_person") or None,
            )
            row.sectors.set(offices)
            return JsonResponse({"message": "Saved", "id": row.id, "code": row.code}, status=201)
        except (KeyError, ValidationError) as e:
            return JsonResponse({"error": str(e)}, status=400)
        except IntegrityError:
            return JsonResponse({"error": "A record with this code may already exist."}, status=400)

    def put(self, request, id):
        try:
            row = BankCashMaster.objects.get(id=id)
        except BankCashMaster.DoesNotExist:
            raise Http404("Record not found")

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        offices, error = _resolve_sectors(data.get("sectors"))
        if error:
            return JsonResponse({"error": error}, status=400)

        try:
            row.is_cash = bool(data.get("is_cash", row.is_cash))
            row.name = data.get("name", row.name)
            row.micr = data.get("micr") or None
            row.address = data.get("address") or None
            row.email = data.get("email") or None
            row.phone = data.get("phone") or None
            row.fax = data.get("fax") or None
            row.contact_person = data.get("contact_person") or None
            row.full_clean(exclude=["code"])
            row.save()
            row.sectors.set(offices)
            return JsonResponse({"message": "Updated"})
        except ValidationError as e:
            return JsonResponse({"error": "; ".join(e.messages)}, status=400)
        except IntegrityError:
            return JsonResponse({"error": "A record with this code may already exist."}, status=400)

    def delete(self, request, id):
        try:
            row = BankCashMaster.objects.get(id=id)
        except BankCashMaster.DoesNotExist:
            raise Http404("Record not found")
        row.delete()
        return JsonResponse({"message": "Deleted"})


# ---------------------------------------------------------------------------
# Cost Center (Account > Master > Cost Center)
# ---------------------------------------------------------------------------

def _default_company():
    return CompanyProfile.objects.filter(pk=1).first()


@login_required
def organization_centre(request):
    company = _default_company()
    return render(request, "organization_centre.html", {
        "centre_types": Sector.objects.order_by("name"),
        "branch_centre_type_code": OrganizationCentre.CENTRE_TYPE_BRANCH,
        "category_choices": OrganizationCentre.CATEGORY_CHOICES,
        "company_name": company.name if company else "",
    })


def _nearest_branch_name(cc):
    """Walk up from *cc* (including itself) to the nearest ancestor that's
    actually linked to a real Branch — so even a Shed/Unit deep in the tree
    can show which Branch it ultimately belongs to, without every node
    needing its own redundant Branch field to keep in sync."""
    node = cc
    seen = set()
    while node is not None and node.pk not in seen:
        if node.branch_id:
            return node.branch.branch_name
        seen.add(node.pk)
        node = node.parent
    return ""


def _organization_centre_dict(cc):
    return {
        "id": cc.id, "code": cc.code, "name": cc.name,
        "centre_type": cc.centre_type_id, "centre_type_name": cc.centre_type.name, "centre_type_code": cc.centre_type.code,
        "category": cc.category,
        "parent": cc.parent_id, "parent_label": f"{cc.parent.name} ({cc.parent.code})" if cc.parent_id else "",
        "level": cc.level,
        "description": cc.description,
        "effective_from": cc.effective_from.isoformat() if cc.effective_from else None,
        "effective_to": cc.effective_to.isoformat() if cc.effective_to else None,
        "is_active": cc.is_active,
        "is_locked": cc.is_locked,
        "allow_manual_selection": cc.allow_manual_selection,
        "allow_children_only": cc.allow_children_only,
        "is_default": cc.is_default,
        "approval_status": cc.approval_status,
        "approval_remarks": cc.approval_remarks,
        "branch": cc.branch_id,
        "branch_name": cc.branch.branch_name if cc.branch_id else "",
        "nearest_branch_name": _nearest_branch_name(cc),
        "company_name": cc.company.name if cc.company_id else "",
        "children_count": cc.children.count(),
        "created_by": cc.created_by.get_username() if cc.created_by_id else "",
        "created_at": cc.created_at.strftime("%d-%b-%Y %H:%M"),
        "updated_by": cc.updated_by.get_username() if cc.updated_by_id else "",
        "updated_at": cc.updated_at.strftime("%d-%b-%Y %H:%M"),
    }


@method_decorator(login_required, name="dispatch")
class OrganizationCentreAPI(View):
    def get(self, request, id=None):
        if id:
            try:
                return JsonResponse(_organization_centre_dict(
                    OrganizationCentre.objects.select_related("parent", "branch", "centre_type", "created_by", "updated_by").get(id=id)
                ))
            except OrganizationCentre.DoesNotExist:
                raise Http404("Organization centre not found")
        rows = OrganizationCentre.objects.select_related("parent", "branch", "centre_type", "created_by", "updated_by").order_by("code")
        return JsonResponse([_organization_centre_dict(r) for r in rows], safe=False)

    def post(self, request):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        if not data.get("name"):
            return JsonResponse({"error": "Name is required"}, status=400)
        centre_type = Sector.objects.filter(id=data.get("centre_type")).first()
        if not centre_type:
            return JsonResponse({"error": "Invalid centre type"}, status=400)
        if centre_type.code == OrganizationCentre.CENTRE_TYPE_BRANCH:
            return JsonResponse(
                {"error": "Branch centres are created automatically from Broiler > Master > Branch — add the Branch there instead."},
                status=400,
            )
        category = data.get("category") or OrganizationCentre.CATEGORY_BOTH
        if category not in dict(OrganizationCentre.CATEGORY_CHOICES):
            return JsonResponse({"error": "Invalid category"}, status=400)
        parent_id = data.get("parent") or None
        if parent_id and OrganizationCentre.objects.filter(id=parent_id, is_locked=True).exists():
            return JsonResponse({"error": "The selected parent centre is locked."}, status=400)

        try:
            cc = OrganizationCentre(
                company=_default_company(), name=data["name"], centre_type=centre_type,
                category=category,
                parent_id=parent_id,
                description=data.get("description", ""),
                effective_from=data.get("effective_from") or None,
                effective_to=data.get("effective_to") or None,
                is_active=bool(data.get("is_active", True)),
                allow_manual_selection=bool(data.get("allow_manual_selection", True)),
                allow_children_only=bool(data.get("allow_children_only", False)),
                is_default=bool(data.get("is_default", False)),
                created_by=request.user, updated_by=request.user,
            )
            cc.full_clean(exclude=["code"])
            cc.save()
            return JsonResponse({"message": "Saved", "id": cc.id, "code": cc.code}, status=201)
        except (KeyError, ValidationError) as e:
            message = "; ".join(e.messages) if hasattr(e, "messages") else str(e)
            return JsonResponse({"error": message}, status=400)
        except IntegrityError:
            return JsonResponse({"error": "An organization centre with this code may already exist."}, status=400)

    def put(self, request, id):
        try:
            cc = OrganizationCentre.objects.select_related("centre_type").get(id=id)
        except OrganizationCentre.DoesNotExist:
            raise Http404("Organization centre not found")
        if cc.is_locked:
            return JsonResponse({"error": "This organization centre is locked."}, status=400)

        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        centre_type = Sector.objects.filter(id=data.get("centre_type", cc.centre_type_id)).first()
        if not centre_type:
            return JsonResponse({"error": "Invalid centre type"}, status=400)
        if not cc.branch_id and centre_type.code == OrganizationCentre.CENTRE_TYPE_BRANCH:
            return JsonResponse(
                {"error": "Branch centres are created automatically from Broiler > Master > Branch — this one isn't linked to a Branch."},
                status=400,
            )
        category = data.get("category", cc.category)
        if category not in dict(OrganizationCentre.CATEGORY_CHOICES):
            return JsonResponse({"error": "Invalid category"}, status=400)
        parent_id = data.get("parent") or None
        if parent_id and int(parent_id) == cc.id:
            return JsonResponse({"error": "An organization centre cannot be its own parent"}, status=400)
        if parent_id and OrganizationCentre.objects.filter(id=parent_id, is_locked=True).exists():
            return JsonResponse({"error": "The selected parent centre is locked."}, status=400)

        try:
            cc.name = data.get("name", cc.name)
            cc.centre_type = centre_type
            cc.category = category
            cc.parent_id = parent_id
            cc.description = data.get("description", cc.description)
            cc.effective_from = data.get("effective_from") or None
            cc.effective_to = data.get("effective_to") or None
            cc.is_active = bool(data.get("is_active", cc.is_active))
            cc.allow_manual_selection = bool(data.get("allow_manual_selection", cc.allow_manual_selection))
            cc.allow_children_only = bool(data.get("allow_children_only", cc.allow_children_only))
            cc.is_default = bool(data.get("is_default", cc.is_default))
            cc.updated_by = request.user
            cc.full_clean(exclude=["code"])
            cc.save()
            return JsonResponse({"message": "Updated"})
        except ValidationError as e:
            return JsonResponse({"error": "; ".join(e.messages)}, status=400)
        except IntegrityError:
            return JsonResponse({"error": "An organization centre with this code may already exist."}, status=400)

    def delete(self, request, id):
        try:
            cc = OrganizationCentre.objects.get(id=id)
        except OrganizationCentre.DoesNotExist:
            raise Http404("Organization centre not found")
        if cc.is_locked:
            return JsonResponse({"error": "This organization centre is locked."}, status=400)
        try:
            cc.delete()
        except ProtectedError as e:
            related = sorted({str(obj._meta.verbose_name) for obj in list(e.protected_objects)[:50]})
            names = ", ".join(related) if related else "other records"
            return JsonResponse(
                {"error": f"Cannot delete: this organization centre is used in {names}."}, status=400
            )
        return JsonResponse({"message": "Deleted"})


@login_required
def organization_centre_duplicate(request, id):
    """Clone one centre — same parent/centre_type/category, a fresh
    auto-generated code, unlocked, approval reset to Draft."""
    if request.method != "POST":
        return JsonResponse({"error": "POST required."}, status=405)
    try:
        source = OrganizationCentre.objects.select_related("centre_type").get(id=id)
    except OrganizationCentre.DoesNotExist:
        return JsonResponse({"error": "Organization centre not found."}, status=404)
    try:
        clone = OrganizationCentre(
            company=source.company, parent_id=source.parent_id,
            name=f"{source.name} (Copy)", centre_type=source.centre_type,
            category=source.category, description=source.description,
            effective_from=source.effective_from, effective_to=source.effective_to,
            is_active=source.is_active,
            allow_manual_selection=source.allow_manual_selection,
            allow_children_only=source.allow_children_only,
            created_by=request.user, updated_by=request.user,
        )
        clone.full_clean(exclude=["code"])
        clone.save()
        return JsonResponse({"message": "Duplicated", "id": clone.id, "code": clone.code}, status=201)
    except ValidationError as e:
        return JsonResponse({"error": "; ".join(e.messages)}, status=400)
    except IntegrityError:
        return JsonResponse({"error": "An organization centre with this code may already exist."}, status=400)


@login_required
def cost_center_toggle_lock(request, id):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        cc = OrganizationCentre.objects.get(id=id)
    except OrganizationCentre.DoesNotExist:
        return JsonResponse({"error": "Organization centre not found."}, status=404)
    cc.is_locked = not cc.is_locked
    cc.save(update_fields=["is_locked"])
    return JsonResponse({"message": "Organization centre updated", "is_locked": cc.is_locked})


@login_required
def cost_center_approve(request, id):
    """Set approval status (Approved/Rejected/Pending) with optional
    remarks. Gated behind the same 'edit' permission as the rest of the
    master — a distinct Approve permission would need extending the whole
    Web-Access permission matrix, out of scope for this page alone."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        cc = OrganizationCentre.objects.get(id=id)
    except OrganizationCentre.DoesNotExist:
        return JsonResponse({"error": "Organization centre not found."}, status=404)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    status = data.get("approval_status")
    if status not in dict(OrganizationCentre.APPROVAL_CHOICES):
        return JsonResponse({"error": "Invalid approval status"}, status=400)
    cc.approval_status = status
    cc.approval_remarks = data.get("approval_remarks", cc.approval_remarks)
    cc.updated_by = request.user
    cc.save(update_fields=["approval_status", "approval_remarks", "updated_by", "updated_at"])
    return JsonResponse({"message": "Approval status updated"})


def _cost_center_tree_node(cc, children_by_parent):
    return {
        "id": cc.id, "code": cc.code, "name": cc.name, "level": cc.level,
        "centre_type_name": cc.centre_type.name, "centre_type_code": cc.centre_type.code, "category": cc.category,
        "is_active": cc.is_active, "is_locked": cc.is_locked,
        "approval_status": cc.approval_status,
        "children": [
            _cost_center_tree_node(child, children_by_parent)
            for child in children_by_parent.get(cc.id, [])
        ],
    }


@login_required
def cost_center_tree(request):
    """Full hierarchy as nested JSON, for the tree panel."""
    rows = list(OrganizationCentre.objects.select_related("centre_type").order_by("code"))
    children_by_parent = {}
    for cc in rows:
        children_by_parent.setdefault(cc.parent_id, []).append(cc)
    roots = children_by_parent.get(None, [])
    return JsonResponse({"tree": [_cost_center_tree_node(cc, children_by_parent) for cc in roots]})


@login_required
def cost_center_children(request, id):
    """Direct children of one organization centre (id=0 for top-level roots)."""
    parent_id = id or None
    rows = OrganizationCentre.objects.select_related("centre_type").filter(parent_id=parent_id).order_by("code")
    return JsonResponse([{
        "id": r.id, "code": r.code, "name": r.name, "level": r.level,
        "centre_type_name": r.centre_type.name, "is_active": r.is_active,
    } for r in rows], safe=False)


@login_required
def cost_center_parent(request, id):
    try:
        cc = OrganizationCentre.objects.select_related("parent__centre_type").get(id=id)
    except OrganizationCentre.DoesNotExist:
        return JsonResponse({"error": "Organization centre not found."}, status=404)
    if not cc.parent_id:
        return JsonResponse({})
    return JsonResponse(_organization_centre_dict(OrganizationCentre.objects.select_related("centre_type", "branch", "created_by", "updated_by").get(id=cc.parent_id)))


@login_required
def cost_center_export_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Organization Centres"
    headers = ["Code", "Name", "Centre Type", "Category", "Level", "Parent", "Branch", "Status", "Locked", "Approval Status", "Effective From", "Effective To"]
    sheet.append(headers)
    for cc in OrganizationCentre.objects.select_related("centre_type", "parent", "branch").order_by("code"):
        sheet.append([
            cc.code, cc.name, cc.centre_type.name, cc.get_category_display(), cc.level,
            cc.parent.code if cc.parent_id else "",
            cc.branch.branch_name if cc.branch_id else "",
            "Active" if cc.is_active else "Inactive",
            "Yes" if cc.is_locked else "No",
            cc.approval_status,
            cc.effective_from.isoformat() if cc.effective_from else "",
            cc.effective_to.isoformat() if cc.effective_to else "",
        ])
    for column_cells in sheet.columns:
        lengths = [len(str(cell.value)) for cell in column_cells if cell.value is not None]
        sheet.column_dimensions[column_cells[0].column_letter].width = max(10, (max(lengths) if lengths else 0) + 2)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="organization_centres.xlsx"'
    workbook.save(response)
    return response


# ---------------------------------------------------------------------------
# Organization Centre Mapping (Account > Master > Organization Centre Mapping)
# ---------------------------------------------------------------------------
# Separate from the Organization Centre master on purpose: that page only
# names/codes a centre; this page is the one place that says which Office(s)
# route their postings to it. Entirely manual — no auto-creation, no
# auto-sync — built on the same generic Mapping table Inventory > Office
# Mapping uses, under its own type so it doesn't show up on that screen.

@login_required
def cost_center_mapping(request):
    return render(request, "organization_centre_mapping.html")


@login_required
def cost_center_mapping_data(request):
    mapped = dict(
        Mapping.objects.filter(type=Mapping.TYPE_OFFICE_COST_CENTER, to_id__isnull=False)
        .values_list("from_id", "to_id")
    )
    cost_center_names = dict(OrganizationCentre.objects.values_list("id", "name"))
    offices = [
        {
            "id": office.id, "name": office.name,
            "cost_center_id": mapped.get(office.id),
            "cost_center_name": cost_center_names.get(mapped.get(office.id), ""),
        }
        for office in Warehouse.objects.order_by("name")
    ]
    cost_centers = list(OrganizationCentre.objects.order_by("code").values("id", "code", "name"))
    return JsonResponse({"offices": offices, "cost_centers": cost_centers})


@login_required
def cost_center_mapping_save(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required."}, status=405)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    office_id = data.get("office")
    cost_center_id = data.get("cost_center") or None
    if not office_id:
        return JsonResponse({"error": "Office is required"}, status=400)
    if not Warehouse.objects.filter(id=office_id).exists():
        return JsonResponse({"error": "Office not found"}, status=404)
    if cost_center_id and not OrganizationCentre.objects.filter(id=cost_center_id).exists():
        return JsonResponse({"error": "Organization centre not found"}, status=404)

    if cost_center_id is None:
        Mapping.objects.filter(type=Mapping.TYPE_OFFICE_COST_CENTER, from_id=office_id).delete()
    else:
        Mapping.objects.update_or_create(
            type=Mapping.TYPE_OFFICE_COST_CENTER, from_id=office_id,
            defaults={"to_id": cost_center_id},
        )
    return JsonResponse({"message": "Mapping saved"})


@login_required
def cost_center_report(request):
    return render(request, "cost_center_report.html")


@login_required
def branch_summary_report(request):
    return render(request, "branch_summary_report.html")


@method_decorator(login_required, name="dispatch")
class FinancialYearAPI(View):
    def get(self, request, id=None):
        """
        Get details of a single financial year by ID or list all financial years.
        """
        if id:
            try:
                fin_year = FinancialYear.objects.get(id=id)
                return JsonResponse(
                    {
                        "id": fin_year.id,
                        "start_date": fin_year.start_date,
                        "end_date": fin_year.end_date,
                        "is_active": fin_year.is_active,
                        "state": fin_year.state,
                    }
                )
            except FinancialYear.DoesNotExist:
                raise Http404("Financial Year not found")
        else:
            fin_years = list(
                FinancialYear.objects.values(
                    "id", "start_date", "end_date", "is_active", "state"
                )
            )
            return JsonResponse(fin_years, safe=False)

    def post(self, request):
        """
        Create a new financial year.
        """
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        try:
            fin_year = FinancialYear.objects.create(
                start_date=data["start_date"],
                end_date=data["end_date"],
                is_active=data.get("is_active", False),
            )
            return JsonResponse(
                {
                    "message": "Financial year created",
                    "id": fin_year.id,
                    "start_date": fin_year.start_date,
                    "end_date": fin_year.end_date,
                    "is_active": fin_year.is_active,
                },
                status=201,
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def put(self, request, id):
        """
        Update an existing financial year.
        """
        try:
            fin_year = FinancialYear.objects.get(id=id)
        except FinancialYear.DoesNotExist:
            raise Http404("Financial Year not found")

        try:
            data = json.loads(request.body)
            fin_year.start_date = data["start_date"]
            fin_year.end_date = data["end_date"]
            fin_year.is_active = data["is_active"]
            if data.get("state") in dict(FinancialYear.STATE_CHOICES):
                fin_year.state = data["state"]
            fin_year.save()
            return JsonResponse(
                {
                    "message": "Financial year updated",
                    "id": fin_year.id,
                    "start_date": fin_year.start_date,
                    "end_date": fin_year.end_date,
                    "is_active": fin_year.is_active,
                    "state": fin_year.state,
                }
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def delete(self, request, id):
        """
        Delete a financial year by ID.
        """
        try:
            fin_year = FinancialYear.objects.get(id=id)
        except FinancialYear.DoesNotExist:
            raise Http404("Financial Year not found")

        fin_year.delete()
        return JsonResponse({"message": "Financial year deleted"})


@method_decorator(login_required, name="dispatch")
class ChartOfAccountsAPI(View):
    def get(self, request, id=None):
        """
        Get details of a single chart of account by ID or list all chart of accounts.
        """
        if id:
            try:
                coa = ChartOfAccount.objects.get(id=id)
                return JsonResponse(
                    {
                        "id": coa.id,
                        "code": coa.code,
                        "description": coa.description,
                        "type": coa.type,
                        "control_type": coa.control_type,
                        "status": coa.status,
                        "schedule": coa.schedule_id,
                        "schedule__name": coa.schedule.name,
                        "created_at": coa.created_at.strftime("%d-%b-%Y %H:%M"),  # Format the datetime
                    }
                )
            except ChartOfAccount.DoesNotExist:
                raise Http404("Chart of Account not found")
        else:
            coas = ChartOfAccount.objects.values(
                "id", "code", "description", "type","control_type", "status", "schedule_id", "schedule__name","created_at"
            )
            formatted_coas = [
                {
                    **coa,
                    "created_at": coa["created_at"].strftime("%d-%b-%Y %H:%M"),  # Format the datetime
                }
                for coa in coas
            ]
            return JsonResponse(formatted_coas, safe=False)

    def post(self, request):
        """
        Create a new chart of account.
        """
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        try:
            coa = ChartOfAccount.objects.create(
                code=data["code"],
                description=data["description"],
                type=data["type"],
                control_type=data["control_type"],
                status=data["status"],
                schedule_id=data["schedule"],
            )
            return JsonResponse(
                {
                    "message": "Chart of Account created",
                    "id": coa.id,
                    "code": coa.code,
                    "description": coa.description,
                    "type": coa.type,
                    "control_type": coa.control_type,
                    "status": coa.status,
                    "schedule": coa.schedule_id,
                },
                status=201,
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def put(self, request, id):
        """
        Update an existing chart of account.
        """
        try:
            coa = ChartOfAccount.objects.get(id=id)
        except ChartOfAccount.DoesNotExist:
            raise Http404("Chart of Account not found")
        
        try:
            data = json.loads(request.body)
            coa.code = data["code"]
            coa.description = data["description"]
            coa.type = data["type"]
            coa.status = data["status"]
            coa.schedule_id = data["schedule"]
            coa.save()
            return JsonResponse(
                {
                    "message": "Chart of Account updated",
                    "id": coa.id,
                    "code": coa.code,
                    "description": coa.description,
                    "type": coa.type,
                    "status": coa.status,
                    "schedule": coa.schedule_id,
                }
            )
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
        
    def delete(self, request, id):

        try:
            coa = ChartOfAccount.objects.get(id=id)
        except ChartOfAccount.DoesNotExist:
            raise Http404("Chart of Account not found")
        
        coa.delete()
        return JsonResponse({"message": "Chart of Account deleted"})


@login_required
def company_profile(request):
    profile = CompanyProfile.get_solo()

    if request.method == "POST":
        profile.name = request.POST.get("name", "").strip()
        profile.address = request.POST.get("address", "").strip()
        profile.state = request.POST.get("state", "").strip()
        profile.mobile = request.POST.get("mobile", "").strip()
        profile.email = request.POST.get("email", "").strip()
        profile.gstin = request.POST.get("gstin", "").strip()
        profile.pan = request.POST.get("pan", "").strip()
        profile.bank_name = request.POST.get("bank_name", "").strip()
        profile.bank_account_no = request.POST.get("bank_account_no", "").strip()
        profile.ifsc_code = request.POST.get("ifsc_code", "").strip()
        profile.bank_branch = request.POST.get("bank_branch", "").strip()
        try:
            profile.full_clean()
            profile.save()
            messages.success(request, "Company profile updated successfully.")
            return redirect("company_profile")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "company_profile.html", {
        "profile": profile,
        "states_and_union_territories": STATES_AND_TERRITORIES,
    })


@login_required()
def terms(request):
    return render(request, "t&c.html", {"party_types": TermsConditions.PartyType.choices})


@method_decorator(login_required, name="dispatch")
class TermsConditionsAPI(View):

    def get(self, request, id=None):
        if id:
            try:
                terms_conditions = TermsConditions.objects.get(id=id)
                return JsonResponse(
                    {
                        "id": terms_conditions.id,
                        "type": terms_conditions.type,
                        "party_type": terms_conditions.party_type,
                        "condition": terms_conditions.condition,
                    }
                )
            except TermsConditions.DoesNotExist:
                raise Http404("TermsConditions not found")
        else:
            terms_conditions = list(
                TermsConditions.objects.values("id", "type", "party_type", "condition")
            )
            return JsonResponse(terms_conditions, safe=False)

    def post(self, request):
        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        TermsConditions.objects.create(
            type=data.get("type"),
            party_type=data.get("party_type") or TermsConditions.PartyType.CUSTOMER,
            condition=data.get("condition"),
        )
        return JsonResponse({"message": "TermsConditions created"}, status=201)

    def put(self, request, id):
        try:
            terms_conditions = TermsConditions.objects.get(id=id)
        except TermsConditions.DoesNotExist:
            raise Http404("TermsConditions not found")

        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        terms_conditions.type = data.get("type", terms_conditions.type)
        terms_conditions.party_type = data.get("party_type", terms_conditions.party_type)
        terms_conditions.condition = data.get("condition", terms_conditions.condition)
        terms_conditions.save()

        return JsonResponse({"message": "TermsConditions updated"})

    def delete(self, request, id):
        try:
            terms_conditions = TermsConditions.objects.get(id=id)
        except TermsConditions.DoesNotExist:
            raise Http404("TermsConditions not found")

        terms_conditions.delete()
        return JsonResponse({"message": "TermsConditions deleted"})

