#pylint: disable=no-member

from typing import Dict, List, Optional, Union
from django.shortcuts import render, get_object_or_404, redirect

# Data scoping: every user-facing option list below is narrowed to the
# branches / farms / warehouses the signed-in user is scoped to.
from user.services.scoping import (branches_for, customers_for, farms_for,
                                   scope_any, scope_multi, scope_or_null,
                                   suppliers_for, supervisors_for,
                                   warehouses_for)
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.http import Http404, JsonResponse
from django.db.models import F, Max, Min, Prefetch, Q, Sum
from django.core.files.storage import default_storage
from django.core.exceptions import ValidationError
from Hitech_BIMS.entry_dates import reject_future_date
from django.db import transaction
from django.core.cache import cache
from django.conf import settings
from .models import (
    BirdSale, BirdSaleReceipt, Branch, Breed, BreedStandard, BroilerBatch, BroilerDisease, BroilerFarm, BroilerFarmImage,
    BroilerFarmShed, BroilerLine, DailyEntry, Farmer, FarmerGroup,
    GrowingChargeScheme, GCProductionCostIncentive, GCSalesIncentive, GCMortalityIncentive,
    GCFCRIncentive, GCSummerIncentive, GCProductionCostDecentive, GCMortalityDecentive,
    GCFCRRecovery, GCFarmerClassification, GrowingChargeSettlement, MedicineVaccineEntry,
    Region, Supervisor, FeedPhaseMaster, FeedPhaseLine, BirdCategory,
    FarmLocationCapture, FarmCaptureFile,
)
from account.models import ChartOfAccount
from inventory.models import Item, Warehouse
from sales.models import Customer
from hatchery_master.models import Hatchery
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import timedelta
from django.utils import timezone
import json
import logging

logger = logging.getLogger(__name__)


def _uom_label(uom):
    """Display text for a UnitOfMeasurement FK (symbol if set, else name)."""
    if not uom:
        return ""
    return uom.symbol or uom.name

# Constants
STATES_AND_TERRITORIES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka",
    "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
    "Nagaland", "Odisha", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu",
    "Telangana", "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal",
    "Andaman and Nicobar Islands", "Chandigarh",
    "Dadra and Nagar Haveli and Daman and Diu", "Delhi", "Jammu and Kashmir",
    "Ladakh", "Lakshadweep", "Puducherry",
]

class BaseAPIView(View):
    """Base class for API views with common functionality."""
    
    def handle_exception(self, e: Exception) -> JsonResponse:
        """Handle common exceptions and return appropriate responses."""
        logger.error(f"Error in {self.__class__.__name__}: {str(e)}", exc_info=True)
        if isinstance(e, Http404):
            return JsonResponse({"error": str(e)}, status=404)
        if isinstance(e, ValidationError):
            if hasattr(e, "message_dict"):
                message = "; ".join(
                    f"{field}: {' '.join(msgs)}" for field, msgs in e.message_dict.items()
                )
            else:
                message = "; ".join(e.messages)
            return JsonResponse({"error": message}, status=400)
        return JsonResponse({"error": "Internal server error"}, status=500)

    def get_cached_data(self, cache_key: str, ttl: int = 300) -> Optional[List]:
        """Get data from cache or return None if not found."""
        return cache.get(cache_key)

    def set_cached_data(self, cache_key: str, data: List, ttl: int = 300) -> None:
        """Set data in cache with specified TTL."""
        cache.set(cache_key, data, ttl)

@method_decorator(login_required, name="dispatch")
class FarmerGroupTemplateView(View):
    """View for rendering the farmer group template."""

    def get(self, request):
        context = {"chart_of_accounts": ChartOfAccount.objects.filter(status='Active')}
        return render(request, "farmer_group.html", context)

@method_decorator(login_required, name="dispatch")
class FarmerGroupAPI(BaseAPIView):
    """API endpoints for FarmerGroup operations."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                fg = FarmerGroup.objects.get(id=id)
                return JsonResponse({
                    "id": fg.id,
                    "code": fg.code,
                    "description": fg.description,
                    "pay_account_id": fg.pay_account_id,
                    "pay_account_name": fg.pay_account.description,
                    "advance_account_id": fg.advance_account_id,
                    "advance_account_name": fg.advance_account.description,
                    "is_active": fg.is_active,
                    "is_locked": fg.is_locked,
                })

            farmer_groups = FarmerGroup.objects.select_related("pay_account", "advance_account").all()
            results = [{
                "id": fg.id,
                "code": fg.code,
                "description": fg.description,
                "pay_account_name": fg.pay_account.description,
                "advance_account_name": fg.advance_account.description,
                "is_active": fg.is_active,
                "is_locked": fg.is_locked,
            } for fg in farmer_groups]
            return JsonResponse(results, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        try:
            data = request.POST
            with transaction.atomic():
                FarmerGroup.objects.create(
                    description=data["description"],
                    pay_account_id=data["pay_account"],
                    advance_account_id=data["advance_account"],
                )
            return JsonResponse({"message": "Farmer group created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            fg = FarmerGroup.objects.get(id=id)
            if fg.is_locked:
                return JsonResponse({"error": "This farmer group is locked."}, status=400)
            data = json.loads(request.body)
            with transaction.atomic():
                fg.description = data["description"]
                fg.pay_account_id = data["pay_account"]
                fg.advance_account_id = data["advance_account"]
                fg.save()
            return JsonResponse({"message": "Farmer group updated"})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            fg = FarmerGroup.objects.get(id=id)
            if fg.is_locked:
                return JsonResponse({"error": "This farmer group is locked."}, status=400)
            with transaction.atomic():
                fg.delete()
            return JsonResponse({"message": "Farmer group deleted"})
        except Exception as e:
            return self.handle_exception(e)


@login_required
def toggle_farmer_group_active(request, id):
    """Toggle a farmer group's active/inactive status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        fg = FarmerGroup.objects.get(id=id)
        if fg.is_locked:
            return JsonResponse({"error": "This farmer group is locked."}, status=400)
        fg.is_active = not fg.is_active
        fg.save(update_fields=["is_active"])
        return JsonResponse({"message": "Farmer group updated", "is_active": fg.is_active})
    except FarmerGroup.DoesNotExist:
        return JsonResponse({"error": "Farmer group not found."}, status=404)


@login_required
def toggle_farmer_group_lock(request, id):
    """Toggle a farmer group's locked status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        fg = FarmerGroup.objects.get(id=id)
        fg.is_locked = not fg.is_locked
        fg.save(update_fields=["is_locked"])
        return JsonResponse({"message": "Farmer group updated", "is_locked": fg.is_locked})
    except FarmerGroup.DoesNotExist:
        return JsonResponse({"error": "Farmer group not found."}, status=404)


@method_decorator(login_required, name="dispatch")
class RegionTemplateView(View):
    """View for rendering the region template."""

    def get(self, request):
        context = {"states_and_union_territories": STATES_AND_TERRITORIES}
        return render(request, "region.html", context)

@method_decorator(login_required, name="dispatch")
class RegionAPI(BaseAPIView):
    """API endpoints for Region operations."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                region = Region.objects.get(id=id)
                return JsonResponse({
                    "id": region.id,
                    "code": region.code,
                    "description": region.description,
                    "is_active": region.is_active,
                    "is_locked": region.is_locked,
                })

            regions = Region.objects.all()
            results = [{
                "id": region.id,
                "code": region.code,
                "description": region.description,
                "is_active": region.is_active,
                "is_locked": region.is_locked,
            } for region in regions]
            return JsonResponse(results, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        try:
            data = request.POST
            with transaction.atomic():
                Region.objects.create(description=data["description"])
            return JsonResponse({"message": "Region created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            region = Region.objects.get(id=id)
            if region.is_locked:
                return JsonResponse({"error": "This region is locked."}, status=400)
            data = json.loads(request.body)
            with transaction.atomic():
                region.description = data["description"]
                region.save()
            return JsonResponse({"message": "Region updated"})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            region = Region.objects.get(id=id)
            if region.is_locked:
                return JsonResponse({"error": "This region is locked."}, status=400)
            with transaction.atomic():
                region.delete()
            return JsonResponse({"message": "Region deleted"})
        except Exception as e:
            return self.handle_exception(e)


@login_required
def toggle_region_active(request, id):
    """Toggle a region's active/inactive status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        region = Region.objects.get(id=id)
        if region.is_locked:
            return JsonResponse({"error": "This region is locked."}, status=400)
        region.is_active = not region.is_active
        region.save(update_fields=["is_active"])
        return JsonResponse({"message": "Region updated", "is_active": region.is_active})
    except Region.DoesNotExist:
        return JsonResponse({"error": "Region not found."}, status=404)


@login_required
def toggle_region_lock(request, id):
    """Toggle a region's locked status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        region = Region.objects.get(id=id)
        region.is_locked = not region.is_locked
        region.save(update_fields=["is_locked"])
        return JsonResponse({"message": "Region updated", "is_locked": region.is_locked})
    except Region.DoesNotExist:
        return JsonResponse({"error": "Region not found."}, status=404)


@method_decorator(login_required, name="dispatch")
class BreedTemplateView(View):
    """View for rendering the breed template."""

    def get(self, request):
        return render(request, "breed.html", {
            "categories": BirdCategory.objects.filter(is_active=True).order_by("sort_order", "name"),
        })


@method_decorator(login_required, name="dispatch")
class BreedAPI(BaseAPIView):
    """API endpoints for Breed operations."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                breed = Breed.objects.select_related("bird_category").get(id=id)
                return JsonResponse({
                    "id": breed.id,
                    "code": breed.code,
                    "description": breed.description,
                    "bird_category": breed.bird_category_id,
                    "bird_category_name": breed.bird_category.name if breed.bird_category_id else "",
                    "is_active": breed.is_active,
                    "is_locked": breed.is_locked,
                })

            breeds = Breed.objects.select_related("bird_category").all()
            results = [{
                "id": breed.id,
                "code": breed.code,
                "description": breed.description,
                "bird_category": breed.bird_category_id,
                "bird_category_name": breed.bird_category.name if breed.bird_category_id else "",
                "is_active": breed.is_active,
                "is_locked": breed.is_locked,
            } for breed in breeds]
            return JsonResponse(results, safe=False)
        except Breed.DoesNotExist:
            return JsonResponse({"error": "Breed not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        try:
            data = request.POST
            with transaction.atomic():
                Breed.objects.create(description=data["description"],
                                     bird_category_id=data.get("bird_category") or None)
            return JsonResponse({"message": "Breed created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            breed = Breed.objects.get(id=id)
            if breed.is_locked:
                return JsonResponse({"error": "This breed is locked."}, status=400)
            data = json.loads(request.body)
            with transaction.atomic():
                breed.description = data["description"]
                if "bird_category" in data:
                    breed.bird_category_id = data["bird_category"] or None
                breed.save()
            return JsonResponse({"message": "Breed updated"})
        except Breed.DoesNotExist:
            return JsonResponse({"error": "Breed not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            breed = Breed.objects.get(id=id)
            if breed.is_locked:
                return JsonResponse({"error": "This breed is locked."}, status=400)
            with transaction.atomic():
                breed.delete()
            return JsonResponse({"message": "Breed deleted"})
        except Breed.DoesNotExist:
            return JsonResponse({"error": "Breed not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)


@login_required
def toggle_breed_active(request, id):
    """Toggle a breed's active/inactive status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        breed = Breed.objects.get(id=id)
        if breed.is_locked:
            return JsonResponse({"error": "This breed is locked."}, status=400)
        breed.is_active = not breed.is_active
        breed.save(update_fields=["is_active"])
        return JsonResponse({"message": "Breed updated", "is_active": breed.is_active})
    except Breed.DoesNotExist:
        return JsonResponse({"error": "Breed not found."}, status=404)


@login_required
def toggle_breed_lock(request, id):
    """Toggle a breed's locked status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        breed = Breed.objects.get(id=id)
        breed.is_locked = not breed.is_locked
        breed.save(update_fields=["is_locked"])
        return JsonResponse({"message": "Breed updated", "is_locked": breed.is_locked})
    except Breed.DoesNotExist:
        return JsonResponse({"error": "Breed not found."}, status=404)


def _to_decimal(value):
    """Parse a user-supplied number, treating blanks as 0."""
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def _fmt(value):
    """Render a Decimal without trailing zeros (12.00 -> '12', 0.240 -> '0.24')."""
    s = f"{Decimal(value):f}"
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"


@method_decorator(login_required, name="dispatch")
class BreedStandardTemplateView(View):
    """View for rendering the breed standard template."""

    def get(self, request):
        context = {"breeds": Breed.objects.filter(is_active=True)}
        return render(request, "breed_standard.html", context)


def _breed_standard_row(bs):
    return {
        "id": bs.id,
        "code": bs.code,
        "age": bs.age,
        "body_weight": _fmt(bs.body_weight),
        "feed_intake": _fmt(bs.feed_intake),
        "avg_daily_gain": _fmt(bs.avg_daily_gain),
        "fcr": _fmt(bs.fcr),
        "cum_feed": _fmt(bs.cum_feed),
        "is_active": bs.is_active,
        "is_locked": bs.is_locked,
    }


@method_decorator(login_required, name="dispatch")
class BreedStandardAPI(BaseAPIView):
    """API endpoints for Breed Standard operations. The list is grouped
    into one folder per breed."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                bs = BreedStandard.objects.select_related("breed").get(id=id)
                data = _breed_standard_row(bs)
                data.update({"breed_id": bs.breed_id, "breed_name": bs.breed.description})
                return JsonResponse(data)

            # Flat list, ordered by breed then age (keeps each breed's rows together).
            rows = (
                BreedStandard.objects.select_related("breed")
                .order_by("breed__code", "age")
            )
            results = []
            for bs in rows:
                data = _breed_standard_row(bs)
                data.update({
                    "breed_id": bs.breed_id,
                    "breed_code": bs.breed.code,
                    "breed_name": bs.breed.description,
                })
                results.append(data)
            return JsonResponse(results, safe=False)
        except BreedStandard.DoesNotExist:
            return JsonResponse({"error": "Breed standard not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            bs = BreedStandard.objects.get(id=id)
            if bs.is_locked:
                return JsonResponse({"error": "This breed standard row is locked."}, status=400)
            with transaction.atomic():
                bs.delete()
            return JsonResponse({"message": "Breed standard deleted"})
        except BreedStandard.DoesNotExist:
            return JsonResponse({"error": "Breed standard not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)


@login_required
def breed_standard_by_breed(request, breed_id: int):
    """Return all standard rows for a breed (used to load the edit form)."""
    try:
        breed = Breed.objects.get(id=breed_id)
    except Breed.DoesNotExist:
        return JsonResponse({"error": "Breed not found."}, status=404)
    rows = [
        _breed_standard_row(bs)
        for bs in BreedStandard.objects.filter(breed=breed).order_by("age")
    ]
    return JsonResponse({
        "breed_id": breed.id,
        "breed_name": breed.description,
        "rows": rows,
    })


@login_required
def save_breed_standard(request):
    """Create/replace the whole standard curve for a breed. Rows are
    re-numbered by age (1..N) in submitted order."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        data = json.loads(request.body or "{}")
        breed_id = data.get("breed_id")
        rows = data.get("rows") or []
        if not breed_id:
            return JsonResponse({"error": "Please select a breed."}, status=400)
        if not rows:
            return JsonResponse({"error": "Add at least one row."}, status=400)
        try:
            breed = Breed.objects.get(id=breed_id)
        except Breed.DoesNotExist:
            return JsonResponse({"error": "Invalid breed."}, status=400)

        existing = BreedStandard.objects.filter(breed=breed)
        if existing.filter(is_locked=True).exists():
            return JsonResponse(
                {"error": "This breed has locked rows. Unlock them before editing."},
                status=400,
            )

        # Resolve ages first (Age is editable; blanks default to row position) and
        # reject duplicates before we touch the DB, so nothing is half-written.
        ages = []
        seen = set()
        for idx, row in enumerate(rows, start=1):
            try:
                age = int(row.get("age"))
            except (TypeError, ValueError):
                age = idx
            if age <= 0:
                age = idx
            if age in seen:
                return JsonResponse(
                    {"error": f"Duplicate age {age}. Each row must have a unique age."},
                    status=400,
                )
            seen.add(age)
            ages.append(age)

        with transaction.atomic():
            existing.delete()
            cum_running = Decimal("0")  # running total, anchored to whatever is stored
            prev_body_weight = None     # previous row's body weight, for Avg Daily Gain
            for age, row in zip(ages, rows):
                body_weight = _to_decimal(row.get("body_weight"))
                feed_intake = _to_decimal(row.get("feed_intake"))

                # Cum. Feed: manual override if given, else previous cumulative + feed intake
                cum_raw = row.get("cum_feed")
                if cum_raw not in (None, ""):
                    cum_feed = _to_decimal(cum_raw)
                else:
                    cum_feed = cum_running + feed_intake
                cum_running = cum_feed

                # FCR: manual override if given, else cumulative feed / body weight
                fcr_raw = row.get("fcr")
                if fcr_raw not in (None, ""):
                    fcr = _to_decimal(fcr_raw)
                elif body_weight > 0:
                    fcr = (cum_feed / body_weight).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                else:
                    fcr = Decimal("0")

                # Avg Daily Gain: manual override if given, else this row's body
                # weight minus the previous row's (the first row has no previous
                # weight to diff against, so it falls back to 0 unless overridden).
                gain_raw = row.get("avg_daily_gain")
                if gain_raw not in (None, ""):
                    avg_daily_gain = _to_decimal(gain_raw)
                elif prev_body_weight is not None:
                    avg_daily_gain = body_weight - prev_body_weight
                else:
                    avg_daily_gain = Decimal("0")
                prev_body_weight = body_weight

                BreedStandard.objects.create(
                    breed=breed,
                    age=age,
                    body_weight=body_weight,
                    feed_intake=feed_intake,
                    avg_daily_gain=avg_daily_gain,
                    fcr=fcr,
                    cum_feed=cum_feed,
                )
        return JsonResponse({"message": "Breed standard saved"}, status=201)
    except Exception as e:
        logger.error(f"Error in save_breed_standard: {str(e)}", exc_info=True)
        return JsonResponse({"error": "Internal server error"}, status=500)


@login_required
def toggle_breed_standard_active(request, id):
    """Toggle a breed standard row's active/inactive status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        bs = BreedStandard.objects.get(id=id)
        if bs.is_locked:
            return JsonResponse({"error": "This breed standard row is locked."}, status=400)
        bs.is_active = not bs.is_active
        bs.save(update_fields=["is_active"])
        return JsonResponse({"message": "Breed standard updated", "is_active": bs.is_active})
    except BreedStandard.DoesNotExist:
        return JsonResponse({"error": "Breed standard not found."}, status=404)


@login_required
def toggle_breed_standard_lock(request, id):
    """Toggle a breed standard row's locked status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        bs = BreedStandard.objects.get(id=id)
        bs.is_locked = not bs.is_locked
        bs.save(update_fields=["is_locked"])
        return JsonResponse({"message": "Breed standard updated", "is_locked": bs.is_locked})
    except BreedStandard.DoesNotExist:
        return JsonResponse({"error": "Breed standard not found."}, status=404)


@method_decorator(login_required, name="dispatch")
class BranchTemplateView(View):
    """View for rendering the branch template."""

    def get(self, request):
        context = {
            "regions": Region.objects.filter(is_active=True),
        }
        return render(request, "branch.html", context)

def _branch_to_dict(branch: Branch) -> dict:
    return {
        "id": branch.id,
        "code": branch.code,
        "branch_name": branch.branch_name,
        "region_id": branch.region_id,
        "region_name": branch.region.description,
        "prefix": branch.prefix,
        "is_active": branch.is_active,
        "is_locked": branch.is_locked,
    }


@method_decorator(login_required, name="dispatch")
class BranchAPI(BaseAPIView):
    """API endpoints for Branch operations."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                branch = Branch.objects.select_related("region").get(id=id)
                return JsonResponse(_branch_to_dict(branch))

            branches = Branch.objects.select_related("region").all()
            return JsonResponse([_branch_to_dict(b) for b in branches], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        """Create one or more branches against a single region in one submit."""
        try:
            data = request.POST
            region = Region.objects.filter(id=data.get("region")).first()
            if not region:
                return JsonResponse({"error": "Select a valid region."}, status=400)

            branch_names = data.getlist("branch_name[]")
            prefixes = data.getlist("prefix[]")

            # Validate every row before creating anything.
            rows = []
            for branch_name, prefix in zip(branch_names, prefixes):
                branch_name = branch_name.strip()
                prefix = prefix.strip()
                if not branch_name and not prefix:
                    continue
                if not branch_name or not prefix:
                    return JsonResponse({"error": "Enter both a branch name and a prefix for every row."}, status=400)
                rows.append((branch_name, prefix))

            if not rows:
                return JsonResponse({"error": "Enter at least one branch name and prefix."}, status=400)

            created = 0
            with transaction.atomic():
                for branch_name, prefix in rows:
                    Branch.objects.create(region=region, branch_name=branch_name, prefix=prefix)
                    created += 1
                cache.delete("branch_list")

            return JsonResponse({"message": f"{created} branch(es) created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        """Update the branch."""
        try:
            branch = Branch.objects.get(id=id)
            if branch.is_locked:
                return JsonResponse({"error": "This branch is locked."}, status=400)
            data = json.loads(request.body)
            region = Region.objects.filter(id=data.get("region")).first()
            if not region:
                return JsonResponse({"error": "Select a valid region."}, status=400)

            with transaction.atomic():
                branch.region = region
                branch.branch_name = data["branch_name"]
                branch.prefix = data["prefix"]
                branch.save()
                cache.delete("branch_list")
            return JsonResponse({"message": "Branch updated"})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            branch = Branch.objects.get(id=id)
            if branch.is_locked:
                return JsonResponse({"error": "This branch is locked."}, status=400)
            with transaction.atomic():
                branch.delete()
                cache.delete("branch_list")
            return JsonResponse({"message": "Branch deleted"})
        except Exception as e:
            return self.handle_exception(e)


@login_required
def toggle_branch_active(request, id):
    """Toggle a branch's active/inactive status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        branch = Branch.objects.get(id=id)
        if branch.is_locked:
            return JsonResponse({"error": "This branch is locked."}, status=400)
        branch.is_active = not branch.is_active
        branch.save(update_fields=["is_active"])
        cache.delete("branch_list")
        return JsonResponse({"message": "Branch updated", "is_active": branch.is_active})
    except Branch.DoesNotExist:
        return JsonResponse({"error": "Branch not found."}, status=404)


@login_required
def toggle_branch_lock(request, id):
    """Toggle a branch's locked status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        branch = Branch.objects.get(id=id)
        branch.is_locked = not branch.is_locked
        branch.save(update_fields=["is_locked"])
        cache.delete("branch_list")
        return JsonResponse({"message": "Branch updated", "is_locked": branch.is_locked})
    except Branch.DoesNotExist:
        return JsonResponse({"error": "Branch not found."}, status=404)

@method_decorator(login_required, name="dispatch")
class SupervisorTemplateView(View):
    """View for rendering the supervisor template."""
    
    def get(self, request):
        from hr.models import Employee

        # Cached per scope, not globally: a single "branch_list" key would hand
        # one user's branches to the next request from another.
        from user.services.dashboard_widgets import _scope_signature

        cache_key = "branch_list:%s" % _scope_signature(request.user)
        branches = cache.get(cache_key)
        if not branches:
            branches = list(branches_for(request.user, Branch.objects.all()).values())
            cache.set(cache_key, branches)
        context = {
            "branches": branches,
            "employees": Employee.objects.filter(relieve=False).order_by("full_name"),
        }
        return render(request, "supervisor.html", context)

@method_decorator(login_required, name="dispatch")
class BroilerLineTemplateView(View):
    """View for rendering the broiler line template."""

    def get(self, request):
        context = {"regions": Region.objects.filter(is_active=True)}
        return render(request, "broiler_line.html", context)

@method_decorator(login_required, name="dispatch")
class BroilerFarmTemplateView(View):
    """View for rendering the broiler farm template (Add Farmer + Add Farm tabs)."""

    def get(self, request):
        farmers = list(Farmer.objects.values("id", "farmer_name"))
        context = {
            "regions": Region.objects.filter(is_active=True),
            "farmers": farmers,
            "farmer_groups": FarmerGroup.objects.filter(is_active=True),
            "shed_types": BroilerFarmShed.SHED_TYPE_CHOICES,
        }
        return render(request, "broiler_farm.html", context)

def batch_shed_options(farm_id=None) -> list[dict]:
    """The Shed / Unit picker for a new batch, as both front ends need it.

    Every shed, deliberately NOT filtered on ``is_active``.
    ``BroilerFarmShed.is_active`` is occupancy-driven ("has birds in it"), and
    birds only arrive through chicks placement on a batch, so filtering on it
    would hide exactly the vacant sheds a new batch needs. Sheds already
    holding an open batch are flagged ``occupied`` instead, and the caller
    renders them disabled — the same rule the create endpoint enforces, so a
    picker cannot offer what the save will refuse.

    Lives here rather than in each front end because the browser and the phone
    have to agree about which units are free; two copies of that would not
    stay in step.
    """
    occupied = dict(
        BroilerBatch.objects.filter(
            shed__isnull=False, end_date__isnull=True, is_closed=False
        ).values_list("shed_id", "batch_name")
    )
    sheds = BroilerFarmShed.objects.order_by("farm__farm_code", "unit_no")
    if farm_id:
        sheds = sheds.filter(farm_id=farm_id)
    return [
        {
            "id": s.id,
            "farm_id": s.farm_id,
            "label": (s.shed_name or s.shed_code or f"Unit {s.unit_no}")
            + (f" · Unit {s.unit_no}" if s.unit_no else ""),
            "occupied": s.id in occupied,
            # Which batch is in the way, so the phone can say so rather than
            # leaving someone to guess why a unit is greyed out.
            "occupied_by": occupied.get(s.id, ""),
        }
        for s in sheds
    ]


@method_decorator(login_required, name="dispatch")
class BroilerBatchTemplateView(View):
    """View for rendering the broiler batch template."""

    def get(self, request):
        cache_key = "broiler_farm_list"
        broiler_farms = cache.get(cache_key)
        if not broiler_farms:
            broiler_farms = list(BroilerFarm.objects.values())
            cache.set(cache_key, broiler_farms)
        context = {
            "broiler_farms": broiler_farms,
            "breeds": Breed.objects.filter(is_active=True).order_by("description"),
            "sheds": batch_shed_options(),
        }
        return render(request, "broiler_batch.html", context)

@method_decorator(login_required, name="dispatch")
class BroilerDiseaseTemplateView(View):
    """View for rendering the broiler disease template."""
    
    def get(self, request):
        cache_key = "broiler_farm_list"
        broiler_farms = cache.get(cache_key)
        if not broiler_farms:
            broiler_farms = list(BroilerFarm.objects.values())
            cache.set(cache_key, broiler_farms)
        context = {"broiler_farms": broiler_farms}
        return render(request, "broiler_disease.html", context)

@method_decorator(login_required, name="dispatch")
class SupervisorAPI(BaseAPIView):
    """API endpoints for Supervisor operations."""
    
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                supervisor = Supervisor.objects.select_related("branch", "employee").get(id=id)
                return JsonResponse({
                    "id": supervisor.id,
                    "supervisor": supervisor.name,
                    "branch": supervisor.branch_id,
                    "branch_name": supervisor.branch.branch_name,
                    "employee": supervisor.employee_id,
                })

            cache_key = "supervisor_list"
            cached_data = self.get_cached_data(cache_key)
            if cached_data:
                return JsonResponse(cached_data, safe=False)

            supervisors = list(
                Supervisor.objects.select_related("branch")
                .annotate(branch_name=F("branch__branch_name"))
                .values("id", "name", "branch_name", "branch", "employee")
            )
            self.set_cached_data(cache_key, supervisors)
            return JsonResponse(supervisors, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @staticmethod
    def _apply_employee(supervisor, employee):
        """Supervisor's name/phone/address are always a copy of the chosen
        HR Employee record — Supervisor is picked from the Employee list,
        not typed in freehand."""
        supervisor.employee = employee
        supervisor.name = employee.full_name or ""
        supervisor.phone_no = str(employee.personal_contact) if employee.personal_contact else ""
        supervisor.address = employee.correspondence_address or ""

    def post(self, request) -> JsonResponse:
        from hr.models import Employee
        try:
            data = json.loads(request.body or "{}")
            try:
                branch = Branch.objects.get(id=data.get("branch"))
            except Branch.DoesNotExist:
                return JsonResponse({"error": "Invalid branch"}, status=400)
            try:
                employee = Employee.objects.get(id=data.get("employee"))
            except Employee.DoesNotExist:
                return JsonResponse({"error": "Select an employee"}, status=400)
            if Supervisor.objects.filter(employee=employee).exists():
                return JsonResponse({"error": f"{employee.full_name} is already mapped to a supervisor."}, status=400)
            with transaction.atomic():
                supervisor = Supervisor(branch=branch)
                self._apply_employee(supervisor, employee)
                supervisor.save()
                cache.delete("supervisor_list")
            return JsonResponse({"message": "Supervisor created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        from hr.models import Employee
        try:
            supervisor = Supervisor.objects.get(id=id)
            data = json.loads(request.body or "{}")
            with transaction.atomic():
                if data.get("branch"):
                    supervisor.branch_id = data["branch"]
                if data.get("employee"):
                    try:
                        employee = Employee.objects.get(id=data["employee"])
                    except Employee.DoesNotExist:
                        return JsonResponse({"error": "Invalid employee"}, status=400)
                    if Supervisor.objects.filter(employee=employee).exclude(pk=supervisor.pk).exists():
                        return JsonResponse({"error": f"{employee.full_name} is already mapped to another supervisor."}, status=400)
                    self._apply_employee(supervisor, employee)
                supervisor.save()
                cache.delete("supervisor_list")
            return JsonResponse({"message": "Supervisor updated"})
        except Supervisor.DoesNotExist:
            raise Http404("Supervisor not found")
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            supervisor = Supervisor.objects.get(id=id)
            with transaction.atomic():
                supervisor.delete()
                cache.delete("supervisor_list")
            return JsonResponse({"message": "Supervisor deleted"})
        except Exception as e:
            return self.handle_exception(e)

def _broiler_line_to_dict(line: BroilerLine) -> dict:
    return {
        "id": line.id,
        "code": line.code,
        "description": line.description,
        "region_id": line.region_id,
        "region_name": line.region.description,
        "branch_id": line.branch_id,
        "branch_name": line.branch.branch_name,
        "is_active": line.is_active,
        "is_locked": line.is_locked,
    }

@method_decorator(login_required, name="dispatch")
class BroilerLineAPI(BaseAPIView):
    """API endpoints for BroilerLine operations."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                line = BroilerLine.objects.select_related("region", "branch").get(id=id)
                return JsonResponse(_broiler_line_to_dict(line))

            lines = BroilerLine.objects.select_related("region", "branch").all()
            return JsonResponse([_broiler_line_to_dict(l) for l in lines], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        """Create one or more lines against a single region/branch in one submit."""
        try:
            data = request.POST
            region = Region.objects.filter(id=data.get("region")).first()
            branch = Branch.objects.filter(id=data.get("branch")).first()
            if not region or not branch:
                return JsonResponse({"error": "Select a valid region and branch."}, status=400)

            descriptions = data.getlist("description[]")

            created = 0
            with transaction.atomic():
                for description in descriptions:
                    description = description.strip()
                    if not description:
                        continue
                    BroilerLine.objects.create(region=region, branch=branch, description=description)
                    created += 1

            if not created:
                return JsonResponse({"error": "Enter at least one line."}, status=400)
            return JsonResponse({"message": f"{created} line(s) created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            line = BroilerLine.objects.get(id=id)
            if line.is_locked:
                return JsonResponse({"error": "This line is locked."}, status=400)
            data = json.loads(request.body)
            region = Region.objects.filter(id=data.get("region")).first()
            branch = Branch.objects.filter(id=data.get("branch")).first()
            if not region or not branch:
                return JsonResponse({"error": "Select a valid region and branch."}, status=400)
            with transaction.atomic():
                line.region = region
                line.branch = branch
                line.description = data["description"]
                line.save()
            return JsonResponse({"message": "Line updated"})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            line = BroilerLine.objects.get(id=id)
            if line.is_locked:
                return JsonResponse({"error": "This line is locked."}, status=400)
            with transaction.atomic():
                line.delete()
            return JsonResponse({"message": "Line deleted"})
        except Exception as e:
            return self.handle_exception(e)


@login_required
def toggle_broiler_line_active(request, id):
    """Toggle a broiler line's active/inactive status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        line = BroilerLine.objects.get(id=id)
        if line.is_locked:
            return JsonResponse({"error": "This line is locked."}, status=400)
        line.is_active = not line.is_active
        line.save(update_fields=["is_active"])
        return JsonResponse({"message": "Line updated", "is_active": line.is_active})
    except BroilerLine.DoesNotExist:
        return JsonResponse({"error": "Line not found."}, status=404)


@login_required
def toggle_broiler_line_lock(request, id):
    """Toggle a broiler line's locked status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        line = BroilerLine.objects.get(id=id)
        line.is_locked = not line.is_locked
        line.save(update_fields=["is_locked"])
        return JsonResponse({"message": "Line updated", "is_locked": line.is_locked})
    except BroilerLine.DoesNotExist:
        return JsonResponse({"error": "Line not found."}, status=404)


@login_required
def get_branches_by_region(request) -> JsonResponse:
    """Get branches for a specific region."""
    try:
        region_id = request.GET.get('region_id')
        # Shared by many forms, so it stays reachable for any signed-in user —
        # but it must not list branches outside their scope.
        branches = list(branches_for(request.user, Branch.objects.filter(
            region_id=region_id, is_active=True)).values('id', 'branch_name'))
        return JsonResponse({'branches': branches})
    except Exception as e:
        logger.error(f"Error in get_branches_by_region: {str(e)}", exc_info=True)
        return JsonResponse({"error": str(e)}, status=500)

@login_required
def get_lines_by_branch(request) -> JsonResponse:
    """Get broiler lines for a specific branch."""
    try:
        branch_id = request.GET.get('branch_id')
        if not branches_for(request.user).filter(id=branch_id).exists():
            return JsonResponse({'lines': []})
        lines = list(BroilerLine.objects.filter(branch_id=branch_id, is_active=True).values('id', 'description'))
        return JsonResponse({'lines': lines})
    except Exception as e:
        logger.error(f"Error in get_lines_by_branch: {str(e)}", exc_info=True)
        return JsonResponse({"error": str(e)}, status=500)

@method_decorator(login_required, name="dispatch")
class FarmerAPI(BaseAPIView):
    """API endpoints for Farmer operations."""

    FILE_FIELDS = ["farmer_photo", "pan_upload", "aadhar_upload_front", "aadhar_upload_back"]
    FORM_FIELDS = [
        "farmer_name", "phone_no", "mobile_no", "mobile_2", "pan_no", "aadhar_no",
        "national_id", "usc", "service_no", "tds_percent",
        "account_holder_name", "acc_no", "ifsc_code", "bank_name", "bank_branch", "address",
    ]

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                farmer = Farmer.objects.select_related("farmer_group").get(id=id)
                data = {field: getattr(farmer, field) for field in self.FORM_FIELDS}
                data["id"] = farmer.id
                data["tds_percent"] = str(farmer.tds_percent) if farmer.tds_percent is not None else None
                data["farmer_group_id"] = farmer.farmer_group_id
                data["farmer_group"] = farmer.farmer_group.description if farmer.farmer_group_id else None
                for field in self.FILE_FIELDS:
                    file_obj = getattr(farmer, field)
                    data[field] = file_obj.url if file_obj else None
                return JsonResponse(data)

            cache_key = "farmer_list"
            cached_data = self.get_cached_data(cache_key)
            if cached_data:
                return JsonResponse(cached_data, safe=False)

            farmers = list(
                Farmer.objects.select_related("farmer_group").values(
                    "id", "farmer_name", "mobile_no", "usc", "service_no",
                    farmer_group_name=F("farmer_group__description"),
                )
            )
            self.set_cached_data(cache_key, farmers)
            return JsonResponse(farmers, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            data = request.POST
            with transaction.atomic():
                farmer = Farmer.objects.get(id=id) if id else Farmer()
                for field in self.FORM_FIELDS:
                    if field not in data:
                        continue
                    setattr(farmer, field, data[field] or None if field == "tds_percent" else data[field])
                if "farmer_group" in data:
                    farmer.farmer_group_id = data["farmer_group"] or None
                for field in self.FILE_FIELDS:
                    if field in request.FILES:
                        setattr(farmer, field, request.FILES[field])
                farmer.full_clean(exclude=self.FILE_FIELDS)
                farmer.save()
                cache.delete("farmer_list")
            return JsonResponse(
                {"message": "Farmer updated" if id else "Farmer created", "id": farmer.id},
                status=200 if id else 201,
            )
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            farmer = Farmer.objects.get(id=id)
            with transaction.atomic():
                farmer.delete()
                cache.delete("farmer_list")
            return JsonResponse({"message": "Farmer deleted"})
        except Exception as e:
            return self.handle_exception(e)

def _active_batch_on_shed(shed_id, exclude_batch_id=None):
    """Return the open/active batch occupying `shed_id`, if any.
    A batch is 'active' while it is still growing — end_date not set and not
    yet closed by a growing-charge settlement. A shed can hold only one at a
    time, so this gates creation / re-assignment onto an occupied unit."""
    if not shed_id:
        return None
    qs = BroilerBatch.objects.filter(
        shed_id=shed_id, end_date__isnull=True, is_closed=False
    )
    if exclude_batch_id:
        qs = qs.exclude(id=exclude_batch_id)
    return qs.first()


@method_decorator(login_required, name="dispatch")
class BroilerBatchAPI(BaseAPIView):
    """API endpoints for BroilerBatch operations."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                broiler_batch = BroilerBatch.objects.select_related("broiler_farm").get(id=id)
                return JsonResponse({
                    "id": broiler_batch.id,
                    "name": broiler_batch.batch_name,
                    "book_number": broiler_batch.book_number,
                    "lot_no": broiler_batch.lot_no,
                    "breed_id": broiler_batch.breed_id,
                    "shed_id": broiler_batch.shed_id,
                    "broiler_farm_id": broiler_batch.broiler_farm_id,
                    "broiler_farm_name": broiler_batch.broiler_farm.farm_name,
                })

            cache_key = "broiler_batch_list"
            cached_data = self.get_cached_data(cache_key)
            if cached_data:
                return JsonResponse(cached_data, safe=False)

            broiler_batches = list(
                BroilerBatch.objects.select_related("broiler_farm", "breed", "shed")
                .annotate(broiler_farm_name=F("broiler_farm__farm_name"),
                          breed_name=F("breed__description"),
                          shed_name=F("shed__shed_name"),
                          shed_code=F("shed__shed_code"),
                          shed_unit_no=F("shed__unit_no"))
                .values("id", "batch_name", "book_number", "lot_no", "broiler_farm_name",
                        "broiler_farm_id", "breed_id", "breed_name",
                        "shed_id", "shed_name", "shed_code", "shed_unit_no",
                        "created_at", "end_date", "is_closed", "closed_on")
            )
            self.set_cached_data(cache_key, broiler_batches)
            return JsonResponse(broiler_batches, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        try:
            data = request.POST
            farm_obj = BroilerFarm.objects.get(id=data["broiler_farm_id"])
            shed_id = data.get("shed") or None
            # A flock is housed somewhere and is of some breed: the shed is
            # what occupancy, placement and the growing charge all hang off,
            # and the breed is what the daily numbers are judged against.
            # Enforced here rather than only in the markup, so the browser
            # form, the inline edit and the phone all get the same answer.
            missing = [name for name, value in
                       (("shed", shed_id), ("breed", data.get("breed") or None))
                       if not value]
            if missing:
                return JsonResponse(
                    {"error": "Choose the "
                              + " and the ".join({"shed": "shed / unit",
                                                  "breed": "breed"}[m]
                                                 for m in missing) + "."},
                    status=400,
                )
            occupied_by = _active_batch_on_shed(shed_id)
            if occupied_by:
                return JsonResponse(
                    {"error": f"This shed/unit already has an active batch "
                              f"({occupied_by.batch_name}). Close it before starting a new one."},
                    status=400,
                )
            with transaction.atomic():
                # batch_name is auto-generated (<farm code minus FRM/>-<n>,
                # e.g. BAH-0201-1) in BroilerBatch.save() — never accepted
                # from the form.
                batch = BroilerBatch.objects.create(
                    broiler_farm=farm_obj,
                    shed_id=shed_id,
                    book_number=data.get("book_number") or "",
                    lot_no=data.get("lot_no") or "",
                    breed_id=data.get("breed") or None,
                )
                cache.delete("broiler_batch_list")
            return JsonResponse({"message": "BroilerBatch created", "batch_name": batch.batch_name}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            broiler_batch = BroilerBatch.objects.get(id=id)
            data = json.loads(request.body or "{}")
            with transaction.atomic():
                # Only Book Number / Lot No are editable — batch_name is an
                # auto-generated code and the farm is fixed at creation.
                if "book_number" in data:
                    broiler_batch.book_number = data["book_number"] or ""
                if "lot_no" in data:
                    broiler_batch.lot_no = data["lot_no"] or ""
                # Sent-and-empty is someone clearing a field that a batch
                # cannot be without; a key simply absent is a partial update
                # and leaves what is already there alone.
                if "breed" in data and not data["breed"]:
                    return JsonResponse({"error": "Choose the breed."}, status=400)
                if "shed" in data and not data["shed"]:
                    return JsonResponse({"error": "Choose the shed / unit."},
                                        status=400)
                if "breed" in data:
                    broiler_batch.breed_id = data["breed"] or None
                if "shed" in data:
                    new_shed_id = data["shed"] or None
                    if new_shed_id and str(new_shed_id) != str(broiler_batch.shed_id):
                        occupied_by = _active_batch_on_shed(new_shed_id, exclude_batch_id=broiler_batch.id)
                        if occupied_by:
                            return JsonResponse(
                                {"error": f"This shed/unit already has an active batch "
                                          f"({occupied_by.batch_name}). Close it before moving a batch here."},
                                status=400,
                            )
                    broiler_batch.shed_id = new_shed_id
                broiler_batch.save()
                cache.delete("broiler_batch_list")
            return JsonResponse({"message": "BroilerBatch updated"})
        except BroilerBatch.DoesNotExist:
            raise Http404("Broiler batch not found")
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            broiler_batch = BroilerBatch.objects.get(id=id)
            with transaction.atomic():
                broiler_batch.delete()
                cache.delete("broiler_batch_list")
            return JsonResponse({"message": "BroilerBatch deleted"})
        except Exception as e:
            return self.handle_exception(e)

@method_decorator(login_required, name="dispatch")
class BroilerDiseaseAPI(BaseAPIView):
    """API endpoints for BroilerDisease operations."""
    
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                broiler_disease = BroilerDisease.objects.get(id=id)
                return JsonResponse({
                    "id": broiler_disease.id,
                    "disease_code": broiler_disease.disease_code,
                    "disease_name": broiler_disease.disease_name,
                    "symptoms": broiler_disease.symptoms,
                    "diagnosis": broiler_disease.diagnosis,
                    "image": broiler_disease.image.url if broiler_disease.image else None,
                })
            
            cache_key = "broiler_disease_list"
            cached_data = self.get_cached_data(cache_key)
            if cached_data:
                return JsonResponse(cached_data, safe=False)
            
            broiler_diseases = list(
                BroilerDisease.objects.values(
                    "id", "disease_code", "disease_name", "symptoms", "diagnosis", "image"
                )
            )
            for disease in broiler_diseases:
                disease["image"] = request.build_absolute_uri(disease["image"]) if disease["image"] else None
            
            self.set_cached_data(cache_key, broiler_diseases)
            return JsonResponse(broiler_diseases, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        try:
            data = request.POST
            image = request.FILES.get("image")
            with transaction.atomic():
                BroilerDisease.objects.create(
                    disease_code=data["disease_code"],
                    disease_name=data["disease_name"],
                    symptoms=data["symptoms"],
                    diagnosis=data["diagnosis"],
                    image=image,
                )
                cache.delete("broiler_disease_list")
            return JsonResponse({"message": "BroilerDisease created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            broiler_disease = BroilerDisease.objects.get(id=id)
            data = json.loads(request.body.decode("utf-8"))
            with transaction.atomic():
                broiler_disease.disease_code = data.get("disease_code", broiler_disease.disease_code)
                broiler_disease.disease_name = data.get("disease_name", broiler_disease.disease_name)
                broiler_disease.symptoms = data.get("symptoms", broiler_disease.symptoms)
                broiler_disease.diagnosis = data.get("diagnosis", broiler_disease.diagnosis)
                broiler_disease.save()
                cache.delete("broiler_disease_list")
            return JsonResponse({"message": "BroilerDisease updated"})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            broiler_disease = BroilerDisease.objects.get(id=id)
            with transaction.atomic():
                if broiler_disease.image:
                    default_storage.delete(broiler_disease.image.path)
                broiler_disease.delete()
                cache.delete("broiler_disease_list")
            return JsonResponse({"message": "BroilerDisease deleted"})
        except Exception as e:
            return self.handle_exception(e)

@method_decorator(login_required, name="dispatch")
class BroilerFarmShedTemplateView(View):
    """Broiler > Master > Farm Shed — manage the sheds that belong to each farm.
    Organization Centre / Branch / Company auto-fill from the chosen farm."""

    def get(self, request):
        from account.models import OrganizationCentre, CompanyProfile
        company_name = CompanyProfile.get_solo().name
        oc_by_branch = {oc.branch_id: oc for oc
                        in OrganizationCentre.objects.select_related("company").all()
                        if oc.branch_id}
        farms = []
        for f in BroilerFarm.objects.select_related("branch").order_by("farm_code"):
            oc = oc_by_branch.get(f.branch_id)
            farms.append({
                "id": f.id, "farm_code": f.farm_code, "farm_name": f.farm_name,
                "branch_name": f.branch.branch_name if f.branch_id else "",
                "org_centre_name": oc.name if oc else "",
                "company_name": (oc.company.name if oc and oc.company_id else company_name),
            })
        return render(request, "broiler_farm_shed.html", {
            "farms": farms,
            "shed_types": BroilerFarmShed.SHED_TYPE_CHOICES,
        })


def _farm_shed_to_dict(s: BroilerFarmShed) -> dict:
    from account.models import CompanyProfile
    oc = s.organization_centre
    if oc and oc.company_id:
        company_name = oc.company.name
    else:
        company_name = CompanyProfile.get_solo().name
    return {
        "id": s.id,
        "shed_code": s.shed_code,
        "shed_name": s.shed_name or "",
        "farm_id": s.farm_id,
        "farm_code": s.farm.farm_code if s.farm_id else "",
        "farm_name": s.farm.farm_name if s.farm_id else "",
        "org_centre_id": s.organization_centre_id,
        "org_centre_name": oc.name if oc else "",
        "branch_name": s.farm.branch.branch_name if s.farm_id and s.farm.branch_id else "",
        "company_name": company_name,
        "shed_type": s.shed_type,
        "shed_type_display": s.get_shed_type_display(),
        "unit_no": s.unit_no,
        "is_active": s.is_active,
        "length": str(s.length) if s.length is not None else "",
        "width": str(s.width) if s.width is not None else "",
        "dimensions": s.dimensions or "",
        "sq_feet": s.sq_feet or "",
        "capacity": s.capacity,
        "occupied": s.occupied,
        "free_space": s.free_space,
        "utilization_pct": s.utilization_pct,
    }


@method_decorator(login_required, name="dispatch")
class BroilerFarmShedAPI(BaseAPIView):
    """CRUD for BroilerFarmShed (Broiler > Master > Farm Shed)."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                s = BroilerFarmShed.objects.select_related("farm").get(id=id)
                return JsonResponse(_farm_shed_to_dict(s))
            sheds = (BroilerFarmShed.objects.select_related("farm")
                     .order_by("farm__farm_code", "shed_no"))
            return JsonResponse([_farm_shed_to_dict(s) for s in sheds], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    _VALID_TYPES = {c[0] for c in BroilerFarmShed.SHED_TYPE_CHOICES}

    def _clean(self, data):
        farm = BroilerFarm.objects.filter(id=data.get("farm")).first()
        if not farm:
            return None, "Select a valid farm."
        shed_name = (data.get("shed_name") or "").strip()  # blank -> auto in model.save()
        shed_type = (data.get("shed_type") or "broiler").strip()
        if shed_type not in self._VALID_TYPES:
            return None, "Select a valid shed type."

        def _int(v):
            try:
                return max(int(float(v)), 0)
            except (TypeError, ValueError):
                return 0

        def _dec(v):
            v = (str(v) if v is not None else "").strip()
            if not v:
                return None
            try:
                return max(Decimal(v), Decimal("0"))
            except (InvalidOperation, ValueError):
                return None
        capacity = _int(data.get("capacity"))
        occupied = _int(data.get("occupied"))
        if occupied > capacity:
            return None, "Occupied cannot exceed capacity."
        return {"farm": farm, "shed_name": shed_name, "shed_type": shed_type,
                "is_active": str(data.get("is_active", "true")).lower() in ("true", "1", "on", "yes"),
                "length": _dec(data.get("length")), "width": _dec(data.get("width")),
                "capacity": capacity, "occupied": occupied}, None

    def post(self, request) -> JsonResponse:
        try:
            vals, err = self._clean(request.POST)
            if err:
                return JsonResponse({"error": err}, status=400)
            with transaction.atomic():
                # shed_code, unit_no, organization_centre are auto-set in .save()
                BroilerFarmShed.objects.create(
                    farm=vals["farm"], shed_name=vals["shed_name"],
                    shed_no=vals["shed_name"], shed_type=vals["shed_type"],
                    is_active=vals["is_active"], length=vals["length"],
                    width=vals["width"], capacity=vals["capacity"],
                    occupied=vals["occupied"])
            return JsonResponse({"message": "Shed created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            s = BroilerFarmShed.objects.get(id=id)
            vals, err = self._clean(json.loads(request.body.decode("utf-8")))
            if err:
                return JsonResponse({"error": err}, status=400)
            with transaction.atomic():
                farm_changed = s.farm_id != vals["farm"].id
                s.farm = vals["farm"]
                s.shed_name = vals["shed_name"]
                s.shed_no = vals["shed_name"]
                s.shed_type = vals["shed_type"]
                s.is_active = vals["is_active"]
                s.length = vals["length"]
                s.width = vals["width"]
                s.capacity = vals["capacity"]
                s.occupied = vals["occupied"]
                if farm_changed:  # re-derive centre for the new farm's branch
                    s.organization_centre = None
                s.save()
            return JsonResponse({"message": "Shed updated"})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            s = BroilerFarmShed.objects.get(id=id)
            with transaction.atomic():
                s.delete()
            return JsonResponse({"message": "Shed deleted"})
        except Exception as e:
            return self.handle_exception(e)


@method_decorator(login_required, name="dispatch")
class BroilerFarmAPI(BaseAPIView):
    """API endpoints for BroilerFarm operations."""

    FILE_FIELDS = [
        "agreement_copy", "other_documents",
        "cheque_1_file", "cheque_2_file", "cheque_3_file", "cheque_4_file",
    ]
    # farm_code is auto-generated (<branch prefix>-<branch suffix><serial>,
    # e.g. AKB-0203) in BroilerFarm.save() — never accepted from the form.
    FORM_FIELDS = [
        "farm_name", "region", "line", "farm_pincode", "farm_capacity",
        "farm_type", "state", "district", "area", "farm_address", "farm_latitude",
        "farm_longitude", "agreement_start_date", "agreement_end_date", "agreement_months",
        "farm_sqft", "cheque_1_no", "cheque_2_no", "cheque_3_no", "cheque_4_no", "remarks",
    ]

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                broiler_farm = BroilerFarm.objects.select_related(
                    "branch", "supervisor", "farmer"
                ).prefetch_related("sheds", "images").get(id=id)
                data = {field: getattr(broiler_farm, field) for field in self.FORM_FIELDS}
                data.update({
                    "id": broiler_farm.id,
                    "farm_code": broiler_farm.farm_code,
                    "branch_id": broiler_farm.branch_id,
                    "supervisor_id": broiler_farm.supervisor_id,
                    "farmer_id": broiler_farm.farmer_id,
                    "agreement_start_date": broiler_farm.agreement_start_date.isoformat() if broiler_farm.agreement_start_date else None,
                    "agreement_end_date": broiler_farm.agreement_end_date.isoformat() if broiler_farm.agreement_end_date else None,
                    "sheds": list(broiler_farm.sheds.values(
                        "id", "shed_code", "shed_name", "shed_type",
                        "length", "width", "capacity", "sq_feet")),
                    "images": [{"id": img.id, "url": img.image.url} for img in broiler_farm.images.all()],
                })
                for field in self.FILE_FIELDS:
                    file_obj = getattr(broiler_farm, field)
                    data[field] = file_obj.url if file_obj else None
                return JsonResponse(data)

            cache_key = "broiler_farm_list"
            cached_data = self.get_cached_data(cache_key)
            if cached_data:
                return JsonResponse(cached_data, safe=False)

            broiler_farms = list(
                BroilerFarm.objects.select_related("branch", "supervisor", "farmer")
                .values(
                    "id", "farm_code", "farm_name", "region", "line", "farm_type",
                    "agreement_start_date", "agreement_end_date",
                    branch_name=F("branch__branch_name"),
                    supervisor_name=F("supervisor__name"),
                    farmer_name=F("farmer__farmer_name"),
                )
            )
            self.set_cached_data(cache_key, broiler_farms)
            return JsonResponse(broiler_farms, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def _save_sheds(self, broiler_farm, sheds_json: str) -> None:
        """Upsert this farm's sheds from the form rows, sharing the Farm Shed
        master's model logic (auto code / unit / name / org-centre / sq-ft /
        status). Existing rows are matched by id so master-only fields (occupied,
        organization centre) are preserved; rows removed from the editor are
        deleted. Non-destructive to sheds the editor still shows."""
        sheds = json.loads(sheds_json) if sheds_json else []
        valid_types = {c[0] for c in BroilerFarmShed.SHED_TYPE_CHOICES}

        def _dec(v):
            v = (str(v) if v is not None else "").strip()
            if not v:
                return None
            try:
                return max(Decimal(v), Decimal("0"))
            except (InvalidOperation, ValueError):
                return None

        def _int(v):
            try:
                return max(int(float(v)), 0)
            except (TypeError, ValueError):
                return 0

        kept_ids = []
        for shed in sheds:
            shed_name = (shed.get("shed_name") or "").strip()
            length, width = _dec(shed.get("length")), _dec(shed.get("width"))
            capacity = _int(shed.get("capacity"))
            shed_type = shed.get("shed_type") or "broiler"
            if shed_type not in valid_types:
                shed_type = "broiler"
            sid = shed.get("id")
            if not sid and not (shed_name or length or width or capacity):
                continue  # skip a blank new row
            obj = broiler_farm.sheds.filter(id=sid).first() if sid else None
            if obj is None:
                obj = BroilerFarmShed(farm=broiler_farm)
            obj.shed_name = shed_name
            obj.shed_type = shed_type
            obj.length, obj.width = length, width
            obj.capacity = capacity
            obj.save()  # auto: shed_code, unit_no, shed_name, org centre, sq_ft, status
            kept_ids.append(obj.id)
        broiler_farm.sheds.exclude(id__in=kept_ids).delete()

    def post(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            data = request.POST
            with transaction.atomic():
                broiler_farm = BroilerFarm.objects.get(id=id) if id else BroilerFarm()

                broiler_farm.branch = Branch.objects.get(id=data["branch_id"])
                broiler_farm.supervisor = Supervisor.objects.get(id=data["supervisor_id"])
                broiler_farm.farmer = Farmer.objects.get(id=data["farmer_id"])

                for field in self.FORM_FIELDS:
                    if field in data:
                        value = data[field]
                        if value == "" and BroilerFarm._meta.get_field(field).null:
                            value = None
                        setattr(broiler_farm, field, value)

                for field in self.FILE_FIELDS:
                    if field in request.FILES:
                        setattr(broiler_farm, field, request.FILES[field])

                broiler_farm.full_clean(exclude=self.FILE_FIELDS)
                broiler_farm.save()

                self._save_sheds(broiler_farm, data.get("sheds", "[]"))

                for picture in request.FILES.getlist("farm_pictures"):
                    BroilerFarmImage.objects.create(farm=broiler_farm, image=picture)

                cache.delete("broiler_farm_list")
            return JsonResponse(
                {"message": "BroilerFarm updated" if id else "BroilerFarm created", "id": broiler_farm.id},
                status=200 if id else 201,
            )
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            broiler_farm = BroilerFarm.objects.get(id=id)
            with transaction.atomic():
                broiler_farm.delete()
                cache.delete("broiler_farm_list")
            return JsonResponse({"message": "BroilerFarm deleted"})
        except Exception as e:
            return self.handle_exception(e)

@login_required()
def get_supervisors(request) -> JsonResponse:
    """Get supervisors for a specific branch."""
    try:
        branch_id = request.GET.get('branch_id')
        # A branch the user cannot see has no supervisors as far as they are
        # concerned; checked before the cache so a warm entry cannot leak one.
        if branch_id and not branches_for(request.user).filter(id=branch_id).exists():
            return JsonResponse({'supervisors': []})
        cache_key = f"supervisors_branch_{branch_id}"
        
        cached_data = cache.get(cache_key)
        if cached_data:
            return JsonResponse({'supervisors': cached_data})
        
        supervisors = list(Supervisor.objects.filter(branch_id=branch_id).values('id', 'name'))
        cache.set(cache_key, supervisors, 300)  # Cache for 5 minutes
        return JsonResponse({'supervisors': supervisors})
    except Exception as e:
        logger.error(f"Error in get_supervisors: {str(e)}", exc_info=True)
        return JsonResponse({"error": str(e)}, status=500)



# ---------------------------------------------------------------------------
# Daily Entry (Broiler > Transactions)
# ---------------------------------------------------------------------------

def _daily_entry_to_dict(row):
    return {
        "id": row.id, "date": row.date.isoformat(), "entry_no": row.entry_no,
        "branch_name": row.farm.branch.branch_name, "farm": row.farm_id,
        "farm_name": row.farm.farm_name,
        "batch": row.batch_id, "batch_name": row.batch.batch_name if row.batch_id else "",
        "age_days": row.age_days, "mortality": row.mortality, "culls": row.culls,
        "feed_1": row.feed_1_id, "feed_1_name": row.feed_1.item_code if row.feed_1_id else "",
        "feed_1_qty": str(row.feed_1_qty), "feed_1_stock": str(row.feed_1_stock),
        "feed_2": row.feed_2_id, "feed_2_name": row.feed_2.item_code if row.feed_2_id else "",
        "feed_2_qty": str(row.feed_2_qty), "feed_2_stock": str(row.feed_2_stock),
        "avg_weight_gms": str(row.avg_weight_gms), "remarks": row.remarks,
        "is_active": row.is_batch_active,
        "entry_by": row.entry_by.username if row.entry_by_id else "",
        "entry_time": timezone.localtime(row.entry_time).strftime("%Y-%m-%d %H:%M") if row.entry_time else "",
    }


def _active_batch_for_farm(farm_id):
    return (_open_batches_for_farm(farm_id).first()
            or BroilerBatch.objects.filter(broiler_farm_id=farm_id).order_by('-start_date', '-id').first())


def _open_batches_for_farm(farm_id):
    """Every batch still running on a farm, newest first.

    A farm normally has one, but nothing stops it having two — and until this
    existed the forms silently took whichever sorted first, so a second open
    flock could take another's entries with no way to tell from the screen.
    """
    return (BroilerBatch.objects
            .filter(broiler_farm_id=farm_id, end_date__isnull=True, is_closed=False)
            .order_by('-start_date', '-id'))


def _resolve_batch(farm_id, batch_id=None):
    """The batch an entry belongs to: the one explicitly chosen, else the
    single open one.

    A chosen batch is checked against the farm rather than trusted — the id
    arrives from the browser, and one belonging to someone else's farm would
    otherwise post entries straight into it.
    """
    if not farm_id:
        return None
    if batch_id:
        chosen = BroilerBatch.objects.filter(id=batch_id,
                                             broiler_farm_id=farm_id).first()
        if chosen:
            return chosen
    return _active_batch_for_farm(farm_id)


def placement_context(farm_id, date_str=None, batch_id=None):
    """Which batch a farm is on, which day it is due, and its age that day.

    Extracted because the mobile ``farm-lookup`` endpoint had grown its own
    copy of it and drifted: it read the raw column rather than the resolved
    placement, so a batch placed by stock transfer reported age 0; it dated from the
    farm's last entry rather than the batch's, so a new flock inherited the
    previous one's; and it fell back to today instead of the day after
    placement. Every one of those was fixed on the web and not here.
    """
    from django.utils.dateparse import parse_date

    batch = _resolve_batch(farm_id, batch_id)
    placed_on = _placement_date(batch)

    # Scoped to the batch, not the farm: a farm re-used for a new flock would
    # otherwise inherit the previous batch's last entry and start weeks late.
    if batch:
        last_entry = DailyEntry.objects.filter(batch=batch).order_by("-date", "-id").first()
    elif farm_id:
        last_entry = DailyEntry.objects.filter(farm_id=farm_id).order_by("-date", "-id").first()
    else:
        last_entry = None

    if last_entry:
        next_date = last_entry.date + timedelta(days=1)
    elif placed_on:
        # Placement day is Age 0, so the first entry is the day after it.
        # Falling back to today skipped every day in between.
        next_date = placed_on + timedelta(days=1)
    else:
        next_date = timezone.localdate()

    entry_date = parse_date(date_str or "") or next_date
    age_days = max((entry_date - placed_on).days, 0) if placed_on else 0
    return {
        "batch": batch,
        "placed_on": placed_on,
        "next_date": next_date,
        "entry_date": entry_date,
        "age_days": age_days,
    }


def _batch_options(farm_id):
    """Open batches as the forms' Batch dropdown wants them. One entry means
    the form fills it in and stays quiet; more means the user has to pick."""
    if not farm_id:
        return []
    return [{"id": b.id, "name": b.batch_name,
             "placed_on": (_placement_date(b).isoformat()
                           if _placement_date(b) else None)}
            for b in _open_batches_for_farm(farm_id)]


def _apply_daily_entry_row(instance, row, entry_date, user, default_farm_id=None):
    # Daily Entry sends farm per row; Single Batch Daily Entry fixes one farm
    # for the whole submission and only varies date/mortality/etc per row —
    # default_farm_id backs a row that omits its own "farm".
    farm_id = row.get("farm") or default_farm_id
    if row.get("date"):
        entry_date = timezone.datetime.fromisoformat(row["date"]).date()
    reject_future_date(entry_date)
    batch = _resolve_batch(farm_id, row.get("batch"))
    instance.date = entry_date
    instance.farm_id = farm_id
    instance.batch = batch
    placed_on = _placement_date(batch)
    if placed_on:
        # Placement day is Age 0; the first entry day (the day after
        # placement) is Age 1. This is the age that gets *stored*, and every
        # figure on the advisory panel — phase, standard feed, cumulative cap —
        # is derived from it, so a batch with no start_date recorded age 0 and
        # took the wrong numbers with it.
        instance.age_days = max((entry_date - placed_on).days, 0)
    else:
        instance.age_days = 0
    instance.mortality = int(row.get("mortality") or 0)
    instance.culls = int(row.get("culls") or 0)
    instance.feed_1_id = row.get("feed_1") or None
    instance.feed_1_qty = Decimal(str(row.get("feed_1_qty") or 0))
    instance.feed_2_id = row.get("feed_2") or None
    instance.feed_2_qty = Decimal(str(row.get("feed_2_qty") or 0))
    instance.avg_weight_gms = Decimal(str(row.get("avg_weight_gms") or 0))
    instance.remarks = row.get("remarks") or ""
    if not instance.pk:
        instance.entry_by = user
    prev1 = DailyEntry.previous_stock(farm_id, instance.feed_1_id, entry_date, instance.pk)
    instance.feed_1_stock = Decimal(str(prev1)) - instance.feed_1_qty if instance.feed_1_id else 0
    prev2 = DailyEntry.previous_stock(farm_id, instance.feed_2_id, entry_date, instance.pk)
    instance.feed_2_stock = Decimal(str(prev2)) - instance.feed_2_qty if instance.feed_2_id else 0


def _recompute_stock_chain(farm_id, item_id):
    """Recomputes feed_1_stock/feed_2_stock for every entry of this farm
    that touches item_id, walking chronologically from an opening balance of
    0. Needed after an edit changes a row's date, Kgs, or feed item, since
    every later row's opening balance is the previous row's closing balance
    — without this, only the edited row itself would reflect the change and
    every row after it would silently keep its stale, pre-edit stock."""
    if not farm_id or not item_id:
        return
    qs = (DailyEntry.objects.filter(farm_id=farm_id)
          .filter(Q(feed_1_id=item_id) | Q(feed_2_id=item_id))
          .order_by('date', 'id'))
    running = Decimal('0')
    for r in qs:
        changed = False
        if r.feed_1_id == item_id:
            running -= r.feed_1_qty
            if r.feed_1_stock != running:
                r.feed_1_stock = running
                changed = True
        if r.feed_2_id == item_id:
            running -= r.feed_2_qty
            if r.feed_2_stock != running:
                r.feed_2_stock = running
                changed = True
        if changed:
            r.save(update_fields=["feed_1_stock", "feed_2_stock"])


def _scope_rows(user, qs, farm_field="farm"):
    """Narrow a transaction queryset to the farms and branches a user may see.

    Scoping the dropdowns only narrows what can be *asked for*; the rows behind
    them still have to be filtered, or the grid answers a question the filter
    bar would not let you type.
    """
    from user.services.scoping import scope_multi

    return scope_multi(user, qs, farms=f"{farm_field}_id",
                       branches=f"{farm_field}__branch_id")


@method_decorator(login_required, name="dispatch")
class DailyEntryListTemplateView(View):
    def get(self, request):
        return render(request, "daily_entry_list.html", {
            "farms": farms_for(request.user, BroilerFarm.objects.order_by("farm_name")),
            "items": Item.objects.order_by("item_code"),
        })


@method_decorator(login_required, name="dispatch")
class SingleBatchDailyEntryListTemplateView(View):
    def get(self, request):
        return render(request, "daily_entry_single_list.html", {
            "items": Item.objects.order_by("item_code"),
        })


@method_decorator(login_required, name="dispatch")
class SingleBatchDailyEntryFormTemplateView(View):
    def get(self, request):
        return render(request, "daily_entry_single_form.html", {
            "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
            "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
            "items": Item.objects.order_by("item_code"),
            "today": timezone.localdate().isoformat(),
        })


@method_decorator(login_required, name="dispatch")
class DailyEntryFormTemplateView(View):
    def get(self, request):
        return render(request, "daily_entry_form.html", {
            "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
            "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
            "items": Item.objects.order_by("item_code"),
            "today": timezone.localdate().isoformat(),
        })


@method_decorator(login_required, name="dispatch")
class DailyEntryAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                # Scoped on the single fetch too: a row id in the url is not a
                # permission, and the list below is the only thing that would
                # otherwise have hidden it.
                row = _scope_rows(request.user, DailyEntry.objects.select_related(
                    "farm__branch", "batch", "feed_1", "feed_2")).get(id=id)
                return JsonResponse(_daily_entry_to_dict(row))

            qs = _scope_rows(request.user, DailyEntry.objects.select_related(
                "farm__branch", "batch", "feed_1", "feed_2"))
            from_date = (request.GET.get("from_date") or "").strip()
            to_date = (request.GET.get("to_date") or "").strip()
            status = (request.GET.get("status") or "").strip()
            farm_id = (request.GET.get("farm") or "").strip()
            batch_id = (request.GET.get("batch") or "").strip()
            nobatch_farm_id = (request.GET.get("nobatch_farm") or "").strip()
            if from_date:
                qs = qs.filter(date__gte=from_date)
            if to_date:
                qs = qs.filter(date__lte=to_date)
            if farm_id:
                qs = qs.filter(farm_id=farm_id)
            if batch_id:
                qs = qs.filter(batch_id=batch_id)
            if nobatch_farm_id:
                qs = qs.filter(farm_id=nobatch_farm_id, batch__isnull=True)
            if status == "Active":
                qs = qs.filter(batch__end_date__isnull=True)
            elif status == "Inactive":
                qs = qs.filter(batch__end_date__isnull=False)
            return JsonResponse([_daily_entry_to_dict(r) for r in qs.order_by("-date", "-id")], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body)
            supervisor_id = data.get("supervisor")
            default_farm_id = data.get("farm")  # Single Batch Daily Entry: one farm for the whole submission
            rows = data.get("rows") or []
            if not supervisor_id:
                return JsonResponse({"error": "Supervisor is required"}, status=400)
            if not rows:
                return JsonResponse({"error": "Add at least one entry row"}, status=400)
            entry_date = data.get("date") or timezone.localdate().isoformat()
            created = []
            for row in rows:
                if not (row.get("farm") or default_farm_id):
                    continue
                instance = DailyEntry(supervisor_id=supervisor_id)
                _apply_daily_entry_row(instance, row, timezone.datetime.fromisoformat(entry_date).date(),
                                       request.user, default_farm_id=default_farm_id)
                instance.full_clean(exclude=["entry_no", "batch"])
                instance.save()
                created.append(instance.id)
                # A backdated row can land ahead of already-saved later rows
                # in the chain; recompute so every row's stored stock reflects
                # its true chronological position, not just the ones after it
                # at the moment each was individually inserted.
                _recompute_stock_chain(instance.farm_id, instance.feed_1_id)
                _recompute_stock_chain(instance.farm_id, instance.feed_2_id)
            if not created:
                return JsonResponse({"error": "Add at least one entry row with a Farm selected"}, status=400)
            return JsonResponse({"message": "Daily entries created", "ids": created}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id: int) -> JsonResponse:
        try:
            instance = DailyEntry.objects.get(id=id)
            old_feed_ids = {instance.feed_1_id, instance.feed_2_id}
            data = json.loads(request.body)
            if data.get("supervisor"):
                instance.supervisor_id = data["supervisor"]
            entry_date = instance.date
            if data.get("date"):
                entry_date = timezone.datetime.fromisoformat(data["date"]).date()
            # Farm isn't editable from the edit modal (it's the grouping key),
            # so fall back to the row's existing farm rather than letting a
            # payload without "farm" wipe it out.
            _apply_daily_entry_row(instance, data, entry_date, request.user, default_farm_id=instance.farm_id)
            instance.full_clean(exclude=["entry_no", "batch"])
            instance.save()
            # Editing this row's date/Kgs/feed item invalidates the stored
            # closing stock of every later row chained after it — recompute
            # both the old and new feed items' full chains, not just this row.
            for item_id in old_feed_ids | {instance.feed_1_id, instance.feed_2_id}:
                _recompute_stock_chain(instance.farm_id, item_id)
            return JsonResponse({"message": "Daily entry updated"})
        except DailyEntry.DoesNotExist:
            raise Http404("Daily entry not found")
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            instance = DailyEntry.objects.get(id=id)
            # Deleting a middle entry would leave every later row's stored
            # stock chained off a balance that no longer exists — restrict
            # deletion to the most recent entry in the batch (or, for entries
            # with no batch, the most recent for that farm) so entries can
            # only be removed newest-first, back toward older ones.
            group_filter = ({"batch_id": instance.batch_id} if instance.batch_id
                             else {"farm_id": instance.farm_id, "batch_id": None})
            newer_exists = (DailyEntry.objects.filter(**group_filter).exclude(id=instance.id)
                             .filter(Q(date__gt=instance.date) | (Q(date=instance.date) & Q(id__gt=instance.id)))
                             .exists())
            if newer_exists:
                return JsonResponse(
                    {"error": "Only the most recent entry in this batch can be deleted. Delete newer entries first."},
                    status=400)
            instance.delete()
            return JsonResponse({"message": "Daily entry deleted"})
        except DailyEntry.DoesNotExist:
            raise Http404("Daily entry not found")
        except Exception as e:
            return self.handle_exception(e)


def resolve_feed_phase(batch, on_date, age_days):
    """The feed-phase line applicable to a batch at `age_days` on `on_date`.
    Matches the active Feed Phase Master by the batch's breed (then its bird
    category), whose effective window covers the date, then the phase whose
    From/To age range contains age_days. Returns a dict or None."""
    if not batch or not batch.breed_id:
        return None
    breed = batch.breed

    def eff_ok(m):
        if on_date and m.effective_from and on_date < m.effective_from:
            return False
        if on_date and m.effective_to and on_date > m.effective_to:
            return False
        return True

    masters = (FeedPhaseMaster.objects.filter(status="active")
               .prefetch_related("lines__feed_item"))
    master = next((m for m in masters.filter(breed=breed) if eff_ok(m)), None)
    if not master and breed.bird_category_id:
        master = next((m for m in masters.filter(bird_category_id=breed.bird_category_id) if eff_ok(m)), None)
    if not master:
        return None

    lines = sorted(master.lines.all(), key=lambda x: x.from_age)
    # feed item -> its phase name + the age range(s) it's used in, so the form
    # can label & validate whichever feed is selected in either slot.
    phase_by_item = {}
    for l in lines:
        if not l.feed_item_id:
            continue
        e = phase_by_item.setdefault(str(l.feed_item_id),
                                     {"name": l.feed_item.description, "ranges": [], "max": 0})
        e["ranges"].append([l.from_age, l.to_age])
        # Max Feed Qty (kg/bird) — the feed-quantity changeover trigger; keep the
        # largest cap if the item spans several phase lines.
        try:
            m = float(l.max_feed_qty or 0)
        except (TypeError, ValueError):
            m = 0
        if m > e["max"]:
            e["max"] = m

    for i, l in enumerate(lines):
        if l.from_age <= age_days and (l.to_age is None or age_days <= l.to_age):
            nxt = lines[i + 1] if i + 1 < len(lines) else None
            return {
                "program": master.program,
                "phase_name": l.feed_item.description if l.feed_item_id else "",
                "phase_code": l.phase_code,
                "feed_item": l.feed_item_id,
                "max_feed_qty": str(l.max_feed_qty),
                # the next (changeover) phase — the Feed 2 hint when it's blank
                "next_name": (nxt.feed_item.description if nxt and nxt.feed_item_id else "") if nxt else "",
                "next_feed_item": (nxt.feed_item_id if nxt else None),
                "phase_by_item": phase_by_item,
            }
    return None


def _flock_counts(batch, as_of=None):
    """Head count for a batch as of `as_of` (default: all-time), broken into
    the parts rather than just the total: chicks placed, mortality, culls,
    birds sold, and what is left alive.

    The total alone answers the feed-cap warning, but the phone's Daily Entry
    also shows each loss against the birds placed ("0.42% of opening"), and
    deriving those from a single number is not possible. Same one pass over the
    same records, so the parts and the total cannot disagree.
    """
    from inventory.models import Item, StockTransfer
    from django.db.models import Sum
    empty = {"placed": 0, "mortality": 0, "culls": 0, "sold": 0, "live": 0}
    if not batch:
        return empty
    chick_ids = list(Item.objects.filter(category__name__icontains="chick").values_list("id", flat=True))
    placed_q = StockTransfer.objects.filter(to_batch_id=batch.id, item_id__in=chick_ids)
    de_q = DailyEntry.objects.filter(batch_id=batch.id)
    sold_q = BirdSale.objects.filter(batch_id=batch.id)
    if as_of:
        placed_q = placed_q.filter(date__lte=as_of)
        de_q = de_q.filter(date__lte=as_of)
        sold_q = sold_q.filter(date__lte=as_of)
    placed = int(placed_q.aggregate(t=Sum("quantity"))["t"] or 0)
    de = de_q.aggregate(m=Sum("mortality"), c=Sum("culls"))
    mortality, culls = int(de["m"] or 0), int(de["c"] or 0)
    sold = int(sold_q.aggregate(b=Sum("birds"))["b"] or 0)
    return {
        "placed": placed,
        "mortality": mortality,
        "culls": culls,
        "sold": sold,
        "live": max(placed - mortality - culls - sold, 0),
    }


def _placement_date(batch):
    """The day the chicks went in, or None.

    ``BroilerBatch.start_date`` is the obvious source but it is not always
    filled: a batch can be created from a chicks placement without it, and one
    such batch is enough for every date on the form to silently fall back to
    today. The placement itself is a chick-category stock transfer into the
    batch, which is the same definition the Live Flock figures use, so the
    earliest of those stands in when start_date is blank.

    Where both exist they agree — checked against the live data before relying
    on it.
    """
    if batch is None:
        return None
    if batch.start_date:
        return batch.start_date
    from inventory.models import Item, StockTransfer

    chick_ids = Item.objects.filter(
        category__name__icontains="chick").values_list("id", flat=True)
    return (StockTransfer.objects
            .filter(to_batch=batch, item_id__in=chick_ids)
            .order_by("date").values_list("date", flat=True).first())


def daily_entry_lookup_payload(farm_id, date_str=None, batch_id=None):
    """Returns the active batch/age for a farm, for the Add form's
    auto-filled Batch/Age fields as soon as a Farm is picked. ``next_date``
    continues the day after this farm's most recently saved entry (so
    backfilling picks up where it left off), falling back to today when
    there's no prior entry.

    Plain dict rather than a response, because the mobile API serves the same
    payload (``broiler.api.DailyEntryLookupView``). The feed-phase, breed
    standard and live-bird figures behind every warning on the form are
    computed here once, so the two clients cannot drift apart.
    """
    # Imported up here, not in the branches below that use them: the live-bird
    # walk needs Sum whenever there is a batch at all, while the breed-standard
    # block that used to import it only runs for a batch whose breed has a
    # standard. A batch with no breed (or no standard) reached that Sum unbound
    # and 500'd the lookup — for the web form as well as the API.
    from django.db.models import F, Max, Min, Sum
    from django.utils.dateparse import parse_date
    # Which batch, which day, what age — shared with the mobile endpoints via
    # placement_context so the two cannot answer differently. Everything below
    # (phase, standards, live birds) is then resolved as of that entry date, so
    # a backfilled or edited row gets the right phase and bird count.
    ctx = placement_context(farm_id, date_str, batch_id)
    batch = ctx["batch"]
    placed_on = ctx["placed_on"]
    next_date = ctx["next_date"]
    entry_date = ctx["entry_date"]

    age_days = 0
    if placed_on:
        # Placement day is Age 0; the first entry day is Age 1.
        age_days = max((entry_date - placed_on).days, 0)

    phase = resolve_feed_phase(batch, entry_date, age_days) if batch else None

    # Expected daily feed per bird from the Breed Standard master (feed_intake is
    # grams/bird/day at that age) — drives the day's over-feed cap check. Only
    # trusted while the age is within the breed's defined curve; beyond it we
    # flag the gap instead of carrying forward a stale (too-low) value.
    std_feed_kg, std_weight_g, std_note = None, None, None
    if batch and batch.breed_id:
        max_age = BreedStandard.objects.filter(breed_id=batch.breed_id).aggregate(m=Max("age"))["m"]
        min_age = BreedStandard.objects.filter(breed_id=batch.breed_id).aggregate(m=Min("age"))["m"]
        if max_age is None:
            std_note = "No breed standard for this breed"
        elif age_days > max_age:
            std_note = f"No breed standard beyond age {max_age} - add rows to Breed Standard"
        elif min_age is not None and age_days < min_age:
            # The other end of the same gap, and the one that used to pass
            # silently: the curve was read at its first row instead.
            std_note = f"No breed standard below age {min_age} - add rows to Breed Standard"
        else:
            std = _breed_standard_at(batch.breed_id, age_days)
            if std and std.feed_intake:
                std_feed_kg = str((_num(std.feed_intake) / 1000).quantize(Decimal("0.001")))
            if std and std.body_weight:
                std_weight_g = str(_num(std.body_weight).quantize(Decimal("0.1")))

    # Breed-standard curve + this batch's feed-to-date, so the form can show
    # "Std Feed @ B.Wt" / "Std B.Wt @ Feed" (like the Live Flock report).
    bs_curve, cum_feed_before_kg = [], None
    if batch and batch.breed_id and not std_note:
        bs_curve = [
            {"a": r.age, "w": float(r.body_weight), "cf": float(r.cum_feed)}
            for r in BreedStandard.objects.filter(breed_id=batch.breed_id, is_active=True).order_by("age")
        ]
        prior = (DailyEntry.objects.filter(batch=batch, date__lt=entry_date)
                 .aggregate(t=Sum(F("feed_1_qty") + F("feed_2_qty")))["t"])
        cum_feed_before_kg = str(_num(prior).quantize(Decimal("0.01")))

    # Actual feed consumed to date (saved entries before this date), per item.
    consumed_by_item, consumed_total_kg = [], None
    if batch:
        by_item = {}
        de_qs = DailyEntry.objects.filter(batch=batch, date__lt=entry_date).select_related("feed_1", "feed_2")
        for de in de_qs:
            if de.feed_1_id:
                by_item[de.feed_1.description] = by_item.get(de.feed_1.description, Decimal("0")) + _num(de.feed_1_qty)
            if de.feed_2_id:
                by_item[de.feed_2.description] = by_item.get(de.feed_2.description, Decimal("0")) + _num(de.feed_2_qty)
        consumed_by_item = [{"name": k, "kg": str(v.quantize(Decimal("0.01")))}
                            for k, v in sorted(by_item.items(), key=lambda x: -x[1]) if v > 0]
        consumed_total_kg = str(sum((v for v in by_item.values()), Decimal("0")).quantize(Decimal("0.01")))

    # Actual feed eaten per SURVIVING bird (bird-day weighted): each day's feed
    # is shared among the birds alive that day, so — unlike total ÷ current-live
    # — it excludes the extra share that later-dead/culled/sold birds ate.
    # Sum over saved days of (that day's feed / birds alive that day), in g/bird.
    consumed_per_bird_actual_g = None
    if batch:
        from inventory.models import Item as _Item, StockTransfer as _ST
        chick_ids = list(_Item.objects.filter(category__name__icontains="chick").values_list("id", flat=True))
        placed = _ST.objects.filter(to_batch_id=batch.id, item_id__in=chick_ids, date__lt=entry_date)\
                            .aggregate(t=Sum("quantity"))["t"] or 0
        sold_by_date = {}
        for bs in BirdSale.objects.filter(batch=batch, date__lt=entry_date).values("date").annotate(t=Sum("birds")):
            sold_by_date[bs["date"]] = bs["t"] or 0
        alive = int(placed)
        cum_pb = Decimal("0")
        for de in DailyEntry.objects.filter(batch=batch, date__lt=entry_date).order_by("date", "id"):
            feed_d = _num(de.feed_1_qty) + _num(de.feed_2_qty)   # birds present at start of day eat it
            if alive > 0 and feed_d:
                cum_pb += feed_d / Decimal(alive) * Decimal("1000")
            alive -= (de.mortality or 0) + (de.culls or 0) + sold_by_date.get(de.date, 0)   # end-of-day losses
        consumed_per_bird_actual_g = str(cum_pb.quantize(Decimal("0.01")))

    # Who this flock is — shed, breed, bird type. The phone heads the entry with
    # it so a supervisor can see they are recording the flock in front of them;
    # the web form has the same facts on screen in its Batch panel.
    shed = batch.shed if batch and batch.shed_id else None
    breed = batch.breed if batch and batch.breed_id else None
    counts = _flock_counts(batch, as_of=entry_date) if batch else _flock_counts(None)

    return {
        "batch": batch.id if batch else None,
        "batch_name": batch.batch_name if batch else "",
        "shed_name": (shed.shed_name or shed.shed_no or shed.shed_code) if shed else "",
        "breed_name": breed.description if breed else "",
        "bird_type": (breed.bird_category.name if breed and breed.bird_category_id else ""),
        # Chicks placed, and the losses booked against them before this entry —
        # the phone shows each as a share of the opening count.
        "opening_birds": counts["placed"],
        "mortality_to_date": counts["mortality"],
        "culls_to_date": counts["culls"],
        "sold_to_date": counts["sold"],
        # Every open batch on the farm, so the form can fill the Batch box in
        # when there is one and ask when there is more than one.
        "batches": _batch_options(farm_id),
        "age_days": age_days,
        # The resolved placement, not the raw column: the form computes Age
        # from this in the browser, so sending None leaves the Age box empty
        # on exactly the batches this fix is for.
        "start_date": placed_on.isoformat() if placed_on else None,
        "next_date": next_date.isoformat(),
        "feed_phase": phase,
        "std_feed_kg": std_feed_kg,
        "std_weight_g": std_weight_g,
        "std_note": std_note,
        "bs_curve": bs_curve,
        "cum_feed_before_kg": cum_feed_before_kg,
        "consumed_by_item": consumed_by_item,
        "consumed_total_kg": consumed_total_kg,
        "consumed_per_bird_actual_g": consumed_per_bird_actual_g,
        "live_birds": counts["live"],
    }


@login_required
def daily_entry_farm_lookup(request):
    """GET ?farm=&date= — the web form's wrapper around
    ``daily_entry_lookup_payload``."""
    return JsonResponse(daily_entry_lookup_payload(request.GET.get("farm"),
                                                   request.GET.get("date"),
                                                   request.GET.get("batch")))


@login_required
def daily_entry_stock_lookup(request):
    """Opening stock for a farm+feed item as of a given date — i.e. the
    closing balance of the most recent saved entry before that date (0 if
    none). Used to seed the Add form's live running-stock preview; the grid
    itself then subtracts each row's own Kgs client-side as you type."""
    farm_id = request.GET.get("farm")
    item_id = request.GET.get("item")
    entry_date = request.GET.get("date")
    if not farm_id or not item_id or not entry_date:
        return JsonResponse({"stock": "0"})
    d = timezone.datetime.fromisoformat(entry_date).date()
    stock = DailyEntry.previous_stock(farm_id, int(item_id), d, None)
    return JsonResponse({"stock": str(stock)})


# ---------------------------------------------------------------------------
# Medicine Vaccine Consumption (Broiler > Transactions)
# ---------------------------------------------------------------------------

def _medicine_entry_to_dict(row):
    return {
        "id": row.id, "date": row.date.isoformat(), "entry_no": row.entry_no,
        "branch_name": row.farm.branch.branch_name, "farm": row.farm_id,
        "farm_name": row.farm.farm_name,
        "batch": row.batch_id, "batch_name": row.batch.batch_name if row.batch_id else "",
        "age_days": row.age_days,
        "item": row.item_id, "item_name": row.item.item_code if row.item_id else "",
        "unit": _uom_label(row.item.consumption_uom) if row.item_id else "",
        "qty": str(row.qty), "stock": str(row.stock),
        "remarks": row.remarks,
        "is_active": row.is_batch_active,
        "entry_by": row.entry_by.username if row.entry_by_id else "",
        "entry_time": timezone.localtime(row.entry_time).strftime("%Y-%m-%d %H:%M") if row.entry_time else "",
    }


def _apply_medicine_entry_row(instance, row, entry_date, user):
    farm_id = row.get("farm") or instance.farm_id
    if row.get("date"):
        entry_date = timezone.datetime.fromisoformat(row["date"]).date()
    reject_future_date(entry_date)
    batch = _resolve_batch(farm_id, row.get("batch"))
    instance.date = entry_date
    instance.farm_id = farm_id
    instance.batch = batch
    placed_on = _placement_date(batch)
    if placed_on:
        # Placement day is Age 0; the first entry day (the day after
        # placement) is Age 1. This is the age that gets *stored*, and every
        # figure on the advisory panel — phase, standard feed, cumulative cap —
        # is derived from it, so a batch with no start_date recorded age 0 and
        # took the wrong numbers with it.
        instance.age_days = max((entry_date - placed_on).days, 0)
    else:
        instance.age_days = 0
    instance.item_id = row.get("item") or None
    instance.qty = Decimal(str(row.get("qty") or 0))
    instance.remarks = row.get("remarks") or ""
    if not instance.pk:
        instance.entry_by = user
    prev = MedicineVaccineEntry.previous_stock(farm_id, instance.item_id, entry_date, instance.pk)
    instance.stock = Decimal(str(prev)) - instance.qty if instance.item_id else 0


def _recompute_medicine_stock_chain(farm_id, item_id):
    """Recomputes stock for every medicine/vaccine entry of this farm that
    touches item_id, walking chronologically from an opening balance of 0.
    See _recompute_stock_chain (Daily Entry) for the full rationale."""
    if not farm_id or not item_id:
        return
    qs = (MedicineVaccineEntry.objects.filter(farm_id=farm_id, item_id=item_id)
          .order_by('date', 'id'))
    running = Decimal('0')
    for r in qs:
        running -= r.qty
        if r.stock != running:
            r.stock = running
            r.save(update_fields=["stock"])


@method_decorator(login_required, name="dispatch")
class MedicineEntryListTemplateView(View):
    def get(self, request):
        return render(request, "medicine_entry_list.html", {
            "farms": farms_for(request.user, BroilerFarm.objects.order_by("farm_name")),
            "items": Item.objects.order_by("item_code"),
        })


@method_decorator(login_required, name="dispatch")
class MedicineEntryFormTemplateView(View):
    def get(self, request):
        return render(request, "medicine_entry_form.html", {
            "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
            "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
            "items": Item.objects.order_by("item_code"),
            "today": timezone.localdate().isoformat(),
        })


@method_decorator(login_required, name="dispatch")
class MedicineEntryAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                row = MedicineVaccineEntry.objects.select_related("farm__branch", "batch", "item").get(id=id)
                return JsonResponse(_medicine_entry_to_dict(row))

            qs = MedicineVaccineEntry.objects.select_related("farm__branch", "batch", "item")
            from_date = (request.GET.get("from_date") or "").strip()
            to_date = (request.GET.get("to_date") or "").strip()
            status = (request.GET.get("status") or "").strip()
            farm_id = (request.GET.get("farm") or "").strip()
            batch_id = (request.GET.get("batch") or "").strip()
            nobatch_farm_id = (request.GET.get("nobatch_farm") or "").strip()
            if from_date:
                qs = qs.filter(date__gte=from_date)
            if to_date:
                qs = qs.filter(date__lte=to_date)
            if farm_id:
                qs = qs.filter(farm_id=farm_id)
            if batch_id:
                qs = qs.filter(batch_id=batch_id)
            if nobatch_farm_id:
                qs = qs.filter(farm_id=nobatch_farm_id, batch__isnull=True)
            if status == "Active":
                qs = qs.filter(batch__end_date__isnull=True)
            elif status == "Inactive":
                qs = qs.filter(batch__end_date__isnull=False)
            return JsonResponse([_medicine_entry_to_dict(r) for r in qs.order_by("-date", "-id")], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body)
            supervisor_id = data.get("supervisor")
            rows = data.get("rows") or []
            if not supervisor_id:
                return JsonResponse({"error": "Supervisor is required"}, status=400)
            if not rows:
                return JsonResponse({"error": "Add at least one entry row"}, status=400)
            entry_date = data.get("date") or timezone.localdate().isoformat()
            created = []
            for row in rows:
                if not row.get("farm"):
                    continue
                instance = MedicineVaccineEntry(supervisor_id=supervisor_id)
                _apply_medicine_entry_row(instance, row, timezone.datetime.fromisoformat(entry_date).date(), request.user)
                instance.full_clean(exclude=["entry_no", "batch"])
                instance.save()
                created.append(instance.id)
                _recompute_medicine_stock_chain(instance.farm_id, instance.item_id)
            if not created:
                return JsonResponse({"error": "Add at least one entry row with a Farm selected"}, status=400)
            return JsonResponse({"message": "Medicine/Vaccine entries created", "ids": created}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id: int) -> JsonResponse:
        try:
            instance = MedicineVaccineEntry.objects.get(id=id)
            old_item_id = instance.item_id
            data = json.loads(request.body)
            if data.get("supervisor"):
                instance.supervisor_id = data["supervisor"]
            entry_date = instance.date
            if data.get("date"):
                entry_date = timezone.datetime.fromisoformat(data["date"]).date()
            _apply_medicine_entry_row(instance, data, entry_date, request.user)
            instance.full_clean(exclude=["entry_no", "batch"])
            instance.save()
            for item_id in {old_item_id, instance.item_id}:
                _recompute_medicine_stock_chain(instance.farm_id, item_id)
            return JsonResponse({"message": "Medicine/Vaccine entry updated"})
        except MedicineVaccineEntry.DoesNotExist:
            raise Http404("Medicine/Vaccine entry not found")
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            instance = MedicineVaccineEntry.objects.get(id=id)
            # Same tail-only rule as Daily Entry: deleting a middle entry
            # would leave every later row's stored stock chained off a
            # balance that no longer exists.
            group_filter = ({"batch_id": instance.batch_id} if instance.batch_id
                             else {"farm_id": instance.farm_id, "batch_id": None})
            newer_exists = (MedicineVaccineEntry.objects.filter(**group_filter).exclude(id=instance.id)
                             .filter(Q(date__gt=instance.date) | (Q(date=instance.date) & Q(id__gt=instance.id)))
                             .exists())
            if newer_exists:
                return JsonResponse(
                    {"error": "Only the most recent entry in this batch can be deleted. Delete newer entries first."},
                    status=400)
            instance.delete()
            return JsonResponse({"message": "Medicine/Vaccine entry deleted"})
        except MedicineVaccineEntry.DoesNotExist:
            raise Http404("Medicine/Vaccine entry not found")
        except Exception as e:
            return self.handle_exception(e)


@login_required
def medicine_entry_farm_lookup(request):
    """Returns the active batch/age for a farm, for the Add form's
    auto-filled Batch/Age fields as soon as a Farm is picked."""
    farm_id = request.GET.get("farm")
    batch = _resolve_batch(farm_id, request.GET.get("batch"))
    age_days = 0
    placed_on = _placement_date(batch)
    if placed_on:
        age_days = max((timezone.localdate() - placed_on).days, 0)
    return JsonResponse({
        "batch": batch.id if batch else None,
        "batch_name": batch.batch_name if batch else "",
        "batches": _batch_options(farm_id),
        "age_days": age_days,
        "start_date": placed_on.isoformat() if placed_on else None,
    })


@login_required
def medicine_entry_item_lookup(request):
    """Unit (consumption UOM) for a selected Medicine/Vaccine item, for the
    Add form's auto-filled Unit field."""
    item_id = request.GET.get("item")
    item = Item.objects.filter(id=item_id).first() if item_id else None
    return JsonResponse({"unit": _uom_label(item.consumption_uom) if item else ""})


@login_required
def medicine_entry_stock_lookup(request):
    """Opening stock for a farm+item as of a given date — the closing
    balance of the most recent saved entry before that date (0 if none)."""
    farm_id = request.GET.get("farm")
    item_id = request.GET.get("item")
    entry_date = request.GET.get("date")
    if not farm_id or not item_id or not entry_date:
        return JsonResponse({"stock": "0"})
    d = timezone.datetime.fromisoformat(entry_date).date()
    stock = MedicineVaccineEntry.previous_stock(farm_id, int(item_id), d, None)
    return JsonResponse({"stock": str(stock)})


@login_required
def daily_entry_group_delete(request):
    """Bulk-deletes every Daily Entry / Single Batch Daily Entry row for one
    batch (or, for entries with no active batch, one farm) at once — the
    register's group-level Delete action. Safe to do in one shot regardless
    of the tail-only rule on individual deletes, since wiping the whole
    group at once leaves no later row anywhere still chained to it."""
    if request.method != "DELETE":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    batch_id = (request.GET.get("batch") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()
    if batch_id:
        qs = DailyEntry.objects.filter(batch_id=batch_id)
    elif farm_id:
        qs = DailyEntry.objects.filter(farm_id=farm_id, batch__isnull=True)
    else:
        return JsonResponse({"error": "batch or farm is required"}, status=400)
    count = qs.count()
    if not count:
        return JsonResponse({"error": "No entries found for this batch"}, status=404)
    qs.delete()
    return JsonResponse({"message": f"Deleted {count} entries"})


@login_required
def medicine_entry_group_delete(request):
    """Bulk-deletes every Medicine Vaccine Consumption row for one batch (or,
    for entries with no active batch, one farm) at once. See
    daily_entry_group_delete for the rationale."""
    if request.method != "DELETE":
        return JsonResponse({"error": "Method not allowed"}, status=405)
    batch_id = (request.GET.get("batch") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()
    if batch_id:
        qs = MedicineVaccineEntry.objects.filter(batch_id=batch_id)
    elif farm_id:
        qs = MedicineVaccineEntry.objects.filter(farm_id=farm_id, batch__isnull=True)
    else:
        return JsonResponse({"error": "batch or farm is required"}, status=400)
    count = qs.count()
    if not count:
        return JsonResponse({"error": "No entries found for this batch"}, status=404)
    qs.delete()
    return JsonResponse({"message": f"Deleted {count} entries"})


# ---------------------------------------------------------------------------
# Bird Sale (Broiler > Transactions)
# ---------------------------------------------------------------------------

def _bird_sale_to_dict(row):
    buyer_name = row.customer.name if row.customer_id else (row.farmer.farmer_name if row.farmer_id else "")
    return {
        "id": row.id, "sale_no": row.sale_no, "date": row.date.isoformat(), "doc_no": row.doc_no,
        "sale_type": row.sale_type,
        "customer": row.customer_id, "customer_name": row.customer.name if row.customer_id else "",
        "farmer": row.farmer_id, "farmer_name": row.farmer.farmer_name if row.farmer_id else "",
        "buyer_name": buyer_name,
        "farm": row.farm_id, "farm_name": row.farm.farm_name,
        "batch": row.batch_id, "batch_name": row.batch.batch_name if row.batch_id else "",
        "birds": row.birds, "net_weight": str(row.net_weight), "avg_weight": str(row.avg_weight),
        "rate": str(row.rate), "round_off": str(row.round_off), "amount": str(row.amount),
        "lifting_supervisor": row.lifting_supervisor_id,
        "lifting_supervisor_name": str(row.lifting_supervisor) if row.lifting_supervisor_id else "",
        "vehicle": row.vehicle, "driver": row.driver, "remarks": row.remarks,
        # Field evidence, captured by the phone. A desk raising the same sale
        # from a slip brought in has none of it, so every one of these is
        # allowed to be empty — the register says which liftings were
        # witnessed and which were typed up afterwards.
        "lift_latitude": row.lift_latitude, "lift_longitude": row.lift_longitude,
        "lift_place": row.lift_place,
        "photos": [
            {"kind": p.kind, "label": p.get_kind_display(), "url": p.image.url}
            for p in row.photos.all() if p.image
        ],
    }


def _apply_bird_sale(instance, data):
    if data.get("date"):
        instance.date = reject_future_date(timezone.datetime.fromisoformat(data["date"]).date())
    instance.doc_no = data.get("doc_no") or ""
    sale_type = data.get("sale_type") or "customer"
    instance.sale_type = sale_type
    farm_id = data.get("farm") or None
    instance.farm_id = farm_id
    instance.batch = _resolve_batch(farm_id, data.get("batch"))
    if sale_type == "customer":
        instance.customer_id = data.get("customer") or None
        instance.farmer_id = None
    else:
        # A Farmer Sale is always the same farmer who grew these birds on
        # this farm (buying their own birds back) — never a free pick from
        # the whole Farmer list, so it's derived from the farm, not the
        # client payload.
        instance.customer_id = None
        farm = BroilerFarm.objects.filter(id=farm_id).first() if farm_id else None
        instance.farmer_id = farm.farmer_id if farm else None
    instance.birds = int(data.get("birds") or 0)
    instance.net_weight = Decimal(str(data.get("net_weight") or 0))
    instance.rate = Decimal(str(data.get("rate") or 0))
    # round_off and amount are derived in BirdSale.save(); a value sent for
    # either is ignored rather than trusted.
    instance.lifting_supervisor_id = data.get("lifting_supervisor") or None
    instance.vehicle = data.get("vehicle") or ""
    instance.driver = data.get("driver") or ""
    instance.remarks = data.get("remarks") or ""


@method_decorator(login_required, name="dispatch")
class BirdSaleListTemplateView(View):
    def get(self, request):
        return render(request, "bird_sale_list.html")


@method_decorator(login_required, name="dispatch")
class BirdSaleFormTemplateView(View):
    def get(self, request, id=None):
        from hr.models import Employee
        return render(request, "bird_sale_form.html", {
            "instance": BirdSale.objects.filter(id=id).first() if id else None,
            "farms": farms_for(request.user, BroilerFarm.objects.order_by("farm_name")),
            "customers": customers_for(request.user, Customer.objects.order_by("name")),
            "farmers": Farmer.objects.order_by("farmer_name"),
            # Lifting supervisor can be any active employee.
            "supervisors": Employee.objects.filter(relieve=False).order_by("full_name"),
            "today": timezone.localdate().isoformat(),
        })


@method_decorator(login_required, name="dispatch")
class BirdSaleAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                row = _scope_rows(request.user, BirdSale.objects.select_related(
                    "customer", "farmer", "farm", "batch", "lifting_supervisor")
                    .prefetch_related("photos")).get(id=id)
                return JsonResponse(_bird_sale_to_dict(row))

            # Evidence photos are rendered per row, so they are fetched with the
            # page rather than one query per sale.
            qs = _scope_rows(request.user, BirdSale.objects.select_related(
                "customer", "farmer", "farm", "batch", "lifting_supervisor")
                .prefetch_related("photos"))
            from_date = (request.GET.get("from_date") or "").strip()
            to_date = (request.GET.get("to_date") or "").strip()
            if from_date:
                qs = qs.filter(date__gte=from_date)
            if to_date:
                qs = qs.filter(date__lte=to_date)
            return JsonResponse([_bird_sale_to_dict(r) for r in qs.order_by("-date", "-id")], safe=False)
        except BirdSale.DoesNotExist:
            raise Http404("Bird sale not found")
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body or "{}")
            rows = data.get("rows") or []
            if not rows:
                return JsonResponse({"error": "Add at least one sale row"}, status=400)
            created = []
            for row in rows:
                if not row.get("farm"):
                    continue
                instance = BirdSale(entry_by=request.user)
                _apply_bird_sale(instance, row)
                instance.full_clean(exclude=["sale_no", "batch"])
                instance.save()
                created.append(instance.id)
            if not created:
                return JsonResponse({"error": "Add at least one sale row with a Farm selected"}, status=400)
            return JsonResponse({"message": "Bird sale(s) created", "ids": created}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id: int) -> JsonResponse:
        try:
            instance = BirdSale.objects.get(id=id)
            data = json.loads(request.body or "{}")
            _apply_bird_sale(instance, data)
            instance.full_clean(exclude=["sale_no", "batch"])
            instance.save()
            return JsonResponse({"message": "Bird sale updated"})
        except BirdSale.DoesNotExist:
            raise Http404("Bird sale not found")
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            instance = BirdSale.objects.get(id=id)
            instance.delete()
            return JsonResponse({"message": "Bird sale deleted"})
        except BirdSale.DoesNotExist:
            raise Http404("Bird sale not found")
        except Exception as e:
            return self.handle_exception(e)


@login_required
def bird_sale_farm_lookup(request):
    """Returns the active batch and owning farmer for a farm, for the Add
    form's auto-filled Batch field and Farmer Sale buyer (always the same
    farmer who grew the birds on this farm) as soon as a Farm is picked."""
    farm_id = request.GET.get("farm")
    batch = _resolve_batch(farm_id, request.GET.get("batch"))
    farm = BroilerFarm.objects.filter(id=farm_id).select_related("farmer").first() if farm_id else None
    return JsonResponse({
        "batch": batch.id if batch else None,
        "batch_name": batch.batch_name if batch else "",
        # Every open batch, so the form can fill the box in when there is one
        # and ask when there is more — a sale filed against the wrong flock
        # takes birds off it, which the stock and the settlement both carry.
        "batches": _batch_options(farm_id),
        "farmer": farm.farmer_id if farm else None,
        "farmer_name": farm.farmer.farmer_name if farm and farm.farmer_id else "",
    })


# ---------------------------------------------------------------------------
# Bird Sale Receipt (Broiler > Transactions)
# ---------------------------------------------------------------------------

def _bird_sale_receipt_to_dict(row):
    buyer_name = row.customer.name if row.customer_id else (row.farmer.farmer_name if row.farmer_id else "")
    return {
        "id": row.id, "receipt_no": row.receipt_no, "date": row.date.isoformat(),
        "location": row.location_id, "location_name": row.location.name,
        "sale_type": row.sale_type,
        "customer": row.customer_id, "customer_name": row.customer.name if row.customer_id else "",
        "farmer": row.farmer_id, "farmer_name": row.farmer.farmer_name if row.farmer_id else "",
        "buyer_name": buyer_name,
        "mode": row.mode,
        "receipt_account": row.receipt_account_id,
        "receipt_account_name": (f"{row.receipt_account.code} - {row.receipt_account.description}"
                                 if row.receipt_account_id else ""),
        "amount": str(row.amount), "reference_no": row.reference_no, "remarks": row.remarks,
    }


def _apply_bird_sale_receipt(instance, data):
    if data.get("date"):
        instance.date = reject_future_date(timezone.datetime.fromisoformat(data["date"]).date())
    instance.location_id = data.get("location") or None
    sale_type = data.get("sale_type") or "customer"
    instance.sale_type = sale_type
    if sale_type == "customer":
        instance.customer_id = data.get("customer") or None
        instance.farmer_id = None
    else:
        instance.customer_id = None
        instance.farmer_id = data.get("farmer") or None
    instance.mode = data.get("mode") or "Cash"
    instance.receipt_account_id = data.get("receipt_account") or None
    instance.amount = Decimal(str(data.get("amount") or 0))
    instance.reference_no = data.get("reference_no") or ""
    instance.remarks = data.get("remarks") or ""


@method_decorator(login_required, name="dispatch")
class BirdSaleReceiptListTemplateView(View):
    def get(self, request):
        return render(request, "bird_sale_receipt_list.html")


@method_decorator(login_required, name="dispatch")
class BirdSaleReceiptFormTemplateView(View):
    def get(self, request, id=None):
        from account.services.bank_cash import bank_cash_accounts, active_payment_modes, payment_mode_map
        import json as _json
        return render(request, "bird_sale_receipt_form.html", {
            "instance": BirdSaleReceipt.objects.filter(id=id).first() if id else None,
            "locations": Warehouse.objects.order_by("name"),
            "customers": customers_for(request.user, Customer.objects.order_by("name")),
            "farmers": Farmer.objects.order_by("farmer_name"),
            "accounts": bank_cash_accounts(),   # receipt into a Bank/Cash master account
            "payment_modes": active_payment_modes("receipt"),
            "payment_mode_map_json": _json.dumps(payment_mode_map("receipt")),
            "today": timezone.localdate().isoformat(),
        })


@method_decorator(login_required, name="dispatch")
class BirdSaleReceiptAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                row = BirdSaleReceipt.objects.select_related(
                    "customer", "farmer", "location", "receipt_account").get(id=id)
                return JsonResponse(_bird_sale_receipt_to_dict(row))

            qs = BirdSaleReceipt.objects.select_related("customer", "farmer", "location", "receipt_account")
            from_date = (request.GET.get("from_date") or "").strip()
            to_date = (request.GET.get("to_date") or "").strip()
            if from_date:
                qs = qs.filter(date__gte=from_date)
            if to_date:
                qs = qs.filter(date__lte=to_date)
            return JsonResponse([_bird_sale_receipt_to_dict(r) for r in qs.order_by("-date", "-id")], safe=False)
        except BirdSaleReceipt.DoesNotExist:
            raise Http404("Receipt not found")
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body or "{}")
            rows = data.get("rows") or []
            if not rows:
                return JsonResponse({"error": "Add at least one receipt row"}, status=400)
            created = []
            for row in rows:
                if not row.get("location"):
                    continue
                instance = BirdSaleReceipt(entry_by=request.user)
                _apply_bird_sale_receipt(instance, row)
                instance.full_clean(exclude=["receipt_no"])
                instance.save()
                created.append(instance.id)
            if not created:
                return JsonResponse({"error": "Add at least one receipt row with a Location selected"}, status=400)
            return JsonResponse({"message": "Receipt(s) created", "ids": created}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id: int) -> JsonResponse:
        try:
            instance = BirdSaleReceipt.objects.get(id=id)
            data = json.loads(request.body or "{}")
            _apply_bird_sale_receipt(instance, data)
            instance.full_clean(exclude=["receipt_no"])
            instance.save()
            return JsonResponse({"message": "Receipt updated"})
        except BirdSaleReceipt.DoesNotExist:
            raise Http404("Receipt not found")
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            instance = BirdSaleReceipt.objects.get(id=id)
            instance.delete()
            return JsonResponse({"message": "Receipt deleted"})
        except BirdSaleReceipt.DoesNotExist:
            raise Http404("Receipt not found")
        except Exception as e:
            return self.handle_exception(e)


@login_required
def bird_sale_receipt_balance_lookup(request):
    """Outstanding balance for a Customer/Farmer at a cost centre (rolled up
    across every Farm sharing that Location's Branch), for the Add form's
    auto-filled Balance field as soon as Location and Customer/Farmer are
    picked."""
    location_id = request.GET.get("location")
    sale_type = request.GET.get("sale_type") or "customer"
    customer_id = request.GET.get("customer")
    farmer_id = request.GET.get("farmer")
    exclude_id = request.GET.get("exclude_id")
    if sale_type == "customer":
        # Customer receipts show the full customer ledger balance (all modules),
        # matching the Customer Ledger; farmer receipts keep the farm/branch
        # cost-centre balance since farmers aren't customers.
        from sales.views import _customer_current_balance
        balance = _customer_current_balance(customer_id, exclude_bird_receipt_id=exclude_id)
    else:
        balance = BirdSaleReceipt.balance_due(location_id, sale_type, customer_id, farmer_id, exclude_id=exclude_id)
    return JsonResponse({"balance": str(balance)})


# ---------------------------------------------------------------------------
# Batch History Report (Broiler > Reports)
# ---------------------------------------------------------------------------

def _transfer_location_name(header):
    if header.from_warehouse_id:
        return header.from_warehouse.name
    if header.from_farm_id:
        return header.from_farm.farm_name
    return ""


def _transfer_to_location_name(header):
    if header.to_warehouse_id:
        return header.to_warehouse.name
    if header.to_farm_id:
        return header.to_farm.farm_name
    return ""


def _div(numerator, denominator):
    """Safe division returning Decimal('0') when the denominator is zero/None."""
    n = Decimal(str(numerator or 0))
    d = Decimal(str(denominator or 0))
    return (n / d) if d else Decimal("0")


def _match_growing_charge_scheme(batch, on_date):
    """The GrowingChargeScheme whose master-defined date range covers the
    placement date, for this batch's region — branch-specific preferred over
    an all-branches scheme. Returns None when no scheme covers the date (the
    caller then shows 'No Data' for scheme-driven cost/price figures). Never
    falls back to a scheme outside the date range or another branch."""
    if not on_date:
        return None
    branch = batch.broiler_farm.branch
    qs = GrowingChargeScheme.objects.filter(
        region_id=branch.region_id, is_active=True,
        from_date__lte=on_date, to_date__gte=on_date,
    )
    return (qs.filter(branch=branch).order_by("-from_date").first()
            or qs.filter(branch__isnull=True).order_by("-from_date").first())


def _build_batch_costing(batch, placement_total, cum_mortality, cum_culls, mortality_rows,
                         chick_rows, feed_rows, feed_summary_rows, feed_return_rows,
                         medicine_transfer_rows, medicine_consumption_rows, medicine_return_rows,
                         bird_sale_rows, fetch_type="farmer", scheme_override=None):
    """Batch Costing Information and Summary block of the Growing Charge
    Statement — bird/feed/weight KPIs plus the cost roll-up. Cost drivers
    (Admin Cost, Grade) come from the batch's applicable GrowingChargeScheme
    (or ``scheme_override`` when a Schema is hand-picked in the filter bar).

    ``fetch_type`` chooses the statement perspective: 'farmer' bills only the
    farmer's admin share, 'management' bills the full admin cost."""
    q2, q3 = Decimal("0.01"), Decimal("0.001")

    # --- dates ---
    placement_date = min((r["date"] for r in chick_rows), default=None) or batch.start_date
    sale_start_date = min((r["date"] for r in bird_sale_rows), default=None)

    # --- birds ---
    sold_birds = sum(r["birds"] for r in bird_sale_rows)
    sold_weight = sum((r["net_weight"] or 0) for r in bird_sale_rows)
    sold_amount = sum((r["amount"] or 0) for r in bird_sale_rows)
    shortage_birds = 0  # not tracked in this system
    excess_birds = placement_total - cum_mortality - cum_culls - sold_birds - shortage_birds

    # --- mortality % (denominator = chicks placed) ---
    mort_upto_7 = mort_upto_30 = 0
    for r in mortality_rows:
        if r["age"] <= 7:
            mort_upto_7 = r["cum_mortality"]
        if r["age"] <= 30:
            mort_upto_30 = r["cum_mortality"]
    first_week_mort_pct = _div(mort_upto_7 * 100, placement_total)
    upto_30_mort_pct = _div(mort_upto_30 * 100, placement_total)
    after_30_mort_pct = _div((cum_mortality - mort_upto_30) * 100, placement_total)
    total_mort_pct = _div(cum_mortality * 100, placement_total)

    # --- weights / age ---
    avg_body_weight = _div(sold_weight, sold_birds).quantize(q2)        # kg
    weighted_age = sum(
        ((r["date"] - placement_date).days + 1) * r["birds"]
        for r in bird_sale_rows) if placement_date else 0
    mean_age = _div(weighted_age, sold_birds).quantize(q2)
    day_gain = _div(avg_body_weight * 1000, mean_age)                   # g/day (avg wt in g / mean age)

    # --- feed ---
    feed_in_kg = sum((r["quantity"] or 0) for r in feed_rows)
    feed_in_amount = sum((r["amount"] or 0) for r in feed_rows)
    avg_feed_rate = _div(feed_in_amount, feed_in_kg)
    feed_consumed = sum((r["consumed"] or 0) for r in feed_summary_rows)
    feed_return_kg = sum((r["quantity"] or 0) for r in feed_return_rows)
    feed_cost = feed_consumed * avg_feed_rate

    # --- medicine / vaccine ---
    med_in_qty = sum((r["quantity"] or 0) for r in medicine_transfer_rows)
    med_in_amount = sum((r["amount"] or 0) for r in medicine_transfer_rows)
    med_consumed = sum((r["quantity"] or 0) for r in medicine_consumption_rows)
    med_return_qty = sum((r["quantity"] or 0) for r in medicine_return_rows)
    med_cost = med_consumed * _div(med_in_amount, med_in_qty)

    # --- FCR / CFCR / Livability / EEF (per client KPI definitions) ---
    # Mortality Weight = weight of birds that died, from each day's mortality
    # valued at that day's average body weight (early deaths weigh less).
    mortality_weight = sum(
        (Decimal(str(r["mortality"])) * Decimal(str(r["avg_weight_kg"])) for r in mortality_rows),
        Decimal("0"),
    )
    fcr = _div(feed_consumed, sold_weight)                              # Feed / Live Weight
    cfcr = _div(feed_consumed, sold_weight + mortality_weight)          # Feed / (Live Wt + Mortality Wt)
    livability_pct = _div(sold_birds * 100, placement_total)           # (Birds Sold / Chicks Placed) x 100
    eef = _div(livability_pct * avg_body_weight * 100, mean_age * fcr)  # (Livability x Avg Wt x 100) / (Age x FCR)

    # --- costs from the applicable Growing Charge Scheme ---
    scheme = scheme_override or _match_growing_charge_scheme(batch, placement_date)
    chick_cost = sum((r["amount"] or 0) for r in chick_rows)
    if scheme:
        farmer_rate = scheme.farmer_admin_cost or 0
        mgmt_rate = scheme.management_admin_cost or 0
        # Farmer report bills only the farmer's admin share; Management report
        # bills the full admin cost (farmer + management).
        admin_rate = farmer_rate if fetch_type == "farmer" else (farmer_rate + mgmt_rate)
        admin_cost = admin_rate * placement_total
    else:
        admin_cost = Decimal("0")
    total_production_cost = feed_cost + chick_cost + med_cost + admin_cost
    production_cost_per_kg = _div(total_production_cost, sold_weight)

    # Cost Distribution donut (dashboard widget) — captured as plain floats
    # BEFORE the "no scheme -> No Data" string override below, so the chart
    # always has real numbers (admin_cost is simply 0 without a scheme).
    cost_breakdown = {
        "feed_cost": float(feed_cost.quantize(q2)),
        "chick_cost": float(Decimal(str(chick_cost)).quantize(q2)),
        "med_cost": float(med_cost.quantize(q2)),
        "admin_cost": float(admin_cost.quantize(q2)),
    }
    _cb_total = sum(cost_breakdown.values())
    cost_breakdown["total"] = round(_cb_total, 2)
    for _k in ("feed_cost", "chick_cost", "med_cost", "admin_cost"):
        cost_breakdown[f"{_k}_pct"] = round(cost_breakdown[_k] / _cb_total * 100, 1) if _cb_total else 0

    # Performance Overview (dashboard widget) — actual vs the Growing Charge
    # Master's standard rates, for the 3 metrics the master actually defines
    # a target for (standard_fcr, standard_mortality, std_production_cost).
    # Avg Live Weight / Feed-per-Bird / CFCR have no master-defined target in
    # this system, so they're deliberately left off rather than inventing one.
    def _perf_status(actual, target):
        if not target:
            return "No Target"
        actual, target = Decimal(str(actual)), Decimal(str(target))
        if actual <= target:
            return "Good"
        if actual <= target * Decimal("1.05"):
            return "Medium"
        return "High"

    performance_overview = [
        {"kpi": "Mortality %", "actual": total_mort_pct.quantize(q2),
         "target": scheme.standard_mortality if scheme else None,
         "status": _perf_status(total_mort_pct, scheme.standard_mortality if scheme else None)},
        {"kpi": "FCR", "actual": fcr.quantize(q2),
         "target": scheme.standard_fcr if scheme else None,
         "status": _perf_status(fcr, scheme.standard_fcr if scheme else None)},
        {"kpi": "Production Cost / Kg", "actual": production_cost_per_kg.quantize(q2),
         "target": scheme.std_production_cost if scheme else None,
         "status": _perf_status(production_cost_per_kg, scheme.std_production_cost if scheme else None)},
    ]

    # --- Grade: Farmer-Classification band matched on production cost/kg ---
    grade = ""
    if scheme:
        for fc in scheme.farmer_classifications.all():
            if fc.production_cost_from <= production_cost_per_kg <= fc.production_to:
                grade = fc.grade
                break

    # --- Financials (Management view) ---
    revenue = Decimal(str(sold_amount))
    gross_profit = revenue - total_production_cost
    # No overhead ledger (labour/electricity/fuel/...) exists, so Net == Gross here.
    net_profit = gross_profit
    prod_cost_per_bird = _div(total_production_cost, placement_total)
    margin_per_bird = _div(gross_profit, sold_birds)
    margin_per_kg = _div(gross_profit, sold_weight)
    roi = _div(gross_profit * 100, total_production_cost)

    # --- Standard rates from the Growing Charge Master (settlement basis;
    #     NEVER actual company purchase costs) ---
    _weight = Decimal(str(sold_weight))
    std_chick_rate = Decimal(str(scheme.chick_cost)) if scheme else Decimal("0")
    std_feed_rate = Decimal(str(scheme.feed_cost)) if scheme else Decimal("0")
    std_med_rate = Decimal(str(scheme.medicine_cost)) if scheme else Decimal("0")
    std_prod_per_kg = Decimal(str(scheme.std_production_cost)) if scheme else Decimal("0")
    base_gc_rate = Decimal(str(scheme.standard_gc_cost)) if scheme else Decimal("0")
    base_gc_amount = base_gc_rate * _weight
    std_prod_total = std_prod_per_kg * _weight
    prod_cost_variance_kg = std_prod_per_kg - production_cost_per_kg
    prod_cost_variance_total = std_prod_total - total_production_cost

    data = {
        "revenue": revenue.quantize(q2),
        "gross_profit": gross_profit.quantize(q2),
        "net_profit": net_profit.quantize(q2),
        "prod_cost_per_bird": prod_cost_per_bird.quantize(q2),
        "margin_per_bird": margin_per_bird.quantize(q2),
        "margin_per_kg": margin_per_kg.quantize(q2),
        "roi": roi.quantize(q2),
        "std_chick_rate": std_chick_rate.quantize(q2),
        "std_feed_rate": std_feed_rate.quantize(q2),
        "std_med_rate": std_med_rate.quantize(q2),
        "std_prod_per_kg": std_prod_per_kg.quantize(q2),
        "base_gc_rate": base_gc_rate.quantize(q2),
        "base_gc_amount": base_gc_amount.quantize(q2),
        "std_prod_total": std_prod_total.quantize(q2),
        "prod_cost_variance_kg": prod_cost_variance_kg.quantize(q2),
        "prod_cost_variance_total": prod_cost_variance_total.quantize(q2),
        "has_scheme": bool(scheme),
        "placement_date": placement_date,
        "sale_start_date": sale_start_date,
        "chicks_placed": placement_total,
        "mortality": cum_mortality,
        "culls": cum_culls,
        "excess_birds": excess_birds,
        "shortage_birds": shortage_birds,
        "grade": grade,
        "first_week_mort_pct": first_week_mort_pct.quantize(q2),
        "upto_30_mort_pct": upto_30_mort_pct.quantize(q2),
        "after_30_mort_pct": after_30_mort_pct.quantize(q2),
        "total_mort_pct": total_mort_pct.quantize(q2),
        "avg_body_weight": avg_body_weight.quantize(q2),
        "mean_age": mean_age.quantize(q2),
        "day_gain": day_gain.quantize(q2),
        "fcr": fcr.quantize(q2),
        "cfcr": cfcr.quantize(q2),
        "livability_pct": livability_pct.quantize(q2),
        "mortality_weight": mortality_weight.quantize(q2),
        "eef": eef.quantize(q2),
        "sold_birds": sold_birds,
        "sold_weight": Decimal(str(sold_weight)).quantize(q2),
        "sold_amount": Decimal(str(sold_amount)).quantize(q2),
        "avg_sale_rate": _div(sold_amount, sold_weight).quantize(q2),
        "feed_sent": Decimal(str(feed_in_kg)).quantize(q2),
        "feed_consumed": Decimal(str(feed_consumed)).quantize(q2),
        "feed_return": Decimal(str(feed_return_kg)).quantize(q2),
        "feed_cost": feed_cost.quantize(q2),
        "chick_cost": Decimal(str(chick_cost)).quantize(q2),
        "med_sent": Decimal(str(med_in_qty)).quantize(q2),
        "med_consumed": Decimal(str(med_consumed)).quantize(q2),
        "med_return": Decimal(str(med_return_qty)).quantize(q2),
        "med_cost": med_cost.quantize(q2),
        "admin_cost": admin_cost.quantize(q2),
        "total_production_cost": total_production_cost.quantize(q2),
        "production_cost_per_kg": production_cost_per_kg.quantize(q2),
        "scheme_code": scheme.scheme_code if scheme else "",
        "cost_breakdown": cost_breakdown,
        "performance_overview": performance_overview,
        "mortality_status": performance_overview[0]["status"],
        "fcr_status": performance_overview[1]["status"],
        "cost_status": performance_overview[2]["status"],
    }

    # No Growing Charge Scheme covers the placement date -> the scheme-driven
    # cost/settlement/variance figures are undefined; show "No Data" for them
    # (operational + performance figures still show real values).
    if not scheme:
        for key in ("admin_cost", "grade", "total_production_cost", "production_cost_per_kg",
                    "prod_cost_per_bird", "gross_profit", "net_profit", "margin_per_bird",
                    "margin_per_kg", "roi", "std_chick_rate", "std_feed_rate", "std_med_rate",
                    "std_prod_per_kg", "base_gc_rate", "base_gc_amount", "std_prod_total",
                    "prod_cost_variance_kg", "prod_cost_variance_total"):
            data[key] = "No Data"
    return data


def _build_batch_report(batch, fetch_type="farmer", scheme_override=None):
    from inventory.models import StockTransfer, MedicineTransfer, Mapping
    from purchase.models import GeneralPurchaseItem

    # Feed Purchase: purchases have no batch/farm FK at all, only a
    # Warehouse — sometimes that Warehouse directly *is* the farm's own
    # cost centre (feed bought straight to the farm), sometimes it's a
    # general warehouse the feed reaches the farm from later via a Stock
    # Transfer (already captured as Feed Transfer-In). Either way it's
    # booked to a Warehouse sharing this farm's Branch, so — like Bird Sale
    # Receipt's balance — it's rolled up at the Branch level, bounded to
    # this batch's growing window, rather than claimed to be exact-to-batch.
    # start_date/end_date are both nullable (the Batch form sets neither),
    # so each bound is applied only when it exists — an unbounded batch
    # simply shows every purchase at the branch.
    # A Warehouse (Office) no longer has a direct Branch FK — its Branch is
    # resolved through inventory.Mapping (TYPE_SECTOR_BRANCH: from_id=office,
    # to_id=branch), so we first gather the offices mapped to this branch.
    branch_id = batch.broiler_farm.branch_id
    branch_office_ids = list(
        Mapping.objects.filter(type=Mapping.TYPE_SECTOR_BRANCH, to_id=branch_id)
        .values_list("from_id", flat=True)
    )
    purchase_items = GeneralPurchaseItem.objects.filter(farm_warehouse_id__in=branch_office_ids)
    _placed = _placement_date(batch)
    if _placed:
        purchase_items = purchase_items.filter(purchase__date__gte=_placed)
    if batch.end_date:
        purchase_items = purchase_items.filter(purchase__date__lte=batch.end_date)
    purchase_items = (purchase_items
                      .select_related("purchase", "item__category", "farm_warehouse")
                      .order_by("purchase__date", "id"))
    feed_purchase_rows = []
    for pi in purchase_items:
        category_name = pi.item.category.name if pi.item.category_id else ""
        if "chick" in category_name.lower():
            continue
        feed_purchase_rows.append({
            "date": pi.purchase.date, "trnum": pi.purchase.purchase_no, "dc_no": pi.purchase.dc_no,
            "from_location": pi.farm_warehouse.name, "item": str(pi.item),
            "quantity": pi.rcv_qty, "rate": pi.rate, "amount": pi.amount,
        })

    transfers = (StockTransfer.objects.filter(to_batch=batch)
                 .select_related("item__category", "from_warehouse", "from_farm")
                 .order_by("date", "id"))
    chick_rows, feed_rows = [], []
    feed_cum = Decimal("0")
    for t in transfers:
        row = {
            "date": t.date, "trnum": t.trnum, "dc_no": t.dc_no,
            "from_location": _transfer_location_name(t),
            "item": str(t.item), "quantity": t.quantity, "rate": t.rate,
            "amount": (t.quantity or 0) * (t.rate or 0),
        }
        category_name = t.item.category.name if t.item.category_id else ""
        # Chick-placement transfers are ordinary Stock Transfers of a
        # "chicks" item; everything else transferred to a batch is treated
        # as feed (this model has no dedicated chick-placement transaction).
        if "chick" in category_name.lower():
            chick_rows.append(row)
        else:
            feed_cum += row["quantity"] or 0
            row["cumulative"] = feed_cum
            feed_rows.append(row)

    med_transfers = (MedicineTransfer.objects.filter(to_batch=batch)
                     .prefetch_related("items__item").select_related("from_warehouse", "from_farm")
                     .order_by("date", "id"))
    medicine_transfer_rows = []
    for mt in med_transfers:
        location_name = _transfer_location_name(mt)
        for line in mt.items.all():
            medicine_transfer_rows.append({
                "date": mt.date, "trnum": mt.trnum, "dc_no": mt.dc_no,
                "from_location": location_name, "item": str(line.item),
                "quantity": line.quantity, "rate": line.rate,
                "amount": (line.quantity or 0) * (line.rate or 0),
            })

    # Feed/Medicine Return and Transfer-to-Other-Farm are the same "moved
    # OUT of this batch" leg (from_batch) of the same Stock/Medicine
    # Transfer transactions — split by destination: back to a warehouse
    # (Return) vs on to a different farm/batch (Transfer to Other Farm).
    # Neither is a distinct "return" transaction type in this system.
    outgoing_transfers = (StockTransfer.objects.filter(from_batch=batch)
                          .select_related("item__category", "to_warehouse", "to_farm")
                          .order_by("date", "id"))
    feed_return_rows, feed_transfer_out_rows = [], []
    for t in outgoing_transfers:
        row = {
            "date": t.date, "trnum": t.trnum, "dc_no": t.dc_no,
            "to_location": _transfer_to_location_name(t),
            "item": str(t.item), "quantity": t.quantity, "rate": t.rate,
            "amount": (t.quantity or 0) * (t.rate or 0),
        }
        (feed_return_rows if t.to_warehouse_id else feed_transfer_out_rows).append(row)

    outgoing_med_transfers = (MedicineTransfer.objects.filter(from_batch=batch)
                              .prefetch_related("items__item").select_related("to_warehouse", "to_farm")
                              .order_by("date", "id"))
    medicine_return_rows, medicine_transfer_out_rows = [], []
    for mt in outgoing_med_transfers:
        location_name = _transfer_to_location_name(mt)
        target = medicine_return_rows if mt.to_warehouse_id else medicine_transfer_out_rows
        for line in mt.items.all():
            target.append({
                "date": mt.date, "trnum": mt.trnum, "dc_no": mt.dc_no,
                "to_location": location_name, "item": str(line.item),
                "quantity": line.quantity, "rate": line.rate,
                "amount": (line.quantity or 0) * (line.rate or 0),
            })

    # Placement baseline for Opening Birds / Cum Mort% / Feed-per-bird below
    # is the Chick Placement total itself — BroilerBatch has no count field,
    # but every placement is a real Chick Placement Stock Transfer, so the
    # baseline is just the sum of those quantities, not a new data point.
    placement_total = sum((r["quantity"] or 0) for r in chick_rows)

    def _bags(qty, item):
        return (qty / item.kg_per_bag) if item and item.kg_per_bag else Decimal("0")

    sold_by_date = {}
    for bs in BirdSale.objects.filter(batch=batch).values("date").annotate(total=Sum("birds")):
        sold_by_date[bs["date"]] = bs["total"] or 0

    daily_entries = DailyEntry.objects.filter(batch=batch).select_related("feed_1", "feed_2").order_by("date", "id")
    mortality_rows = []
    cum_mortality = cum_culls = 0
    cum_feed_kg = Decimal("0")
    opening_birds = placement_total
    for de in daily_entries:
        cum_mortality += de.mortality
        cum_culls += de.culls
        sold = sold_by_date.get(de.date, 0)
        closing_birds = opening_birds - de.mortality - de.culls - sold
        feed_1_kg = de.feed_1_qty or Decimal("0")
        feed_2_kg = de.feed_2_qty or Decimal("0")
        cum_feed_kg += feed_1_kg + feed_2_kg
        avg_bw_kg = (de.avg_weight_gms or Decimal("0")) / Decimal("1000")
        mortality_rows.append({
            "date": de.date, "age": de.age_days,
            "opening_birds": opening_birds,
            "mortality": de.mortality,
            "mortality_pct": round(de.mortality / opening_birds * 100, 2) if opening_birds else 0,
            "culls": de.culls, "cum_mortality": cum_mortality,
            "cum_mortality_pct": round(cum_mortality / placement_total * 100, 2) if placement_total else 0,
            "sold_birds": sold, "closing_birds": closing_birds,
            "avg_weight_kg": round(avg_bw_kg, 3),
            "fcr": round(cum_feed_kg / (closing_birds * avg_bw_kg), 2) if closing_birds and avg_bw_kg else 0,
            "feed_1_name": de.feed_1.description if de.feed_1_id else "", "feed_1_kg": feed_1_kg,
            "feed_1_bags": round(_bags(feed_1_kg, de.feed_1), 2),
            "feed_2_name": de.feed_2.description if de.feed_2_id else "", "feed_2_kg": feed_2_kg,
            "feed_2_bags": round(_bags(feed_2_kg, de.feed_2), 2),
            "cum_feed_kg": cum_feed_kg,
            "feed_per_bird_g": round(cum_feed_kg * 1000 / placement_total, 2) if placement_total else 0,
            "balance_feed_kg": (de.feed_1_stock or 0) + (de.feed_2_stock or 0),
            "balance_feed_bags": round(_bags(de.feed_1_stock or Decimal("0"), de.feed_1)
                                       + _bags(de.feed_2_stock or Decimal("0"), de.feed_2), 2),
            "remarks": de.remarks,
        })
        opening_birds = closing_birds

    # Weekly subtotal rows, interleaved after every 7th daily row: state
    # columns (opening/closing/cumulative/balance/%) repeat the week's LAST
    # day; only the flow quantities (mortality, culls, sold, feed consumed)
    # are summed across the week.
    mortality_display = []
    for i in range(0, len(mortality_rows), 7):
        week_rows = mortality_rows[i:i + 7]
        mortality_display.extend({**r, "is_total": False} for r in week_rows)
        last = week_rows[-1]
        mortality_display.append({
            "is_total": True, "label": f"Week {i // 7 + 1} Total",
            "opening_birds": week_rows[0]["opening_birds"],
            "mortality": sum(r["mortality"] for r in week_rows),
            "mortality_pct": last["cum_mortality_pct"],
            "culls": sum(r["culls"] for r in week_rows),
            "cum_mortality": last["cum_mortality"], "cum_mortality_pct": last["cum_mortality_pct"],
            "sold_birds": sum(r["sold_birds"] for r in week_rows), "closing_birds": last["closing_birds"],
            "avg_weight_kg": last["avg_weight_kg"], "fcr": last["fcr"],
            "feed_1_kg": sum(r["feed_1_kg"] for r in week_rows), "feed_1_bags": sum(r["feed_1_bags"] for r in week_rows),
            "feed_2_kg": sum(r["feed_2_kg"] for r in week_rows), "feed_2_bags": sum(r["feed_2_bags"] for r in week_rows),
            "cum_feed_kg": last["cum_feed_kg"], "feed_per_bird_g": last["feed_per_bird_g"],
            "balance_feed_kg": last["balance_feed_kg"], "balance_feed_bags": last["balance_feed_bags"],
        })

    medicine_entries = MedicineVaccineEntry.objects.filter(batch=batch).select_related("item").order_by("date", "id")
    medicine_consumption_rows = [{
        "date": me.date, "age": me.age_days, "item": str(me.item) if me.item_id else "",
        "quantity": me.qty, "stock": me.stock, "remarks": me.remarks,
    } for me in medicine_entries]

    bird_sales = (BirdSale.objects.filter(batch=batch).select_related("customer", "farmer")
                  .order_by("date", "id"))
    bird_sale_rows = [{
        "date": bs.date, "sale_no": bs.sale_no, "doc_no": bs.doc_no,
        "buyer_name": bs.customer.name if bs.customer_id else (bs.farmer.farmer_name if bs.farmer_id else ""),
        "birds": bs.birds, "net_weight": bs.net_weight, "avg_weight": bs.avg_weight,
        "rate": bs.rate, "round_off": bs.round_off, "amount": bs.amount,
        "vehicle": bs.vehicle, "driver": bs.driver, "remarks": bs.remarks,
    } for bs in bird_sales]

    # Feed Summary: per feed item, Purchase + Transfer In - Consumed -
    # Return - Transfer to Other Farms = Balance (Kg). Bucketed by item_code
    # (not the row tables' display string) so purchases, transfers, and
    # DailyEntry consumption of the same item always land in one bucket.
    feed_summary = {}

    def _feed_bucket(item_code):
        return feed_summary.setdefault(item_code, {"purchased": Decimal("0"), "transfer_in": Decimal("0"),
                                                    "consumed": Decimal("0"), "returned": Decimal("0"),
                                                    "transferred_out": Decimal("0")})

    for pi in purchase_items:
        category_name = pi.item.category.name if pi.item.category_id else ""
        if "chick" not in category_name.lower():
            _feed_bucket(pi.item.item_code)["purchased"] += (pi.rcv_qty or 0) + (pi.free_qty or 0)
    for t in transfers:
        if not t.item.category_id or "chick" not in t.item.category.name.lower():
            _feed_bucket(t.item.item_code)["transfer_in"] += t.quantity or 0
    for t in outgoing_transfers:
        if t.to_warehouse_id:
            _feed_bucket(t.item.item_code)["returned"] += t.quantity or 0
        else:
            _feed_bucket(t.item.item_code)["transferred_out"] += t.quantity or 0
    for de in daily_entries:
        if de.feed_1_id:
            _feed_bucket(de.feed_1.item_code)["consumed"] += de.feed_1_qty or 0
        if de.feed_2_id:
            _feed_bucket(de.feed_2.item_code)["consumed"] += de.feed_2_qty or 0
    # Bucketed by item_code (stable key), but display the item name.
    feed_name_by_code = dict(
        Item.objects.filter(item_code__in=feed_summary.keys()).values_list("item_code", "description")
    )
    feed_summary_rows = [{
        "item": feed_name_by_code.get(item_code, item_code), **b,
        "balance": b["purchased"] + b["transfer_in"] - b["consumed"] - b["returned"] - b["transferred_out"],
    } for item_code, b in feed_summary.items()]

    # Bird-sale totals (reused by the Total row and the costing block)
    _bs_birds = sum(r["birds"] for r in bird_sale_rows)
    _bs_weight = sum(r["net_weight"] for r in bird_sale_rows)
    _bs_amount = sum(r["amount"] for r in bird_sale_rows)

    batch_costing = _build_batch_costing(
        batch, placement_total, cum_mortality, cum_culls, mortality_rows,
        chick_rows, feed_rows, feed_summary_rows, feed_return_rows,
        medicine_transfer_rows, medicine_consumption_rows, medicine_return_rows,
        bird_sale_rows, fetch_type=fetch_type, scheme_override=scheme_override,
    )

    # Dashboard KPI tiles + trend charts (Feed Consumption / Mortality) — as
    # of the batch's latest DailyEntry, from the same per-day mortality_rows
    # already computed above (pre weekly-interleave), not a new data source.
    last_entry = mortality_rows[-1] if mortality_rows else None
    dashboard = {
        "current_live_birds": last_entry["closing_birds"] if last_entry else placement_total,
        "avg_live_weight_kg": last_entry["avg_weight_kg"] if last_entry else Decimal("0"),
        "as_of_date": last_entry["date"] if last_entry else None,
        "feed_per_bird_g": last_entry["feed_per_bird_g"] if last_entry else 0,
        "chart": {
            "dates": [r["date"].strftime("%d %b") for r in mortality_rows],
            "daily_feed_kg": [float(r["feed_1_kg"] + r["feed_2_kg"]) for r in mortality_rows],
            "cum_feed_kg": [float(r["cum_feed_kg"]) for r in mortality_rows],
            "daily_mortality_pct": [float(r["mortality_pct"]) for r in mortality_rows],
            "cum_mortality_pct": [float(r["cum_mortality_pct"]) for r in mortality_rows],
            "fcr": [float(r["fcr"]) for r in mortality_rows],
        },
    }

    return {
        "batch_costing": batch_costing,
        "dashboard": dashboard,
        "chick_placement": chick_rows,
        "feed_purchase": feed_purchase_rows,
        "feed_transfer_in": feed_rows,
        "feed_return": feed_return_rows,
        "feed_transfer_out": feed_transfer_out_rows,
        "feed_summary": feed_summary_rows,
        "medicine_transfer_in": medicine_transfer_rows,
        "medicine_return": medicine_return_rows,
        "medicine_transfer_out": medicine_transfer_out_rows,
        "mortality": mortality_display,
        "placement_total": placement_total,
        "medicine_consumption": medicine_consumption_rows,
        "bird_sales": bird_sale_rows,
        "totals": {
            "birds": _bs_birds,
            "net_weight": _bs_weight,
            "amount": _bs_amount,
            # Overall averages for the Total row: Avg Wt = weight/birds, Rate = amount/weight
            "avg_weight": _div(_bs_weight, _bs_birds).quantize(Decimal("0.01")),
            "avg_sale_rate": _div(_bs_amount, _bs_weight).quantize(Decimal("0.01")),
            "mortality": cum_mortality, "culls": cum_culls,
        },
    }


@login_required
def broiler_batch_report(request):
    """One Batch's full growing history — feed purchase, chick placement,
    feed/medicine transfers-in and returns/transfers-out, a feed summary,
    daily mortality & feed consumption, medicine consumption, and bird
    sales (Broiler > Reports > Batch History Report).

    Feed Purchase has no batch/farm FK at all (only a Warehouse), so it's
    rolled up at the Branch level and bounded to the batch's growing
    window rather than claimed to be exact-to-batch. The Batch Costing
    Information and Summary block uses the Chick Placement total as the
    placement baseline and pulls Admin Cost / Grade from the batch's
    applicable GrowingChargeScheme (see _build_batch_costing).
    """
    from account.models import CompanyProfile

    batch_id = (request.GET.get("batch") or "").strip()
    # Filter-bar selections
    fetch_type = (request.GET.get("fetch_type") or "farmer").strip().lower()
    if fetch_type not in ("farmer", "management"):
        fetch_type = "farmer"
    book_no = (request.GET.get("book_no") or "").strip()
    export = (request.GET.get("export") or "display").strip().lower()
    schema_id = (request.GET.get("schema") or "").strip()

    batch = (scope_multi(request.user,
                         BroilerBatch.objects
                         .select_related("broiler_farm__branch",
                                         "broiler_farm__supervisor",
                                         "broiler_farm__farmer"),
                         farms="broiler_farm_id",
                         branches="broiler_farm__branch_id")
             .filter(id=batch_id).first()) if batch_id else None
    # A hand-picked Schema overrides the auto-matched Growing Charge Scheme.
    scheme_override = (GrowingChargeScheme.objects.filter(id=schema_id).first()
                       if schema_id.isdigit() else None)

    # Schema dropdown: only schemes whose master-defined date range covers the
    # batch's placement date (region-matched); auto-select the applicable one.
    schemes = GrowingChargeScheme.objects.filter(is_active=True)
    selected_schema_id = int(schema_id) if schema_id.isdigit() else None
    if batch:
        placement = _placement_date(batch)
        region_id = batch.broiler_farm.branch.region_id
        if placement:
            schemes = schemes.filter(region_id=region_id,
                                     from_date__lte=placement, to_date__gte=placement)
        else:
            schemes = schemes.filter(region_id=region_id)
        if selected_schema_id is None:
            matched = _match_growing_charge_scheme(batch, placement)
            selected_schema_id = matched.id if matched else None
    schemes = schemes.order_by("schema_name")

    return render(request, "broiler_batch_report.html", {
        "farms": farms_for(request.user, BroilerFarm.objects.order_by("farm_name")),
        "batches": scope_multi(request.user,
                               BroilerBatch.objects.select_related("broiler_farm")
                               .order_by("-start_date", "-id"),
                               farms="broiler_farm_id",
                               branches="broiler_farm__branch_id"),
        "schemes": schemes,
        "batch": batch,
        "batch_requested": bool(batch_id),
        "report": _build_batch_report(batch, fetch_type=fetch_type,
                                      scheme_override=scheme_override) if batch else None,
        "company": CompanyProfile.get_solo(),
        "fetch_type": fetch_type,
        "fetch_type_label": "Management" if fetch_type == "management" else "Farmer",
        "book_no": book_no,
        "export": export,
        "selected_schema_id": selected_schema_id,
    })


# ---------------------------------------------------------------------------
# Live Flock Summary Report (Broiler > Reports)
# ---------------------------------------------------------------------------

def _breed_standard_at(breed_id, age):
    """The breed's standard row at `age` — exact, else the nearest row at or
    below it, carrying the curve forward.

    None when the age is below the curve's first row. It used to return that
    first row instead, so a flock on day 12 against a breed whose standards
    start at day 20 was measured against a 20-day-old bird and read as badly
    behind. A blank says "no standard for this age", which is true; a wrong
    number does not announce itself.
    """
    if not breed_id or age is None:
        return None
    return (BreedStandard.objects.filter(breed_id=breed_id, age__lte=age)
            .order_by("-age").first())


def _interp_standard(breed_id, key_field, target, out_field):
    """Read the breed's standard curve at a *value* of `key_field` instead of at
    an age — e.g. "what's the standard cum_feed when body_weight = X", or the
    reverse. Walks the curve ordered by age (both body_weight and cum_feed rise
    with age) and linearly interpolates `out_field` where `key_field` crosses
    `target`; clamps to the end rows outside the curve. Returns a Decimal, or
    None when the breed has no usable rows."""
    if not breed_id or target is None:
        return None
    rows = [(Decimal(str(k)), Decimal(str(v)))
            for k, v in BreedStandard.objects.filter(breed_id=breed_id)
            .order_by("age").values_list(key_field, out_field)
            if k is not None and v is not None]
    if not rows:
        return None
    target = Decimal(str(target))
    if target <= rows[0][0]:
        return rows[0][1]
    if target >= rows[-1][0]:
        return rows[-1][1]
    for (k0, v0), (k1, v1) in zip(rows, rows[1:]):
        if k0 <= target <= k1:
            if k1 == k0:
                return v0
            return v0 + (v1 - v0) * (target - k0) / (k1 - k0)
    return rows[-1][1]


def _flock_valuation_remark(*, not_started, avg_bwt, std_bwt, fcr, std_fcr,
                            feed_con, std_feed_at_bwt, mort_pct, gap_days):
    """Auto-compose a one-line valuation of a live flock from its own numbers:
    an overall verdict (On Track / Watch / Behind / Critical) plus the drivers
    (weight vs std, feed vs the standard-for-that-weight, FCR vs std, mortality,
    data freshness). Purely derived — no stored field."""
    if not_started:
        return "Not Started — chicks placed, no daily entry yet"

    f = lambda x: float(x) if x is not None else None
    avg_bwt, std_bwt, fcr, std_fcr = f(avg_bwt), f(std_bwt), f(fcr), f(std_fcr)
    feed_con, std_feed_at_bwt, mort_pct = f(feed_con), f(std_feed_at_bwt), f(mort_pct)

    parts, penalty = [], 0
    # Body weight vs the age-standard
    if std_bwt and avg_bwt:
        dev = (avg_bwt - std_bwt) / std_bwt * 100
        if dev <= -20:
            parts.append(f"B.Wt {abs(dev):.0f}% below std"); penalty += 3
        elif dev <= -10:
            parts.append(f"B.Wt {abs(dev):.0f}% below std"); penalty += 2
        elif dev >= 10:
            parts.append(f"B.Wt {dev:.0f}% above std")
        else:
            parts.append("B.Wt on target")
    # Feed used vs what the standard bird eats to reach the SAME weight
    if std_feed_at_bwt and feed_con:
        fe = (feed_con - std_feed_at_bwt) / std_feed_at_bwt * 100
        if fe >= 15:
            parts.append(f"over-fed {fe:.0f}% for weight"); penalty += 2
        elif fe >= 8:
            parts.append(f"over-fed {fe:.0f}% for weight"); penalty += 1
        elif fe <= -10:
            parts.append(f"under-fed {abs(fe):.0f}% for weight")
    # FCR vs std
    if std_fcr and fcr:
        if fcr > std_fcr * 1.10:
            parts.append("FCR well above std"); penalty += 2
        elif fcr > std_fcr * 1.03:
            parts.append("FCR above std"); penalty += 1
        elif fcr < std_fcr * 0.97:
            parts.append("FCR better than std")
    # Mortality
    if mort_pct is not None:
        if mort_pct >= 8:
            parts.append(f"high mortality {mort_pct:.1f}%"); penalty += 3
        elif mort_pct >= 5:
            parts.append(f"mortality {mort_pct:.1f}%"); penalty += 1
    # Data freshness
    if gap_days and gap_days >= 2:
        parts.append(f"entries {gap_days}d stale"); penalty += 1

    verdict = ("Critical" if penalty >= 5 else "Behind" if penalty >= 3
               else "Watch" if penalty >= 1 else "On Track")
    return f"{verdict} — " + ("; ".join(parts) if parts else "all metrics near standard")


def _live_flock_row(batch, today):
    """One report row for a live flock, reusing the batch report engine plus
    the age/feed-outlook columns the fleet view needs."""
    from inventory.models import StockTransfer

    report = _build_batch_report(batch, fetch_type="management")
    bc = report["batch_costing"]
    farm = batch.broiler_farm

    placed = _num(bc.get("chicks_placed"))
    mort = _num(bc.get("mortality"))
    culls = _num(bc.get("culls"))
    sold = _num(bc.get("sold_birds"))
    placement_date = bc.get("placement_date") or batch.start_date
    actual_age = (today - placement_date).days if placement_date else 0

    # latest daily entry + gap
    entries = list(DailyEntry.objects.filter(batch=batch).order_by("-date")[:3])
    latest = entries[0].date if entries else None
    gap_days = (today - latest).days if latest else None

    # last *body-weight* reading (skip days with no weight taken) + its gap
    last_wt = (DailyEntry.objects.filter(batch=batch, avg_weight_gms__gt=0)
               .order_by("-date").first())
    avg_bwt = last_wt.avg_weight_gms if last_wt else Decimal("0")
    last_bwt_date = last_wt.date if last_wt else None
    last_bwt_gap = (today - last_bwt_date).days if last_bwt_date else None

    # recent daily feed rate (last <=3 entries), fallback to overall
    feed_consumed = _num(bc.get("feed_consumed"))
    recent = sum((_num(e.feed_1_qty) + _num(e.feed_2_qty)) for e in entries)
    daily_rate = (recent / len(entries)) if entries else (
        feed_consumed / actual_age if actual_age else Decimal("0"))

    # farm<->farm feed movements (subset of the engine's in/out)
    feed_item_ids = list(Item.objects.filter(category__name__icontains="feed").values_list("id", flat=True))
    transfer_in_farms = StockTransfer.objects.filter(
        to_batch=batch, from_location_type="farm", item_id__in=feed_item_ids
    ).aggregate(t=Sum("quantity"))["t"] or Decimal("0")
    transfer_out_farms = sum((_num(r["quantity"]) for r in report.get("feed_transfer_out", [])), Decimal("0"))

    feed_sent = _num(bc.get("feed_sent"))
    feed_return = _num(bc.get("feed_return"))
    feed_balance = feed_sent - feed_consumed - feed_return - transfer_out_farms

    std = _breed_standard_at(batch.breed_id, actual_age)
    std_bwt = std.body_weight if std else None   # grams (matches Avg B.Wt in grams)
    std_fcr = std.fcr if std else None
    # cum_feed is per-bird grams -> total kg = cum_feed(g) x birds / 1000 (matches Feed Con in kg)
    std_feed_con = (std.cum_feed * placed / Decimal("1000")) if std else None

    # Weight-adjusted (age-independent) standards: read the curve at the flock's
    # ACTUAL weight/feed rather than its age. "Std Feed @ B.Wt" = the cum_feed a
    # standard bird eats to reach the birds' current body weight; "Std B.Wt @
    # Feed" = the body weight a standard bird has after eating the feed actually
    # consumed per bird. Lets you judge feed efficiency independent of how far
    # the flock is ahead/behind on age.
    actual_feed_per_bird = (feed_consumed * Decimal("1000") / placed) if placed else None  # g/bird
    std_feed_at_bwt_g = _interp_standard(batch.breed_id, "body_weight", avg_bwt, "cum_feed")   # g/bird
    std_feed_at_bwt = (std_feed_at_bwt_g * placed / Decimal("1000")) if std_feed_at_bwt_g is not None else None  # total kg
    std_bwt_at_feed = _interp_standard(batch.breed_id, "cum_feed", actual_feed_per_bird, "body_weight")  # g/bird

    # Live-flock metrics on the CURRENT weight on hand (sold weight + current
    # live weight = available birds x last body weight), not the sale-only
    # weight — so a still-growing flock shows real FCR/CFCR/PC/Kg instead of 0
    # (no sales) or a blown-up CFCR (feed / tiny mortality weight). For a sold-
    # out flock live_weight is 0, so this reduces to the sale-based figures.
    available = placed - mort - culls - sold
    live_weight_kg = available * (avg_bwt / Decimal("1000"))
    sold_weight = _num(bc.get("sold_weight"))
    mortality_weight = _num(bc.get("mortality_weight"))
    total_weight_now = sold_weight + live_weight_kg
    total_prod_cost = _num(bc.get("total_production_cost"))

    m_pc_kg = (_div(total_prod_cost, total_weight_now) if total_weight_now > 0
               else _num(bc.get("production_cost_per_kg"))).quantize(Decimal("0.01"))
    fcr_val = _div(feed_consumed, total_weight_now).quantize(Decimal("0.001")) if total_weight_now > 0 else Decimal("0")
    cfcr_val = (_div(feed_consumed, total_weight_now + mortality_weight).quantize(Decimal("0.01"))
                if (total_weight_now + mortality_weight) > 0 else Decimal("0"))

    # EEF: keep the engine's sale-based value once the flock has sales (uses
    # mean lifting age); for a still-growing flock compute it on current values.
    if sold_weight > 0:
        eef_val = _num(bc.get("eef"))
    elif actual_age and fcr_val:
        surviving = sold + available
        livability = _div(surviving * 100, placed)
        avg_wt_kg = _div(total_weight_now, surviving)
        eef_val = _div(livability * avg_wt_kg * 100, Decimal(str(actual_age)) * fcr_val).quantize(Decimal("0.01"))
    else:
        eef_val = Decimal("0")

    q2 = Decimal("0.01")
    remark = _flock_valuation_remark(
        not_started=bool(placed > 0 and latest is None),
        avg_bwt=avg_bwt, std_bwt=std_bwt, fcr=fcr_val, std_fcr=std_fcr,
        feed_con=feed_consumed, std_feed_at_bwt=std_feed_at_bwt,
        mort_pct=_num(bc.get("total_mort_pct")), gap_days=gap_days)

    # --- Feed phase + Max Feed Qty (kg/bird) changeover, matching Daily Entry ---
    from broiler.models import BirdSale as _BirdSale
    phase = resolve_feed_phase(batch, today, actual_age)
    phase_name = phase.get("phase_name") if phase else ""
    next_phase = phase.get("next_name") if phase else ""
    phase_cap = phase_cum_bird = None
    if phase:
        cap = _num(phase.get("max_feed_qty"))
        phase_cap = cap if cap > 0 else None
        item_id = phase.get("feed_item")
        if item_id and phase_cap:
            agg = DailyEntry.objects.filter(batch=batch).aggregate(
                f1=Sum("feed_1_qty", filter=Q(feed_1_id=item_id)),
                f2=Sum("feed_2_qty", filter=Q(feed_2_id=item_id)))
            item_kg = _num(agg["f1"]) + _num(agg["f2"])
            denom = available if available > 0 else placed
            phase_cum_bird = (item_kg / denom).quantize(Decimal("0.001")) if denom else None

    # --- Actual feed eaten per SURVIVING bird (bird-day weighted) ---
    sold_by_date = {}
    for _bs in _BirdSale.objects.filter(batch=batch).values("date").annotate(t=Sum("birds")):
        sold_by_date[_bs["date"]] = _bs["t"] or 0
    _alive = int(placed)
    _cum_pb = Decimal("0")
    for _de in DailyEntry.objects.filter(batch=batch).order_by("date", "id"):
        _fd = _num(_de.feed_1_qty) + _num(_de.feed_2_qty)
        if _alive > 0 and _fd:
            _cum_pb += _fd / Decimal(_alive) * Decimal("1000")
        _alive -= (_de.mortality or 0) + (_de.culls or 0) + sold_by_date.get(_de.date, 0)
    act_feed_bird_live = _cum_pb.quantize(q2)
    std_feed_bird_daily = std.feed_intake if std else None

    return {
        "branch": farm.branch.branch_name if farm.branch_id else "",
        "line": farm.line or "",
        "supervisor": farm.supervisor.name if farm.supervisor_id else "",
        "farmer": farm.farmer.farmer_name if farm.farmer_id else "",
        "batch": batch.batch_name,
        "book_no": batch.book_number or "",
        "actual_age": actual_age,
        "placement_date": placement_date,
        "lifting_start": bc.get("sale_start_date"),
        "mean_age": bc.get("mean_age"),
        "latest_entry": latest, "gap_days": gap_days,
        "not_started": bool(placed > 0 and latest is None),
        "housed": placed, "mort": mort,
        "mort_pct": _num(bc.get("total_mort_pct")),
        "cull": culls, "cull_pct": _div(culls * 100, placed).quantize(q2),
        "sold_birds": sold, "sold_weight": _num(bc.get("sold_weight")),
        "available": available, "available_weight": live_weight_kg.quantize(q2),
        "std_bwt": std_bwt, "avg_bwt": avg_bwt,
        "std_bwt_at_feed": std_bwt_at_feed.quantize(q2) if std_bwt_at_feed is not None else None,
        "last_bwt_date": last_bwt_date, "last_bwt_gap": last_bwt_gap,
        "std_fcr": std_fcr, "fcr": fcr_val, "cfcr": cfcr_val, "eef": eef_val,
        "m_pc_kg": m_pc_kg,
        "m_pc_bird": bc.get("prod_cost_per_bird"),
        "feed_transferred": feed_sent, "transfer_in_farms": transfer_in_farms,
        "std_feed_con": std_feed_con, "std_feed_con_bird": std.cum_feed if std else None,
        "feed_con": feed_consumed,
        "feed_con_bird": actual_feed_per_bird.quantize(q2) if actual_feed_per_bird is not None else None,
        "std_feed_at_bwt": std_feed_at_bwt.quantize(q2) if std_feed_at_bwt is not None else None,
        "std_feed_at_bwt_bird": std_feed_at_bwt_g.quantize(q2) if std_feed_at_bwt_g is not None else None,
        "transfer_out_farms": transfer_out_farms, "feed_balance": feed_balance,
        "feed_balance_days": _div(feed_balance, daily_rate).quantize(q2) if daily_rate else Decimal("0"),
        "next_3_days_feed": (daily_rate * 3).quantize(q2),
        # --- mirrors of Daily Entry feed metrics ---
        "feed_phase": phase_name, "next_phase": next_phase,
        "phase_cap": phase_cap, "phase_cum_bird": phase_cum_bird,
        "act_feed_bird_live": act_feed_bird_live,
        "std_feed_bird_daily": std_feed_bird_daily,
        "remark": remark, "remark_verdict": remark.split(" — ")[0],
    }


@login_required
def live_flock_summary_report(request):
    """Broiler > Reports > Live Flock Summary — one row per live/ongoing flock."""
    from account.models import CompanyProfile

    region_id = (request.GET.get("region") or "").strip()
    branch_id = (request.GET.get("branch") or "").strip()
    supervisor_id = (request.GET.get("supervisor") or "").strip()
    breed_id = (request.GET.get("breed") or "").strip()

    batches = (BroilerBatch.objects.filter(end_date__isnull=True, is_closed=False)
               .select_related("broiler_farm__branch", "broiler_farm__supervisor",
                               "broiler_farm__farmer", "breed")
               .order_by("broiler_farm__branch__branch_name", "batch_name"))
    batches = scope_multi(request.user, batches,
                          farms="broiler_farm_id",
                          branches="broiler_farm__branch_id")
    if branch_id.isdigit():
        batches = batches.filter(broiler_farm__branch_id=branch_id)
    elif region_id.isdigit():
        batches = batches.filter(broiler_farm__branch__region_id=region_id)
    if supervisor_id.isdigit():
        batches = batches.filter(broiler_farm__supervisor_id=supervisor_id)
    if breed_id.isdigit():
        batches = batches.filter(breed_id=breed_id)

    today = timezone.localdate()
    rows = [_live_flock_row(b, today) for b in batches]

    return render(request, "live_flock_summary_report.html", {
        "rows": rows,
        "regions": Region.objects.order_by("description"),
        "branches": branches_for(request.user, Branch.objects.order_by("branch_name")),
        "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
        "breeds": Breed.objects.filter(is_active=True).order_by("description"),
        "region_id": region_id, "branch_id": branch_id,
        "supervisor_id": supervisor_id, "breed_id": breed_id,
        "company": CompanyProfile.get_solo(),
    })


def _haversine_km(lat1, lon1, lat2, lon2):
    """Great-circle distance in km between two lat/long points; None if any
    coordinate is missing."""
    if None in (lat1, lon1, lat2, lon2):
        return None
    import math
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return r * 2 * math.asin(math.sqrt(a))


def _day_record_row(e, sel_date, feed_ids, chick_ids, placed_cache, StockTransfer, BirdSale):
    """One Day Record row for a single DailyEntry `e` recorded on `sel_date`.
    Everything on the day (mort/cull/sold/feed) plus the cumulative figures to
    that date and the breed-standard comparison. Image/disease/entry-geo columns
    are returned blank — not yet captured on the daily entry."""
    from django.db.models import Sum as _Sum
    batch, farm = e.batch, e.farm
    q2 = Decimal("0.01")

    # chicks placed for this batch (chick-category stock transfers into it)
    if batch and batch.id not in placed_cache:
        placed_cache[batch.id] = _num(StockTransfer.objects.filter(
            to_batch_id=batch.id, item_id__in=chick_ids).aggregate(t=_Sum("quantity"))["t"])
    placed = placed_cache.get(batch.id, Decimal("0")) if batch else Decimal("0")

    def _de_agg(flt):
        r = DailyEntry.objects.filter(**flt).aggregate(
            m=_Sum("mortality"), c=_Sum("culls"),
            f1=_Sum("feed_1_qty"), f2=_Sum("feed_2_qty"))
        return _num(r["m"]), _num(r["c"]), _num(r["f1"]) + _num(r["f2"])

    def _sale_agg(flt):
        r = BirdSale.objects.filter(**flt).aggregate(b=_Sum("birds"), w=_Sum("net_weight"))
        return _num(r["b"]), _num(r["w"])

    mort_before, cull_before, _feed_before = _de_agg({"batch": batch, "date__lt": sel_date})
    _mort_upto, _cull_upto, feed_upto = _de_agg({"batch": batch, "date__lte": sel_date})
    sold_before, _soldw_before = _sale_agg({"batch": batch, "date__lt": sel_date})
    sold_today, soldw_today = _sale_agg({"batch": batch, "date": sel_date})
    sold_upto, soldw_upto = _sale_agg({"batch": batch, "date__lte": sel_date})

    mort_today = _num(e.mortality)
    cull_today = _num(e.culls)
    cum_mort = mort_before + mort_today          # cumulative mortality to date
    cum_cull = cull_before + cull_today

    opening = placed - mort_before - cull_before - sold_before   # birds alive at day start
    balance = opening - mort_today - cull_today - sold_today      # closing birds for the day

    avg_bwt = _num(e.avg_weight_gms)
    age = e.age_days
    std = _breed_standard_at(batch.breed_id if batch else None, age)
    std_bwt = std.body_weight if std else None
    std_fcr = std.fcr if std else None

    # feed movement on the day
    feed_con = _num(e.feed_1_qty) + _num(e.feed_2_qty)
    feed_stock = _num(e.feed_1_stock) + _num(e.feed_2_stock)     # closing stock
    prev = (DailyEntry.objects.filter(batch=batch, date__lt=sel_date)
            .order_by("-date", "-id").first())
    feed_ob = (_num(prev.feed_1_stock) + _num(prev.feed_2_stock)) if prev else Decimal("0")
    feed_in = _num(StockTransfer.objects.filter(
        to_batch=batch, item_id__in=feed_ids, date=sel_date).aggregate(t=_Sum("quantity"))["t"])
    feed_out = _num(StockTransfer.objects.filter(
        from_batch=batch, item_id__in=feed_ids, date=sel_date).aggregate(t=_Sum("quantity"))["t"])

    # FCR / CFCR on weight produced to date (live + sold), like the live report
    live_weight = balance * avg_bwt / Decimal("1000")
    total_weight = live_weight + soldw_upto
    mort_weight = cum_mort * avg_bwt / Decimal("1000")          # approx (birds ~current wt)
    fcr = _div(feed_upto, total_weight).quantize(Decimal("0.001")) if total_weight > 0 else Decimal("0")
    cfcr = (_div(feed_upto, total_weight + mort_weight).quantize(q2)
            if (total_weight + mort_weight) > 0 else Decimal("0"))

    farmer = farm.farmer if farm and farm.farmer_id else None
    batch_no = ""
    if batch and batch.batch_name and "-" in batch.batch_name:
        tail = batch.batch_name.rsplit("-", 1)[-1]
        batch_no = tail if tail.isdigit() else ""

    # field-captured attachments / geo (mobile app populates; may be blank)
    from broiler.models import BroilerDisease
    diseases = list(BroilerDisease.objects.filter(batch=batch, diagnosed_date=sel_date)
                    .exclude(disease_name="").values_list("disease_name", flat=True)) if batch else []
    farm_lat = farm.farm_latitude if farm else None
    farm_lon = farm.farm_longitude if farm else None
    diff_km = _haversine_km(farm_lat, farm_lon, e.entry_latitude, e.entry_longitude)

    return {
        "farm_code": farm.farm_code if farm else "",
        "farmer": farmer.farmer_name if farmer else "",
        "batch": batch.batch_name if batch else "",
        "batch_no": batch_no,
        "supervisor": (e.supervisor.name if e.supervisor_id
                       else farm.supervisor.name if farm and farm.supervisor_id else ""),
        "age": age,
        "placed": placed, "opening": opening,
        "mort": mort_today,
        "mort_pct": _div(mort_today * 100, opening).quantize(q2),
        "mort_image": e.mort_image.url if e.mort_image else "",
        "cum_mort": cum_mort,
        "cum_mort_pct": _div(cum_mort * 100, placed).quantize(q2),
        "culls": cull_today, "cull_image": e.cull_image.url if e.cull_image else "",
        "sold": sold_today, "sold_wt": soldw_today.quantize(q2),
        "balance": balance,
        "std_bwt": std_bwt, "avg_bwt": avg_bwt,
        "std_fcr": std_fcr, "fcr": fcr, "cfcr": cfcr,
        "feed_ob": feed_ob.quantize(q2), "feed_in": feed_in.quantize(q2),
        "feed_out": feed_out.quantize(q2), "feed_con": feed_con.quantize(q2),
        "feed_stock": feed_stock.quantize(q2),
        "cum_feed": feed_upto.quantize(q2),
        "feed_images": e.feed_image.url if e.feed_image else "",
        # --- extra columns for the Farm Detailed Daily Entry report ---
        "entry_date": e.date,
        "book_no": batch_no,
        # T. Birds / T. Weight = total (cumulative) birds/weight sold to date.
        "t_birds": sold_upto, "t_weight": soldw_upto.quantize(q2),
        "feed_1_item": e.feed_1.description if e.feed_1_id else "",
        "feed_1_con": _num(e.feed_1_qty).quantize(q2),
        "feed_2_item": e.feed_2.description if e.feed_2_id else "",
        "feed_2_con": _num(e.feed_2_qty).quantize(q2),
        "cum_feed_per_bird": _div(feed_upto * 1000, placed).quantize(q2),   # g/bird cumulative
        "std_feed_per_bird": std.feed_intake if std else None,              # g/bird/day standard
        "act_feed_per_bird": _div(feed_con * 1000, opening).quantize(q2),   # g/bird today
        "line": farm.line if farm else "",
        "branch": farm.branch.branch_name if farm and farm.branch_id else "",
        "farmer_contact": (farmer.mobile_no or farmer.phone_no or "") if farmer else "",
        "entry_time": e.entry_time,
        "entry_by": str(e.entry_by) if e.entry_by_id else "",
        "remarks": e.remarks or "",
        "diseases_name": ", ".join(diseases),
        "farm_location": (f"{farm_lat}, {farm_lon}"
                          if farm_lat is not None and farm_lon is not None else ""),
        "entry_location": (f"{e.entry_latitude}, {e.entry_longitude}"
                           if e.entry_latitude is not None and e.entry_longitude is not None else ""),
        "diff_km": round(diff_km, 3) if diff_km is not None else "",
    }


@login_required
def day_record_report(request):
    """Broiler > Reports > Day Record Report — one row per daily entry recorded
    on the selected date across all farms: that day's mortality/culls/sales,
    body weight vs breed standard, and feed movement (opening/in/out/consumed/
    stock/cumulative), with a Total footer. Image, disease and entry-location
    columns are shown but not yet captured on the daily entry (blank for now)."""
    from account.models import CompanyProfile
    from django.utils.dateparse import parse_date
    from inventory.models import StockTransfer, Item
    from broiler.models import BirdSale

    region_id = (request.GET.get("region") or "").strip()
    branch_id = (request.GET.get("branch") or "").strip()
    line = (request.GET.get("line") or "").strip()
    supervisor_id = (request.GET.get("supervisor") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()
    date_str = (request.GET.get("date") or "").strip()

    if date_str:
        sel_date = parse_date(date_str) or timezone.localdate()
    else:  # default to the most recent day that actually has entries
        sel_date = (DailyEntry.objects.order_by("-date")
                    .values_list("date", flat=True).first()) or timezone.localdate()

    entries = (DailyEntry.objects.filter(date=sel_date)
               .select_related("farm__branch", "farm__supervisor", "farm__farmer",
                               "batch__breed", "supervisor", "entry_by", "feed_1", "feed_2")
               .order_by("farm__farm_code", "batch__batch_name", "id"))
    entries = scope_multi(request.user, entries,
                          farms="farm_id", branches="farm__branch_id")
    if region_id:
        entries = entries.filter(farm__branch__region_id=region_id)
    if branch_id:
        entries = entries.filter(farm__branch_id=branch_id)
    if line:
        entries = entries.filter(farm__line=line)
    if supervisor_id:
        entries = entries.filter(Q(supervisor_id=supervisor_id) | Q(farm__supervisor_id=supervisor_id))
    if farm_id:
        entries = entries.filter(farm_id=farm_id)

    feed_ids = list(Item.objects.filter(category__name__icontains="feed").values_list("id", flat=True))
    chick_ids = list(Item.objects.filter(category__name__icontains="chick").values_list("id", flat=True))
    placed_cache = {}

    rows = [_day_record_row(e, sel_date, feed_ids, chick_ids, placed_cache, StockTransfer, BirdSale)
            for e in entries]

    # Total footer over the summable columns
    tkeys = ["placed", "opening", "mort", "cum_mort", "culls", "sold", "sold_wt",
             "balance", "feed_ob", "feed_in", "feed_out", "feed_con", "feed_stock", "cum_feed"]
    totals = {k: sum((_num(r[k]) for r in rows), Decimal("0")) for k in tkeys}
    totals["mort_pct"] = _div(totals["mort"] * 100, totals["opening"]).quantize(Decimal("0.01"))
    totals["cum_mort_pct"] = _div(totals["cum_mort"] * 100, totals["placed"]).quantize(Decimal("0.01"))

    lines = (farms_for(request.user).exclude(line="").order_by("line")
             .values_list("line", flat=True).distinct())

    return render(request, "day_record_report.html", {
        "rows": rows, "totals": totals, "sel_date": sel_date,
        "regions": Region.objects.order_by("description"),
        "branches": branches_for(request.user, Branch.objects.order_by("branch_name")),
        "lines": lines,
        "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
        "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
        "region_id": region_id, "branch_id": branch_id, "line": line,
        "supervisor_id": supervisor_id, "farm_id": farm_id,
        "company": CompanyProfile.get_solo(),
    })


@login_required
def farm_detailed_daily_entry_report(request):
    """Broiler > Reports > Farm Detailed Daily Entry Report — one row per daily
    entry over a From/To date range (farm-focused register): the day's
    mort/cull/sale/feed with cumulative figures, breed-standard comparison and
    per-bird feed. Reuses the Day Record row builder for each entry's own date."""
    from account.models import CompanyProfile
    from django.utils.dateparse import parse_date
    from datetime import timedelta
    from inventory.models import StockTransfer, Item
    from broiler.models import BirdSale

    branch_id = (request.GET.get("branch") or "").strip()
    line = (request.GET.get("line") or "").strip()
    supervisor_id = (request.GET.get("supervisor") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()

    to_date = parse_date(request.GET.get("to_date") or "") or timezone.localdate()
    from_date = parse_date(request.GET.get("from_date") or "") or (to_date - timedelta(days=30))

    entries = (DailyEntry.objects.filter(date__gte=from_date, date__lte=to_date)
               .select_related("farm__branch", "farm__supervisor", "farm__farmer",
                               "batch__breed", "supervisor", "entry_by", "feed_1", "feed_2")
               .order_by("farm__farm_code", "batch__batch_name", "date", "id"))
    entries = scope_multi(request.user, entries,
                          farms="farm_id", branches="farm__branch_id")
    if branch_id:
        entries = entries.filter(farm__branch_id=branch_id)
    if line:
        entries = entries.filter(farm__line=line)
    if supervisor_id:
        entries = entries.filter(Q(supervisor_id=supervisor_id) | Q(farm__supervisor_id=supervisor_id))
    if farm_id:
        entries = entries.filter(farm_id=farm_id)

    feed_ids = list(Item.objects.filter(category__name__icontains="feed").values_list("id", flat=True))
    chick_ids = list(Item.objects.filter(category__name__icontains="chick").values_list("id", flat=True))
    placed_cache = {}

    rows = [_day_record_row(e, e.date, feed_ids, chick_ids, placed_cache, StockTransfer, BirdSale)
            for e in entries]

    # Total footer over the daily flow columns only (stock/cumulative columns
    # don't sum meaningfully across days).
    tkeys = ["mort", "culls", "sold", "sold_wt",
             "feed_in", "feed_out", "feed_con", "feed_1_con", "feed_2_con"]
    totals = {k: sum((_num(r[k]) for r in rows), Decimal("0")) for k in tkeys}

    lines = (farms_for(request.user).exclude(line="").order_by("line")
             .values_list("line", flat=True).distinct())

    columns = [
        ("slno", "Sl.No."), ("supervisor", "Supervisor"), ("farm_code", "Farm Code"),
        ("farmer", "Farmer"), ("batch", "Batch"), ("book_no", "Book No"),
        ("entry_date", "Entry Date"), ("age", "Age"), ("placed", "Placed Birds"),
        ("opening", "Opening Birds"), ("mort", "Mort"), ("mort_pct", "Mort%"),
        ("mort_image", "Mort Image"), ("cum_mort", "Cum Mort"), ("cum_mort_pct", "Cum Mort%"),
        ("culls", "Culls"), ("cull_image", "Cull Image"), ("sold", "Sold"),
        ("sold_wt", "Sold Wt"), ("t_birds", "T. Birds"), ("t_weight", "T. Weight"),
        ("balance", "Balance Birds"), ("std_bwt", "Std B.Wt"), ("avg_bwt", "Avg B.Wt"),
        ("std_fcr", "Std FCR"), ("fcr", "FCR"), ("cfcr", "CFCR"), ("feed_ob", "Feed OB"),
        ("feed_in", "Feed In"), ("feed_out", "Feed Out"), ("feed_con", "Feed Con"),
        ("feed_1_item", "Feed-1 Item"), ("feed_1_con", "Feed-1 Con"),
        ("feed_2_item", "Feed-2 Item"), ("feed_2_con", "Feed-2 Con"),
        ("feed_stock", "Feed Stock"), ("cum_feed", "Cum. Feed"),
        ("cum_feed_per_bird", "Cum. Feed/Bird"), ("std_feed_per_bird", "Std Feed/Bird"),
        ("act_feed_per_bird", "Act Feed/Bird"), ("diseases_name", "Diseases Names"),
        ("remarks", "Remarks"), ("line", "Line"), ("branch", "Branch"),
        ("farmer_contact", "Farmer Contact"), ("entry_time", "Entry Time"),
        ("entry_by", "Entry By"), ("farm_location", "Farm Location"),
        ("entry_location", "Entry Location"),
    ]

    return render(request, "farm_detailed_daily_entry_report.html", {
        "rows": rows, "totals": totals, "from_date": from_date, "to_date": to_date,
        "columns": columns,
        "branches": branches_for(request.user, Branch.objects.order_by("branch_name")),
        "lines": lines,
        "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
        "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
        "branch_id": branch_id, "line": line,
        "supervisor_id": supervisor_id, "farm_id": farm_id,
        "company": CompanyProfile.get_solo(),
    })


@login_required
def lifting_report(request):
    """Broiler > Reports > Lifting Report — a register of every Bird Sale
    (bird "lifting") within a Region/Branch/Line/Supervisor/Farm/Customer/date
    window: one row per lifting with weight, rate, amount, TCS, receipt and the
    farm/batch/line/supervisor context. Receipts (not linked to a single sale)
    are attributed to the customer's first lifting on each date."""
    from account.models import CompanyProfile
    from broiler.models import BirdSale, BirdSaleReceipt
    from sales.models import SalesInvoice
    from hr.models import Employee
    from django.utils.dateparse import parse_date

    q2 = Decimal("0.01")
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    customer_id = (request.GET.get("customer") or "").strip()
    region_id = (request.GET.get("region") or "").strip()
    branch_id = (request.GET.get("branch") or "").strip()
    line = (request.GET.get("line") or "").strip()
    supervisor_id = (request.GET.get("supervisor") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()
    sale_type = (request.GET.get("type") or "").strip()
    fd = parse_date(from_date) if from_date else None
    td = parse_date(to_date) if to_date else None

    sales = (BirdSale.objects
             .select_related("customer", "farmer", "farm__branch", "farm__supervisor",
                             "lifting_supervisor", "batch")
             .order_by("date", "id"))
    sales = scope_multi(request.user, sales,
                        farms="farm_id", branches="farm__branch_id")
    if fd:
        sales = sales.filter(date__gte=fd)
    if td:
        sales = sales.filter(date__lte=td)
    if customer_id.isdigit():
        sales = sales.filter(customer_id=customer_id)
    if region_id.isdigit():
        sales = sales.filter(farm__branch__region_id=region_id)
    if branch_id.isdigit():
        sales = sales.filter(farm__branch_id=branch_id)
    if line:
        sales = sales.filter(farm__line=line)
    if supervisor_id.isdigit():
        sales = sales.filter(lifting_supervisor_id=supervisor_id)
    if farm_id.isdigit():
        sales = sales.filter(farm_id=farm_id)
    if sale_type in ("customer", "farmer"):
        sales = sales.filter(sale_type=sale_type)
    sales = list(sales)

    # Receipts summed per (customer, date), attributed to that customer's first
    # lifting of the day (mirrors the reference report's receipt column).
    rcpt_qs = BirdSaleReceipt.objects.filter(sale_type="customer")
    rcpt_qs = scope_or_null(request.user, rcpt_qs,
                            "customer_groups", "customer__customer_group_id")
    if fd:
        rcpt_qs = rcpt_qs.filter(date__gte=fd)
    if td:
        rcpt_qs = rcpt_qs.filter(date__lte=td)
    if customer_id.isdigit():
        rcpt_qs = rcpt_qs.filter(customer_id=customer_id)
    rcpt_by_key = {}
    for rc in rcpt_qs:
        rcpt_by_key[(rc.customer_id, rc.date)] = rcpt_by_key.get((rc.customer_id, rc.date), Decimal("0")) + _num(rc.amount)

    # Per-customer ledger events, so each lifting row can show that customer's
    # running balance as of its OWN sale date (all postings up to and including
    # that day). opening is the signed receivable (Dr = owes us).
    from sales.models import Customer as _Customer
    cust_ids = {s.customer_id for s in sales if s.customer_id}
    cust_events = {}  # cust_id -> (opening_signed, [(date, delta), ...])
    if cust_ids:
        for cust in _Customer.objects.filter(id__in=cust_ids):
            opening = _num(cust.opening_balance)
            if str(cust.to_pay_to_receive or "").lower().startswith("pay"):
                opening = -opening
            events = []
            for bs in BirdSale.objects.filter(sale_type="customer", customer=cust):
                events.append((bs.date, _num(bs.amount)))
            for inv in SalesInvoice.objects.filter(customer=cust, is_active=True):
                events.append((inv.date, _num(inv.net_amount)))
            for rc in BirdSaleReceipt.objects.filter(sale_type="customer", customer=cust):
                events.append((rc.date, -_num(rc.amount)))
            cust_events[cust.id] = (opening, events)

    def _balance_asof(cust_id, on_date):
        """Customer's ledger balance up to (and including) on_date."""
        opening, events = cust_events[cust_id]
        bal = opening
        for d, delta in events:
            if d and on_date and d <= on_date:
                bal += delta
        return bal

    rows = []
    seen_rcpt_keys = set()
    for s in sales:
        farm = s.farm
        batch = s.batch
        gross = (_num(s.net_weight) * _num(s.rate)).quantize(q2)
        tcs = Decimal("0.00")
        total = (_num(s.amount) + tcs).quantize(q2)

        # attribute the day's receipts to the first lifting of this customer/date
        receipt = Decimal("0.00")
        if s.customer_id:
            key = (s.customer_id, s.date)
            if key not in seen_rcpt_keys:
                seen_rcpt_keys.add(key)
                receipt = rcpt_by_key.get(key, Decimal("0.00")).quantize(q2)

        mean_age = ""
        _placed = _placement_date(batch)
        if _placed and s.date:
            mean_age = (s.date - _placed).days

        party = s.customer.name if s.customer_id else (s.farmer.farmer_name if s.farmer_id else "")
        code = s.customer.code if s.customer_id else ""
        if s.customer_id and s.customer_id in cust_events:
            bal = _balance_asof(s.customer_id, s.date)
            ledger_balance = abs(bal).quantize(q2)
            ledger_cr_dr = "Dr" if bal >= 0 else "Cr"
        else:
            ledger_balance, ledger_cr_dr = "", ""
        # This column is the lifting / weighment supervisor recorded on the sale,
        # not the farm's managing supervisor.
        supervisor = str(s.lifting_supervisor) if s.lifting_supervisor_id else ""

        rows.append({
            "date": s.date, "code": code or "", "customer": party,
            "customer_id": s.customer_id,
            "invoice": s.sale_no, "dc_no": s.doc_no or "",
            "birds": s.birds or 0, "weight": _num(s.net_weight).quantize(q2),
            "avg_wt": _num(s.avg_weight).quantize(q2), "rate": _num(s.rate).quantize(q2),
            "amount": gross, "tcs": tcs, "total": total, "receipt": receipt,
            "ledger_balance": ledger_balance, "ledger_cr_dr": ledger_cr_dr,
            "branch": farm.branch.branch_name if farm and farm.branch_id else "",
            "line": farm.line if farm else "",
            "supervisor": supervisor,
            "farm": farm.farm_name if farm else "",
            "batch": batch.batch_name if batch else "",
            "mean_age": mean_age,
            "vehicle": s.vehicle or "", "driver": s.driver or "", "remarks": s.remarks or "",
        })

    tkeys = ["birds", "weight", "amount", "tcs", "total", "receipt"]
    totals = {k: sum((_num(r[k]) for r in rows), Decimal("0")) for k in tkeys}
    totals["avg_wt"] = _div(totals["weight"], totals["birds"]).quantize(q2)
    totals["rate"] = _div(totals["amount"], totals["weight"]).quantize(q2)
    for k in ("weight", "amount", "tcs", "total", "receipt"):
        totals[k] = totals[k].quantize(q2)
    totals["birds"] = int(totals["birds"])

    lines = (farms_for(request.user).exclude(line="").order_by("line")
             .values_list("line", flat=True).distinct())

    return render(request, "lifting_report.html", {
        "rows": rows, "totals": totals,
        "customers": customers_for(request.user, Customer.objects.order_by("name")),
        "regions": Region.objects.order_by("description"),
        "branches": branches_for(request.user, Branch.objects.order_by("branch_name")),
        "lines": lines,
        "supervisors": Employee.objects.filter(relieve=False).order_by("full_name"),
        "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
        "from_date": from_date, "to_date": to_date, "customer_id": customer_id,
        "region_id": region_id, "branch_id": branch_id, "line": line,
        "supervisor_id": supervisor_id, "farm_id": farm_id, "sale_type": sale_type,
        "company": CompanyProfile.get_solo(),
    })


@login_required
def chicks_placement_report(request):
    """Register of every Chicks Placement transaction (Warehouse -> Farm/Batch
    Stock Transfer of a chicks-category item) within a Branch/Farm/date window
    (Broiler > Reports > Chicks Placement Report). Chicks Ordered/Transit
    Mortality/Shortage/Culls are the same reference-only fields as the Chicks
    Placement transaction itself — only Placement Qty ever reaches inventory.
    """
    from account.models import CompanyProfile
    from inventory.models import StockTransfer, Warehouse

    region_id = (request.GET.get("region") or "").strip()
    branch_id = (request.GET.get("branch") or "").strip()
    line = (request.GET.get("line") or "").strip()
    supervisor_id = (request.GET.get("supervisor") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()
    warehouse_id = (request.GET.get("warehouse") or "").strip()
    hatchery_id = (request.GET.get("hatchery") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    status = (request.GET.get("status") or "").strip().lower()
    export = (request.GET.get("export") or "display").strip().lower()
    submitted = bool(region_id or branch_id or line or supervisor_id or farm_id
                     or warehouse_id or hatchery_id
                     or from_date or to_date or status or request.GET.get("submit"))

    rows, totals = [], None
    if submitted:
        qs = (StockTransfer.objects
              .filter(to_location_type="farm", item__category__name__icontains="chick")
              .select_related("to_farm__branch", "to_farm__supervisor", "to_batch",
                              "from_warehouse", "source_hatchery", "source_supplier", "item")
              .order_by("date", "id"))
        # The dropdowns below are scoped, but nothing stopped a restricted user
        # leaving them on "All" (or editing the query string) and reading every
        # branch's placements. A placement has two ends, so either the receiving
        # farm/branch or the dispatching warehouse being in scope is enough.
        qs = scope_any(request.user, qs,
                       branches="to_farm__branch_id",
                       farms="to_farm_id",
                       sectors="from_warehouse_id")
        if region_id:
            qs = qs.filter(to_farm__branch__region_id=region_id)
        if branch_id:
            qs = qs.filter(to_farm__branch_id=branch_id)
        if line:
            qs = qs.filter(to_farm__line=line)
        if supervisor_id:
            qs = qs.filter(to_farm__supervisor_id=supervisor_id)
        if farm_id:
            qs = qs.filter(to_farm_id=farm_id)
        if warehouse_id:
            qs = qs.filter(from_warehouse_id=warehouse_id)
        if hatchery_id:
            # Prefixed "h:<id>" / "s:<id>" (see _chicks_sources). A bare number
            # is an older bookmarked URL, back when only hatcheries existed.
            kind, _, source_pk = hatchery_id.partition(":")
            if not source_pk:
                qs = qs.filter(source_hatchery_id=kind)
            elif kind == "s":
                qs = qs.filter(source_supplier_id=source_pk)
            else:
                qs = qs.filter(source_hatchery_id=source_pk)
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
        # Status reflects the linked Batch's own state (see batch_status below).
        if status == "active":
            qs = qs.filter(to_batch__isnull=False, to_batch__end_date__isnull=True)
        elif status == "completed":
            qs = qs.filter(to_batch__isnull=False, to_batch__end_date__isnull=False)

        # Free Quantity has no backing field on StockTransfer yet, so it's always
        # 0 (shown, not hidden, so the column stays honest about what isn't
        # tracked) and Total Chicks Placed = Quantity Received + Free Quantity.
        free_quantity = Decimal("0")
        t_ordered = t_mortality = t_shortage = t_culls = t_excess = Decimal("0")
        t_received = t_free = t_placed = t_amount = Decimal("0")
        mort_pct_sum = Decimal("0")
        farm_ids, branch_ids, warehouse_ids = set(), set(), set()
        for t in qs:
            ordered = t.chicks_ordered or Decimal("0")
            mortality = t.transit_mortality or Decimal("0")
            shortage = t.shortage or Decimal("0")
            culls = t.culls or Decimal("0")
            received = t.quantity or Decimal("0")
            placed = received + free_quantity
            amount = (received * (t.rate or Decimal("0"))).quantize(Decimal("0.01"))
            # Excess = birds received beyond what the ordered/loss breakdown expected
            # (only meaningful when Chicks Ordered was actually recorded).
            excess = Decimal("0")
            if ordered > 0:
                expected = max(ordered - mortality - shortage - culls, Decimal("0"))
                excess = max(received - expected, Decimal("0"))
            row_mort_pct = (mortality / ordered * 100).quantize(Decimal("0.01")) if ordered else Decimal("0")
            # Status reflects the linked Batch's own state — a Batch with no
            # end_date is still running (Active); once end_date is set it's
            # Closed. No batch linked means there's nothing to report on.
            if not t.to_batch_id:
                batch_status = ""
            elif t.to_batch.end_date is None:
                batch_status = "Active"
            else:
                batch_status = "Completed"
            rows.append({
                "batch_status": batch_status,
                "date": t.date, "trnum": t.trnum, "dc_no": t.dc_no,
                "branch_name": t.to_farm.branch.branch_name if t.to_farm_id else "",
                "line": t.to_farm.line if t.to_farm_id else "",
                "supervisor_name": t.to_farm.supervisor.name if t.to_farm_id and t.to_farm.supervisor_id else "",
                "farm_code": t.to_farm.farm_code if t.to_farm_id else "",
                "farm_name": t.to_farm.farm_name if t.to_farm_id else "",
                "batch_name": t.to_batch.batch_name if t.to_batch_id else "",
                "source_hatchery_name": t.source_name,
                "warehouse_name": t.from_warehouse.name if t.from_warehouse_id else "",
                "chicks_ordered": ordered, "transit_mortality": mortality,
                "mort_pct": row_mort_pct,
                "shortage": shortage, "culls": culls,
                "culls_pct": (culls / ordered * 100).quantize(Decimal("0.01")) if ordered else Decimal("0"),
                "excess": excess,
                "quantity_received": received, "free_quantity": free_quantity, "total_placed": placed,
                "rate": t.rate or Decimal("0"), "amount": amount,
                "farm_capacity": t.to_farm.farm_capacity if t.to_farm_id else "",
            })
            t_ordered += ordered; t_mortality += mortality; t_shortage += shortage
            t_culls += culls; t_excess += excess
            t_received += received; t_free += free_quantity; t_placed += placed; t_amount += amount
            mort_pct_sum += row_mort_pct
            if t.to_farm_id:
                farm_ids.add(t.to_farm_id)
                branch_ids.add(t.to_farm.branch_id)
            if t.from_warehouse_id:
                warehouse_ids.add(t.from_warehouse_id)

        shortage_pct = (t_shortage / t_ordered * 100).quantize(Decimal("0.01")) if t_ordered else Decimal("0")
        culls_pct = (t_culls / t_ordered * 100).quantize(Decimal("0.01")) if t_ordered else Decimal("0")
        totals = {
            "chicks_ordered": t_ordered, "transit_mortality": t_mortality,
            "mort_pct": (t_mortality / t_ordered * 100).quantize(Decimal("0.01")) if t_ordered else Decimal("0"),
            "shortage": t_shortage, "culls": t_culls, "culls_pct": culls_pct,
            "excess": t_excess,
            "quantity_received": t_received, "free_quantity": t_free, "total_placed": t_placed,
            "amount": t_amount,
        }
        # KPI cards: real, derived-only figures — no fabricated "status" metric,
        # since neither StockTransfer nor BroilerBatch tracks any such state.
        kpi = {
            "total_placed": t_placed,
            "total_ordered": t_ordered,
            "placed_pct": (t_placed / t_ordered * 100).quantize(Decimal("0.1")) if t_ordered else Decimal("100.0"),
            "total_mortality": t_mortality,
            "overall_mort_pct": totals["mort_pct"],
            "total_shortage": t_shortage,
            "shortage_pct": shortage_pct,
            "total_culls": t_culls,
            "culls_pct": culls_pct,
            "average_mort_pct": (mort_pct_sum / len(rows)).quantize(Decimal("0.01")) if rows else Decimal("0"),
            "farms_count": len(farm_ids),
            "branches_count": len(branch_ids),
            "warehouses_count": len(warehouse_ids),
        }

    if not submitted:
        kpi = {"total_placed": 0, "total_ordered": 0, "placed_pct": Decimal("0"),
               "total_mortality": 0, "overall_mort_pct": Decimal("0"), "average_mort_pct": Decimal("0"),
               "total_shortage": 0, "shortage_pct": Decimal("0"),
               "total_culls": 0, "culls_pct": Decimal("0"),
               "farms_count": 0, "branches_count": 0, "warehouses_count": 0}

    lines = (farms_for(request.user).exclude(line="").order_by("line")
             .values_list("line", flat=True).distinct())

    return render(request, "chicks_placement_report.html", {
        "regions": Region.objects.order_by("description"),
        "branches": branches_for(request.user, Branch.objects.order_by("branch_name")),
        "lines": lines,
        "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
        "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
        "warehouses": warehouses_for(request.user, Warehouse.objects.order_by("name")),
        "sources": _chicks_sources(request.user),
        "region_id": region_id, "branch_id": branch_id, "line": line,
        "supervisor_id": supervisor_id, "farm_id": farm_id,
        "warehouse_id": warehouse_id, "hatchery_id": hatchery_id,
        "from_date": from_date, "to_date": to_date, "status": status,
        "submitted": submitted, "rows": rows, "totals": totals, "kpi": kpi,
        "export": export,
        "company": CompanyProfile.get_solo(),
    })


def _feed_phase_master_for(batch, on_date, masters):
    """The Feed Phase Master applying to a batch on a date, picked from an
    already-fetched list (same matching rules as ``resolve_feed_phase``:
    breed first, then the breed's bird category, within the effective
    window) so a whole report can match without re-querying per batch."""
    if not batch.breed_id:
        return None

    def eff_ok(m):
        if on_date and m.effective_from and on_date < m.effective_from:
            return False
        if on_date and m.effective_to and on_date > m.effective_to:
            return False
        return True

    cat_id = batch.breed.bird_category_id
    return (next((m for m in masters if m.breed_id == batch.breed_id and eff_ok(m)), None)
            or (next((m for m in masters if cat_id and m.bird_category_id == cat_id and eff_ok(m)), None)))


def _feed_curve_at(curve, age):
    """Cumulative standard feed (g/bird) at `age`, read off a breed's ordered
    (age, cum_feed, body_weight, feed_intake) curve and carried forward from
    the last defined row. Nothing before the curve starts — cumulative feed
    there is zero, not the first row's value, which is why this doesn't reuse
    ``_breed_standard_at`` (that one carries backwards to keep Std columns
    populated)."""
    if age is None or age < 0 or not curve:
        return Decimal("0")
    value = Decimal("0")
    for row in curve:
        if row[0] > age:
            break
        value = row[1]
    return value


def _std_at(curve, age):
    """(body weight g, daily feed intake g) a standard bird of this breed has
    at `age`, carried forward from the last defined row — (None, None) when
    the breed has no curve."""
    if age is None or age < 0 or not curve:
        return None, None
    found = None
    for row in curve:
        if row[0] > age:
            break
        found = row
    if found is None:
        found = curve[0]
    return found[2], found[3]


def _phase_feed_due(curve, pl, age):
    """Feed (g/bird) a phase should have delivered by `age`, and the basis it
    was read from.

    The phase's Max Feed Qty is a changeover *trigger* — the total per-bird
    feed at which the flock moves on — so comparing it to consumption
    mid-phase always flatters the flock. What's wanted is feed due so far,
    taken from the breed standard's cumulative curve across the phase's own
    age band ("std"); where the breed has no curve, the cap spread evenly
    over the band ("pro"); and for an open-ended last phase with neither, the
    cap itself ("cap"). Returns (None, "") when there's nothing to go on.
    """
    if pl is None or age is None:
        return None, ""
    lo, hi = pl.from_age or 0, pl.to_age
    upto = age if hi is None else min(age, hi)
    if upto < lo:
        return Decimal("0"), "std" if curve else ""
    if curve:
        due = _feed_curve_at(curve, upto) - _feed_curve_at(curve, lo - 1)
        if due > 0:
            return due, "std"
    cap_g = (pl.max_feed_qty or Decimal("0")) * 1000
    if cap_g <= 0:
        return None, ""
    if hi is None:
        return cap_g, "cap"
    span = Decimal(hi - lo + 1)
    elapsed = Decimal(min(upto - lo + 1, hi - lo + 1))
    return (cap_g * elapsed / span), "pro"


def _warehouse_feed_stock(wh_ids, item_ids, as_of):
    """Feed on hand per item across a set of Warehouses, as of a date.

    Reconciled from the movements themselves — the same event set the Feed
    Dispatch & Stock ledger replays — rather than read off
    ``StockTransfer.stock``. That running-balance field is only written by the
    transfer flow, so a warehouse stocked by purchase or stock-receive and not
    since transferred out still carries a stale zero on it; counting the
    movements is the only way this agrees with what is really on the floor.
    """
    from purchase.models import GeneralPurchaseItem
    from inventory.models import (StockTransfer, StockReceiveItem, StockIssueItem,
                                  InventoryAdjustmentItem)
    bal = {}
    if not wh_ids or not item_ids:
        return bal

    def apply(rows, sign, field="t"):
        for r in rows:
            bal[r["item_id"]] = bal.get(r["item_id"], Decimal("0")) + sign * (r[field] or Decimal("0"))

    apply(StockTransfer.objects.filter(item_id__in=item_ids, to_warehouse_id__in=wh_ids, date__lte=as_of)
          .values("item_id").annotate(t=Sum("quantity")), 1)
    apply(StockTransfer.objects.filter(item_id__in=item_ids, from_warehouse_id__in=wh_ids, date__lte=as_of)
          .values("item_id").annotate(t=Sum("quantity")), -1)
    for r in (GeneralPurchaseItem.objects
              .filter(item_id__in=item_ids, farm_warehouse_id__in=wh_ids, purchase__date__lte=as_of)
              .values("item_id").annotate(t=Sum("rcv_qty"), f=Sum("free_qty"))):
        bal[r["item_id"]] = (bal.get(r["item_id"], Decimal("0"))
                             + (r["t"] or Decimal("0")) + (r["f"] or Decimal("0")))
    apply(StockReceiveItem.objects
          .filter(item_id__in=item_ids, location_type="warehouse", warehouse_id__in=wh_ids,
                  receive__date__lte=as_of).values("item_id").annotate(t=Sum("quantity")), 1)
    apply(StockIssueItem.objects
          .filter(item_id__in=item_ids, location_type="warehouse", warehouse_id__in=wh_ids,
                  issue__date__lte=as_of).values("item_id").annotate(t=Sum("quantity")), -1)
    for r in (InventoryAdjustmentItem.objects
              .filter(item_id__in=item_ids, adjustment__location_type="warehouse",
                      adjustment__warehouse_id__in=wh_ids, adjustment__date__lte=as_of)
              .values("item_id", "adjustment_type").annotate(t=Sum("quantity"))):
        sign = 1 if r["adjustment_type"] == "Add" else -1
        bal[r["item_id"]] = bal.get(r["item_id"], Decimal("0")) + sign * (r["t"] or Decimal("0"))
    return bal


def _feed_row_actions(r, *, is_current, next_phase, bag_kg):
    """What to do about one feed item on one flock, worst first.

    Everything here is a decision a scheduler can act on today - send feed,
    change phase, fix a broken stock figure - rather than a verdict on the
    flock. Ordering matters: the list is rendered in place, so the first entry
    is what the row is really telling you.
    """
    out = []
    cover, cap, cons = r["days_cover"], r["cap_bird"], r["cons_bird"]

    if r["avail_qty"] < 0:
        out.append(("urgent", "Stock negative - check transfers"))
    if cover is not None and cover >= 0 and r["daily_rate"] > 0:
        need = r["next_3_days"]
        bags = f" ({(need / bag_kg).quantize(Decimal('0.1'))} bags)" if bag_kg else ""
        if cover < 1.5:
            out.append(("urgent", f"Dispatch {need:.0f} Kg{bags} - {cover:.1f} d cover"))
        elif cover < 3:
            out.append(("warn", f"Send {need:.0f} Kg{bags} - {cover:.1f} d cover"))
    # The cap is the changeover trigger, so reaching it is an instruction to
    # move the flock on, not merely an overshoot to note.
    if is_current and cap:
        if cons >= cap:
            out.append(("urgent", f"Cap reached - change to {next_phase}" if next_phase
                                  else "Cap reached - final phase"))
        elif cons >= cap * Decimal("0.9"):
            out.append(("warn", f"Nearing cap - {next_phase} next" if next_phase
                                else "Nearing cap"))
    return out


def _feed_batch_actions(*, is_live, gap_days, no_programme, mort_pct, avg_bwt, std_bwt):
    """Flock-level notes that belong to the whole batch rather than one feed
    item, drawn from the same signals Live Flock Summary reads."""
    out = []
    if no_programme:
        out.append(("warn", "No feed programme - nothing to schedule against"))
    if is_live and gap_days is not None and gap_days >= 2:
        out.append(("urgent" if gap_days >= 7 else "warn", f"Entries {gap_days} d stale"))
    if mort_pct is not None and mort_pct >= 5:
        out.append(("urgent" if mort_pct >= 8 else "warn", f"Mortality {mort_pct:.1f}%"))
    if std_bwt and avg_bwt:
        dev = (avg_bwt - std_bwt) / std_bwt * 100
        if dev <= -10:
            out.append(("urgent" if dev <= -20 else "warn", f"B.Wt {abs(dev):.0f}% below std"))
    return out


@login_required
def batch_wise_feed_scheduling_report(request):
    """Broiler > Reports > Batch wise Feed Scheduling Report — one row per feed
    phase of a batch's Feed Phase Master: the feed due by now set against what
    the batch actually received, ate and moved out, the feed still lying at the
    farm, and how long that feed will last at the flock's current rate.

    Live flocks are the default; Closed and All are there to look back at a
    finished batch, which is read as of the day it ended rather than today so
    its age, phase and closing burn rate are the ones it actually had.

    "Scheduled to Date (cumulative intake)" is feed due so far, not the phase's
    whole allowance — the breed standard's intake accumulated over the phase's
    age band, where "Day's Intake" is the same curve read for today alone. See
    ``_phase_feed_due`` for how it is read and what the basis marker on each row
    means. Difference is that cumulative figure minus consumption, so a positive
    number is a flock behind its curve and a negative one is a flock ahead of
    it; the Phase Cap column keeps the changeover trigger itself in view.

    Every phase of the batch's programme is listed whatever the flock's age —
    a phase it hasn't reached still carries a Total Required Feed to procure
    against — and each row says in the Phase column whether it is the one being
    fed now, one already passed, or one still to come. Feed items used by a
    batch but absent from its programme also get a row (with no schedule), so
    nothing is dropped.

    Feed quantities are in Kg and per-bird feed in grams. Per-bird figures are
    against birds still alive (placed - mortality - culls - sold), falling back
    to birds placed before any are lost, matching how Live Flock Summary reads
    the same phase cap — scheduling feed for birds that have died would
    over-dispatch, and dividing consumption by them would understate what the
    surviving flock is really eating.

    Daily Rate is the mean of the flock's last three daily entries for that
    feed item within a fortnight, so Days Cover reports on the feed actually
    moving now; a phase the flock has finished, or one whose entries have gone
    stale, gets no rate and no cover rather than a fabricated one.

    Two column meanings worth stating, since both fold several transactions
    into one figure so that In - Consumed - Out always reconciles to Farm
    Stock: "Feed In" counts only Stock Transfers into the batch (feed
    purchases carry no batch/farm, so charging them here would credit every
    batch of the branch with the same feed), and "Transferred Out" counts all
    feed leaving the batch, whether returned to a warehouse or passed to
    another farm.
    """
    from inventory.models import Mapping, StockTransfer

    branch_id = (request.GET.get("branch") or "").strip()
    line = (request.GET.get("line") or "").strip()
    supervisor_id = (request.GET.get("supervisor") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()
    excess_only = bool(request.GET.get("excess"))
    status = (request.GET.get("status") or "live").strip().lower()
    sort = (request.GET.get("sort") or "farm").strip().lower()
    export = (request.GET.get("export") or "display").strip().lower()
    submitted = bool(branch_id or line or supervisor_id or farm_id or excess_only
                     or request.GET.get("status") or request.GET.get("submit"))

    today = timezone.localdate()
    q2 = Decimal("0.01")
    rows, totals, feed_summary = [], None, []

    if submitted:
        batches = (BroilerBatch.objects
                   .select_related("broiler_farm__branch", "broiler_farm__supervisor",
                                   "breed__bird_category")
                   .order_by("broiler_farm__farm_name", "batch_name"))
        batches = scope_multi(request.user, batches,
                              farms="broiler_farm_id",
                              branches="broiler_farm__branch_id")
        if status == "closed":
            batches = batches.filter(Q(end_date__isnull=False) | Q(is_closed=True))
        elif status != "all":
            batches = batches.filter(end_date__isnull=True, is_closed=False)
        if branch_id.isdigit():
            batches = batches.filter(broiler_farm__branch_id=branch_id)
        if line:
            batches = batches.filter(broiler_farm__line=line)
        if supervisor_id.isdigit():
            batches = batches.filter(broiler_farm__supervisor_id=supervisor_id)
        if farm_id.isdigit():
            batches = batches.filter(broiler_farm_id=farm_id)
        batches = list(batches)
        batch_ids = [b.id for b in batches]

        # A finished flock is read as of the day it ended, not today: its age,
        # the feed phase it was on and its closing burn rate all belong to that
        # date, and measuring a batch that ended last winter against today would
        # age it into a phase it never saw.
        live_flag, as_of_by_batch = {}, {}
        for b in batches:
            is_live = b.end_date is None and not b.is_closed
            live_flag[b.id] = is_live
            as_of_by_batch[b.id] = today if is_live else min(b.end_date or b.closed_on or today, today)

        # Everything the rows need, in one aggregate per movement type rather
        # than per batch — this report spans every live flock in a branch.
        is_chick = Q(item__category__name__icontains="chick")
        # Placement also dates the flock: a batch with no start_date still has
        # a real placement — the chick transfer — to take its age from.
        placement = {
            r["to_batch_id"]: (r["t"] or Decimal("0"), r["d"])
            for r in (StockTransfer.objects.filter(to_batch_id__in=batch_ids).filter(is_chick)
                      .values("to_batch_id").annotate(t=Sum("quantity"), d=Min("date")))
        }
        feed_in = {
            (r["to_batch_id"], r["item_id"]): r["t"] or Decimal("0")
            for r in (StockTransfer.objects.filter(to_batch_id__in=batch_ids).exclude(is_chick)
                      .values("to_batch_id", "item_id").annotate(t=Sum("quantity")))
        }
        feed_out = {
            (r["from_batch_id"], r["item_id"]): r["t"] or Decimal("0")
            for r in (StockTransfer.objects.filter(from_batch_id__in=batch_ids).exclude(is_chick)
                      .values("from_batch_id", "item_id").annotate(t=Sum("quantity")))
        }
        consumed = {}
        for slot_item, slot_qty in (("feed_1_id", "feed_1_qty"), ("feed_2_id", "feed_2_qty")):
            for r in (DailyEntry.objects.filter(batch_id__in=batch_ids, **{f"{slot_item}__isnull": False})
                      .values("batch_id", slot_item).annotate(t=Sum(slot_qty))):
                key = (r["batch_id"], r[slot_item])
                consumed[key] = consumed.get(key, Decimal("0")) + (r["t"] or Decimal("0"))

        # Birds lost, so per-bird figures run on the flock that's still eating.
        losses_by_batch = {
            r["batch_id"]: (r["m"] or 0) + (r["c"] or 0)
            for r in (DailyEntry.objects.filter(batch_id__in=batch_ids)
                      .values("batch_id").annotate(m=Sum("mortality"), c=Sum("culls")))
        }
        sold_by_batch = {
            r["batch_id"]: r["b"] or 0
            for r in (BirdSale.objects.filter(batch_id__in=batch_ids)
                      .values("batch_id").annotate(b=Sum("birds")))
        }
        last_entry_by_batch = {
            r["batch_id"]: r["d"]
            for r in (DailyEntry.objects.filter(batch_id__in=batch_ids)
                      .values("batch_id").annotate(d=Max("date")))
        }

        # Each breed's standard curve: cumulative feed drives the feed-due-so-far
        # figure, body weight and daily intake drive the flock notes and the
        # fallback burn rate for a phase with no entries to average.
        curves = {}
        for r in (BreedStandard.objects
                  .filter(breed_id__in={b.breed_id for b in batches if b.breed_id}, is_active=True)
                  .order_by("breed_id", "age")
                  .values("breed_id", "age", "cum_feed", "body_weight", "feed_intake")):
            curves.setdefault(r["breed_id"], []).append(
                (r["age"], r["cum_feed"] or Decimal("0"), r["body_weight"], r["feed_intake"]))

        # Latest weighed entry per batch (days without a weighing are skipped,
        # as in Live Flock Summary) for the body-weight-against-standard note.
        weight_by_batch = {}
        for de in (DailyEntry.objects.filter(batch_id__in=batch_ids, avg_weight_gms__gt=0)
                   .values("batch_id", "avg_weight_gms").order_by("-date", "-id")):
            weight_by_batch.setdefault(de["batch_id"], de["avg_weight_gms"])

        # Burn rate per feed item: the last few daily entries that used it in
        # the fortnight up to each batch's own reference date, so a live flock
        # whose entries have stopped doesn't keep projecting cover off stale
        # numbers and a closed one is rated on the days it was actually fed.
        window = timedelta(days=14)
        floor = min((a - window for a in as_of_by_batch.values()), default=today)
        recent_feed = {}
        for de in (DailyEntry.objects.filter(batch_id__in=batch_ids, date__gte=floor)
                   .values("batch_id", "date", "feed_1_id", "feed_1_qty", "feed_2_id", "feed_2_qty")
                   .order_by("-date", "-id")):
            as_of = as_of_by_batch.get(de["batch_id"], today)
            if not (as_of - window <= de["date"] <= as_of):
                continue
            for slot_item, slot_qty in (("feed_1_id", "feed_1_qty"), ("feed_2_id", "feed_2_qty")):
                if not de[slot_item]:
                    continue
                bucket = recent_feed.setdefault((de["batch_id"], de[slot_item]), [])
                if len(bucket) < 3:
                    bucket.append(de[slot_qty] or Decimal("0"))
        daily_rate = {k: _div(sum(v), len(v)) for k, v in recent_feed.items() if v}

        masters = list(FeedPhaseMaster.objects.filter(status="active")
                       .select_related("breed", "bird_category")
                       .prefetch_related("lines__feed_item"))
        item_meta = {
            r["id"]: r for r in Item.objects.filter(
                id__in={i for _b, i in list(feed_in) + list(feed_out) + list(consumed)}
            ).values("id", "description", "kg_per_bag")
        }

        groups = []
        for batch in batches:
            farm = batch.broiler_farm
            as_of = as_of_by_batch.get(batch.id, today)
            is_live = live_flag.get(batch.id, True)
            placed, placed_on = placement.get(batch.id, (Decimal("0"), None))
            start = batch.start_date or placed_on
            age = (as_of - start).days if start else None
            alive = placed - losses_by_batch.get(batch.id, 0) - sold_by_batch.get(batch.id, 0)
            # Before any bird is lost the two are the same; once a flock is
            # fully sold out, fall back to placed so the phase figures stay
            # readable instead of collapsing to zero.
            birds = alive if alive > 0 else placed
            master = _feed_phase_master_for(batch, as_of, masters)
            curve = curves.get(batch.breed_id) or []
            last_entry = last_entry_by_batch.get(batch.id)
            gap_days = (as_of - last_entry).days if last_entry else None
            mort_pct = (_div(losses_by_batch.get(batch.id, 0) * 100, placed)
                        if placed else None)
            std_bwt, std_intake = _std_at(curve, age)
            avg_bwt = weight_by_batch.get(batch.id)

            # Phases of the batch's own programme, in feeding order, then any
            # other feed item it has actually moved or eaten (no schedule to
            # compare against, but its stock still belongs on the farm).
            phase_items, ordered = [], []
            if master:
                for pl in sorted(master.lines.all(), key=lambda x: (x.from_age, x.seq_no)):
                    if pl.status == "active" and pl.feed_item_id and pl.feed_item_id not in phase_items:
                        phase_items.append(pl.feed_item_id)
                        ordered.append((pl.feed_item_id, pl.feed_item.description,
                                        pl.feed_item.kg_per_bag, pl))
            for _b, item_id in [k for k in list(feed_in) + list(feed_out) + list(consumed) if k[0] == batch.id]:
                if item_id not in phase_items:
                    phase_items.append(item_id)
                    meta = item_meta.get(item_id) or {}
                    ordered.append((item_id, meta.get("description") or "", meta.get("kg_per_bag"), None))

            # Which phase the flock is on now, and what follows it — the
            # changeover instruction needs both.
            phase_lines = [o for o in ordered if o[3] is not None]
            current_item, next_phase = None, ""
            for idx, (iid, _nm, _kg, pl) in enumerate(phase_lines):
                if age is not None and pl.from_age <= age and (pl.to_age is None or age <= pl.to_age):
                    current_item = iid
                    next_phase = phase_lines[idx + 1][1] if idx + 1 < len(phase_lines) else ""
                    break

            batch_rows = []
            for item_id, item_name, kg_per_bag, pl in ordered:
                in_qty = feed_in.get((batch.id, item_id), Decimal("0"))
                cons_qty = consumed.get((batch.id, item_id), Decimal("0"))
                out_qty = feed_out.get((batch.id, item_id), Decimal("0"))
                due_bird, basis = _phase_feed_due(curve, pl, age)
                due_qty = (due_bird / 1000 * birds) if due_bird is not None else None
                cons_bird = _div(cons_qty, birds) * 1000
                avail_qty = in_qty - cons_qty - out_qty
                is_current = pl is not None and item_id == current_item
                rate = daily_rate.get((batch.id, item_id), Decimal("0"))
                # With no entries to average, the phase the flock is on can
                # still be projected from the breed's standard daily intake —
                # which is exactly when cover matters most. Finished phases get
                # no such fallback; they aren't being eaten any more.
                rate_basis = "actual" if rate > 0 else ""
                if rate <= 0 and is_current and std_intake:
                    rate = _div(Decimal(str(std_intake)) * birds, 1000)
                    rate_basis = "std" if rate > 0 else ""
                cap_bird = ((pl.max_feed_qty or Decimal("0")) * 1000) if pl else None
                batch_rows.append({
                    "item": item_name, "item_id": item_id,
                    # Plain hyphen, not an en dash: this string goes out through
                    # CSV, where a non-ASCII character turns to mojibake in Excel.
                    "phase_band": ("" if pl is None else
                                   (f"{pl.from_age}+" if pl.to_age is None else f"{pl.from_age}-{pl.to_age}")),
                    # Every phase of the programme is listed whatever the
                    # flock's age, so each says where it stands: the one being
                    # fed now, one already behind it, or one still to come
                    # (listed so its Total Required can be procured ahead).
                    "phase_state": ("" if pl is None or age is None else
                                    "current" if is_current else
                                    "upcoming" if pl.from_age > age else "done"),
                    # Per-bird feed reads in grams, as everywhere else in the
                    # broiler reports; the phase cap is held per bird in Kg,
                    # so it is scaled here rather than stored that way.
                    "cap_bird": cap_bird.quantize(q2) if cap_bird else None,
                    # The whole phase's requirement for this flock, against
                    # which Scheduled to Date is the part due so far.
                    "req_qty": (cap_bird / 1000 * birds).quantize(q2) if cap_bird else None,
                    # What a standard bird of this breed eats on today's day of
                    # age, and what that comes to across the flock — the day's
                    # requirement, shown against the phase actually being fed
                    # since that is the feed the intake applies to.
                    "intake_bird": (Decimal(str(std_intake)).quantize(q2)
                                    if is_current and std_intake else None),
                    "intake_qty": (_div(Decimal(str(std_intake)) * birds, 1000).quantize(q2)
                                   if is_current and std_intake else None),
                    "due_bird": due_bird.quantize(q2) if due_bird is not None else None,
                    "due_qty": due_qty.quantize(q2) if due_qty is not None else None,
                    "basis": basis,
                    "feed_in": in_qty.quantize(q2),
                    "cons_bird": cons_bird.quantize(q2), "cons_qty": cons_qty.quantize(q2),
                    "out_qty": out_qty.quantize(q2),
                    "avail_qty": avail_qty.quantize(q2),
                    "bags": _div(avail_qty, kg_per_bag).quantize(q2) if kg_per_bag else None,
                    "diff_bird": (due_bird - cons_bird).quantize(q2) if due_bird is not None else None,
                    "diff_kgs": (due_qty - cons_qty).quantize(q2) if due_qty is not None else None,
                    "daily_rate": rate.quantize(q2), "rate_basis": rate_basis,
                    # No rate at all means no honest projection — a finished
                    # phase with no standard to fall back on, not "0 days".
                    "days_cover": _div(avail_qty, rate).quantize(q2) if rate > 0 else None,
                    "next_3_days": (rate * 3).quantize(q2),
                    "is_current": is_current, "next_phase": next_phase,
                    "bag_kg": kg_per_bag,
                    # Feeding order, so the per-feed summary below lists items
                    # the way the programme runs rather than alphabetically.
                    "sort_age": pl.from_age if pl else 9999,
                })

            if excess_only:
                # Excess is measured against the phase cap, not against the
                # standard curve the Difference column uses: the cap is the
                # operational allowance a farm shouldn't feed past, while the
                # curve is a performance benchmark a flock can trail either way.
                batch_rows = [r for r in batch_rows
                              if r["cap_bird"] is not None and r["cons_bird"] > r["cap_bird"]]
            elif not batch_rows:
                # Keep a placed-but-idle flock visible instead of silently
                # dropping it out of the branch's scheduling picture.
                batch_rows = [{"item": "", "item_id": None, "phase_band": "", "phase_state": "",
                               "cap_bird": None, "req_qty": None,
                               "intake_bird": None, "intake_qty": None,
                               "due_bird": None, "due_qty": None, "basis": "",
                               "feed_in": Decimal("0.00"), "cons_bird": Decimal("0.00"),
                               "cons_qty": Decimal("0.00"), "out_qty": Decimal("0.00"),
                               "avail_qty": Decimal("0.00"), "bags": None,
                               "diff_bird": None, "diff_kgs": None,
                               "daily_rate": Decimal("0.00"), "rate_basis": "",
                               "days_cover": None, "next_3_days": Decimal("0.00"),
                               "is_current": False, "next_phase": "", "bag_kg": None,
                               "sort_age": 9999}]
            if not batch_rows:
                continue

            # Anything that would otherwise make a figure look wrong rather
            # than say why: no programme to schedule against, a flock that has
            # stopped reporting, or stock that has gone impossible.
            flags = []
            if master is None:
                flags.append("No feed programme")
            if not is_live:
                flags.append(f"Closed {as_of.strftime('%d.%m.%Y')}")
            elif last_entry and (today - last_entry).days > 14:
                # Only a running flock owes daily entries; a finished one
                # having none lately is simply what "finished" looks like.
                flags.append(f"No entry since {last_entry.strftime('%d.%m.%Y')}")
            if any(r["avail_qty"] < 0 for r in batch_rows):
                flags.append("Negative stock")

            batch_actions = _feed_batch_actions(
                is_live=is_live, gap_days=gap_days, no_programme=master is None,
                mort_pct=mort_pct, avg_bwt=avg_bwt, std_bwt=std_bwt)
            for r in batch_rows:
                r["actions"] = _feed_row_actions(
                    r, is_current=r["is_current"], next_phase=r["next_phase"],
                    bag_kg=r["bag_kg"])
            if batch_rows:
                batch_rows[0]["actions"] = batch_actions + batch_rows[0]["actions"]

            covers = [r["days_cover"] for r in batch_rows if r["days_cover"] is not None]
            groups.append({
                "batch": batch, "farm": farm, "rows": batch_rows, "flags": flags,
                "age": age, "placed": placed, "birds": birds, "as_of": as_of,
                "last_entry": last_entry, "gap_days": gap_days,
                "start": start, "min_cover": min(covers) if covers else None,
            })

        # Urgency ordering is what makes this a dispatch list rather than a
        # register; flocks with no rate to project from sort last either way.
        if sort == "cover":
            groups.sort(key=lambda g: (g["min_cover"] is None,
                                       g["min_cover"] if g["min_cover"] is not None else 0))

        t_req = t_intake = t_due = t_in = t_cons = t_out = t_avail = Decimal("0")
        t_rate = t_next3 = t_bags = Decimal("0")
        t_low_cover = t_negative = t_no_prog = t_urgent = 0
        for sl_no, g in enumerate(groups, start=1):
            batch, farm = g["batch"], g["farm"]
            search_key = " ".join([farm.farm_name or "", farm.farm_code or "", batch.batch_name or "",
                                   farm.line or "", *(r["item"] for r in g["rows"])]).lower()
            if "No feed programme" in g["flags"]:
                t_no_prog += 1
            for i, r in enumerate(g["rows"]):
                r.update({
                    "sl_no": sl_no if i == 0 else "",
                    "farm_name": farm.farm_name if i == 0 else "",
                    "batch_name": batch.batch_name if i == 0 else "",
                    "age": (g["age"] if g["age"] is not None else "") if i == 0 else "",
                    "last_entry": g["last_entry"] if i == 0 else "",
                    "gap_days": (g["gap_days"] if g["gap_days"] is not None else "") if i == 0 else "",
                    "placed": g["placed"] if i == 0 else "",
                    "birds": g["birds"] if i == 0 else "",
                    "flags": g["flags"] if i == 0 else [],
                    "is_first": i == 0, "group": sl_no, "search": search_key,
                    "farm_code": farm.farm_code, "branch": farm.branch.branch_name if farm.branch_id else "",
                    "line": farm.line or "",
                    "supervisor": farm.supervisor.name if farm.supervisor_id else "",
                    # Farm name links through to this flock's day-by-day
                    # register, opened over the batch's own life rather than
                    # that report's default last-30-days.
                    "farm_id": farm.id, "batch_id": batch.id,
                    "entry_from": g["start"].isoformat() if g["start"] else "",
                    "entry_to": g["as_of"].isoformat(),
                })
                t_req += r["req_qty"] or Decimal("0")
                t_intake += r["intake_qty"] or Decimal("0")
                t_due += r["due_qty"] or Decimal("0")
                t_in += r["feed_in"]; t_cons += r["cons_qty"]
                t_out += r["out_qty"]; t_avail += r["avail_qty"]
                t_bags += r["bags"] or Decimal("0")
                t_rate += r["daily_rate"]; t_next3 += r["next_3_days"]
                if r["days_cover"] is not None and r["days_cover"] < 3:
                    t_low_cover += 1
                if r["avail_qty"] < 0:
                    t_negative += 1
                # Flat text so the same actions survive into CSV/Excel.
                r["action_text"] = "; ".join(t for _lvl, t in r["actions"])
                if any(lvl == "urgent" for lvl, _t in r["actions"]):
                    t_urgent += 1
            rows.extend(g["rows"])

        # Per feed type, across everything the filters selected. A single
        # grand total mixes Pre-Starter with Finisher into a figure nobody can
        # order against — what a buyer needs is how much of each feed.
        by_feed = {}
        for r in rows:
            b = by_feed.setdefault(r["item"] or "(no feed item)", {
                "item": r["item"] or "(no feed item)", "item_id": r["item_id"],
                "sort_age": r["sort_age"],
                "batches": set(), "bag_kg": r["bag_kg"],
                "req_qty": Decimal("0"), "intake_qty": Decimal("0"), "next_3_days": Decimal("0"),
                "feed_in": Decimal("0"), "cons_qty": Decimal("0"), "avail_qty": Decimal("0"),
                "bags": Decimal("0"),
            })
            b["sort_age"] = min(b["sort_age"], r["sort_age"])
            b["batches"].add(r["group"])
            b["bag_kg"] = b["bag_kg"] or r["bag_kg"]
            for k in ("req_qty", "intake_qty", "next_3_days", "feed_in", "cons_qty", "avail_qty", "bags"):
                b[k] += r[k] or Decimal("0")
        feed_summary = sorted(by_feed.values(), key=lambda b: (b["sort_age"], b["item"]))
        # What the branch warehouses could actually dispatch against that. A
        # Warehouse reaches its Branch through inventory.Mapping, the same hop
        # the batch report takes.
        branch_ids = {g["farm"].branch_id for g in groups if g["farm"].branch_id}
        wh_ids = list(Mapping.objects.filter(type=Mapping.TYPE_SECTOR_BRANCH, to_id__in=branch_ids)
                      .values_list("from_id", flat=True))
        item_ids = {r["item_id"] for r in rows if r["item_id"]}
        wh_stock = _warehouse_feed_stock(wh_ids, item_ids, today)

        for b in feed_summary:
            b["batches"] = len(b["batches"])
            # What is still to reach the farms for the whole programme; a
            # negative reads as already over-delivered, not as nothing to send.
            b["to_send"] = b["req_qty"] - b["feed_in"]
            b["next_3_bags"] = (b["next_3_days"] / b["bag_kg"]).quantize(q2) if b["bag_kg"] else None
            b["wh_stock"] = wh_stock.get(b["item_id"], Decimal("0"))
            # Nothing can be dispatched out of a negative balance, so the
            # shortfall is measured against stock actually on hand: a warehouse
            # whose books have gone negative is a data problem to fix (flagged
            # in red), not extra feed to go and buy.
            b["short_by"] = max(b["to_send"] - max(b["wh_stock"], Decimal("0")), Decimal("0"))

        totals = {"next_3_bags": sum((b["next_3_bags"] or Decimal("0") for b in feed_summary),
                                     Decimal("0")),
                  "wh_stock": sum((b["wh_stock"] for b in feed_summary), Decimal("0")),
                  "short_by": sum((b["short_by"] for b in feed_summary), Decimal("0")),
                  "req_qty": t_req, "intake_qty": t_intake,
                  "due_qty": t_due, "feed_in": t_in, "cons_qty": t_cons,
                  "out_qty": t_out, "avail_qty": t_avail, "bags": t_bags,
                  "diff_kgs": t_due - t_cons, "daily_rate": t_rate, "next_3_days": t_next3,
                  "low_cover": t_low_cover, "negative": t_negative, "no_programme": t_no_prog,
                  "urgent": t_urgent, "batches": len(groups)}

        if export in ("csv", "excel"):
            return _feed_scheduling_export(rows, totals, export, feed_summary)

    lines = (farms_for(request.user).exclude(line="").order_by("line")
             .values_list("line", flat=True).distinct())

    return render(request, "batch_wise_feed_scheduling_report.html", {
        "branches": branches_for(request.user, Branch.objects.order_by("branch_name")),
        "lines": lines,
        "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
        "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch").order_by("farm_name")),
        "branch_id": branch_id, "line": line, "supervisor_id": supervisor_id, "farm_id": farm_id,
        "excess_only": excess_only, "export": export, "sort": sort, "status": status,
        "submitted": submitted, "rows": rows, "totals": totals, "feed_summary": feed_summary,
    })


FEED_SCHEDULING_COLUMNS = [
    ("sl_no", "Sl.No"), ("farm_name", "Farm"), ("batch_name", "Batch"), ("age", "Age (days)"),
    ("last_entry", "Latest Entry"), ("gap_days", "Gap (days)"),
    ("item", "Item"), ("placed", "Placed Birds"), ("birds", "Available Birds"),
    ("phase_band", "Phase Age (days)"), ("cap_bird", "Phase Cap (gm)"),
    ("req_qty", "Total Required Feed (Kg)"),
    ("intake_bird", "Day's Intake/Bird (gm)"), ("intake_qty", "Day's Intake Total (Kg)"),
    ("due_bird", "Scheduled to Date (Cumulative Intake)/Bird (gm)"),
    ("due_qty", "Scheduled to Date (Cumulative Intake) Qty (Kg)"),
    ("basis", "Schedule Basis"),
    ("feed_in", "Feed In Qty (Kg)"),
    ("cons_bird", "Consumption Feed/Bird (gm)"), ("cons_qty", "Consumed Qty (Kg)"),
    ("out_qty", "Transferred Out Qty (Kg)"), ("avail_qty", "Available Qty (Kg)"),
    ("bags", "Available Bags"),
    ("diff_bird", "Difference Feed/Bird (gm)"), ("diff_kgs", "Difference Feed (Kg)"),
    ("daily_rate", "Daily Rate (Kg)"), ("rate_basis", "Rate Basis"), ("days_cover", "Days Cover"),
    ("next_3_days", "Next 3 Days Req. (Kg)"), ("action_text", "Action Required"),
]


FEED_SUMMARY_COLUMNS = [
    ("item", "Feed Item"), ("batches", "Batches"),
    ("req_qty", "Total Required (Kg)"), ("feed_in", "Already Sent (Kg)"),
    ("to_send", "Still To Send (Kg)"),
    ("wh_stock", "Branch Warehouse Stock (Kg)"), ("short_by", "Short By (Kg)"),
    ("cons_qty", "Consumed (Kg)"),
    ("avail_qty", "Farm Stock (Kg)"), ("bags", "Farm Stock (Bags)"),
    ("intake_qty", "Day's Intake (Kg/day)"),
    ("next_3_days", "Next 3 Days (Kg)"), ("next_3_bags", "Next 3 Days (Bags)"),
]


def _feed_scheduling_export(rows, totals, export, feed_summary=()):
    """Batch wise Feed Scheduling rows as a CSV or Excel download. Excel also
    carries the per-feed-type summary on its own sheet, since that is the sheet
    a buyer orders from; CSV stays one table and keeps only the detail."""
    from django.http import HttpResponse
    import csv as _csv

    headers = [label for _key, label in FEED_SCHEDULING_COLUMNS]
    data = [[("" if r.get(key) is None else r.get(key, "")) for key, _label in FEED_SCHEDULING_COLUMNS]
            for r in rows]
    total_row = ["", "", "", "", "", "", "Total", "", "", "", "", totals["req_qty"],
                 "", totals["intake_qty"], "", totals["due_qty"], "", totals["feed_in"],
                 "", totals["cons_qty"], totals["out_qty"], totals["avail_qty"],
                 totals["bags"], "", totals["diff_kgs"], totals["daily_rate"], "", "",
                 totals["next_3_days"], ""]

    if export == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="batch_wise_feed_scheduling.csv"'
        writer = _csv.writer(response)
        writer.writerow(headers)
        writer.writerows(data)
        if rows:
            writer.writerow(total_row)
        return response

    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Feed Scheduling"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in data:
        ws.append([float(v) if isinstance(v, Decimal) else v for v in row])
    if rows:
        ws.append([float(v) if isinstance(v, Decimal) else v for v in total_row])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for i, label in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(len(label) + 2, 12)

    if feed_summary:
        ws2 = wb.create_sheet("By Feed Type")
        s_headers = [label for _key, label in FEED_SUMMARY_COLUMNS]
        ws2.append(s_headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for b in feed_summary:
            ws2.append([float(b[k]) if isinstance(b[k], Decimal) else ("" if b[k] is None else b[k])
                        for k, _label in FEED_SUMMARY_COLUMNS])
        for i, label in enumerate(s_headers, start=1):
            ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = max(len(label) + 2, 12)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="batch_wise_feed_scheduling.xlsx"'
    wb.save(response)
    return response


@login_required
def feed_dispatch_stock_report(request):
    """Feed Dispatch & Stock ledger for a single Warehouse (Broiler > Reports
    > Feed Dispatch & Stock Report) — one row per dispatch (Warehouse ->
    Farm), return (Farm -> Warehouse) or purchase receipt (Supplier ->
    Warehouse) of a tracked feed item, in chronological order, carrying a
    running per-feed-type stock balance (in bags) forward after every row —
    matching a feed store's own paper stock register.

    Stock is computed independently, event by event, from the very first
    historical transaction for this warehouse: no single transaction type's
    own running-stock field reflects the *combined* physical balance (see
    StockTransfer.stock / InventoryAdjustmentItem.stock — each only chains
    through its own transaction type), and Purchases don't touch any
    running-stock field at all.

    Freight is only ever real on purchase-receipt rows (GeneralPurchase.
    freight_amount) — Stock Transfer has no freight field, so dispatch/return
    rows show it blank rather than a fabricated 0.
    """
    from account.models import CompanyProfile
    from django.utils.dateparse import parse_date
    from inventory.models import StockTransfer, Mapping
    from purchase.models import GeneralPurchaseItem

    region_id = (request.GET.get("region") or "").strip()
    branch_id = (request.GET.get("branch") or "").strip()
    warehouse_id = (request.GET.get("warehouse") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    export = (request.GET.get("export") or "display").strip().lower()
    submitted = bool(from_date or to_date or request.GET.get("submit"))

    # Feed columns are fully dynamic — every Item under a "Feed" category is
    # its own ledger column, keyed by item id and ordered by item_code. Add a
    # new feed Item and it shows up here automatically, no code change needed.
    feed_items_all = list(Item.objects.filter(category__name__icontains="feed").order_by("item_code"))
    # Every figure here is in bags, converted from kg by the item's bag weight.
    # An item without one divides by nothing, so its columns read 0 however much
    # was actually bought or dispatched — which looks like "no activity" rather
    # than "cannot convert". Name those items instead of leaving zeros to be
    # puzzled over.
    items_missing_bag_weight = [it.description for it in feed_items_all if not it.kg_per_bag]
    label_ids = [it.id for it in feed_items_all]
    label_item = {it.id: it for it in feed_items_all}
    label_name = {it.id: it.description for it in feed_items_all}
    # Fixed (non-feed) columns: Date, Txn No.(ERP), Challan No., Branch,
    # Warehouse, Supplier/Farm Name, Farm Batch No. (7) + Total Bags/Total Kg
    # (2) + Net Total Bag Stock (1) + Vehicle No./Freight Paid/Remarks (3) = 13.
    opening_label_colspan = 7  # Date..Farm Batch No.
    total_columns = 13 + 3 * len(label_ids)

    ledger_rows, opening, ledger_totals, trend = [], None, None, None
    kpi = {"total_dispatched_bags": Decimal("0"), "total_received_bags": Decimal("0"),
           "net_total_bag_stock": Decimal("0"), "total_freight": Decimal("0")}
    top_feed = [{"label": label_name[l], "value": Decimal("0")} for l in label_ids]
    recent_activity = {"dispatched": Decimal("0"), "received": Decimal("0"),
                       "adjustment": Decimal("0"), "returned": Decimal("0")}
    warehouse_snapshot = []

    # A Warehouse (Office) has no direct Branch FK — resolved via inventory.
    # Mapping (TYPE_SECTOR_BRANCH: from_id=warehouse, to_id=branch). Region
    # narrows to every Branch in it, Branch narrows to its mapped Warehouse(s).
    from user.services.scoping import (allowed_ids, branches_for, describe,
                                       warehouses_for)

    # Data scoping. The querystring is not a permission: an id outside the
    # user's scope is dropped rather than honoured, so a hand-typed
    # ?warehouse=<other branch> cannot read another branch's ledger.
    scoped_branches = branches_for(request.user)
    scoped_warehouses = warehouses_for(request.user)
    if branch_id and not scoped_branches.filter(id=branch_id).exists():
        branch_id = ""
    if warehouse_id and not scoped_warehouses.filter(id=warehouse_id).exists():
        warehouse_id = ""

    branch_obj = Branch.objects.filter(id=branch_id).first() if branch_id else None
    region_obj = Region.objects.filter(id=region_id).first() if region_id else None

    warehouse_obj = Warehouse.objects.filter(id=warehouse_id).first() if warehouse_id else None
    # "All Warehouses" combines every location's stock into one running total
    # (a company-wide feed pipeline view) — the Warehouse column is only
    # shown in that mode so each row's origin/destination stays traceable;
    # with one Warehouse picked, the ledger matches the paper register 1:1.
    all_warehouses = submitted and not warehouse_obj

    if submitted:
        tracked_item_ids = label_ids
        item_label = {iid: iid for iid in label_ids}  # event item_id -> column key (identity: 1 item = 1 column)

        from_date_obj = parse_date(from_date) if from_date else None
        to_date_obj = parse_date(to_date) if to_date else None

        # Scope every event query to the narrowest filter given: a specific
        # Warehouse wins outright; otherwise Branch (via Mapping) or Region
        # (via every Branch in it) narrows to that set of Warehouses; with
        # none of the three, every Warehouse is in scope (None = no filter).
        if warehouse_obj:
            scoped_warehouse_ids = {warehouse_obj.id}
        elif branch_obj:
            scoped_warehouse_ids = set(Mapping.objects.filter(
                type=Mapping.TYPE_SECTOR_BRANCH, to_id=branch_obj.id).values_list("from_id", flat=True))
        elif region_obj:
            branch_ids_in_region = Branch.objects.filter(region_id=region_obj.id).values_list("id", flat=True)
            scoped_warehouse_ids = set(Mapping.objects.filter(
                type=Mapping.TYPE_SECTOR_BRANCH, to_id__in=branch_ids_in_region).values_list("from_id", flat=True))
        else:
            scoped_warehouse_ids = None

        # ...and "All Warehouses" means all the *user* may see. Without this the
        # unfiltered path would read every warehouse in the company.
        permitted = allowed_ids(request.user, "sectors")
        if permitted is not None:
            scoped_warehouse_ids = (permitted if scoped_warehouse_ids is None
                                    else scoped_warehouse_ids & permitted)

        # Warehouse -> Branch, resolved once per warehouse and cached (a
        # Warehouse/Office has no direct Branch FK — only via inventory.
        # Mapping) so "All Warehouses" mode doesn't re-query per row.
        _branch_cache = {}

        def _branch_for_warehouse(wh):
            if not wh:
                return ""
            if wh.id not in _branch_cache:
                branch_id = (Mapping.objects.filter(type=Mapping.TYPE_SECTOR_BRANCH, from_id=wh.id)
                             .values_list("to_id", flat=True).first())
                branch = Branch.objects.filter(id=branch_id).first() if branch_id else None
                _branch_cache[wh.id] = branch.branch_name if branch else ""
            return _branch_cache[wh.id]

        # ---- gather every event ever recorded for the tracked feed items,
        # scoped to one Warehouse, or every Warehouse when none is chosen ----
        events = []
        transfer_qs = StockTransfer.objects.filter(item_id__in=tracked_item_ids)
        transfer_qs = (transfer_qs.filter(Q(from_warehouse_id__in=scoped_warehouse_ids) | Q(to_warehouse_id__in=scoped_warehouse_ids))
                      if scoped_warehouse_ids is not None
                      else transfer_qs.filter(Q(from_warehouse__isnull=False) | Q(to_warehouse__isnull=False)))
        transfer_qs = transfer_qs.select_related("to_farm__farmer", "from_farm__farmer",
                                                  "from_warehouse", "to_warehouse", "to_batch", "from_batch")
        for t in transfer_qs:
            if t.from_warehouse_id and (scoped_warehouse_ids is None or t.from_warehouse_id in scoped_warehouse_ids):
                events.append({
                    "date": t.date, "sort_key": (t.date, 0, t.id), "kind": "dispatch", "source": "dispatch",
                    "item_id": t.item_id,
                    "qty_kg": t.quantity or Decimal("0"), "challan_no": t.dc_no, "txn_no": t.trnum,
                    "warehouse_id": t.from_warehouse_id,
                    "warehouse_name": t.from_warehouse.name, "branch_name": _branch_for_warehouse(t.from_warehouse),
                    "batch_no": t.to_batch.batch_name if t.to_batch_id else "",
                    "farm_code": t.to_farm.farm_code if t.to_farm_id else "",
                    "name": (t.to_farm.farmer.farmer_name if t.to_farm_id and t.to_farm.farmer_id
                             else (t.to_farm.farm_name if t.to_farm_id else "")),
                    "vehicle_no": t.vehicle_no, "freight": None, "remarks": t.remarks,
                })
            if t.to_warehouse_id and (scoped_warehouse_ids is None or t.to_warehouse_id in scoped_warehouse_ids):
                events.append({
                    "date": t.date, "sort_key": (t.date, 1, t.id), "kind": "receipt", "source": "return",
                    "item_id": t.item_id,
                    "qty_kg": t.quantity or Decimal("0"), "challan_no": t.dc_no, "txn_no": t.trnum,
                    "warehouse_id": t.to_warehouse_id,
                    "warehouse_name": t.to_warehouse.name, "branch_name": _branch_for_warehouse(t.to_warehouse),
                    "batch_no": t.from_batch.batch_name if t.from_batch_id else "",
                    "farm_code": t.from_farm.farm_code if t.from_farm_id else "",
                    "name": (t.from_farm.farmer.farmer_name if t.from_farm_id and t.from_farm.farmer_id
                             else (t.from_farm.farm_name if t.from_farm_id else "")),
                    "vehicle_no": t.vehicle_no, "freight": None, "remarks": t.remarks,
                })

        purchase_qs = GeneralPurchaseItem.objects.filter(item_id__in=tracked_item_ids)
        purchase_qs = (purchase_qs.filter(farm_warehouse_id__in=scoped_warehouse_ids) if scoped_warehouse_ids is not None
                      else purchase_qs.filter(farm_warehouse__isnull=False))
        purchase_qs = purchase_qs.select_related("purchase", "purchase__supplier", "farm_warehouse")
        for p in purchase_qs:
            events.append({
                "date": p.purchase.date, "sort_key": (p.purchase.date, 1, -p.id), "kind": "receipt",
                "source": "purchase", "item_id": p.item_id,
                # Sent or Received per the purchase's own basis: on the Sent
                # basis (the default) rcv_qty stays zero, and reading it alone
                # left every such receipt showing nothing here.
                "qty_kg": (p.effective_qty() or Decimal("0")) + (p.free_qty or Decimal("0")),
                "challan_no": p.purchase.dc_no, "txn_no": p.purchase.purchase_no,
                "warehouse_id": p.farm_warehouse_id,
                "warehouse_name": p.farm_warehouse.name, "branch_name": _branch_for_warehouse(p.farm_warehouse),
                "batch_no": "", "farm_code": "",
                "name": p.purchase.supplier.name if p.purchase.supplier_id else "",
                "vehicle_no": p.purchase.vehicle_no, "freight": p.purchase.freight_amount or Decimal("0"),
                "remarks": p.purchase.remarks,
            })

        # ---- Inventory > Transactions: Stock Received / Stock Issued /
        # Inventory Adjustment — all warehouse-scoped (location_type='warehouse'),
        # none of these carry their own running-stock balance either, so
        # they fold into the same from-scratch reconciliation as everything
        # else here. None of these three have a real Challan/DC field (only
        # an auto ERP trnum), so Challan No. is blank and trnum is shown
        # under Transaction No.(ERP) instead.
        from inventory.models import StockReceiveItem, StockIssueItem, InventoryAdjustmentItem

        receive_qs = StockReceiveItem.objects.filter(item_id__in=tracked_item_ids, location_type="warehouse")
        receive_qs = (receive_qs.filter(warehouse_id__in=scoped_warehouse_ids) if scoped_warehouse_ids is not None
                     else receive_qs.filter(warehouse__isnull=False))
        receive_qs = receive_qs.select_related("receive", "warehouse")
        for r in receive_qs:
            events.append({
                "date": r.receive.date, "sort_key": (r.receive.date, 1, -r.id), "kind": "receipt",
                "source": "stock_receive", "item_id": r.item_id, "qty_kg": r.quantity or Decimal("0"),
                "challan_no": "", "txn_no": r.receive.trnum,
                "warehouse_id": r.warehouse_id,
                "warehouse_name": r.warehouse.name, "branch_name": _branch_for_warehouse(r.warehouse),
                "batch_no": "", "farm_code": "",
                "name": "Stock Received" + (f" — {r.remarks}" if r.remarks else ""),
                "vehicle_no": "", "freight": None, "remarks": r.remarks,
            })

        issue_qs = StockIssueItem.objects.filter(item_id__in=tracked_item_ids, location_type="warehouse")
        issue_qs = (issue_qs.filter(warehouse_id__in=scoped_warehouse_ids) if scoped_warehouse_ids is not None
                   else issue_qs.filter(warehouse__isnull=False))
        issue_qs = issue_qs.select_related("issue", "warehouse")
        for s in issue_qs:
            events.append({
                "date": s.issue.date, "sort_key": (s.issue.date, 0, s.id), "kind": "dispatch",
                "source": "stock_issue", "item_id": s.item_id, "qty_kg": s.quantity or Decimal("0"),
                "challan_no": "", "txn_no": s.issue.trnum,
                "warehouse_id": s.warehouse_id,
                "warehouse_name": s.warehouse.name, "branch_name": _branch_for_warehouse(s.warehouse),
                "batch_no": "", "farm_code": "",
                "name": "Stock Issued",
                "vehicle_no": "", "freight": None, "remarks": "",
            })

        adj_qs = InventoryAdjustmentItem.objects.filter(item_id__in=tracked_item_ids,
                                                         adjustment__location_type="warehouse")
        adj_qs = (adj_qs.filter(adjustment__warehouse_id__in=scoped_warehouse_ids) if scoped_warehouse_ids is not None
                 else adj_qs.filter(adjustment__warehouse__isnull=False))
        adj_qs = adj_qs.select_related("adjustment", "adjustment__warehouse")
        for a in adj_qs:
            is_add = a.adjustment_type == "Add"
            events.append({
                "date": a.adjustment.date, "sort_key": (a.adjustment.date, 1 if is_add else 0, a.id),
                "kind": "receipt" if is_add else "dispatch", "source": "adjustment",
                "item_id": a.item_id, "qty_kg": a.quantity or Decimal("0"),
                "challan_no": "", "txn_no": a.adjustment.trnum,
                "warehouse_id": a.adjustment.warehouse_id,
                "warehouse_name": a.adjustment.warehouse.name, "branch_name": _branch_for_warehouse(a.adjustment.warehouse),
                "batch_no": "", "farm_code": "",
                "name": f"Stock Adjustment ({a.adjustment_type})" + (f" — {a.remarks}" if a.remarks else ""),
                "vehicle_no": "", "freight": None, "remarks": a.remarks,
            })
        events.sort(key=lambda e: e["sort_key"])

        def _bags(qty_kg, item):
            return (qty_kg / item.kg_per_bag).quantize(Decimal("0.01")) if item and item.kg_per_bag else Decimal("0")

        # Opening: replay every event strictly before From Date, from the very
        # start of history. With no From Date given there is no opening period
        # at all — opening stays zero and every event falls in the display
        # window below (guard is essential: without it the loop never breaks
        # and replays all events here, then the display loop replays them
        # again, doubling the running balance).
        running_kg = {label: Decimal("0") for label in label_ids}
        if from_date_obj:
            for e in events:
                if e["date"] >= from_date_obj:
                    break
                label = item_label.get(e["item_id"])
                if not label:
                    continue
                running_kg[label] += e["qty_kg"] if e["kind"] == "receipt" else -e["qty_kg"]

        opening_values = [_bags(running_kg[label], label_item.get(label)) for label in label_ids]
        opening = {"values": opening_values, "net": sum(opening_values)}

        # ---- emit rows within the display window, carrying the running balance ----
        # Events sharing the same Date + Challan No. + Warehouse + direction
        # (dispatch/receipt) are one real document covering several feed
        # types (e.g. one delivery challan with Pre-Starter + Starter both
        # on it) and are merged into a single row — each feed type still
        # gets its own column, but the farm/date/challan aren't repeated.
        # Events with a *different* Challan No. stay on their own row even
        # if same-day/same-farm, since they're genuinely separate documents.
        t_dispatched_bags = t_received_bags = t_freight = Decimal("0")
        dispatch_label_totals = [Decimal("0") for _ in label_ids]
        received_label_totals = [Decimal("0") for _ in label_ids]
        total_bag_sum = total_kg_sum = Decimal("0")
        grouped_rows, group_order = {}, []
        for e in events:
            if from_date_obj and e["date"] < from_date_obj:
                continue
            if to_date_obj and e["date"] > to_date_obj:
                continue
            label = item_label.get(e["item_id"])
            if not label:
                continue
            item = label_item.get(label)
            bags = _bags(e["qty_kg"], item)
            is_receipt = e["kind"] == "receipt"
            running_kg[label] += e["qty_kg"] if is_receipt else -e["qty_kg"]
            bag_weight = item.kg_per_bag if item and item.kg_per_bag else Decimal("0")

            if is_receipt:
                t_received_bags += bags
            else:
                t_dispatched_bags += bags
                total_bag_sum += bags
                total_kg_sum += (bags * bag_weight).quantize(Decimal("0.01"))
            if e["freight"]:
                t_freight += e["freight"]
            label_idx = label_ids.index(label)
            if is_receipt:
                received_label_totals[label_idx] += bags
            else:
                dispatch_label_totals[label_idx] += bags

            # Stock Received/Issued/Adjustment have no real Challan No. (blank
            # for all their lines), so grouping on it directly would merge
            # unrelated same-day documents at the same warehouse together.
            # Fall back to the event's own ERP txn_no (unique per document)
            # as the grouping key whenever there's no real challan.
            doc_key = e["challan_no"] or e["txn_no"]
            group_key = (e["date"], doc_key, e["warehouse_name"], is_receipt)
            row = grouped_rows.get(group_key)
            if row is None:
                row = {
                    "date": e["date"], "challan_no": e["challan_no"], "txn_no": e["txn_no"],
                    "warehouse_name": e["warehouse_name"], "branch_name": e["branch_name"], "batch_no": e["batch_no"],
                    "farm_code": e["farm_code"], "name": e["name"],
                    "dispatch_bags": [Decimal("0")] * len(label_ids),
                    "received_bags": [Decimal("0")] * len(label_ids),
                    "total_bag": Decimal("0"), "total_kg": Decimal("0"),
                    "vehicle_no": e["vehicle_no"], "freight": e["freight"], "remarks": e["remarks"],
                    "is_receipt": is_receipt,
                }
                grouped_rows[group_key] = row
                group_order.append(group_key)
            if is_receipt:
                row["received_bags"][label_idx] += bags
            else:
                row["dispatch_bags"][label_idx] += bags
                row["total_bag"] += bags
                row["total_kg"] += (bags * bag_weight).quantize(Decimal("0.01"))
            if e["freight"]:
                row["freight"] = (row["freight"] or Decimal("0")) + e["freight"]
            # Snapshot the running balance as of the latest event folded into
            # this row, so a merged row shows the state after the whole document.
            row["stock_bags"] = [_bags(running_kg[l], label_item.get(l)) for l in label_ids]
            row["net_stock"] = sum(row["stock_bags"])

        ledger_rows = [grouped_rows[k] for k in group_order]

        current_stock_values = [_bags(running_kg[l], label_item.get(l)) for l in label_ids]
        current_net_stock = sum(current_stock_values)
        kpi = {
            "total_dispatched_bags": t_dispatched_bags, "total_received_bags": t_received_bags,
            "net_total_bag_stock": current_net_stock,
            "total_freight": t_freight,
        }
        ledger_totals = {
            "dispatch_bags": dispatch_label_totals, "received_bags": received_label_totals,
            "total_bag": total_bag_sum, "total_kg": total_kg_sum,
            "stock_bags": current_stock_values, "net_stock": current_net_stock,
        }

        # Top feed by |net stock| — all tracked feed columns, ranked by
        # magnitude (a large negative balance is just as noteworthy as a
        # large positive one).
        top_feed = sorted(
            ({"label": label_name[l], "value": v} for l, v in zip(label_ids, current_stock_values)),
            key=lambda x: abs(x["value"]), reverse=True,
        )

        # Warehouse Snapshot: the same combined balance, split back out per
        # Warehouse — only meaningful in All-Warehouses mode (a single
        # Warehouse view already shows its own net stock in the KPI card).
        # Replays every event up to To Date (same cutoff as the ledger's own
        # current balance), keyed by warehouse this time instead of feed type.
        warehouse_snapshot = []
        if all_warehouses:
            wh_running_kg, wh_names, wh_branches = {}, {}, {}
            for e in events:
                if to_date_obj and e["date"] > to_date_obj:
                    continue
                label = item_label.get(e["item_id"])
                if not label:
                    continue
                wh_id = e["warehouse_id"]
                if wh_id not in wh_running_kg:
                    wh_running_kg[wh_id] = {l: Decimal("0") for l in label_ids}
                    wh_names[wh_id] = e["warehouse_name"]
                    wh_branches[wh_id] = e["branch_name"]
                wh_running_kg[wh_id][label] += e["qty_kg"] if e["kind"] == "receipt" else -e["qty_kg"]
            warehouse_snapshot = sorted(
                ({"name": wh_names[wh_id], "branch_name": wh_branches[wh_id],
                  "net_stock": sum(_bags(kg_map[l], label_item.get(l)) for l in label_ids)}
                 for wh_id, kg_map in wh_running_kg.items()),
                key=lambda x: x["name"],
            )

        # Recent Activity: last 7 days from *today*, independent of the report's
        # own date filter. Classified by each event's explicit `source` tag
        # (not kind/freight heuristics) so Stock Received/Issued/Adjustment
        # from Inventory > Transactions land in the right bucket:
        #   dispatched = Stock Transfer dispatch + Stock Issued
        #   received   = Purchase receipt + Stock Received
        #   returned   = Stock Transfer farm return
        #   adjustment = Inventory Adjustment (Add or Deduct)
        today = timezone.localdate()
        week_ago = today - timedelta(days=6)
        recent_dispatched = recent_received = recent_returned = recent_adjustment = Decimal("0")
        for e in events:
            if not (week_ago <= e["date"] <= today):
                continue
            label = item_label.get(e["item_id"])
            if not label:
                continue
            bags = _bags(e["qty_kg"], label_item.get(label))
            if e["source"] in ("dispatch", "stock_issue"):
                recent_dispatched += bags
            elif e["source"] in ("purchase", "stock_receive"):
                recent_received += bags
            elif e["source"] == "return":
                recent_returned += bags
            elif e["source"] == "adjustment":
                recent_adjustment += bags
        recent_activity = {
            "dispatched": recent_dispatched, "received": recent_received,
            "adjustment": recent_adjustment, "returned": recent_returned,
        }

        # Previous-period trend on the 4 headline KPIs — only computable when
        # both dates are given, since "previous period" needs a defined length.
        trend = None
        if from_date_obj and to_date_obj:
            period_days = (to_date_obj - from_date_obj).days + 1
            prev_to = from_date_obj - timedelta(days=1)
            prev_from = prev_to - timedelta(days=period_days - 1)

            prev_dispatched = prev_received = prev_freight = Decimal("0")
            for e in events:
                if not (prev_from <= e["date"] <= prev_to):
                    continue
                label = item_label.get(e["item_id"])
                if not label:
                    continue
                bags = _bags(e["qty_kg"], label_item.get(label))
                if e["kind"] == "dispatch":
                    prev_dispatched += bags
                else:
                    prev_received += bags
                    if e["freight"]:
                        prev_freight += e["freight"]

            prev_running = {label: Decimal("0") for label in label_ids}
            for e in events:
                if e["date"] > prev_to:
                    break
                label = item_label.get(e["item_id"])
                if not label:
                    continue
                prev_running[label] += e["qty_kg"] if e["kind"] == "receipt" else -e["qty_kg"]
            prev_net_stock = sum(_bags(prev_running[l], label_item.get(l)) for l in label_ids)

            def _pct_change(curr, prev):
                if prev:
                    return ((curr - prev) / abs(prev) * 100).quantize(Decimal("0.01"))
                return Decimal("0") if curr == 0 else Decimal("100.00")

            trend = {
                "dispatched": _pct_change(t_dispatched_bags, prev_dispatched),
                "received": _pct_change(t_received_bags, prev_received),
                "net_stock": _pct_change(current_net_stock, prev_net_stock),
                "freight": _pct_change(t_freight, prev_freight),
            }

    return render(request, "feed_dispatch_stock_report.html", {
        "regions": Region.objects.order_by("description"),
        "branches": scoped_branches.order_by("branch_name"),
        "warehouses": scoped_warehouses.order_by("name"),
        "scope_note": describe(request.user),
        "feed_labels": [label_name[l] for l in label_ids],
        "total_columns": total_columns, "opening_label_colspan": opening_label_colspan,
        "region_id": region_id, "branch_id": branch_id, "warehouse_id": warehouse_id,
        "warehouse_obj": warehouse_obj, "all_warehouses": all_warehouses,
        "from_date": from_date, "to_date": to_date,
        "submitted": submitted, "ledger_rows": ledger_rows, "opening": opening, "kpi": kpi,
        "ledger_totals": ledger_totals, "top_feed": top_feed,
        "recent_activity": recent_activity, "trend": trend, "warehouse_snapshot": warehouse_snapshot,
        "export": export,
        "items_missing_bag_weight": items_missing_bag_weight,
        "company": CompanyProfile.get_solo(),
    })


# ---------------------------------------------------------------------------
# Chicks Placement (Broiler > Transactions)
# ---------------------------------------------------------------------------
# Not a separate model — chick placement is an ordinary inventory Stock
# Transfer (Warehouse -> Farm/Batch) of a "chicks" category item, per the
# same convention the Batch History Report already relies on. This is a
# purpose-built, simplified front end over that same StockTransfer data
# (Warehouse supplier -> Farm, auto-derived active Batch, item restricted to
# the chicks category) so Broiler users don't have to use the generic
# Inventory > Stock Transfer form's full location-type/item-picker.

def _hatcheries_with_warehouse():
    """Hatchery queryset annotated with `warehouse_id` — its mapped Office,
    looked up from inventory.Mapping (Inventory > Office Mapping) rather
    than a direct FK, so templates can keep reading `h.warehouse_id` as
    before."""
    from django.db.models import OuterRef, Subquery
    from inventory.models import Mapping

    mapped_warehouse = Mapping.objects.filter(
        type=Mapping.TYPE_HATCHERY_OFFICE, from_id=OuterRef("pk")
    ).values("to_id")[:1]
    return Hatchery.objects.order_by("hatchery_name").annotate(
        warehouse_id=Subquery(mapped_warehouse)
    )


def _chicks_sources(user=None):
    """Options for the "Source Hatchery / Supplier" picker — every Hatchery
    followed by every Supplier, as one list the templates render in two
    optgroups.

    Values are prefixed ("h:<id>" / "s:<id>") because the two sides are
    separate masters stored in separate columns; the prefix is what lets a
    single select round-trip either without the form tracking which is which.
    Only hatcheries carry a `warehouse_id` (from Inventory > Office Mapping) —
    suppliers have no such mapping, so choosing one just leaves From Warehouse
    to be picked by hand.
    """
    from purchase.models import Supplier

    sources = [{"value": f"h:{h.id}", "label": h.hatchery_name or "",
                "group": "Hatcheries", "warehouse_id": h.warehouse_id or ""}
               for h in _hatcheries_with_warehouse()]
    supplier_qs = Supplier.objects.order_by("name")
    if user is not None:
        supplier_qs = suppliers_for(user, supplier_qs)
    sources += [{"value": f"s:{s.id}", "label": s.name or "",
                 "group": "Suppliers", "warehouse_id": ""}
                for s in supplier_qs]
    return sources


@method_decorator(login_required, name="dispatch")
class ChicksPlacementListTemplateView(View):
    def get(self, request):
        return render(request, "chicks_placement_list.html", {
            "warehouses": warehouses_for(request.user, Warehouse.objects.order_by("name")),
            "farms": farms_for(request.user, BroilerFarm.objects.order_by("farm_name")),
            "chick_items": Item.objects.filter(category__name__icontains="chick").order_by("item_code"),
            "sources": _chicks_sources(request.user),
        })


@method_decorator(login_required, name="dispatch")
class ChicksPlacementFormTemplateView(View):
    def get(self, request):
        return render(request, "chicks_placement_form.html", {
            "warehouses": warehouses_for(request.user, Warehouse.objects.order_by("name")),
            "farms": farms_for(request.user, BroilerFarm.objects.order_by("farm_name")),
            "chick_items": Item.objects.filter(category__name__icontains="chick").order_by("item_code"),
            "sources": _chicks_sources(request.user),
            "today": timezone.localdate().isoformat(),
        })


# ---------------------------------------------------------------------------
# Growing Charges Master (Rearing Charge)
# ---------------------------------------------------------------------------

GROWING_CHARGE_LIST_CACHE_KEY = "growing_charge_scheme_list"

# related_name -> (child model, ordered field list). Rows round-trip as JSON
# arrays; string fields (incentive_on / grade) stay as text, the rest default 0.
GC_CHILD_SPECS = [
    ("production_cost_incentives", GCProductionCostIncentive, ["from_production_cost", "to_production_cost", "rate_pct"]),
    ("sales_incentives", GCSalesIncentive, ["sale_rate_from", "sale_rate_to", "sales_incentive"]),
    ("mortality_incentives", GCMortalityIncentive, ["from_mortality_pct", "to_mortality_pct", "incentive_value"]),
    ("fcr_incentives", GCFCRIncentive, ["cfcr_limit", "body_weight", "incentive_value"]),
    ("summer_incentives", GCSummerIncentive, ["min_production_cost", "max_production_cost", "incentive_on", "from_production_cost", "to_production_cost", "incentive_rate"]),
    ("production_cost_decentives", GCProductionCostDecentive, ["from_production_cost", "to_production_cost", "rate_pct"]),
    ("mortality_decentives", GCMortalityDecentive, ["from_mortality_pct", "to_mortality_pct", "decentive_value"]),
    ("fcr_recoveries", GCFCRRecovery, ["cfcr_limit", "production_limit", "recovery_rate"]),
    ("farmer_classifications", GCFarmerClassification, ["production_cost_from", "production_to", "grade"]),
]
GC_STRING_ROW_FIELDS = {"incentive_on", "grade"}


@method_decorator(login_required, name="dispatch")
class GrowingChargeSchemeTemplateView(View):
    """Renders the Rearing / Growing Charge master page."""

    def get(self, request):
        context = {
            "regions": Region.objects.order_by("description"),
            "branches": branches_for(request.user, Branch.objects.select_related("region").order_by("branch_name")),
            "medicine_basis_choices": GrowingChargeScheme.MedicineCostBasis.choices,
            "shortage_basis_choices": GrowingChargeScheme.ShortageBasis.choices,
            "summer_incentive_on_choices": GCSummerIncentive.IncentiveOn.choices,
        }
        return render(request, "growing_charge.html", context)


@method_decorator(login_required, name="dispatch")
class GrowingChargeSchemeAPI(BaseAPIView):
    """CRUD API for the Rearing / Growing Charge master with its nested rows."""

    SCALAR_FIELDS = [
        "schema_name", "from_date", "to_date",
        "chick_cost", "feed_cost", "medicine_cost_basis", "medicine_cost",
        "farmer_admin_cost", "management_admin_cost", "std_production_cost",
        "standard_gc_cost", "minimum_gc_cost", "standard_fcr", "standard_mortality",
        "unloading_charges", "maximum_prod_cost", "maximum_rate_incentive",
        "mort_dec_first_week_exceeds", "mort_dec_overall_above", "mort_dec_first_week_value",
        "shortage_basis",
    ]
    CHOICE_FIELDS = {"medicine_cost_basis", "shortage_basis"}
    DECIMAL_FIELDS = {
        "chick_cost", "feed_cost", "medicine_cost", "farmer_admin_cost",
        "management_admin_cost", "std_production_cost", "standard_gc_cost",
        "minimum_gc_cost", "standard_fcr", "standard_mortality", "unloading_charges",
        "maximum_prod_cost", "maximum_rate_incentive", "mort_dec_first_week_exceeds",
        "mort_dec_overall_above", "mort_dec_first_week_value",
    }

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                scheme = GrowingChargeScheme.objects.select_related("region", "branch").get(id=id)
                data = {f: getattr(scheme, f) for f in self.SCALAR_FIELDS}
                for f in self.DECIMAL_FIELDS:
                    data[f] = str(data[f])
                data["from_date"] = scheme.from_date.isoformat() if scheme.from_date else None
                data["to_date"] = scheme.to_date.isoformat() if scheme.to_date else None
                data.update({
                    "id": scheme.id,
                    "scheme_code": scheme.scheme_code,
                    "region_id": scheme.region_id,
                    "branch_id": scheme.branch_id,
                    "is_active": scheme.is_active,
                    "is_locked": scheme.is_locked,
                })
                for related_name, _model, fields in GC_CHILD_SPECS:
                    rows = list(getattr(scheme, related_name).values(*fields))
                    for row in rows:
                        for k in fields:
                            if k not in GC_STRING_ROW_FIELDS:
                                row[k] = str(row[k])
                    data[related_name] = rows
                return JsonResponse(data)

            cached = self.get_cached_data(GROWING_CHARGE_LIST_CACHE_KEY)
            if cached:
                return JsonResponse(cached, safe=False)

            schemes = []
            for s in GrowingChargeScheme.objects.select_related("branch"):
                schemes.append({
                    "id": s.id,
                    "scheme_code": s.scheme_code,
                    "from_date": s.from_date.strftime("%d.%m.%Y") if s.from_date else "",
                    "to_date": s.to_date.strftime("%d.%m.%Y") if s.to_date else "",
                    "branch_name": s.branch.branch_name if s.branch else "-All-",
                    "schema_name": s.schema_name,
                    "chick_cost": str(s.chick_cost),
                    "feed_cost": str(s.feed_cost),
                    "medicine_cost": str(s.medicine_cost),
                    "farmer_admin_cost": str(s.farmer_admin_cost),
                    "std_production_cost": str(s.std_production_cost),
                    "minimum_gc_cost": str(s.minimum_gc_cost),
                    "standard_fcr": str(s.standard_fcr),
                    "standard_mortality": str(s.standard_mortality),
                    "is_active": s.is_active,
                    "is_locked": s.is_locked,
                })
            self.set_cached_data(GROWING_CHARGE_LIST_CACHE_KEY, schemes)
            return JsonResponse(schemes, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    def _save_children(self, scheme, data):
        """Replace every nested-row set from the posted JSON arrays."""
        for related_name, model, fields in GC_CHILD_SPECS:
            rows = json.loads(data.get(related_name, "[]") or "[]")
            model.objects.filter(scheme=scheme).delete()
            for row in rows:
                kwargs = {}
                blank = True
                for f in fields:
                    val = row.get(f, "")
                    if f in GC_STRING_ROW_FIELDS:
                        kwargs[f] = val or ""
                        if val:
                            blank = False
                    else:
                        kwargs[f] = val if val not in ("", None) else 0
                        if val not in ("", None, "0", "0.00", 0):
                            blank = False
                if blank:
                    continue
                model.objects.create(scheme=scheme, **kwargs)

    def post(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            data = request.POST
            with transaction.atomic():
                scheme = GrowingChargeScheme.objects.get(id=id) if id else GrowingChargeScheme()
                if id and scheme.is_locked:
                    return JsonResponse({"error": "This scheme is locked."}, status=400)

                scheme.region = Region.objects.get(id=data["region_id"])
                branch_id = data.get("branch_id")
                scheme.branch = Branch.objects.get(id=branch_id) if branch_id else None

                for field in self.SCALAR_FIELDS:
                    if field in data:
                        value = data[field]
                        if value == "" and field in self.DECIMAL_FIELDS:
                            value = 0
                        setattr(scheme, field, value)

                scheme.full_clean(exclude=["scheme_code"])
                scheme.save()

                self._save_children(scheme, data)
                cache.delete(GROWING_CHARGE_LIST_CACHE_KEY)
            return JsonResponse(
                {"message": "Scheme updated" if id else "Scheme created", "id": scheme.id},
                status=200 if id else 201,
            )
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            scheme = GrowingChargeScheme.objects.get(id=id)
            if scheme.is_locked:
                return JsonResponse({"error": "This scheme is locked."}, status=400)
            with transaction.atomic():
                scheme.delete()
                cache.delete(GROWING_CHARGE_LIST_CACHE_KEY)
            return JsonResponse({"message": "Scheme deleted"})
        except Exception as e:
            return self.handle_exception(e)


@login_required
def growing_charge_duplicate(request, id):
    """Clone a scheme (header + all nested rows) into a new draft record."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        src = GrowingChargeScheme.objects.get(id=id)
    except GrowingChargeScheme.DoesNotExist:
        return JsonResponse({"error": "Scheme not found."}, status=404)
    with transaction.atomic():
        children = [(rn, m, list(getattr(src, rn).all())) for rn, m, _f in GC_CHILD_SPECS]
        clone = src
        clone.pk = None
        clone.scheme_code = ""
        clone.schema_name = f"{src.schema_name} (Copy)"
        clone.is_locked = False
        clone._state.adding = True
        clone.save()
        for related_name, model, rows in children:
            for row in rows:
                row.pk = None
                row.scheme = clone
                row._state.adding = True
                row.save()
        cache.delete(GROWING_CHARGE_LIST_CACHE_KEY)
    return JsonResponse({"message": "Scheme duplicated", "id": clone.id}, status=201)


@login_required
def toggle_growing_charge_active(request, id):
    """Toggle a scheme's active/inactive status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        scheme = GrowingChargeScheme.objects.get(id=id)
        if scheme.is_locked:
            return JsonResponse({"error": "This scheme is locked."}, status=400)
        scheme.is_active = not scheme.is_active
        scheme.save(update_fields=["is_active"])
        cache.delete(GROWING_CHARGE_LIST_CACHE_KEY)
        return JsonResponse({"message": "Scheme updated", "is_active": scheme.is_active})
    except GrowingChargeScheme.DoesNotExist:
        return JsonResponse({"error": "Scheme not found."}, status=404)


@login_required
def toggle_growing_charge_lock(request, id):
    """Toggle a scheme's locked status."""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        scheme = GrowingChargeScheme.objects.get(id=id)
        scheme.is_locked = not scheme.is_locked
        scheme.save(update_fields=["is_locked"])
        cache.delete(GROWING_CHARGE_LIST_CACHE_KEY)
        return JsonResponse({"message": "Scheme updated", "is_locked": scheme.is_locked})
    except GrowingChargeScheme.DoesNotExist:
        return JsonResponse({"error": "Scheme not found."}, status=404)


# ---------------------------------------------------------------------------
# Farmer Growing Charge Settlement / Batch Closing (Broiler > Growing Charges)
# ---------------------------------------------------------------------------
# The "Add Rearing Charges" transaction. Auto-loads a batch's computed figures
# from the same engine the Growing Charge Statement report uses
# (_build_batch_report), applies the scheme's incentive/deduction slabs to
# arrive at a Farmer Payable, and — on save — closes the batch. Every field is
# override-able, so the slab units/chain below are sensible defaults, not a
# hard contract (see the plan's Notes/risks).

FEED_KG_PER_BAG = Decimal("50")  # feed items are 50kg/bag across this system

# Manual-entry fields (not slab-derived): default 0 on autofill, user fills.
GC_SETTLEMENT_MANUAL_FIELDS = [
    "other_incentives", "ifft_charges", "farmer_sales_deduction", "feed_transfer_charges",
    "vaccinator_charges", "transportation_charges", "other_deductions", "equipment_charges",
    "advance_deductions",
]
# Every persisted numeric field the POST accepts (the client sends final values;
# the server recomputes the authoritative running totals from them on save).
GC_SETTLEMENT_INPUT_FIELDS = GC_SETTLEMENT_MANUAL_FIELDS + [
    "standard_growing_charges", "actual_growing_charges", "sales_incentives",
    "mortality_incentives", "fcr_incentives", "summer_incentives",
    "birds_shortage_rate", "birds_shortage_amount", "fcr_deduction", "mortality_deduction",
]


def _num(x):
    """Coerce a batch_costing value (Decimal, int, or the 'No Data' string used
    when no scheme matches) to a Decimal — 'No Data'/blank become 0."""
    if x is None or x == "No Data" or x == "":
        return Decimal("0")
    return x if isinstance(x, Decimal) else Decimal(str(x))


def _slab_match(rows, value, lo_attr, hi_attr, val_attr):
    """First row whose [lo, hi] band contains ``value`` → its ``val_attr`` (0 if none)."""
    v = _num(value)
    for r in rows:
        if _num(getattr(r, lo_attr)) <= v <= _num(getattr(r, hi_attr)):
            return _num(getattr(r, val_attr))
    return Decimal("0")


def _actual_gc_rate(scheme, std_cost, actual_cost, base_rate):
    """Actual Growing Charge rate (Rs./kg): the standard GC rate adjusted by the
    gap between the batch's actual production cost/kg and the scheme's standard
    production cost, spread PROGRESSIVELY across the per-rupee Production-Cost
    slab bands (tax-bracket style).

    actual > std -> decentive bands (ascending, keyed on to_production_cost):
        each ₹ segment of the excess above std, up to that band's ceiling, is
        multiplied by band.rate_pct/100 and SUBTRACTED.
    actual < std -> incentive bands (descending, keyed on from_production_cost):
        each ₹ segment of the shortfall below std is multiplied by rate_pct/100
        and ADDED.

    e.g. std=90, actual=91.89, both bands 50%: (1.00 + 0.89) tiered ->
        1.00*0.5 + 0.89*0.5 = 0.945 subtracted, so 7.50 -> 6.555.
    """
    if not scheme or std_cost <= 0:
        return base_rate
    adj = Decimal("0")
    if actual_cost > std_cost:
        decentives = sorted(scheme.production_cost_decentives.all(),
                            key=lambda r: _num(r.to_production_cost))
        # Beyond the highest defined decentive slab, the growing charge is
        # wiped entirely (production cost too high -> no GC).
        if decentives and actual_cost > _num(decentives[-1].to_production_cost):
            return Decimal("0")
        prev = std_cost
        for band in decentives:
            top = _num(band.to_production_cost)
            if top <= std_cost:
                continue
            seg = min(actual_cost, top) - prev
            if seg <= 0:
                break
            adj -= seg * _num(band.rate_pct) / Decimal("100")
            prev = top
            if prev >= actual_cost:
                break
    elif actual_cost < std_cost:
        # Incentive accrues only through the defined slabs; below the lowest
        # one the bonus is capped at the first slab (no further increase).
        prev = std_cost
        for band in sorted(scheme.production_cost_incentives.all(),
                           key=lambda r: _num(r.from_production_cost), reverse=True):
            bottom = _num(band.from_production_cost)
            if bottom >= std_cost:
                continue
            seg = prev - max(actual_cost, bottom)
            if seg <= 0:
                break
            adj += seg * _num(band.rate_pct) / Decimal("100")
            prev = bottom
            if prev <= actual_cost:
                break
    # GC rate never goes negative.
    return max(Decimal("0"), base_rate + adj)


def _sales_incentive_per_kg(scheme, avg_sale_rate):
    """Sales incentive per kg of sold weight. The slab's ``sales_incentive`` is
    a rate PER RUPEE of sale rate above the band floor, accrued progressively
    across bands (like the growing-charge tiers): e.g. band 105-140 @ 0.10 and
    a ₹115 sale rate -> (115-105) x 0.10 = ₹1.00/kg. Below the lowest band's
    floor there is no incentive; above the highest band's ceiling it caps."""
    if not scheme:
        return Decimal("0")
    v = _num(avg_sale_rate)
    per_kg = Decimal("0")
    for band in sorted(scheme.sales_incentives.all(), key=lambda r: _num(r.sale_rate_from)):
        lo, hi, rate = _num(band.sale_rate_from), _num(band.sale_rate_to), _num(band.sales_incentive)
        if v <= lo:
            break
        per_kg += (min(v, hi) - lo) * rate
        if v <= hi:
            break
    return per_kg


def _shortage_rate(scheme, bc):
    """Per-bird shortage recovery rate per the scheme's shortage_basis."""
    if not scheme:
        return Decimal("0")
    basis = scheme.shortage_basis
    prod = _num(bc.get("production_cost_per_kg"))
    std_prod = _num(bc.get("std_prod_per_kg"))
    avg_rate = _num(bc.get("avg_sale_rate"))
    B = GrowingChargeScheme.ShortageBasis
    if basis == B.STD_PRODUCTION_COST:
        return std_prod
    if basis == B.PRODUCTION_COST:
        return prod
    if basis == B.AVG_SALE_RATE:
        return avg_rate
    if basis == B.MAX_SALE_RATE:
        return avg_rate  # no per-sale max tracked; avg is the best available proxy
    return max(std_prod, prod, avg_rate)  # WHICH_IS_HIGHER


def _gc_settlement_autofill(batch, scheme):
    """All settlement field defaults for a batch, keyed by the model's field
    names. Read-only figures come from _build_batch_report; incentive/deduction
    defaults from the scheme's slab tables. Returns Decimals."""
    report = _build_batch_report(batch, fetch_type="farmer", scheme_override=scheme)
    bc = report["batch_costing"]
    q2 = Decimal("0.01")

    # Final liquidation = the LAST bird-sale date; GC closing defaults to the
    # day after it (editable in the form).
    sale_dates = [r["date"] for r in report.get("bird_sales", []) if r.get("date")]
    last_sale_date = max(sale_dates) if sale_dates else bc.get("sale_start_date")
    gc_date_default = (last_sale_date + timedelta(days=1)) if last_sale_date else None

    sold_weight = _num(bc.get("sold_weight"))
    sold_birds = _num(bc.get("sold_birds"))
    placed = _num(bc.get("chicks_placed"))
    live_birds = placed - _num(bc.get("mortality")) - _num(bc.get("culls"))

    # ---- slab-derived incentives (treated as per-kg of sold live weight,
    #      except summer which is per-bird on its `incentive_on` basis) ----
    if scheme:
        sales_rate = _sales_incentive_per_kg(scheme, bc.get("avg_sale_rate"))
        mort_rate = _slab_match(scheme.mortality_incentives.all(), bc.get("total_mort_pct"),
                                "from_mortality_pct", "to_mortality_pct", "incentive_value")
        fcr_rate = _slab_match(scheme.fcr_incentives.all(), bc.get("cfcr"),
                               "cfcr_limit", "cfcr_limit", "incentive_value") \
            if scheme.fcr_incentives.exists() else Decimal("0")
        # FCR incentive slab is a limit, not a band: reward when CFCR <= limit.
        fcr_rate = Decimal("0")
        for r in scheme.fcr_incentives.all().order_by("cfcr_limit"):
            if _num(bc.get("cfcr")) <= _num(r.cfcr_limit):
                fcr_rate = _num(r.incentive_value)
                break

        summer_amount = Decimal("0")
        for r in scheme.summer_incentives.all():
            if _num(r.from_production_cost) <= _num(bc.get("production_cost_per_kg")) <= _num(r.to_production_cost):
                basis_birds = {"sold_birds": sold_birds, "placed_birds": placed,
                               "live_birds": live_birds}.get(r.incentive_on, sold_birds)
                summer_amount = _num(r.incentive_rate) * basis_birds
                break

        sales_incentives = (sales_rate * sold_weight)
        mortality_incentives = (mort_rate * sold_weight)
        fcr_incentives = (fcr_rate * sold_weight)
        summer_incentives = summer_amount

        # ---- slab-derived deductions ----
        fcr_recovery_rate = Decimal("0")
        for r in scheme.fcr_recoveries.all().order_by("cfcr_limit"):
            if _num(bc.get("cfcr")) >= _num(r.cfcr_limit):
                fcr_recovery_rate = _num(r.recovery_rate)
        fcr_deduction = fcr_recovery_rate * sold_weight
        mort_dec_rate = _slab_match(scheme.mortality_decentives.all(), bc.get("total_mort_pct"),
                                    "from_mortality_pct", "to_mortality_pct", "decentive_value")
        mortality_deduction = mort_dec_rate * sold_weight

        # ---- Standard vs Actual GC ----
        # Standard GC rate (Rs./kg) comes from the master; the Actual GC rate
        # adjusts it by how far the batch's actual production cost/kg is from the
        # scheme's standard production cost, distributed PROGRESSIVELY across the
        # per-rupee Production-Cost slab bands (like tax brackets):
        #   actual > std  -> walk the decentive bands upward, subtract each
        #                    segment x (band rate/100) from the GC rate
        #   actual < std  -> walk the incentive bands downward, add each segment
        # Amounts are rate x sold weight.
        std_gc_rate = _num(bc.get("base_gc_rate"))               # scheme.standard_gc_cost
        std_cost = _num(bc.get("std_prod_per_kg"))               # scheme.std_production_cost
        actual_cost = _num(bc.get("production_cost_per_kg"))
        actual_gc_rate = _actual_gc_rate(scheme, std_cost, actual_cost, std_gc_rate)
        # Standard + Incentive/Decentive = Actual (the net the farmer is paid).
        incdec_rate = actual_gc_rate - std_gc_rate               # +ve incentive, -ve decentive
        standard_gc = std_gc_rate * sold_weight
        gc_incdec = incdec_rate * sold_weight
        actual_gc = actual_gc_rate * sold_weight
    else:
        sales_incentives = mortality_incentives = fcr_incentives = summer_incentives = Decimal("0")
        fcr_deduction = mortality_deduction = standard_gc = actual_gc = Decimal("0")
        std_gc_rate = incdec_rate = actual_gc_rate = gc_incdec = Decimal("0")

    shortage_birds = _num(bc.get("shortage_birds"))
    shortage_rate = _shortage_rate(scheme, bc)
    shortage_amount = shortage_rate * shortage_birds

    data = {
        "placement_date": bc.get("placement_date"),
        "liquidation_date": last_sale_date,
        "gc_date_default": gc_date_default,
        # Bird details
        "placed_birds": placed, "mortality": _num(bc.get("mortality")),
        "sold_birds": sold_birds, "sold_weight": sold_weight,
        "excess": _num(bc.get("excess_birds")), "shortage": shortage_birds,
        "sale_amount": _num(bc.get("sold_amount")), "sale_rate": _num(bc.get("avg_sale_rate")),
        "age": _num(bc.get("mean_age")),
        # Performance
        "first_week_mortality_pct": _num(bc.get("first_week_mort_pct")),
        "days30_mortality_pct": _num(bc.get("upto_30_mort_pct")),
        "after30_mortality_pct": _num(bc.get("after_30_mort_pct")),
        "total_mortality_pct": _num(bc.get("total_mort_pct")),
        "fcr": _num(bc.get("fcr")), "cfcr": _num(bc.get("cfcr")),
        "avg_weight": _num(bc.get("avg_body_weight")), "mean_age": _num(bc.get("mean_age")),
        "day_gain": _num(bc.get("day_gain")), "eef": _num(bc.get("eef")),
        "grade": bc.get("grade") if bc.get("grade") not in (None, "No Data") else "",
        # Feed / medicine
        "feed_in": _num(bc.get("feed_sent")), "feed_consumption": _num(bc.get("feed_consumed")),
        "feed_out": _num(bc.get("feed_return")),
        "feed_balance": _num(bc.get("feed_sent")) - _num(bc.get("feed_consumed")) - _num(bc.get("feed_return")),
        "med_transfer_in": _num(bc.get("med_sent")), "med_consumption": _num(bc.get("med_consumed")),
        "med_transfer_out": _num(bc.get("med_return")),
        "med_closing": _num(bc.get("med_sent")) - _num(bc.get("med_consumed")) - _num(bc.get("med_return")),
        # Costing (per-unit = per kg of sold live weight)
        "chick_cost": _num(bc.get("chick_cost")), "chick_cost_per_unit": _div(_num(bc.get("chick_cost")), sold_weight),
        "feed_cost": _num(bc.get("feed_cost")), "feed_cost_per_unit": _div(_num(bc.get("feed_cost")), sold_weight),
        "admin_cost": _num(bc.get("admin_cost")), "admin_cost_per_unit": _div(_num(bc.get("admin_cost")), sold_weight),
        "medicine_cost": _num(bc.get("med_cost")), "medicine_cost_per_unit": _div(_num(bc.get("med_cost")), sold_weight),
        "total_cost": _num(bc.get("total_production_cost")),
        "total_cost_per_unit": _num(bc.get("production_cost_per_kg")),
        # Production Cost is quoted per kg of live weight (Rs./kg) — standard
        # comes from the scheme's per-kg rate, actual is this batch's own
        # per-kg cost (== Total Cost's per-unit).
        "standard_production_cost": _num(bc.get("std_prod_per_kg")),
        "actual_production_cost": _num(bc.get("production_cost_per_kg")),
        # Rearing charges (incentives) — Standard + Incentive/Decentive = Actual
        "standard_growing_charges": standard_gc, "gc_incentive_decentive": gc_incdec,
        "actual_growing_charges": actual_gc,
        # per-kg rates for the "Rs." column of the three GC rows
        "std_gc_rate": std_gc_rate, "gc_incdec_rate": incdec_rate, "actual_gc_rate": actual_gc_rate,
        "sales_incentives": sales_incentives, "mortality_incentives": mortality_incentives,
        "fcr_incentives": fcr_incentives, "summer_incentives": summer_incentives,
        # Deductions (slab-derived)
        "birds_shortage_rate": shortage_rate, "birds_shortage_amount": shortage_amount,
        "fcr_deduction": fcr_deduction, "mortality_deduction": mortality_deduction,
    }
    for f in GC_SETTLEMENT_MANUAL_FIELDS:
        data[f] = Decimal("0")

    # Farmer sales deduction: birds this farmer bought from the batch
    # (sale_type='farmer') that are still UNPAID — i.e. the batch's farmer
    # bird-sale amount capped by the farmer's overall unpaid balance
    # (total farmer sales minus receipts; receipts aren't batch-scoped).
    # Auto-picked here; still editable on the form.
    from django.db.models import Sum
    farmer = batch.broiler_farm.farmer
    Z = Decimal("0")

    def _sum(qs):
        return qs.aggregate(t=Sum("amount"))["t"] or Z

    batch_farmer_sales = _sum(BirdSale.objects.filter(batch=batch, sale_type="farmer", farmer=farmer))
    total_farmer_sales = _sum(BirdSale.objects.filter(sale_type="farmer", farmer=farmer))
    total_farmer_receipts = _sum(BirdSaleReceipt.objects.filter(sale_type="farmer", farmer=farmer))
    unpaid = total_farmer_sales - total_farmer_receipts
    data["farmer_sales_deduction"] = max(Z, min(batch_farmer_sales, unpaid))

    # Running totals (recomputed identically on save from whatever values stand).
    data.update(_gc_settlement_totals(data, sold_birds, sold_weight))
    return {k: (v.quantize(q2) if isinstance(v, Decimal) else v) for k, v in data.items()}


def _gc_settlement_totals(d, sold_birds, sold_weight):
    """The dependent running totals of the settlement, from the current field
    values — the single source of truth used by both autofill and save."""
    g = lambda k: _num(d.get(k))
    total_incentives = (g("sales_incentives") + g("mortality_incentives") + g("fcr_incentives")
                        + g("summer_incentives") + g("other_incentives") + g("ifft_charges"))
    total_deduction = g("birds_shortage_amount") + g("fcr_deduction") + g("mortality_deduction")
    amount_payable = g("actual_growing_charges") + total_incentives - total_deduction
    total_amount_payable = (amount_payable - g("farmer_sales_deduction") - g("feed_transfer_charges")
                            - g("vaccinator_charges") - g("other_deductions") + g("transportation_charges"))
    tds = (total_amount_payable * Decimal("0.01"))
    farmer_payable = total_amount_payable - tds - g("equipment_charges") - g("advance_deductions")
    return {
        "total_incentives": total_incentives,
        "gc_paid_per_kg": _div(g("actual_growing_charges"), _num(sold_weight)),
        "total_deduction": total_deduction,
        "amount_payable": amount_payable,
        "total_amount_payable": total_amount_payable,
        "tds": tds,
        "farmer_payable": farmer_payable,
        "per_bird_cost": _div(farmer_payable, _num(sold_birds)),
    }


@method_decorator(login_required, name="dispatch")
class GCSettlementTemplateView(View):
    """Renders the Farmer GC Settlement / batch-closing form page."""

    def get(self, request):
        return render(request, "gc_settlement_form.html", {
            "farms": farms_for(request.user, BroilerFarm.objects.select_related("branch", "supervisor").order_by("farm_name")),
            "today": timezone.localdate().isoformat(),
        })


@login_required
def gc_settlement_batches(request):
    """Open (not-yet-closed) batches for a farm, for the Batch dropdown."""
    farm_id = (request.GET.get("farm") or "").strip()
    if not farm_id:
        return JsonResponse([], safe=False)
    batches = (BroilerBatch.objects.filter(broiler_farm_id=farm_id, is_closed=False)
               .order_by("-start_date", "-id"))
    return JsonResponse([{"id": b.id, "batch_name": b.batch_name} for b in batches], safe=False)


@login_required
def gc_settlement_schemes(request):
    """Schemes whose date range covers the batch's placement date (region
    matched), plus the auto-matched one — for the Scheme Name dropdown."""
    batch_id = (request.GET.get("batch") or "").strip()
    batch = BroilerBatch.objects.select_related("broiler_farm__branch").filter(id=batch_id).first()
    if not batch:
        return JsonResponse({"schemes": [], "selected": None}, safe=False)
    placement = _placement_date(batch)
    region_id = batch.broiler_farm.branch.region_id
    qs = GrowingChargeScheme.objects.filter(region_id=region_id, is_active=True)
    if placement:
        qs = qs.filter(from_date__lte=placement, to_date__gte=placement)
    matched = _match_growing_charge_scheme(batch, placement)
    return JsonResponse({
        "schemes": [{"id": s.id, "name": f"{s.scheme_code} - {s.schema_name}"}
                    for s in qs.order_by("schema_name")],
        "selected": matched.id if matched else None,
    })


@login_required
def gc_settlement_autofill_api(request):
    """All auto-computed settlement figures for a batch + scheme, as JSON."""
    batch_id = (request.GET.get("batch") or "").strip()
    scheme_id = (request.GET.get("scheme") or "").strip()
    batch = (BroilerBatch.objects
             .select_related("broiler_farm__branch", "broiler_farm__supervisor")
             .filter(id=batch_id).first())
    if not batch:
        return JsonResponse({"error": "Batch not found"}, status=404)
    if batch.is_closed:
        return JsonResponse({"error": "This batch is already closed/settled."}, status=400)
    scheme = GrowingChargeScheme.objects.filter(id=scheme_id).first() if scheme_id.isdigit() else \
        _match_growing_charge_scheme(batch, _placement_date(batch))

    farm = batch.broiler_farm
    data = _gc_settlement_autofill(batch, scheme)
    out = {}
    for k, v in data.items():
        if hasattr(v, "isoformat"):
            out[k] = v.isoformat()
        elif isinstance(v, Decimal):
            out[k] = str(v)
        else:
            out[k] = v
    out.update({
        "branch": farm.branch.branch_name if farm.branch_id else "",
        "line": farm.line or "",
        "supervisor": farm.supervisor.name if farm.supervisor_id else "",
        "batch_name": batch.batch_name,
        "scheme_id": scheme.id if scheme else None,
        "scheme_name": (f"{scheme.scheme_code} - {scheme.schema_name}") if scheme else "",
        "kg_per_bag": str(FEED_KG_PER_BAG),
    })
    return JsonResponse(out)


@method_decorator(login_required, name="dispatch")
class GCSettlementAPI(View):
    """List / retrieve / create / delete Farmer GC settlements."""

    def get(self, request, id=None):
        if id:
            s = get_object_or_404(GrowingChargeSettlement.objects.select_related(
                "batch__broiler_farm", "farm", "scheme"), id=id)
            return JsonResponse(_gc_settlement_detail(s))
        rows = (GrowingChargeSettlement.objects
                .select_related("batch__broiler_farm", "farm", "scheme").order_by("-gc_date", "-id"))
        from_date = (request.GET.get("from_date") or "").strip()
        to_date = (request.GET.get("to_date") or "").strip()
        if from_date:
            rows = rows.filter(gc_date__gte=from_date)
        if to_date:
            rows = rows.filter(gc_date__lte=to_date)
        def fmt(d):
            return d.strftime("%d.%m.%Y") if d else ""
        return JsonResponse([{
            "id": s.id,
            "closed_date": fmt(timezone.localtime(s.created_at).date() if s.created_at else None),
            "settlement_code": s.settlement_code,
            "gc_date": fmt(s.gc_date),
            "farm_name": s.farm.farm_name, "batch_name": s.batch.batch_name,
            "start_date": fmt(s.placement_date or s.batch.start_date),
            "liquidation_date": fmt(s.liquidation_date),
            "gc_amount": str(s.farmer_payable),
            "grade": s.grade or "-",
        } for s in rows], safe=False)

    @transaction.atomic
    def post(self, request):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        batch = (BroilerBatch.objects.select_related("broiler_farm")
                 .filter(id=data.get("batch")).first())
        if not batch:
            return JsonResponse({"error": "Select a Batch."}, status=400)
        if batch.is_closed or GrowingChargeSettlement.objects.filter(batch=batch).exists():
            return JsonResponse({"error": "This batch is already settled/closed."}, status=400)

        scheme = GrowingChargeScheme.objects.filter(id=data.get("scheme")).first()
        gc_date = timezone.datetime.fromisoformat(data["gc_date"]).date() if data.get("gc_date") \
            else timezone.localdate()

        # Start from a fresh autofill (authoritative read-only figures), then
        # overlay the user's editable inputs, then recompute the running totals.
        fields = _gc_settlement_autofill(batch, scheme)
        for f in GC_SETTLEMENT_INPUT_FIELDS:
            if f in data and data[f] not in (None, ""):
                fields[f] = Decimal(str(data[f]))
        sold_birds = fields.get("sold_birds") or Decimal("0")
        sold_weight = fields.get("sold_weight") or Decimal("0")
        fields.update(_gc_settlement_totals(fields, sold_birds, sold_weight))

        settlement = GrowingChargeSettlement(
            batch=batch, farm=batch.broiler_farm, scheme=scheme,
            gc_date=gc_date, remarks=data.get("remarks") or "", created_by=request.user,
        )
        model_field_names = {f.name for f in GrowingChargeSettlement._meta.get_fields()}
        for k, v in fields.items():
            if k in model_field_names:
                setattr(settlement, k, v)
        settlement.save()

        # Close the batch.
        batch.is_closed = True
        batch.closed_on = gc_date
        if not batch.end_date:
            batch.end_date = gc_date
        batch.save(update_fields=["is_closed", "closed_on", "end_date"])
        # The Batch tab caches this list; its Status/active-guard depend on the
        # closed flag we just changed, so drop the stale snapshot.
        cache.delete("broiler_batch_list")

        return JsonResponse({"message": "Settlement saved and batch closed",
                             "id": settlement.id, "code": settlement.settlement_code}, status=201)

    @transaction.atomic
    def put(self, request, id):
        s = get_object_or_404(GrowingChargeSettlement, id=id)
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        # Editing only touches the manual/override inputs + gc_date/remarks; the
        # auto-computed read-only figures stay as stored. Totals recompute from
        # the current stored values overlaid with the edited inputs.
        fields = {f.name: getattr(s, f.name) for f in GrowingChargeSettlement._meta.fields}
        for f in GC_SETTLEMENT_INPUT_FIELDS:
            if f in data and data[f] not in (None, ""):
                fields[f] = Decimal(str(data[f]))
        totals = _gc_settlement_totals(fields, s.sold_birds, s.sold_weight)
        for k, v in {**{f: fields[f] for f in GC_SETTLEMENT_INPUT_FIELDS}, **totals}.items():
            setattr(s, k, v)
        if data.get("gc_date"):
            s.gc_date = timezone.datetime.fromisoformat(data["gc_date"]).date()
        s.remarks = data.get("remarks", s.remarks) or ""
        s.save()
        return JsonResponse({"message": "Settlement updated", "id": s.id, "code": s.settlement_code})

    @transaction.atomic
    def delete(self, request, id):
        s = get_object_or_404(GrowingChargeSettlement.objects.select_related("batch"), id=id)
        batch = s.batch
        # Reopen the batch — clear the end/closed marks the settlement set so
        # the flock counts as live again.
        s.delete()
        batch.is_closed = False
        batch.closed_on = None
        batch.end_date = None
        batch.save(update_fields=["is_closed", "closed_on", "end_date"])
        # Reopening flips the batch back to active — invalidate the Batch tab's
        # cached list so it no longer shows the flock as Closed.
        cache.delete("broiler_batch_list")
        return JsonResponse({"message": "Settlement deleted and batch reopened"})


def _gc_settlement_detail(s):
    """Full field dict of a saved settlement (all numeric fields as strings)."""
    farm = s.farm
    out = {"id": s.id, "settlement_code": s.settlement_code,
           "farm_id": s.farm_id, "batch_id": s.batch_id, "scheme_id": s.scheme_id,
           "farm_name": farm.farm_name, "batch_name": s.batch.batch_name,
           "scheme_name": (f"{s.scheme.scheme_code} - {s.scheme.schema_name}") if s.scheme_id else "",
           "branch": farm.branch.branch_name if farm.branch_id else "",
           "line": farm.line or "", "supervisor": farm.supervisor.name if farm.supervisor_id else "",
           "gc_date": s.gc_date.isoformat() if s.gc_date else "",
           "placement_date": s.placement_date.isoformat() if s.placement_date else "",
           "liquidation_date": s.liquidation_date.isoformat() if s.liquidation_date else "",
           "grade": s.grade, "remarks": s.remarks}
    for f in GrowingChargeSettlement._meta.get_fields():
        val = getattr(s, f.name, None)
        if isinstance(val, Decimal):
            out[f.name] = str(val)
        elif isinstance(val, int) and not isinstance(val, bool):
            out[f.name] = val
    return out


@login_required
def gc_settlement_print(request, id):
    """Farmer-facing printable Growing Charges / Batch Closing report for a
    saved settlement (the "Shalimar" field set). Shows performance + the
    farmer's growing-charge payment breakdown ONLY — no company revenue,
    profitability or margin (in contract growing the company owns/sells the
    birds; the farmer is paid growing charges)."""
    from account.models import CompanyProfile
    s = get_object_or_404(GrowingChargeSettlement.objects.select_related(
        "batch__broiler_farm__branch", "batch__broiler_farm__supervisor",
        "batch__broiler_farm__farmer", "scheme"), id=id)
    batch = s.batch
    farm = batch.broiler_farm

    sw = s.sold_weight or Decimal("0")
    sb = Decimal(str(s.sold_birds or 0))
    q3 = Decimal("0.001")

    def rate(amount, by):
        return (Decimal(str(amount)) / by).quantize(q3) if by else Decimal("0")

    # Derived per-kg / per-bird rearing-charge rates (not stored on the model).
    d = {
        "rc_per_kg": rate(s.actual_growing_charges, sw),           # Rearing Charges/Kg
        "std_rc_per_kg": rate(s.standard_growing_charges, sw),     # Std Rearing Charges/Kg
        "prod_cost_incentive_rate": rate(s.gc_incentive_decentive, sw),  # Prod Cost Incentives
        "rc_per_bird": rate(s.actual_growing_charges, sb),         # Rearing Charges/Bird
        "std_fcr": s.scheme.standard_fcr if s.scheme_id else None,
        # net earning (+) / deduction (-) for the FCR & mortality lines
        "fcr_net": s.fcr_incentives - s.fcr_deduction,
        "mortality_net": s.mortality_incentives - s.mortality_deduction,
    }
    return render(request, "gc_settlement_print.html", {
        "s": s, "batch": batch, "farm": farm, "farmer": farm.farmer,
        "supervisor": farm.supervisor.name if farm.supervisor_id else "", "branch": farm.branch, "scheme": s.scheme,
        "company": CompanyProfile.get_solo(), "d": d,
    })


# ---------------------------------------------------------------------------
# Feed Phase Master (Broiler > Master) — a feeding program header with feed
# phase line items (Pre-Starter, Starter, Grower, Finisher…).
# ---------------------------------------------------------------------------
def _feed_phase_line_to_dict(line):
    return {
        "id": line.id, "seq_no": line.seq_no,
        "from_age": line.from_age, "to_age": line.to_age,
        "category": line.category_id,
        "category_name": line.category.name if line.category_id else "",
        "feed_item": line.feed_item_id,
        "feed_phase": line.feed_item.description if line.feed_item_id else "",
        "phase_code": line.phase_code,
        "max_feed_qty": str(line.max_feed_qty), "priority": line.priority,
        "status": line.status,
    }


def _feed_phase_master_to_dict(m, with_lines=False):
    data = {
        "id": m.id, "program": m.program,
        "bird_category": m.bird_category_id,
        "bird_category_name": m.bird_category.name if m.bird_category_id else "",
        "breed": m.breed_id, "breed_name": m.breed.description if m.breed_id else "",
        "effective_from": m.effective_from.isoformat() if m.effective_from else "",
        "effective_to": m.effective_to.isoformat() if m.effective_to else "",
        "status": m.status, "description": m.description,
        "line_count": m.lines.count(),
        "phases": [l.feed_item.description for l in m.lines.all() if l.feed_item_id],
    }
    if with_lines:
        data["lines"] = [_feed_phase_line_to_dict(l) for l in m.lines.all()]
    return data


def _validate_feed_phase(data):
    """Return an error message if the feed phase master is invalid, else None:
    effective dates ordered, at least one item line, each From<=To, at most one
    open-ended ('& above') phase and it must be last, and no overlapping ages."""
    ef, et = data.get("effective_from"), data.get("effective_to")
    if ef and et and str(ef) > str(et):
        return "Effective From must be on or before Effective To."
    lines = [r for r in (data.get("lines") or []) if r.get("feed_item")]
    if not lines:
        return "Add at least one feed phase with an item selected."
    ranges, open_count = [], 0
    for r in lines:
        try:
            fa = int(r.get("from_age") or 0)
        except (TypeError, ValueError):
            fa = 0
        to_raw = r.get("to_age")
        ta = None if to_raw in (None, "", 0, "0") else int(to_raw)
        if ta is not None and ta < fa:
            return f"A phase starting at day {fa} has To Age ({ta}) before From Age ({fa})."
        if ta is None:
            open_count += 1
        ranges.append((fa, ta))
    if open_count > 1:
        return "Only one phase can be open-ended ('& above')."
    ranges.sort(key=lambda x: x[0])
    for i, (fa, ta) in enumerate(ranges):
        if ta is None and i != len(ranges) - 1:
            return "The open-ended ('& above') phase must be the last one."
        if i > 0:
            prev_ta = ranges[i - 1][1]
            if prev_ta is None or fa <= prev_ta:
                return f"Age ranges overlap around day {fa}. Each phase must start after the previous one ends."
    return None


def _apply_feed_phase_master(instance, data):
    instance.program = (data.get("program") or "").strip()
    instance.bird_category_id = data.get("bird_category") or None
    instance.breed_id = data.get("breed") or None
    instance.effective_from = data.get("effective_from") or None
    instance.effective_to = data.get("effective_to") or None
    instance.status = data.get("status") or "active"
    instance.description = data.get("description") or ""


def _save_feed_phase_lines(master, rows):
    master.lines.all().delete()
    for i, r in enumerate(rows, start=1):
        if not r.get("feed_item"):
            continue
        FeedPhaseLine.objects.create(
            master=master,
            seq_no=r.get("seq_no") or i,
            from_age=r.get("from_age") or 0,
            to_age=r.get("to_age") if r.get("to_age") not in (None, "", 0, "0") else None,
            category_id=r.get("category") or None,
            feed_item_id=r.get("feed_item"),
            phase_code=(r.get("phase_code") or "").strip(),
            max_feed_qty=Decimal(str(r.get("max_feed_qty") or 0)),
            priority=r.get("priority") or i,
            status=r.get("status") or "active",
        )


@method_decorator(login_required, name="dispatch")
class FeedPhaseMasterListTemplateView(View):
    def get(self, request):
        return render(request, "feed_phase_master_list.html")


@method_decorator(login_required, name="dispatch")
class FeedPhaseMasterFormTemplateView(View):
    def get(self, request, id=None):
        from inventory.models import Item, ItemCategory
        instance = FeedPhaseMaster.objects.filter(id=id).first() if id else None
        return render(request, "feed_phase_master_form.html", {
            "instance": instance,
            "lines_json": json.dumps([_feed_phase_line_to_dict(l)
                                      for l in instance.lines.select_related("feed_item", "category").all()]) if instance else "[]",
            "breeds": Breed.objects.filter(is_active=True).select_related("bird_category").order_by("description"),
            "bird_categories": BirdCategory.objects.filter(is_active=True).order_by("sort_order", "name"),
            "categories": ItemCategory.objects.order_by("name"),
            # every item with its category, so the form can filter items by the chosen category
            "items_json": json.dumps(list(Item.objects.order_by("description")
                                          .values("id", "description", "category_id", "item_code"))),
            "programs": list(FeedPhaseMaster.objects.order_by("program")
                             .values_list("program", flat=True).distinct()),
            "today": timezone.localdate().isoformat(),
        })


@method_decorator(login_required, name="dispatch")
class FeedPhaseMasterAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                m = FeedPhaseMaster.objects.prefetch_related("lines").select_related("breed").get(id=id)
                return JsonResponse(_feed_phase_master_to_dict(m, with_lines=True))
            masters = (FeedPhaseMaster.objects.select_related("breed")
                       .prefetch_related("lines__feed_item").order_by("-created_at"))
            return JsonResponse([_feed_phase_master_to_dict(m) for m in masters], safe=False)
        except FeedPhaseMaster.DoesNotExist:
            raise Http404("Feed Phase Master not found")
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body or "{}")
            if not (data.get("program") or "").strip():
                return JsonResponse({"error": "Program is required"}, status=400)
            err = _validate_feed_phase(data)
            if err:
                return JsonResponse({"error": err}, status=400)
            with transaction.atomic():
                instance = FeedPhaseMaster()
                _apply_feed_phase_master(instance, data)
                instance.save()
                _save_feed_phase_lines(instance, data.get("lines") or [])
            return JsonResponse({"message": "Feed Phase Master created", "id": instance.id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            instance = FeedPhaseMaster.objects.get(id=id)
            data = json.loads(request.body or "{}")
            err = _validate_feed_phase(data)
            if err:
                return JsonResponse({"error": err}, status=400)
            with transaction.atomic():
                _apply_feed_phase_master(instance, data)
                instance.save()
                _save_feed_phase_lines(instance, data.get("lines") or [])
            return JsonResponse({"message": "Feed Phase Master updated"})
        except FeedPhaseMaster.DoesNotExist:
            raise Http404("Feed Phase Master not found")
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            FeedPhaseMaster.objects.get(id=id).delete()
            return JsonResponse({"message": "Feed Phase Master deleted"})
        except FeedPhaseMaster.DoesNotExist:
            raise Http404("Feed Phase Master not found")
        except Exception as e:
            return self.handle_exception(e)


@method_decorator(login_required, name="dispatch")
class FeedPhaseMasterDuplicateAPI(BaseAPIView):
    """Clone a Feed Phase Master and all its lines into a new '(Copy)' record."""
    def post(self, request, id: int) -> JsonResponse:
        try:
            src = FeedPhaseMaster.objects.prefetch_related("lines").get(id=id)
            with transaction.atomic():
                dup = FeedPhaseMaster.objects.create(
                    program=f"{src.program} (Copy)", bird_category_id=src.bird_category_id,
                    breed_id=src.breed_id, effective_from=src.effective_from,
                    effective_to=src.effective_to, status=src.status,
                    description=src.description,
                )
                for l in src.lines.all():
                    FeedPhaseLine.objects.create(
                        master=dup, seq_no=l.seq_no, from_age=l.from_age, to_age=l.to_age,
                        category_id=l.category_id, feed_item_id=l.feed_item_id,
                        phase_code=l.phase_code, max_feed_qty=l.max_feed_qty,
                        priority=l.priority, status=l.status,
                    )
            return JsonResponse({"message": "Feed Phase Master duplicated", "id": dup.id}, status=201)
        except FeedPhaseMaster.DoesNotExist:
            raise Http404("Feed Phase Master not found")
        except Exception as e:
            return self.handle_exception(e)


# ---------------------------------------------------------------------------
# Bird Category master (Broiler > Growing Charges) — simple lookup.
# ---------------------------------------------------------------------------
@method_decorator(login_required, name="dispatch")
class BirdCategoryTemplateView(View):
    def get(self, request):
        return render(request, "bird_category.html")


@method_decorator(login_required, name="dispatch")
class BirdCategoryAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                c = BirdCategory.objects.get(id=id)
                return JsonResponse({"id": c.id, "name": c.name, "is_active": c.is_active})
            cats = BirdCategory.objects.all()
            return JsonResponse([{"id": c.id, "name": c.name, "is_active": c.is_active} for c in cats], safe=False)
        except BirdCategory.DoesNotExist:
            return JsonResponse({"error": "Bird category not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request) -> JsonResponse:
        try:
            name = (request.POST.get("name") or "").strip()
            if not name:
                return JsonResponse({"error": "Name is required."}, status=400)
            with transaction.atomic():
                nxt = (BirdCategory.objects.order_by("-sort_order").values_list("sort_order", flat=True).first() or 0) + 1
                BirdCategory.objects.create(name=name, sort_order=nxt)
            return JsonResponse({"message": "Bird category created"}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    def put(self, request, id: int) -> JsonResponse:
        try:
            c = BirdCategory.objects.get(id=id)
            data = json.loads(request.body)
            name = (data.get("name") or "").strip()
            if not name:
                return JsonResponse({"error": "Name is required."}, status=400)
            with transaction.atomic():
                c.name = name
                c.save(update_fields=["name", "updated_at"])
            return JsonResponse({"message": "Bird category updated"})
        except BirdCategory.DoesNotExist:
            return JsonResponse({"error": "Bird category not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            with transaction.atomic():
                BirdCategory.objects.get(id=id).delete()
            return JsonResponse({"message": "Bird category deleted"})
        except BirdCategory.DoesNotExist:
            return JsonResponse({"error": "Bird category not found."}, status=404)
        except Exception as e:
            return self.handle_exception(e)


@login_required
def toggle_bird_category_active(request, id):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method."}, status=400)
    try:
        c = BirdCategory.objects.get(id=id)
        c.is_active = not c.is_active
        c.save(update_fields=["is_active"])
        return JsonResponse({"message": "Bird category updated", "is_active": c.is_active})
    except BirdCategory.DoesNotExist:
        return JsonResponse({"error": "Bird category not found."}, status=404)


# ------------------------------------------- farm location & photo captures

def _capture_row(c):
    farmer = c.farmer
    files = list(c.files.all())
    return {
        "id": c.id,
        "capture_no": c.capture_no,
        "date": c.date.isoformat() if c.date else "",
        "farm": c.farm.farm_name if c.farm_id else "",
        "farm_code": c.farm.farm_code if c.farm_id else "",
        "farmer": farmer.farmer_name if farmer else "",
        "branch": (c.farm.branch.branch_name
                   if c.farm_id and c.farm.branch_id else ""),
        "latitude": c.latitude, "longitude": c.longitude,
        "has_location": c.has_location,
        "state": c.state or "", "district": c.district or "", "area": c.area or "",
        "address": c.address or "",
        "photos": sum(1 for f in files if f.kind == FarmCaptureFile.KIND_PHOTO),
        "documents": sum(1 for f in files if f.kind != FarmCaptureFile.KIND_PHOTO),
        "remarks": c.remarks or "",
        "captured_by": ((c.captured_by.get_full_name() or c.captured_by.username)
                        if c.captured_by_id else ""),
        "files": [{
            "id": f.id, "kind": f.kind, "label": f.get_kind_display(),
            "url": f.file.url if f.file else "",
            "name": f.file.name.rsplit("/", 1)[-1] if f.file else "",
            "is_image": (f.file.name or "").lower().endswith(
                (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp")),
            "caption": f.caption or "",
        } for f in files],
    }


@login_required(login_url="login")
def farm_location_capture_list(request):
    """Broiler > Transactions > Farm Location & Photos — the register."""
    return render(request, "farm_location_capture_list.html", {
        "farms": farms_for(request.user, BroilerFarm.objects.select_related("farmer", "branch").order_by("farm_name")),
        "slots": [(k, dict(FarmCaptureFile.KIND_CHOICES)[k])
                  for k in FarmCaptureFile.SLOT_TARGETS
                  if k != FarmCaptureFile.KIND_DOCUMENT],
    })


@login_required
def farm_location_capture_api(request):
    """JSON rows for the register, filtered by date range / farm / farmer."""
    qs = (FarmLocationCapture.objects
          .select_related("farm", "farm__farmer", "farm__branch", "captured_by")
          .prefetch_related("files"))
    qs = scope_multi(request.user, qs,
                     farms="farm_id", branches="farm__branch_id")
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    farm_id = (request.GET.get("farm") or "").strip()
    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    if farm_id.isdigit():
        qs = qs.filter(farm_id=farm_id)
    return JsonResponse([_capture_row(c) for c in qs.order_by("-date", "-id")], safe=False)


def _capture_form_context(user, instance=None):
    grouped = _capture_files_by_slot(instance)
    return {
        "capture": instance,
        "next_no": FarmLocationCapture._next_no() if not instance else None,
        "farms": farms_for(user, BroilerFarm.objects.select_related("farmer", "branch").order_by("farm_name")),
        "branches": branches_for(user, Branch.objects.order_by("branch_name")),
        "today": timezone.localdate().isoformat(),
        # (code, label, files) per slot: carrying the files in the row lets the
        # template show each one under its own input without indexing a dict by
        # a loop variable, which Django templates cannot do.
        "slot_rows": [
            (k, dict(FarmCaptureFile.KIND_CHOICES)[k], grouped.get(k, []))
            for k in FarmCaptureFile.SLOT_TARGETS
            if k != FarmCaptureFile.KIND_DOCUMENT
        ],
        "farm_pictures": grouped.get(FarmCaptureFile.KIND_PHOTO, []),
        "other_documents": grouped.get(FarmCaptureFile.KIND_DOCUMENT, []),
    }


def _capture_files_by_slot(instance):
    grouped = {}
    if instance is None:
        return grouped
    for f in instance.files.all():
        grouped.setdefault(f.kind, []).append({
            "id": f.id, "kind": f.kind, "label": f.get_kind_display(),
            "name": f.file.name.rsplit("/", 1)[-1] if f.file else "",
            "url": f.file.url if f.file else "",
            "is_image": (f.file.name or "").lower().endswith(
                (".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp")),
        })
    return grouped


def _save_capture(request, instance):
    """Create or update a capture, attaching any newly uploaded files."""
    def blank_to_none(value):
        value = (value or "").strip()
        return value or None

    instance.date = request.POST.get("date") or timezone.localdate()
    instance.farm_id = request.POST.get("farm") or None
    instance.latitude = blank_to_none(request.POST.get("latitude"))
    instance.longitude = blank_to_none(request.POST.get("longitude"))
    instance.state = (request.POST.get("state") or "").strip()
    instance.district = (request.POST.get("district") or "").strip()
    instance.area = (request.POST.get("area") or "").strip()
    instance.address = (request.POST.get("address") or "").strip()
    instance.remarks = (request.POST.get("remarks") or "").strip()
    if instance.captured_by_id is None:
        instance.captured_by = request.user
    if not instance.farm_id:
        raise ValidationError("Select a farm.")
    instance.full_clean(exclude=["capture_no", "captured_by"])
    instance.save()

    # Farm pictures and other documents take any number of files; the master
    # fields behind the remaining slots hold one each, so those take one.
    for upload in request.FILES.getlist("photos"):
        FarmCaptureFile.objects.create(
            capture=instance, kind=FarmCaptureFile.KIND_PHOTO, file=upload)
    for upload in request.FILES.getlist("documents"):
        FarmCaptureFile.objects.create(
            capture=instance, kind=FarmCaptureFile.KIND_DOCUMENT, file=upload)
    for kind in FarmCaptureFile.SLOT_TARGETS:
        if kind == FarmCaptureFile.KIND_DOCUMENT:
            continue                      # handled above, many allowed
        upload = request.FILES.get("slot_%s" % kind)
        if upload:
            FarmCaptureFile.objects.create(capture=instance, kind=kind, file=upload)
    return instance


@login_required(login_url="login")
def farm_location_capture_add(request):
    if request.method == "POST":
        try:
            with transaction.atomic():
                _save_capture(request, FarmLocationCapture())
            messages.success(request, "Farm location capture saved successfully.")
            return redirect("farm_location_capture_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))
    return render(request, "farm_location_capture_form.html", _capture_form_context(request.user))


@login_required(login_url="login")
def farm_location_capture_edit(request, id):
    instance = get_object_or_404(FarmLocationCapture, id=id)
    if request.method == "POST":
        try:
            with transaction.atomic():
                _save_capture(request, instance)
            messages.success(request, "Farm location capture updated successfully.")
            return redirect("farm_location_capture_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))
    return render(request, "farm_location_capture_form.html", _capture_form_context(request.user, instance))


@login_required(login_url="login")
@require_POST
def farm_location_capture_clear(request, id):
    """Clear = drop the location only. Photos and documents are kept.

    Deliberately narrow: a wrong GPS reading is the thing that needs undoing,
    and the pictures taken on that visit are still good. Removing a file is a
    separate action on the file itself, and Delete removes the whole visit.
    The farm falls back to whatever earlier capture still holds a reading.
    """
    capture = get_object_or_404(FarmLocationCapture, id=id)
    with transaction.atomic():
        capture.latitude = None
        capture.longitude = None
        capture.address = ""
        capture.save(update_fields=["latitude", "longitude", "address", "updated_at"])
        FarmLocationCapture.sync_farm_from_latest(capture.farm)
    messages.success(request, "Location cleared; the attached files were kept.")
    return redirect("farm_location_capture_list")


@login_required(login_url="login")
@require_POST
def farm_location_capture_delete(request, id):
    capture = get_object_or_404(FarmLocationCapture, id=id)
    capture.delete()
    messages.success(request, "Farm location capture deleted successfully.")
    return redirect("farm_location_capture_list")


@login_required
@require_POST
def farm_capture_file_delete(request, id):
    """Remove one attachment from a capture (used by the View dialog)."""
    f = get_object_or_404(FarmCaptureFile, id=id)
    f.delete()
    return JsonResponse({"message": "File removed."})

@login_required
@require_POST
def farm_location_capture_complete(request, id):
    """Fill in what a capture is still missing, from the register's + button.

    Only blanks are written. The dialog locks what is already there, but the
    check lives here as well: the browser deciding what may be overwritten
    would let a hand-made request quietly replace a document or a GPS reading.
    """
    capture = get_object_or_404(FarmLocationCapture, id=id)
    filled = []

    with transaction.atomic():
        # Location goes in as a pair or not at all, and only if none is held.
        if capture.latitude is None and capture.longitude is None:
            lat = (request.POST.get("latitude") or "").strip()
            lng = (request.POST.get("longitude") or "").strip()
            if lat and lng:
                capture.latitude, capture.longitude = float(lat), float(lng)
                filled.append("location")
        for field in ("state", "district", "area", "address"):
            posted = (request.POST.get(field) or "").strip()
            if posted and not getattr(capture, field):
                setattr(capture, field, posted)
                filled.append(field)
        if filled:
            capture.full_clean(exclude=["capture_no", "captured_by"])
            capture.save()

        # A slot already holding a file on this capture is left alone.
        taken = set(capture.files.values_list("kind", flat=True))
        for kind in FarmCaptureFile.SLOT_TARGETS:
            if kind == FarmCaptureFile.KIND_DOCUMENT or kind in taken:
                continue
            upload = request.FILES.get("slot_%s" % kind)
            if upload:
                FarmCaptureFile.objects.create(capture=capture, kind=kind, file=upload)
                filled.append(kind)

        # Pictures and other documents are never "full" — more can always come.
        for upload in request.FILES.getlist("photos"):
            FarmCaptureFile.objects.create(
                capture=capture, kind=FarmCaptureFile.KIND_PHOTO, file=upload)
            filled.append("photo")
        for upload in request.FILES.getlist("documents"):
            FarmCaptureFile.objects.create(
                capture=capture, kind=FarmCaptureFile.KIND_DOCUMENT, file=upload)
            filled.append("document")

    if not filled:
        return JsonResponse({"error": "Nothing pending was filled in."}, status=400)
    return JsonResponse({"message": "Capture updated (%d item(s) filled in)." % len(filled)})


# --- Farmer & Farm combined report ------------------------------------------

# The report's columns are the master's inputs, in the master's own order, so
# what a clerk typed on Broiler > Master > Farm is what the register shows
# back. Anything not collected there has been dropped: a column that nothing
# can ever fill is a column that teaches people to ignore blanks.
FFR_COLUMNS = [
    "Sr No",
    # --- farmer, as the master collects it ---
    "Farmer Name", "Phone No", "Mobile-1", "Mobile-2", "Address",
    "PAN No", "PAN Upload", "Aadhar No", "Aadhar Front", "Aadhar Back",
    "National ID", "USC", "Service No", "Farmer Photo", "Farmer Group",
    "TDS %", "Account Holder", "Acc No", "IFSC Code", "Bank Name", "Bank Branch",
    # --- farm ---
    "Region", "Branch", "Line", "Supervisor", "Farm Code", "Farm Name",
    "Farm Type", "Farm Capacity", "Farm Sqft", "Pincode", "State", "District",
    "Area Name", "Farm Address", "Farm Location", "Farm Image",
    "Agreement Start", "Agreement End", "Agreement Months", "Agreement Copy",
    "Security Cheque-1", "Security Cheque-2", "Security Cheque-3",
    "Security Cheque-4", "Other Docs", "Remarks",
]


@login_required(login_url="login")
def farmer_farm_report(request):
    """Broiler > Reports > Farmer & Farm.

    One row per farm, carrying its farmer's details alongside — the two
    registers merged, because the question is always about the pair: whose
    farm is this, is their paperwork in order, is the agreement live.

    A farmer with no farm still gets a row, with the farm half blank. "Signed
    up and nothing placed" is exactly the gap a register like this is read to
    find, and dropping those rows would hide it.
    """
    def g(key):
        return (request.GET.get(key) or "").strip()

    group_id = g("farmer_group")
    branch_id, line, supervisor_id, farm_id = (
        g("branch"), g("line"), g("supervisor"), g("farm"))
    farm_filtered = any([branch_id, line, supervisor_id, farm_id])

    farms = (BroilerFarm.objects
             .select_related("branch", "supervisor", "farmer", "farmer__farmer_group")
             .prefetch_related("images")
             .order_by("farmer__farmer_name", "farm_name"))
    farms = scope_multi(request.user, farms, branches="branch_id", farms="id")
    if branch_id.isdigit():
        farms = farms.filter(branch_id=branch_id)
    if line:
        farms = farms.filter(line=line)
    if supervisor_id.isdigit():
        farms = farms.filter(supervisor_id=supervisor_id)
    if farm_id.isdigit():
        farms = farms.filter(id=farm_id)
    if group_id.isdigit():
        farms = farms.filter(farmer__farmer_group_id=group_id)

    farms = list(farms)
    rows = [{"farmer": f.farmer, "farm": f} for f in farms]

    # Farmers with nothing on the ground, appended only while no farm filter is
    # in play — under one they are not "missing", they simply do not match, and
    # listing them there would read as a gap that is not there.
    if not farm_filtered:
        loose = Farmer.objects.select_related("farmer_group").exclude(
            id__in={f.farmer_id for f in farms if f.farmer_id})
        if group_id.isdigit():
            loose = loose.filter(farmer_group_id=group_id)
        rows += [{"farmer": f, "farm": None} for f in loose.order_by("farmer_name")]

    criteria = [c for c in (
        f"Farmer Group: {FarmerGroup.objects.filter(id=group_id).first()}" if group_id.isdigit() else "",
        f"Branch: {Branch.objects.filter(id=branch_id).first()}" if branch_id.isdigit() else "",

    ) if c]

    return render(request, "farmer_farm_report.html", {
        "rows": rows,
        "columns": FFR_COLUMNS,
        "summary": {
            "rows": len(rows),
            "farmers": len({r["farmer"].id for r in rows if r["farmer"]}),
            "farms": len(farms),
            "capacity": sum(f.farm_capacity or 0 for f in farms),
            "unplaced": sum(1 for r in rows if r["farm"] is None),
        },
        "criteria": " | ".join(criteria),
        "group_id": group_id,
        "branch_id": branch_id, "line": line, "supervisor_id": supervisor_id,
        "farm_id": farm_id,
        "farmer_groups": FarmerGroup.objects.order_by("description"),
        "branches": branches_for(request.user),
        "supervisors": supervisors_for(request.user, Supervisor.objects.order_by("name")),
        "all_farms": farms_for(request.user, BroilerFarm.objects.order_by("farm_name")),
        "lines": (BroilerFarm.objects.exclude(line="")
                  .values_list("line", flat=True).distinct().order_by("line")),
    })
