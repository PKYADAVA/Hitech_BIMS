#pylint: disable=no-member

from decimal import Decimal
from typing import Optional

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.http import require_POST
from django.http import Http404, JsonResponse
from django.db.models import F
from django.core.files.storage import default_storage
from django.utils import timezone
from hatchery_master.models import STATES_AND_TERRITORIES
from account.models import ChartOfAccount
from account.services.bank_cash import bank_cash_accounts, active_payment_modes, payment_mode_map
from inventory.models import Item, ItemCategory, Warehouse
from picklist.services import validate_value
from .models import (ChicksPurchase, ChicksPurchaseItem, GeneralPurchase, GeneralPurchaseItem,
                     Supplier, SupplierPayment, SupplierPaymentLine, SupplierShippingAddress, TaxMaster)
import json

# Used only by the billing/shipping address modals (state field itself is
# picklist-bound, see picklist.bindable_fields.BINDABLE_FIELDS).
states_and_union_territories = STATES_AND_TERRITORIES

@login_required()
def supplier(request):
    return render(request, "supplier.html", {"suppliers": Supplier.objects.all()})


def _supplier_form_context(supplier=None):
    return {
        "supplier": supplier,
        "next_code": Supplier.next_code() if not supplier else None,
        "states_and_union_territories": states_and_union_territories,
        "to_pay_to_receive_choices": Supplier.ToPayToReceive.choices,
        "today": timezone.localdate().isoformat(),
    }


def _apply_posted_supplier_fields(instance, request):
    instance.name = request.POST.get("name", "").strip()
    instance.address = request.POST.get("address", "").strip()
    instance.place = request.POST.get("place", "").strip()
    instance.mobile = request.POST.get("mobile", "").strip()
    instance.mobile_2 = request.POST.get("mobile_2", "").strip()
    instance.email = request.POST.get("email", "").strip() or None
    instance.aadhar = request.POST.get("aadhar", "").strip()
    instance.contact_type = request.POST.get("contact_type") or Supplier.ContactType.BOTH
    instance.party_category = request.POST.get("party_category") or None
    instance.pan = request.POST.get("pan", "").strip()
    instance.supplier_group = request.POST.get("supplier_group", "").strip()
    instance.gstin = request.POST.get("gstin", "").strip()
    instance.state = request.POST.get("state", "").strip()
    instance.credit_term = request.POST.get("credit_term") or None
    instance.credit_limit = request.POST.get("credit_limit") or 0
    instance.opening_balance = request.POST.get("opening_balance") or None
    instance.to_pay_to_receive = request.POST.get("to_pay_to_receive") or None
    instance.as_on_date = request.POST.get("as_on_date") or None
    instance.note = request.POST.get("note", "").strip()
    instance.country = request.POST.get("country", "").strip()
    instance.currency = request.POST.get("currency", "").strip()
    instance.account_no = request.POST.get("account_no", "").strip()
    instance.ifsc_code = request.POST.get("ifsc_code", "").strip()
    instance.bank_details = request.POST.get("bank_details", "").strip()
    instance.terms = request.POST.get("terms", "").strip()
    instance.agreement_start_date = request.POST.get("agreement_start_date") or None
    instance.agreement_months = request.POST.get("agreement_months") or None
    if request.FILES.get("agreement_copy"):
        instance.agreement_copy = request.FILES["agreement_copy"]
    if request.FILES.get("other_documents"):
        instance.other_documents = request.FILES["other_documents"]
    for field in ("state", "contact_type", "party_category", "supplier_group"):
        validate_value("purchase", "Supplier", field, getattr(instance, field))


def _create_posted_shipping_addresses(instance, request):
    try:
        addresses = json.loads(request.POST.get("shipping_addresses_json") or "[]")
    except json.JSONDecodeError:
        addresses = []
    if not addresses and instance.address:
        addresses = [{"label": instance.address[:100], "address": instance.address, "is_default": True}]
    default_assigned = False
    for entry in addresses:
        label = (entry.get("label") or "").strip()
        address_text = (entry.get("address") or "").strip()
        if not label or not address_text:
            continue
        is_default = bool(entry.get("is_default")) and not default_assigned
        default_assigned = default_assigned or is_default
        SupplierShippingAddress.objects.create(
            supplier=instance, label=label, address=address_text,
            contact_person=(entry.get("contact_person") or "").strip(),
            mobile=(entry.get("mobile") or "").strip(),
            is_default=is_default,
        )

def _sync_default_shipping_address(instance, previous_address):
    """Keep the default shipping address in step with the billing address.

    On create the default shipping address is seeded from billing, so the two
    start out identical. Editing billing therefore has to move it along too —
    but only while it is still a mirror of the old billing text. Once someone
    has given the shipping address its own content, it is left alone.
    """
    new_address = (instance.address or "").strip()
    if not new_address or new_address == (previous_address or "").strip():
        return
    default = instance.shipping_addresses.filter(is_default=True).first()
    if default is None:
        SupplierShippingAddress.objects.create(
            supplier=instance, label=new_address[:100], address=new_address, is_default=True)
        return
    if (default.address or "").strip() == (previous_address or "").strip():
        default.label = new_address[:100]
        default.address = new_address
        default.save(update_fields=["label", "address"])


@login_required(login_url="login")
def create_supplier(request):
    """Add a new supplier master record."""
    if request.method == "POST":
        instance = Supplier()
        try:
            _apply_posted_supplier_fields(instance, request)
            instance.full_clean()
            instance.save()
            _create_posted_shipping_addresses(instance, request)
            messages.success(request, "Supplier added successfully.")
            return redirect("supplier")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "supplier_form.html", _supplier_form_context())


@login_required(login_url="login")
def edit_supplier(request, id):
    """Edit an existing supplier master record."""
    instance = get_object_or_404(Supplier, id=id)

    if request.method == "POST":
        previous_address = instance.address
        try:
            _apply_posted_supplier_fields(instance, request)
            instance.full_clean()
            instance.save()
            _sync_default_shipping_address(instance, previous_address)
            messages.success(request, "Supplier updated successfully.")
            return redirect("supplier")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "supplier_form.html", _supplier_form_context(instance))


@login_required(login_url="login")
@require_POST
def delete_supplier(request, id):
    """Delete a supplier master record."""
    instance = get_object_or_404(Supplier, id=id)
    instance.delete()
    messages.success(request, "Supplier deleted successfully.")
    return redirect("supplier")


@method_decorator(login_required, name="dispatch")
class SupplierShippingAddressAPI(View):
    """Supplier Master addresses, also usable by transaction forms."""
    def get(self, request, supplier_id, id=None):
        supplier = Supplier.objects.get(id=supplier_id)
        addresses = supplier.shipping_addresses.all()
        if id:
            addresses = addresses.filter(id=id)
        result = [{
            "id": address.id, "label": address.label, "address": address.address,
            "contact_person": address.contact_person, "mobile": address.mobile,
            "is_default": address.is_default,
        } for address in addresses]
        if id:
            if not result:
                return JsonResponse({"error": "Shipping address not found"}, status=404)
            return JsonResponse(result[0])
        return JsonResponse(result, safe=False)

    def post(self, request, supplier_id):
        try:
            data = json.loads(request.body)
            supplier = Supplier.objects.get(id=supplier_id)
            if not data.get("label") or not data.get("address"):
                return JsonResponse({"error": "Address label and address are required"}, status=400)
            if data.get("is_default"):
                supplier.shipping_addresses.update(is_default=False)
            address = SupplierShippingAddress.objects.create(
                supplier=supplier, label=data["label"], address=data["address"],
                contact_person=data.get("contact_person", ""), mobile=data.get("mobile", ""),
                is_default=bool(data.get("is_default")),
            )
            return JsonResponse({"id": address.id, "message": "Shipping address saved"}, status=201)
        except Supplier.DoesNotExist:
            return JsonResponse({"error": "Supplier not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def put(self, request, supplier_id, id):
        try:
            data = json.loads(request.body)
            address = SupplierShippingAddress.objects.get(id=id, supplier_id=supplier_id)
            if data.get("is_default"):
                SupplierShippingAddress.objects.filter(supplier_id=supplier_id).exclude(id=id).update(is_default=False)
            for field in ("label", "address", "contact_person", "mobile"):
                if field in data:
                    setattr(address, field, data[field])
            address.is_default = bool(data.get("is_default", address.is_default))
            address.full_clean(); address.save()
            return JsonResponse({"message": "Shipping address updated"})
        except SupplierShippingAddress.DoesNotExist:
            return JsonResponse({"error": "Shipping address not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def delete(self, request, supplier_id, id):
        try:
            SupplierShippingAddress.objects.get(id=id, supplier_id=supplier_id).delete()
            return JsonResponse({"message": "Shipping address deleted"})
        except SupplierShippingAddress.DoesNotExist:
            return JsonResponse({"error": "Shipping address not found"}, status=404)




@login_required()
def vendor_groups(request):
    from account.models import ChartOfAccount
    return render(request, "vendor_group.html", {
        "coa_accounts": ChartOfAccount.objects.filter(status="Active").order_by("code"),
    })


@login_required()
def tax_master(request):
    return render(request, "tax_master.html")


from .models import VendorGroup


@method_decorator(login_required, name="dispatch")
class VendorGroupAPI(View):

    @staticmethod
    def _serialize(group):
        return {
            "id": group.id,
            "code": group.code,
            "description": group.description,
            "currency": group.currency,
            "control_account": group.control_account_id,
            "control_account_display": str(group.control_account) if group.control_account else "",
            "prepayment_account": group.prepayment_account_id,
            "prepayment_account_display": str(group.prepayment_account) if group.prepayment_account else "",
        }

    def get(self, request, id=None):
        if id:
            try:
                vendor_group = VendorGroup.objects.select_related("control_account", "prepayment_account").get(id=id)
                return JsonResponse(self._serialize(vendor_group))
            except VendorGroup.DoesNotExist:
                raise Http404("VendorGroup not found")
        else:
            vendor_groups = [
                self._serialize(group)
                for group in VendorGroup.objects.select_related("control_account", "prepayment_account")
            ]
            return JsonResponse(vendor_groups, safe=False)

    def post(self, request):
        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError as e:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        VendorGroup.objects.create(
            code=data.get("code"),
            description=data.get("description"),
            currency=data.get("currency"),
            control_account_id=data.get("control_account") or None,
            prepayment_account_id=data.get("prepayment_account") or None,
        )
        return JsonResponse({"message": "VendorGroup created"}, status=201)

    def put(self, request, id):
        try:
            vendor_group = VendorGroup.objects.get(id=id)
        except VendorGroup.DoesNotExist:
            raise Http404("VendorGroup not found")

        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        vendor_group.code = data.get("code", vendor_group.code)
        vendor_group.description = data.get("description", vendor_group.description)
        vendor_group.currency = data.get("currency", vendor_group.currency)
        if "control_account" in data:
            vendor_group.control_account_id = data["control_account"] or None
        if "prepayment_account" in data:
            vendor_group.prepayment_account_id = data["prepayment_account"] or None
        vendor_group.save()
        return JsonResponse({"message": "VendorGroup updated"})

    def delete(self, request, id):
        try:
            vendor_group = VendorGroup.objects.get(id=id)
        except VendorGroup.DoesNotExist:
            raise Http404("VendorGroup not found")

        vendor_group.delete()
        return JsonResponse({"message": "VendorGroup deleted"})


@method_decorator(login_required, name="dispatch")
class TaxMasterAPI(View):

    def get(self, request, id=None):
        if id:
            try:
                tax_master = TaxMaster.objects.get(id=id)
                return JsonResponse(
                    {
                        "id": tax_master.id,
                        "tax_code": tax_master.tax_code,
                        "description": tax_master.description,
                        "tax_percentage": tax_master.tax_percentage,
                        "rule": tax_master.rule,
                        "coa": tax_master.coa,
                    }
                )
            except TaxMaster.DoesNotExist:
                raise Http404("TaxMaster not found")
        else:
            tax_masters = list(
                TaxMaster.objects.values(
                    "id", "tax_code", "description", "tax_percentage", "rule", "coa"
                )
            )
            return JsonResponse(tax_masters, safe=False)

    def post(self, request):
        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        TaxMaster.objects.create(
            tax_code=data.get("tax_code"),
            description=data.get("description"),
            tax_percentage=data.get("tax_percentage"),
            rule=data.get("rule"),
            coa=data.get("coa"),
        )
        return JsonResponse({"message": "TaxMaster created"}, status=201)

    def put(self, request, id):
        try:
            tax_master = TaxMaster.objects.get(id=id)
        except TaxMaster.DoesNotExist:
            raise Http404("TaxMaster not found")

        try:
            data = json.loads(request.body)  # Expect JSON payload
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        tax_master.tax_code = data.get("tax_code", tax_master.tax_code)
        tax_master.description = data.get("description", tax_master.description)
        tax_master.tax_percentage = data.get("tax_percentage", tax_master.tax_percentage)
        tax_master.rule = data.get("rule", tax_master.rule)
        tax_master.coa = data.get("coa", tax_master.coa)
        tax_master.save()

        return JsonResponse({"message": "TaxMaster updated"})

    def delete(self, request, id):
        try:
            tax_master = TaxMaster.objects.get(id=id)
        except TaxMaster.DoesNotExist:
            raise Http404("TaxMaster not found")

        tax_master.delete()
        return JsonResponse({"message": "TaxMaster deleted"})




# ---------------------------------------------------------------------------
# General Purchase (Purchase > Transactions)
# ---------------------------------------------------------------------------

def _general_purchase_to_item_dict(row):
    return {
        "item": row.item_id, "item_code": row.item.item_code,
        "item_description": row.item.description, "unit": row.unit,
        "sent_qty": str(row.sent_qty), "rcv_qty": str(row.rcv_qty), "free_qty": str(row.free_qty),
        "rate": str(row.rate), "discount_percent": str(row.discount_percent),
        "discount_amount": str(row.discount_amount), "gst_percent": str(row.gst_percent),
        "amount": str(row.amount), "farm_warehouse": row.farm_warehouse_id,
        "farm_warehouse_name": row.farm_warehouse.name,
    }


def _general_purchase_list_dict(gp):
    warehouses = ", ".join(dict.fromkeys(
        n for n in gp.items.values_list("farm_warehouse__name", flat=True) if n
    ))
    return {
        "id": gp.id, "date": gp.date.isoformat(), "bill_no": gp.bill_no, "dc_no": gp.dc_no,
        "supplier_name": gp.supplier.name, "item_names": gp.item_names(),
        "quantity": str(gp.total_quantity()), "no_of_bags": str(gp.no_of_bags),
        "avg_rate": str(gp.avg_rate()), "net_amount": str(gp.net_amount),
        "farm_warehouse_names": warehouses, "batch_no": gp.batch_no,
        "vehicle_no": gp.vehicle_no, "driver_name": gp.driver_name,
    }


def _general_purchase_form_context(gp=None):
    return {
        "general_purchase": gp,
        "next_purchase_no": GeneralPurchase._next_purchase_no() if not gp else None,
        "suppliers": Supplier.objects.order_by("name"),
        "items": Item.objects.order_by("item_code"),
        "warehouses": Warehouse.objects.order_by("name"),
        "accounts": ChartOfAccount.objects.order_by("code"),
        "bank_accounts": bank_cash_accounts(),   # Pay Account = Bank/Cash master only
        "tax_masters": TaxMaster.objects.exclude(tax_percentage__isnull=True).order_by("tax_code"),
        "today": timezone.localdate().isoformat(),
        "existing_items_json": json.dumps(
            [_general_purchase_to_item_dict(row) for row in gp.items.select_related("item", "farm_warehouse")]
        ) if gp else "[]",
        "payment_terms_choices": GeneralPurchase.PAYMENT_TERMS_CHOICES,
        "freight_type_choices": GeneralPurchase.FREIGHT_TYPE_CHOICES,
        "bag_type_choices": GeneralPurchase.BAG_TYPE_CHOICES,
        "other_charges_type_choices": GeneralPurchase.OTHER_CHARGES_TYPE_CHOICES,
        "round_off_type_choices": GeneralPurchase.ROUND_OFF_TYPE_CHOICES,
    }


def _apply_posted_general_purchase_fields(instance, request):
    instance.date = request.POST.get("date") or timezone.localdate()
    instance.supplier_id = request.POST.get("supplier") or None
    instance.bill_no = request.POST.get("bill_no", "").strip()
    # DC No. was dropped from the form (it duplicated Bill No.); it is left
    # untouched here so historic values survive an edit.
    instance.vehicle_no = request.POST.get("vehicle_no", "").strip()
    instance.driver_name = request.POST.get("driver_name", "").strip()
    instance.driver_mobile = request.POST.get("driver_mobile", "").strip()
    instance.calculation_based_on = request.POST.get("calculation_based_on") or "Sent Quantity"
    instance.payment_terms = request.POST.get("payment_terms") or "Cash"
    instance.freight_type = request.POST.get("freight_type") or "Extra"
    instance.payment_mode = request.POST.get("payment_mode") or "pay_later"
    instance.pay_account_id = request.POST.get("pay_account") or None
    instance.freight_account_id = request.POST.get("freight_account") or None
    instance.freight_amount = request.POST.get("freight_amount") or 0
    instance.bag_type = request.POST.get("bag_type", "").strip()
    instance.no_of_bags = request.POST.get("no_of_bags") or 0
    instance.batch_no = request.POST.get("batch_no", "").strip()
    instance.expiry_date = request.POST.get("expiry_date") or None
    instance.tds_code = request.POST.get("tds_code", "").strip()
    instance.tds_applicable = request.POST.get("tds_applicable") == "on"
    instance.tds_amount = request.POST.get("tds_amount") or 0
    instance.other_charges_account_id = request.POST.get("other_charges_account") or None
    instance.other_charges_type = request.POST.get("other_charges_type") or "Add"
    instance.other_charges_amount = request.POST.get("other_charges_amount") or 0
    # round_off / round_off_type are auto-derived in compute_net_amount(),
    # never taken from the posted form.
    instance.remarks = request.POST.get("remarks", "").strip()
    if request.FILES.get("reference_document_1"):
        instance.reference_document_1 = request.FILES["reference_document_1"]
    if request.FILES.get("reference_document_2"):
        instance.reference_document_2 = request.FILES["reference_document_2"]
    if request.FILES.get("reference_document_3"):
        instance.reference_document_3 = request.FILES["reference_document_3"]
    validate_value("purchase", "GeneralPurchase", "calculation_based_on", instance.calculation_based_on)


def _save_general_purchase_items(instance, request):
    try:
        rows = json.loads(request.POST.get("items_json") or "[]")
    except json.JSONDecodeError:
        rows = []
    instance.items.all().delete()
    for row in rows:
        if not row.get("item") or not row.get("farm_warehouse"):
            continue
        GeneralPurchaseItem.objects.create(
            purchase=instance, item_id=row["item"], unit=row.get("unit") or "",
            sent_qty=Decimal(str(row.get("sent_qty") or 0)),
            rcv_qty=Decimal(str(row.get("rcv_qty") or 0)),
            free_qty=Decimal(str(row.get("free_qty") or 0)),
            rate=Decimal(str(row.get("rate") or 0)),
            discount_percent=Decimal(str(row.get("discount_percent") or 0)),
            discount_amount=Decimal(str(row.get("discount_amount") or 0)),
            gst_percent=Decimal(str(row.get("gst_percent") or 0)),
            farm_warehouse_id=row["farm_warehouse"],
        )
    instance.net_amount = instance.compute_net_amount()
    # "remarks" is included so an auto-generated description picks up the total
    # that only became known once the line items were saved.
    instance.save(update_fields=["net_amount", "round_off", "round_off_type", "remarks"])


@login_required(login_url="login")
def general_purchase_list(request):
    return render(request, "general_purchase_list.html", {
        "categories": ItemCategory.objects.order_by("name"),
        "warehouses": Warehouse.objects.order_by("name"),
    })


@login_required(login_url="login")
def create_general_purchase(request):
    """Add a new General Purchase transaction."""
    if request.method == "POST":
        instance = GeneralPurchase()
        try:
            _apply_posted_general_purchase_fields(instance, request)
            instance.full_clean(exclude=["purchase_no"])
            with transaction.atomic():
                instance.save()
                _save_general_purchase_items(instance, request)
            messages.success(request, "General purchase added successfully.")
            return redirect("general_purchase_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "general_purchase_form.html", _general_purchase_form_context())


@login_required(login_url="login")
def edit_general_purchase(request, id):
    """Edit an existing General Purchase transaction."""
    instance = get_object_or_404(GeneralPurchase, id=id)

    if request.method == "POST":
        try:
            _apply_posted_general_purchase_fields(instance, request)
            instance.full_clean(exclude=["purchase_no"])
            with transaction.atomic():
                instance.save()
                _save_general_purchase_items(instance, request)
            messages.success(request, "General purchase updated successfully.")
            return redirect("general_purchase_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "general_purchase_form.html", _general_purchase_form_context(instance))


@login_required(login_url="login")
@require_POST
def delete_general_purchase(request, id):
    """Delete a General Purchase transaction."""
    instance = get_object_or_404(GeneralPurchase, id=id)
    instance.delete()
    messages.success(request, "General purchase deleted successfully.")
    return redirect("general_purchase_list")


@login_required
def general_purchase_api_list(request):
    """JSON rows for the General Purchase register's DataTable + filter bar."""
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    category = (request.GET.get("category") or "").strip()
    warehouse = (request.GET.get("warehouse") or "").strip()

    qs = GeneralPurchase.objects.select_related("supplier").prefetch_related(
        "items__item", "items__farm_warehouse")
    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    if category:
        qs = qs.filter(items__item__category_id=category)
    if warehouse:
        qs = qs.filter(items__farm_warehouse_id=warehouse)
    qs = qs.distinct().order_by("-date", "-id")
    return JsonResponse([_general_purchase_list_dict(gp) for gp in qs], safe=False)



# ---------------------------------------------------------------------------
# Chicks Purchase (Purchase > Transactions)
# ---------------------------------------------------------------------------

def _chicks_purchase_to_item_dict(row):
    return {
        "sent_qty": str(row.sent_qty), "sent_free_percent": str(row.sent_free_percent),
        "rcv_free_percent": str(row.rcv_free_percent),
        "mortality": str(row.mortality), "shortage": str(row.shortage), "weaks": str(row.weaks),
        "excess_qty": str(row.excess_qty), "rcv_qty": str(row.rcv_qty),
        "free_qty": str(row.free_qty), "total_qty": str(row.total_qty),
        "rate": str(row.rate), "amount": str(row.amount),
        "farm_warehouse": row.farm_warehouse_id, "farm_warehouse_name": row.farm_warehouse.name,
        "batch": row.batch,
    }


def _chicks_purchase_list_dict(cp):
    warehouses = ", ".join(dict.fromkeys(
        n for n in cp.items.values_list("farm_warehouse__name", flat=True) if n
    ))
    return {
        "id": cp.id, "date": cp.date.isoformat(), "bill_no": cp.bill_no, "dc_no": cp.dc_no,
        "supplier_name": cp.supplier.name,
        "item_code": cp.item.item_code,
        "quantity": str(cp.total_quantity()), "avg_rate": str(cp.avg_rate()),
        "net_amount": str(cp.net_amount), "farm_warehouse_names": warehouses,
    }


def _chicks_purchase_form_context(cp=None):
    return {
        "chicks_purchase": cp,
        "next_purchase_no": ChicksPurchase._next_purchase_no() if not cp else None,
        "suppliers": Supplier.objects.order_by("name"),
        "items": Item.objects.order_by("item_code"),
        "warehouses": Warehouse.objects.order_by("name"),
        "accounts": ChartOfAccount.objects.order_by("code"),
        "bank_accounts": bank_cash_accounts(),   # Pay Account = Bank/Cash master only
        "today": timezone.localdate().isoformat(),
        "existing_items_json": json.dumps(
            [_chicks_purchase_to_item_dict(row) for row in cp.items.select_related("farm_warehouse")]
        ) if cp else "[]",
        "freight_type_choices": ChicksPurchase.FREIGHT_TYPE_CHOICES,
        "bag_type_choices": ChicksPurchase.BAG_TYPE_CHOICES,
        "other_charges_type_choices": ChicksPurchase.OTHER_CHARGES_TYPE_CHOICES,
        "round_off_type_choices": ChicksPurchase.ROUND_OFF_TYPE_CHOICES,
    }


def _apply_posted_chicks_purchase_fields(instance, request):
    instance.date = request.POST.get("date") or timezone.localdate()
    instance.supplier_id = request.POST.get("supplier") or None
    # Hatchery was dropped from the form (the Supplier identifies it); it is
    # left untouched here so historic values survive an edit.
    instance.item_id = request.POST.get("item") or None
    instance.bill_no = request.POST.get("bill_no", "").strip()
    # DC No. was dropped from the form (it duplicated Bill No.); it is left
    # untouched here so historic values survive an edit.
    instance.vehicle_no = request.POST.get("vehicle_no", "").strip()
    instance.driver_name = request.POST.get("driver_name", "").strip()
    instance.freight_type = request.POST.get("freight_type") or "Extra"
    instance.payment_mode = request.POST.get("payment_mode") or "pay_later"
    instance.pay_account_id = request.POST.get("pay_account") or None
    instance.freight_account_id = request.POST.get("freight_account") or None
    instance.freight_amount = request.POST.get("freight_amount") or 0
    instance.bag_type = request.POST.get("bag_type", "").strip()
    instance.no_of_bags = request.POST.get("no_of_bags") or 0
    instance.batch_no = request.POST.get("batch_no", "").strip()
    instance.expiry_date = request.POST.get("expiry_date") or None
    instance.tds_code = request.POST.get("tds_code", "").strip()
    instance.tds_applicable = request.POST.get("tds_applicable") == "on"
    instance.tds_amount = request.POST.get("tds_amount") or 0
    instance.other_charges_account_id = request.POST.get("other_charges_account") or None
    instance.other_charges_type = request.POST.get("other_charges_type") or "Add"
    instance.other_charges_amount = request.POST.get("other_charges_amount") or 0
    # round_off / round_off_type are auto-derived in compute_net_amount(),
    # never taken from the posted form.
    instance.remarks = request.POST.get("remarks", "").strip()
    if request.FILES.get("reference_document_1"):
        instance.reference_document_1 = request.FILES["reference_document_1"]
    if request.FILES.get("reference_document_2"):
        instance.reference_document_2 = request.FILES["reference_document_2"]
    if request.FILES.get("reference_document_3"):
        instance.reference_document_3 = request.FILES["reference_document_3"]


def _save_chicks_purchase_items(instance, request):
    try:
        rows = json.loads(request.POST.get("items_json") or "[]")
    except json.JSONDecodeError:
        rows = []
    instance.items.all().delete()
    for row in rows:
        if not row.get("farm_warehouse"):
            continue
        ChicksPurchaseItem.objects.create(
            purchase=instance,
            sent_qty=Decimal(str(row.get("sent_qty") or 0)),
            sent_free_percent=Decimal(str(row.get("sent_free_percent") or 0)),
            rcv_free_percent=Decimal(str(row.get("rcv_free_percent") or 0)),
            mortality=Decimal(str(row.get("mortality") or 0)),
            shortage=Decimal(str(row.get("shortage") or 0)),
            weaks=Decimal(str(row.get("weaks") or 0)),
            excess_qty=Decimal(str(row.get("excess_qty") or 0)),
            rate=Decimal(str(row.get("rate") or 0)),
            farm_warehouse_id=row["farm_warehouse"],
            batch=row.get("batch") or "",
        )
    instance.net_amount = instance.compute_net_amount()
    # "remarks" is included so an auto-generated description picks up the total
    # that only became known once the line items were saved.
    instance.save(update_fields=["net_amount", "round_off", "round_off_type", "remarks"])


@login_required(login_url="login")
def chicks_purchase_list(request):
    return render(request, "chicks_purchase_list.html", {
        "warehouses": Warehouse.objects.order_by("name"),
    })


@login_required(login_url="login")
def create_chicks_purchase(request):
    """Add a new Chicks Purchase transaction."""
    if request.method == "POST":
        instance = ChicksPurchase()
        try:
            _apply_posted_chicks_purchase_fields(instance, request)
            instance.full_clean(exclude=["purchase_no"])
            with transaction.atomic():
                instance.save()
                _save_chicks_purchase_items(instance, request)
            messages.success(request, "Chicks purchase added successfully.")
            return redirect("chicks_purchase_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "chicks_purchase_form.html", _chicks_purchase_form_context())


@login_required(login_url="login")
def edit_chicks_purchase(request, id):
    """Edit an existing Chicks Purchase transaction."""
    instance = get_object_or_404(ChicksPurchase, id=id)

    if request.method == "POST":
        try:
            _apply_posted_chicks_purchase_fields(instance, request)
            instance.full_clean(exclude=["purchase_no"])
            with transaction.atomic():
                instance.save()
                _save_chicks_purchase_items(instance, request)
            messages.success(request, "Chicks purchase updated successfully.")
            return redirect("chicks_purchase_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "chicks_purchase_form.html", _chicks_purchase_form_context(instance))


@login_required(login_url="login")
@require_POST
def delete_chicks_purchase(request, id):
    """Delete a Chicks Purchase transaction."""
    instance = get_object_or_404(ChicksPurchase, id=id)
    instance.delete()
    messages.success(request, "Chicks purchase deleted successfully.")
    return redirect("chicks_purchase_list")


@login_required
def chicks_purchase_api_list(request):
    """JSON rows for the Chicks Purchase register's DataTable + filter bar."""
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()

    qs = ChicksPurchase.objects.select_related("supplier", "item").prefetch_related(
        "items__farm_warehouse")
    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    qs = qs.order_by("-date", "-id")
    return JsonResponse([_chicks_purchase_list_dict(cp) for cp in qs], safe=False)



# ---------------------------------------------------------------------------
# Supplier Payment (Purchase > Transactions)
# ---------------------------------------------------------------------------

def _payment_to_line_dict(row):
    return {
        "supplier": row.supplier_id, "mode": row.mode, "pay_account": row.pay_account_id,
        "pay_account_name": f"{row.pay_account.code} - {row.pay_account.description}",
        "amount": str(row.amount), "bank_charges": str(row.bank_charges),
        "reference_no": row.reference_no, "remarks": row.remarks,
    }


def _payment_list_dict(p):
    return {
        "id": p.id, "date": p.date.isoformat(), "payment_no": p.payment_no,
        "supplier_name": p.supplier_summary(), "mode": p.mode_summary(),
        "method": p.method_summary(), "amount": str(p.total_amount()),
    }


def _bank_cash_accounts():
    """Chart-of-Account ledgers backed by a Bank/Cash Master entry — the only
    accounts a payment's cash/bank should be picked from (excludes generic COAs
    and the internal __VERIFY_* anchors that aren't tied to a master row)."""
    from account.models import BankCashMaster
    from django.contrib.contenttypes.models import ContentType
    ct = ContentType.objects.get_for_model(BankCashMaster)
    return (ChartOfAccount.objects
            .filter(source_content_type=ct,
                    source_object_id__in=BankCashMaster.objects.values("id"))
            .order_by("code"))


def _payment_form_context(p=None):
    return {
        "payment": p,
        "next_payment_no": SupplierPayment._next_payment_no() if not p else None,
        "suppliers": Supplier.objects.order_by("name"),
        "locations": Warehouse.objects.order_by("name"),
        "accounts": _bank_cash_accounts(),
        "today": timezone.localdate().isoformat(),
        "mode_choices": SupplierPaymentLine.MODE_CHOICES,
        "payment_modes": active_payment_modes("payment"),
        "payment_mode_map_json": json.dumps(payment_mode_map("payment")),
        "existing_lines_json": json.dumps(
            [_payment_to_line_dict(row) for row in p.lines.select_related("pay_account", "supplier")]
        ) if p else "[]",
    }


def _apply_posted_payment_fields(instance, request):
    instance.date = request.POST.get("date") or timezone.localdate()
    instance.location_id = request.POST.get("location") or None


def _save_payment_lines(instance, request):
    try:
        rows = json.loads(request.POST.get("lines_json") or "[]")
    except json.JSONDecodeError:
        rows = []
    instance.lines.all().delete()
    for row in rows:
        if not row.get("supplier") or not row.get("pay_account") or not row.get("amount"):
            continue
        SupplierPaymentLine.objects.create(
            payment=instance, supplier_id=row["supplier"], mode=row.get("mode") or "Cash",
            pay_account_id=row["pay_account"],
            amount=Decimal(str(row.get("amount") or 0)),
            bank_charges=Decimal(str(row.get("bank_charges") or 0)),
            reference_no=row.get("reference_no") or "",
            remarks=row.get("remarks") or "",
        )


@login_required(login_url="login")
def payment_list(request):
    return render(request, "payment_list.html")


@login_required(login_url="login")
def create_payment(request):
    """Add a new Supplier Payment voucher."""
    if request.method == "POST":
        instance = SupplierPayment()
        try:
            _apply_posted_payment_fields(instance, request)
            instance.full_clean(exclude=["payment_no"])
            with transaction.atomic():
                instance.save()
                _save_payment_lines(instance, request)
                if not instance.lines.exists():
                    raise ValidationError("Add at least one payment line.")
            messages.success(request, "Payment added successfully.")
            return redirect("payment_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "payment_form.html", _payment_form_context())


@login_required(login_url="login")
def edit_payment(request, id):
    """Edit an existing Supplier Payment voucher."""
    instance = get_object_or_404(SupplierPayment, id=id)

    if request.method == "POST":
        try:
            _apply_posted_payment_fields(instance, request)
            instance.full_clean(exclude=["payment_no"])
            with transaction.atomic():
                instance.save()
                _save_payment_lines(instance, request)
                if not instance.lines.exists():
                    raise ValidationError("Add at least one payment line.")
            messages.success(request, "Payment updated successfully.")
            return redirect("payment_list")
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))

    return render(request, "payment_form.html", _payment_form_context(instance))


@login_required(login_url="login")
@require_POST
def delete_payment(request, id):
    """Delete a Supplier Payment voucher."""
    instance = get_object_or_404(SupplierPayment, id=id)
    instance.delete()
    messages.success(request, "Payment deleted successfully.")
    return redirect("payment_list")


@login_required
def payment_api_list(request):
    """JSON rows for the Payment register's DataTable + filter bar."""
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()

    qs = SupplierPayment.objects.prefetch_related("lines__pay_account", "lines__supplier")
    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    qs = qs.order_by("-date", "-id")
    return JsonResponse([_payment_list_dict(p) for p in qs], safe=False)


# ---------------------------------------------------------------------------
# Debit / Credit Notes (Purchase > Transactions)
# ---------------------------------------------------------------------------
from .models import DebitNote, CreditNote


def _note_list_dict(n):
    return {
        "id": n.id, "date": n.date.isoformat(), "note_no": n.note_no,
        "supplier_name": n.supplier.name if n.supplier_id else "",
        "against_bill": n.against_bill, "reason": n.reason,
        "amount": str(n.amount), "remarks": n.remarks,
    }


def _note_form_context(model, instance=None):
    return {
        "note": instance,
        "next_no": model._next_no() if not instance else None,
        "note_kind": "Debit Note" if model is DebitNote else "Credit Note",
        "suppliers": Supplier.objects.order_by("name"),
        "accounts": ChartOfAccount.objects.order_by("code"),
        "today": timezone.localdate().isoformat(),
    }


def _apply_note_fields(instance, request):
    instance.date = request.POST.get("date") or timezone.localdate()
    instance.supplier_id = request.POST.get("supplier") or None
    instance.against_bill = request.POST.get("against_bill") or ""
    instance.reason = request.POST.get("reason") or ""
    instance.amount = Decimal(str(request.POST.get("amount") or 0))
    instance.account_id = request.POST.get("account") or None
    instance.remarks = request.POST.get("remarks") or ""


def _save_note(request, model, instance, kind, template, redirect_name):
    if request.method == "POST":
        try:
            _apply_note_fields(instance, request)
            if not instance.supplier_id:
                raise ValidationError("Select a supplier.")
            if instance.amount <= 0:
                raise ValidationError("Enter an amount greater than zero.")
            instance.full_clean(exclude=["note_no"])
            was_edit = bool(instance.pk)
            with transaction.atomic():
                instance.save()
            messages.success(request, f"{kind} {'updated' if was_edit else 'added'} successfully.")
            return redirect(redirect_name)
        except ValidationError as e:
            messages.error(request, " ".join(e.messages) if hasattr(e, "messages") else str(e))
    return render(request, template, _note_form_context(model, instance if instance.pk else None))


def _note_api(request, model):
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    supplier_id = (request.GET.get("supplier") or "").strip()
    qs = model.objects.select_related("supplier")
    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    if supplier_id.isdigit():
        qs = qs.filter(supplier_id=supplier_id)
    return JsonResponse([_note_list_dict(n) for n in qs.order_by("-date", "-id")], safe=False)


# --- Debit Note ---
@login_required(login_url="login")
def debit_note_list(request):
    return render(request, "debit_note_list.html")


@login_required(login_url="login")
def create_debit_note(request):
    return _save_note(request, DebitNote, DebitNote(), "Debit Note", "debit_note_form.html", "debit_note_list")


@login_required(login_url="login")
def edit_debit_note(request, id):
    return _save_note(request, DebitNote, get_object_or_404(DebitNote, id=id), "Debit Note", "debit_note_form.html", "debit_note_list")


@login_required(login_url="login")
@require_POST
def delete_debit_note(request, id):
    get_object_or_404(DebitNote, id=id).delete()
    messages.success(request, "Debit Note deleted successfully.")
    return redirect("debit_note_list")


@login_required
def debit_note_api_list(request):
    return _note_api(request, DebitNote)


# --- Credit Note ---
@login_required(login_url="login")
def credit_note_list(request):
    return render(request, "credit_note_list.html")


@login_required(login_url="login")
def create_credit_note(request):
    return _save_note(request, CreditNote, CreditNote(), "Credit Note", "credit_note_form.html", "credit_note_list")


@login_required(login_url="login")
def edit_credit_note(request, id):
    return _save_note(request, CreditNote, get_object_or_404(CreditNote, id=id), "Credit Note", "credit_note_form.html", "credit_note_list")


@login_required(login_url="login")
@require_POST
def delete_credit_note(request, id):
    get_object_or_404(CreditNote, id=id).delete()
    messages.success(request, "Credit Note deleted successfully.")
    return redirect("credit_note_list")


@login_required
def credit_note_api_list(request):
    return _note_api(request, CreditNote)


def _sl_num(v):
    """Decimal-safe coercion for ledger math."""
    try:
        return Decimal(str(v)) if v is not None else Decimal("0")
    except Exception:
        return Decimal("0")


def _purchase_credit(obj, kind):
    """Invoice value owed to the supplier: the stored net_amount when set, else
    the purchase's own compute_net_amount() (same definition the purchase form
    persists), so the ledger matches the rest of the system."""
    if kind == "EP":  # hatchery Egg Purchase — net_amount() is a method
        return _sl_num(obj.net_amount())
    net = _sl_num(obj.net_amount)
    if net > 0:
        return net
    try:
        return _sl_num(obj.compute_net_amount())
    except Exception:
        return Decimal("0")


@login_required
def supplier_ledger_report(request):
    """Purchase > Reports > Supplier Ledger (Statement) — a supplier's running
    account: purchases (General + Chicks) are credits (payable rises), payments
    are debits (payable falls), carried forward from an opening/previous balance.
    One detail row per purchase item; payments are single rows."""
    from account.models import CompanyProfile
    from django.utils.dateparse import parse_date

    from hatchery.models import EggPurchase
    q2 = Decimal("0.01")
    supplier_id = (request.GET.get("supplier") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    export = (request.GET.get("export") or "").strip().lower()
    fd = parse_date(from_date) if from_date else None
    td = parse_date(to_date) if to_date else None

    supplier = Supplier.objects.filter(id=supplier_id).first() if supplier_id.isdigit() else None
    groups, totals = [], {"credit": Decimal("0"), "debit": Decimal("0")}
    prev_balance = Decimal("0")
    running = Decimal("0")
    purchase_count = payment_count = dn_count = cn_count = 0
    purchases_total = payments_total = dn_total = cn_total = Decimal("0")

    if supplier:
        # opening balance — payable (to pay) counts as credit (Cr)
        opening = _sl_num(supplier.opening_balance)
        if str(supplier.to_pay_to_receive or "").lower().startswith("receive"):
            opening = -opening

        gps = list(GeneralPurchase.objects.filter(supplier=supplier)
                   .prefetch_related("items__item", "items__farm_warehouse").order_by("date", "id"))
        cps = list(ChicksPurchase.objects.filter(supplier=supplier)
                   .select_related("item").prefetch_related("items__farm_warehouse").order_by("date", "id"))
        eps = list(EggPurchase.objects.filter(supplier=supplier)
                   .select_related("warehouse").prefetch_related("items__item").order_by("date", "id"))
        pay_lines = list(SupplierPaymentLine.objects.filter(supplier=supplier)
                         .select_related("payment", "pay_account").order_by("payment__date", "id"))
        dns = list(DebitNote.objects.filter(supplier=supplier).order_by("date", "id"))
        cns = list(CreditNote.objects.filter(supplier=supplier).order_by("date", "id"))

        # previous balance = opening + (purchases+debit notes) - (payments+credit
        # notes) strictly before the window
        prev_balance = opening
        for g in gps:
            if fd and g.date and g.date < fd:
                prev_balance += _purchase_credit(g, "GP")
        for ch in cps:
            if fd and ch.date and ch.date < fd:
                prev_balance += _purchase_credit(ch, "CP")
        for ep in eps:
            if fd and ep.date and ep.date < fd:
                prev_balance += _purchase_credit(ep, "EP")
        for dn in dns:
            if fd and dn.date and dn.date < fd:
                prev_balance += _sl_num(dn.amount)
        for pl in pay_lines:
            d = pl.payment.date
            if fd and d and d < fd:
                prev_balance -= _sl_num(pl.amount)
        for cn in cns:
            if fd and cn.date and cn.date < fd:
                prev_balance -= _sl_num(cn.amount)

        events = []
        for g in gps:
            if g.date and ((fd and g.date < fd) or (td and g.date > td)):
                continue
            events.append((g.date, 0, "GP", g))
        for ch in cps:
            if ch.date and ((fd and ch.date < fd) or (td and ch.date > td)):
                continue
            events.append((ch.date, 1, "CP", ch))
        for ep in eps:
            if ep.date and ((fd and ep.date < fd) or (td and ep.date > td)):
                continue
            events.append((ep.date, 1, "EP", ep))
        for pl in pay_lines:
            d = pl.payment.date
            if d and ((fd and d < fd) or (td and d > td)):
                continue
            events.append((d, 2, "PAY", pl))
        for dn in dns:
            if dn.date and ((fd and dn.date < fd) or (td and dn.date > td)):
                continue
            events.append((dn.date, 3, "DN", dn))
        for cn in cns:
            if cn.date and ((fd and cn.date < fd) or (td and cn.date > td)):
                continue
            events.append((cn.date, 4, "CN", cn))
        events.sort(key=lambda e: (e[0] or parse_date("1900-01-01"), e[1]))

        from collections import OrderedDict
        month_groups = OrderedDict()

        def _grp(d):
            key = d.strftime("%B %Y") if d else "Undated"
            if key not in month_groups:
                month_groups[key] = {"label": key, "rows": [], "debit": Decimal("0"),
                                     "credit": Decimal("0"), "closing": Decimal("0"), "vouchers": 0}
            return month_groups[key]

        running = prev_balance
        purchase_count = payment_count = dn_count = cn_count = 0
        for d, _o, kind, obj in events:
            grp = _grp(d)
            grp["vouchers"] += 1
            if kind in ("GP", "CP", "EP"):
                purchase_count += 1
                amt = _purchase_credit(obj, kind)
                running += amt                 # a purchase raises what we owe (Dr)
                grp["debit"] += amt
                type_label = {"GP": "Purchase Invoice", "CP": "Chicks Purchase", "EP": "Egg Purchase"}[kind]
                type_slug = {"GP": "purchase-invoice", "CP": "chicks-purchase", "EP": "egg-purchase"}[kind]
                # header fields differ across the three purchase sources
                if kind == "EP":
                    h_trnum, h_doc = obj.transaction_no, (obj.dc_no or "")
                    h_freight, h_tds = _sl_num(obj.freight_amount), _sl_num(obj.tcs_amount())
                    h_vehicle, h_boxes = (obj.vehicle or ""), None
                else:
                    h_trnum, h_doc = obj.purchase_no, (obj.bill_no or "")
                    h_freight, h_tds = _sl_num(obj.freight_amount), _sl_num(obj.tds_amount)
                    h_vehicle, h_boxes = (obj.vehicle_no or ""), _sl_num(obj.no_of_bags)
                items = list(obj.items.all()) or [None]
                for i, it in enumerate(items):
                    first = i == 0
                    if kind == "GP":
                        item_name = it.item.description if it and it.item_id else ""
                        gst_amt = (_sl_num(it.amount) * _sl_num(it.gst_percent) / 100).quantize(q2) if it else Decimal("0")
                        it_amount = _sl_num(it.amount) if it else ""
                        sector = it.farm_warehouse.name if it and it.farm_warehouse_id else ""
                        boxes = h_boxes if first else ""
                    elif kind == "CP":
                        item_name = obj.item.description if obj.item_id else "Chicks"
                        gst_amt = Decimal("0")
                        it_amount = _sl_num(it.amount) if it else ""
                        sector = it.farm_warehouse.name if it and it.farm_warehouse_id else ""
                        boxes = h_boxes if first else ""
                    else:  # EP — Egg Purchase
                        item_name = it.item.description if it and it.item_id else "Hatching Eggs"
                        gst_amt = Decimal("0")
                        it_amount = _sl_num(it.total_amount) if it else ""
                        sector = obj.warehouse.name if obj.warehouse_id else ""
                        boxes = _sl_num(it.no_of_boxes) if it else ""
                    grp["rows"].append({
                        "date": d if first else None,
                        "trnum": h_trnum if first else "",
                        "doc_no": h_doc if first else "",
                        "type": type_label if first else "", "type_slug": type_slug if first else "",
                        "item": item_name,
                        "boxes_bags": boxes,
                        "sent_qty": _sl_num(it.sent_qty) if it else "",
                        "rcv_qty": _sl_num(it.rcv_qty) if it else "",
                        "free_qty": _sl_num(it.free_qty) if it else "",
                        "rate": _sl_num(it.rate) if it else "",
                        "amount": it_amount,
                        "freight": h_freight.quantize(q2) if first else "",
                        "gst": gst_amt,
                        "tds": h_tds.quantize(q2) if first else "",
                        "debit": amt.quantize(q2) if first else "",
                        "credit": "",
                        "balance": abs(running).quantize(q2) if first else "",
                        "cr_dr": ("Dr" if running >= 0 else "Cr") if first else "",
                        "sector": sector,
                        "farm_code": "",
                        "remarks": (obj.remarks or "") if first else "",
                        "vehicle": h_vehicle if first else "",
                    })
                totals["debit"] += amt
                purchases_total += amt
            elif kind == "PAY":
                payment_count += 1
                amt = _sl_num(obj.amount)
                running -= amt                 # a payment lowers what we owe
                grp["credit"] += amt
                grp["rows"].append({
                    "date": d, "trnum": obj.payment.payment_no, "doc_no": obj.reference_no or "",
                    "type": "Payment", "type_slug": "payment", "item": obj.get_mode_display(), "boxes_bags": "",
                    "sent_qty": "", "rcv_qty": "", "free_qty": "", "rate": "", "amount": "", "freight": "", "gst": "", "tds": "",
                    "debit": "", "credit": amt.quantize(q2),
                    "balance": abs(running).quantize(q2), "cr_dr": "Dr" if running >= 0 else "Cr",
                    # Sector column shows the cash/bank account the payment came from.
                    "sector": (obj.pay_account.description if obj.pay_account_id else obj.mode) or "",
                    "farm_code": "", "remarks": obj.remarks or "", "vehicle": "",
                })
                totals["credit"] += amt
                payments_total += amt
            elif kind in ("DN", "CN"):
                amt = _sl_num(obj.amount)
                is_dn = kind == "DN"
                if is_dn:
                    dn_count += 1
                    running += amt             # debit note raises payable (Dr)
                    grp["debit"] += amt
                    totals["debit"] += amt
                    dn_total += amt
                else:
                    cn_count += 1
                    running -= amt             # credit note lowers payable
                    grp["credit"] += amt
                    totals["credit"] += amt
                    cn_total += amt
                grp["rows"].append({
                    "date": d, "trnum": obj.note_no, "doc_no": obj.against_bill or "",
                    "type": "Debit Note" if is_dn else "Credit Note",
                    "type_slug": "debit-note" if is_dn else "credit-note",
                    "item": obj.reason or "", "boxes_bags": "",
                    "sent_qty": "", "rcv_qty": "", "free_qty": "", "rate": "", "amount": "", "freight": "", "gst": "", "tds": "",
                    "debit": amt.quantize(q2) if is_dn else "",
                    "credit": "" if is_dn else amt.quantize(q2),
                    "balance": abs(running).quantize(q2), "cr_dr": "Dr" if running >= 0 else "Cr",
                    "sector": "", "farm_code": "", "remarks": obj.remarks or "", "vehicle": "",
                })
            grp["closing"] = running  # running after the latest event in the month

        for g in month_groups.values():
            groups.append({
                "label": g["label"], "count": g["vouchers"],
                "debit": g["debit"].quantize(q2), "credit": g["credit"].quantize(q2),
                "closing": abs(g["closing"]).quantize(q2),
                "cr_dr": "Dr" if g["closing"] >= 0 else "Cr",
                "rows": g["rows"],
            })

    kpi = {
        "opening": abs(prev_balance).quantize(q2), "opening_cr_dr": "Dr" if prev_balance >= 0 else "Cr",
        "purchases": purchases_total.quantize(q2), "purchase_count": purchase_count,
        "payments": payments_total.quantize(q2), "payment_count": payment_count,
        "debit_notes": dn_total.quantize(q2), "debit_note_count": dn_count,
        "credit_notes": cn_total.quantize(q2), "credit_note_count": cn_count,
        "closing": abs(running).quantize(q2), "closing_cr_dr": "Dr" if running >= 0 else "Cr",
    }
    company = CompanyProfile.get_solo()
    ctx = {
        "suppliers": Supplier.objects.order_by("name"),
        "supplier": supplier, "supplier_id": supplier_id,
        "from_date": from_date, "to_date": to_date,
        "groups": groups,
        "totals": {"credit": totals["credit"].quantize(q2), "debit": totals["debit"].quantize(q2)},
        "prev_balance": abs(prev_balance).quantize(q2), "prev_cr_dr": "Dr" if prev_balance >= 0 else "Cr",
        "closing": abs(running).quantize(q2), "closing_cr_dr": "Dr" if running >= 0 else "Cr",
        "kpi": kpi, "company": company,
    }
    if export == "excel" and supplier:
        return _ledger_excel(company, supplier, from_date, to_date, ctx["prev_balance"],
                             ctx["prev_cr_dr"], groups, ctx["totals"], ctx["closing"], ctx["closing_cr_dr"])
    return render(request, "supplier_ledger_report.html", ctx)


def _ledger_excel(company, supplier, from_date, to_date, prev_balance, prev_cr_dr,
                  groups, totals, closing, closing_cr_dr):
    """Stream the Supplier Ledger as an .xlsx workbook (openpyxl)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    headers = ["Date", "Transaction No.", "Transaction Type", "Purchase Bill No.", "Item",
               "Boxes/bags", "Sent Qty", "Received Qty", "Free", "Rate", "Amount", "Freight",
               "GST", "TDS", "Debit", "Credit", "Balance", "Sector", "Farm Code", "Vehicle", "Remarks"]

    def num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None if v in (None, "") else v

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Ledger"
    ws.append([company.name if company else ""])
    ws.append(["Supplier Statement Report"])
    ws.append([f"Supplier: {supplier.name}", "",
               f"Period: {from_date or '—'} to {to_date or '—'}"])
    ws.append([])
    ws.append(headers)
    hdr = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(hdr, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B3A6B")
        cell.alignment = Alignment(horizontal="center")

    ws.append(["Previous Balance"] + [""] * 15 + [f"{prev_balance} {prev_cr_dr}"] + [""] * 4)
    for g in groups:
        ws.append([f"{g['label']} ({g['count']} txn)"] + [""] * 13
                  + [num(g["debit"]), num(g["credit"]), f"{g['closing']} {g['cr_dr']}"] + [""] * 4)
        for r in g["rows"]:
            ws.append([
                r["date"].strftime("%d.%m.%Y") if r["date"] else "",
                r["trnum"], r["type"], r["doc_no"], r["item"],
                num(r["boxes_bags"]), num(r["sent_qty"]), num(r["rcv_qty"]), num(r["free_qty"]),
                num(r["rate"]), num(r["amount"]), num(r["freight"]), num(r["gst"]), num(r["tds"]),
                num(r["debit"]), num(r["credit"]),
                (f"{r['balance']} {r['cr_dr']}" if r["balance"] != "" else ""),
                r["sector"], r["farm_code"], r["vehicle"], r["remarks"],
            ])
    ws.append(["Grand Total"] + [""] * 13
              + [num(totals["debit"]), num(totals["credit"]), f"{closing} {closing_cr_dr}"] + [""] * 4)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="supplier_ledger_{supplier.code or supplier.id}.xlsx"'
    wb.save(resp)
    return resp


def _supplier_balance_row(sup, fd, td, ref_date):
    """One Supplier Balance row: opening (signed payable before the window),
    period Amount (purchases + debit notes) and Receipt (payments + credit
    notes), the between-days movement, the closing split into Credit (payable)
    / Debit (advance), and days since the last payment."""
    q2 = Decimal("0.01")
    signed = _sl_num(sup.opening_balance)
    if str(sup.to_pay_to_receive or "").lower().startswith("receive"):
        signed = -signed

    from hatchery.models import EggPurchase
    gps = list(GeneralPurchase.objects.filter(supplier=sup).prefetch_related("items"))
    cps = list(ChicksPurchase.objects.filter(supplier=sup).prefetch_related("items"))
    eps = list(EggPurchase.objects.filter(supplier=sup).prefetch_related("items"))
    pays = list(SupplierPaymentLine.objects.filter(supplier=sup).select_related("payment"))
    dns = list(DebitNote.objects.filter(supplier=sup))
    cns = list(CreditNote.objects.filter(supplier=sup))

    def before(d):
        return d and fd and d < fd

    def within(d):
        return d and (not fd or d >= fd) and (not td or d <= td)

    opening = signed
    amount = Decimal("0")   # payable-raising in period (purchases + debit notes)
    receipt = Decimal("0")  # payable-lowering in period (payments + credit notes)
    for g in gps:
        v = _purchase_credit(g, "GP")
        opening += v if before(g.date) else 0
        amount += v if within(g.date) else 0
    for ch in cps:
        v = _purchase_credit(ch, "CP")
        opening += v if before(ch.date) else 0
        amount += v if within(ch.date) else 0
    for ep in eps:
        v = _purchase_credit(ep, "EP")
        opening += v if before(ep.date) else 0
        amount += v if within(ep.date) else 0
    for dn in dns:
        v = _sl_num(dn.amount)
        opening += v if before(dn.date) else 0
        amount += v if within(dn.date) else 0
    for pl in pays:
        d = pl.payment.date
        v = _sl_num(pl.amount)
        opening -= v if before(d) else 0
        receipt += v if within(d) else 0
    for cn in cns:
        v = _sl_num(cn.amount)
        opening -= v if before(cn.date) else 0
        receipt += v if within(cn.date) else 0

    bw = amount - receipt
    closing = opening + bw

    pay_dates = [pl.payment.date for pl in pays if pl.payment.date and (not td or pl.payment.date <= td)]
    if pay_dates:
        gap = (ref_date - max(pay_dates)).days
    else:
        txn_dates = ([g.date for g in gps if g.date] + [ch.date for ch in cps if ch.date]
                     + [ep.date for ep in eps if ep.date])
        base = min(txn_dates) if txn_dates else sup.as_on_date
        gap = (ref_date - base).days if base else 0

    return {
        "id": sup.id, "name": sup.name,
        "opening": opening.quantize(q2), "amount": amount.quantize(q2),
        "receipt": receipt.quantize(q2), "bw": bw.quantize(q2),
        "credit": closing.quantize(q2) if closing >= 0 else Decimal("0.00"),
        "debit": (-closing).quantize(q2) if closing < 0 else Decimal("0.00"),
        "gap": max(gap, 0),
        "has_activity": bool(opening or amount or receipt or closing),
    }


@login_required
def supplier_balance_report(request):
    """Purchase > Reports > Supplier Balance — every supplier's payable position:
    opening, this period's Amount/Receipt movement, and closing Credit (payable)
    / Debit (advance), with the last-payment gap."""
    from account.models import CompanyProfile
    from django.utils.dateparse import parse_date

    q2 = Decimal("0.01")
    group = (request.GET.get("supplier_group") or "").strip()
    from_date = (request.GET.get("from_date") or "").strip()
    to_date = (request.GET.get("to_date") or "").strip()
    fd = parse_date(from_date) if from_date else None
    td = parse_date(to_date) if to_date else None
    ref_date = td or timezone.localdate()

    suppliers = Supplier.objects.order_by("name")
    if group:
        suppliers = suppliers.filter(supplier_group=group)

    rows = [_supplier_balance_row(s, fd, td, ref_date) for s in suppliers]

    tkeys = ["opening", "amount", "receipt", "bw", "credit", "debit"]
    totals = {k: sum((r[k] for r in rows), Decimal("0")).quantize(q2) for k in tkeys}

    groups = list(Supplier.objects.exclude(supplier_group__isnull=True)
                  .exclude(supplier_group="").values_list("supplier_group", flat=True)
                  .distinct().order_by("supplier_group"))

    return render(request, "supplier_balance_report.html", {
        "rows": rows, "totals": totals,
        "supplier_groups": groups, "group": group,
        "from_date": from_date, "to_date": to_date,
        "company": CompanyProfile.get_solo(),
    })


# --------------------------------------------------------------------------
# Purchase > Reports > Purchase Report
# --------------------------------------------------------------------------

# Columns rendered by the report, in order. Kept here so the page, the Excel
# export and the totals row can never drift apart.
PURCHASE_REPORT_COLUMNS = [
    ("date", "Date"), ("invoice", "Transaction No."),
    ("txn_type", "Transaction Type"), ("dc_no", "DC No."),
    ("supplier", "Supplier"), ("hsn", "HSN Code"), ("item_code", "Item Code"),
    ("item", "Item"), ("sent_qty", "Sent Qty (Bags/Kg)"),
    ("rcv_qty", "Received Qty (Bags/Qty)"), ("free_qty", "Free Qty"),
    ("rate", "Rate"), ("disc_percent", "Disc%"), ("disc_amount", "Disc Amount"),
    ("amount", "Amount"), ("gst_percent", "Gst%"), ("total_amount", "Total Amount"),
    ("warehouse", "Farm/Warehouse"), ("warehouse_code", "Farm/Warehouse Code"),
    ("farm_batch", "Farm Batch"), ("vehicle", "Vehicle No."),
    ("driver", "Driver Name"), ("remarks", "Remarks"),
    ("upload_status", "Upload Status"), ("added_by", "Added By"),
    ("added_time", "Added Time"),
]


def _pr_bags(item, qty):
    """Bag equivalent of a quantity for items that define a kg-per-bag."""
    kpb = getattr(item, "kg_per_bag", None) if item else None
    if kpb and Decimal(str(kpb)) > 0:
        return Decimal(str(qty or 0)) / Decimal(str(kpb))
    return Decimal("0")


def _pr_branch_by_sector():
    """{sector_id: branch_name} from the Sector -> Branch mapping, so a
    purchase's warehouse can be resolved to a branch."""
    from inventory.models import Mapping
    from broiler.models import Branch
    pairs = dict(Mapping.objects.filter(type=Mapping.TYPE_SECTOR_BRANCH)
                 .values_list("from_id", "to_id"))
    branches = {b.id: b.branch_name for b in Branch.objects.all()}
    return {sector_id: branches.get(branch_id, "")
            for sector_id, branch_id in pairs.items() if branch_id}


def _pr_upload_status(purchase):
    """Purchases carry up to three optional reference documents; the report
    reports whether any of them was actually attached."""
    for field in ("reference_document_1", "reference_document_2", "reference_document_3"):
        if getattr(purchase, field, None):
            return "Uploaded"
    return "Not Uploaded"


def _pr_added_by(model_names, object_ids):
    """{(model_name, object_id): actor} for the *create* audit entries of the
    given purchases — the purchase models carry no created_by of their own, so
    the audit trail is the record of who entered them."""
    from alerts.models import AuditLog
    rows = (AuditLog.objects
            .filter(model_name__in=model_names, action="create",
                    object_id__in=[str(i) for i in object_ids])
            .values_list("model_name", "object_id", "actor_label"))
    return {(m, str(o)): (a or "") for m, o, a in rows}


@login_required
def purchase_report(request):
    """Purchase > Reports > Purchase Report — one row per purchased item line
    across General and Chicks purchases, with supplier, document references,
    quantities, rate/value, the receiving farm/warehouse and who entered it."""
    from django.utils.dateparse import parse_date
    from account.models import CompanyProfile
    from broiler.models import Branch

    def g(key):
        return (request.GET.get(key) or "").strip()

    from_date, to_date = g("from_date"), g("to_date")
    supplier_id, category, item_id = g("supplier"), g("category"), g("item")
    branch_id, warehouse_id = g("branch"), g("warehouse")
    upload_status, export = g("upload_status"), g("export").lower()
    purchase_type = g("purchase_type")

    fd = parse_date(from_date) if from_date else None
    td = parse_date(to_date) if to_date else None
    branch_of_sector = _pr_branch_by_sector()
    branch_name_wanted = ""
    if branch_id:
        b = Branch.objects.filter(id=branch_id).first()
        branch_name_wanted = b.branch_name if b else ""

    def window(qs, field):
        if fd:
            qs = qs.filter(**{f"{field}__gte": fd})
        if td:
            qs = qs.filter(**{f"{field}__lte": td})
        return qs

    # ---- General purchases (feed, medicine, consumables) ----
    gp = window(GeneralPurchaseItem.objects.select_related(
        "purchase", "purchase__supplier", "item", "item__category",
        "farm_warehouse", "farm_warehouse__sector"), "purchase__date")
    if supplier_id:
        gp = gp.filter(purchase__supplier_id=supplier_id)
    if category:
        gp = gp.filter(item__category_id=category)
    if item_id:
        gp = gp.filter(item_id=item_id)
    if warehouse_id:
        gp = gp.filter(farm_warehouse_id=warehouse_id)
    gp = list(gp)

    # ---- Chicks purchases (the item lives on the header) ----
    cp = window(ChicksPurchaseItem.objects.select_related(
        "purchase", "purchase__supplier", "purchase__item",
        "purchase__item__category", "purchase__hatchery",
        "farm_warehouse", "farm_warehouse__sector"), "purchase__date")
    if supplier_id:
        cp = cp.filter(purchase__supplier_id=supplier_id)
    if category:
        cp = cp.filter(purchase__item__category_id=category)
    if item_id:
        cp = cp.filter(purchase__item_id=item_id)
    if warehouse_id:
        cp = cp.filter(farm_warehouse_id=warehouse_id)
    cp = list(cp)

    # ---- Egg purchases (hatchery module, same supplier ledger) ----
    from hatchery.models import EggPurchaseItem
    ep = window(EggPurchaseItem.objects.select_related(
        "egg_purchase", "egg_purchase__supplier", "egg_purchase__warehouse",
        "egg_purchase__warehouse__sector", "item", "item__category"),
        "egg_purchase__date")
    if supplier_id:
        ep = ep.filter(egg_purchase__supplier_id=supplier_id)
    if category:
        ep = ep.filter(item__category_id=category)
    if item_id:
        ep = ep.filter(item_id=item_id)
    if warehouse_id:
        ep = ep.filter(egg_purchase__warehouse_id=warehouse_id)
    ep = list(ep)

    # A purchase type filter short-circuits the sources it excludes.
    if purchase_type == "General Purchase":
        cp, ep = [], []
    elif purchase_type == "Chicks Purchase":
        gp, ep = [], []
    elif purchase_type == "Egg Purchase":
        gp, cp = [], []

    added_by = _pr_added_by(
        ["purchase.GeneralPurchase", "purchase.ChicksPurchase", "hatchery.EggPurchase"],
        [r.purchase_id for r in gp] + [r.purchase_id for r in cp]
        + [r.egg_purchase_id for r in ep])

    rows = []

    def build(line, purchase, item, model_name, txn_type, wh, amount, total_amount):
        """One report row. The three purchase models name their fields
        differently and value their lines differently, so the caller passes in
        the resolved warehouse and the pre-GST / final amounts."""
        sent = Decimal(str(line.sent_qty or 0))
        rcv = Decimal(str(line.rcv_qty or 0))
        return {
            "date": purchase.date,
            "invoice": (getattr(purchase, "purchase_no", None)
                        or getattr(purchase, "transaction_no", "")),
            "txn_type": txn_type,
            "dc_no": purchase.dc_no or getattr(purchase, "bill_no", "") or "",
            "supplier": purchase.supplier.name if purchase.supplier_id else "",
            "hsn": (item.hsn_code or "") if item else "",
            "item_code": item.item_code if item else "",
            "item": item.description if item else "",
            "sent_qty": sent, "sent_bags": _pr_bags(item, sent),
            "rcv_qty": rcv, "rcv_bags": _pr_bags(item, rcv),
            "free_qty": line.free_qty,
            "rate": Decimal(str(line.rate or 0)),
            "disc_percent": Decimal(str(getattr(line, "discount_percent", 0) or 0)),
            "disc_amount": Decimal(str(getattr(line, "discount_amount", 0) or 0)),
            "amount": amount,
            "gst_percent": Decimal(str(getattr(line, "gst_percent", 0) or 0)),
            "total_amount": total_amount,
            "warehouse": wh.name if wh else "",
            "warehouse_code": wh.code if wh else "",
            "farm_batch": getattr(line, "batch", "") or getattr(purchase, "batch_no", "") or "",
            "vehicle": (getattr(purchase, "vehicle_no", None)
                        or getattr(purchase, "vehicle", "") or ""),
            "driver": (getattr(purchase, "driver_name", None)
                       or getattr(purchase, "driver", "") or ""),
            "remarks": purchase.remarks,
            "upload_status": _pr_upload_status(purchase),
            "added_by": added_by.get((model_name, str(purchase.id)), ""),
            "added_time": purchase.created_at,
            "branch": branch_of_sector.get(wh.sector_id, "") if wh else "",
        }

    for r in gp:
        p = r.purchase
        # The stored amount already includes GST, so the pre-GST subtotal is
        # rebuilt exactly the way GeneralPurchaseItem.save() computes it.
        basis = getattr(p, "calculation_based_on", "Sent Quantity")
        qty = Decimal(str((r.rcv_qty if basis == "Received Quantity" else r.sent_qty) or 0))
        subtotal = ((qty * Decimal(str(r.rate or 0)))
                    * (1 - Decimal(str(r.discount_percent or 0)) / 100)
                    - Decimal(str(r.discount_amount or 0)))
        rows.append(build(r, p, r.item, "purchase.GeneralPurchase",
                          "General Purchase", r.farm_warehouse, subtotal, r.amount))

    for r in cp:
        # Chicks lines carry no discount or GST of their own.
        rows.append(build(r, r.purchase, r.purchase.item, "purchase.ChicksPurchase",
                          "Chicks Purchase", r.farm_warehouse, r.amount, r.amount))

    for r in ep:
        # Egg lines: `amount` is gross, `total_amount` is net of discount and
        # there is no GST, so both money columns show the net figure.
        rows.append(build(r, r.egg_purchase, r.item, "hatchery.EggPurchase",
                          "Egg Purchase", r.egg_purchase.warehouse,
                          r.total_amount, r.total_amount))

    if branch_name_wanted:
        rows = [r for r in rows if r["branch"] == branch_name_wanted]
    if upload_status:
        rows = [r for r in rows if r["upload_status"] == upload_status]

    rows.sort(key=lambda r: (r["date"] or parse_date("1900-01-01"), r["invoice"]))

    totals = {k: sum((Decimal(str(r[k] or 0)) for r in rows), Decimal("0"))
              for k in ("sent_qty", "sent_bags", "rcv_qty", "rcv_bags", "free_qty",
                        "disc_amount", "amount", "total_amount")}
    totals["rate"] = (totals["total_amount"] / totals["rcv_qty"]) if totals["rcv_qty"] else Decimal("0")

    # Headline figures for the KPI strip. Bills are counted per invoice, not
    # per line, so a multi-item bill counts once.
    bills = {r["invoice"] for r in rows if r["invoice"]}
    pending_bills = {r["invoice"] for r in rows if r["upload_status"] != "Uploaded"}
    kpi = {
        "bills": len(bills),
        "quantity": totals["rcv_qty"],
        "value": totals["total_amount"],
        "avg_rate": totals["rate"],
        "pending_upload": len(pending_bills),
        "suppliers": len({r["supplier"] for r in rows if r["supplier"]}),
    }

    criteria = "From: %s   To: %s   %d line(s)" % (
        from_date or "Beginning", to_date or "Date", len(rows))

    context = {
        "rows": rows, "totals": totals, "criteria": criteria, "kpi": kpi,
        "from_date": from_date, "to_date": to_date,
        "supplier_id": supplier_id, "category": category, "item_id": item_id,
        "branch_id": branch_id, "warehouse_id": warehouse_id,
        "upload_status": upload_status, "purchase_type": purchase_type,
        "suppliers": Supplier.objects.order_by("name"),
        "categories": ItemCategory.objects.order_by("name"),
        "items": Item.objects.order_by("description"),
        "branches": Branch.objects.order_by("branch_name"),
        "warehouses": Warehouse.objects.order_by("name"),
        "company": CompanyProfile.get_solo(),
    }
    if export == "excel":
        return _purchase_report_excel(context)
    return render(request, "purchase_report.html", context)


def _purchase_report_excel(ctx):
    """Stream the Purchase Report as an .xlsx workbook (openpyxl)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Report"
    bold = Font(bold=True)

    ws.append([ctx["company"].name if ctx["company"] else ""])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Purchase Report"])
    ws["A2"].font = bold
    ws.append([ctx["criteria"]])
    ws.append([])

    ws.append([label for _key, label in PURCHASE_REPORT_COLUMNS])
    for col in range(1, len(PURCHASE_REPORT_COLUMNS) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1B3A6B")
        cell.alignment = Alignment(horizontal="center")

    numeric = {"sent_qty", "rcv_qty", "free_qty", "rate", "disc_percent",
               "disc_amount", "amount", "gst_percent", "total_amount"}

    for r in ctx["rows"]:
        line = []
        for key, _label in PURCHASE_REPORT_COLUMNS:
            value = r.get(key)
            if key in numeric:
                line.append(float(value or 0))
            elif key == "date":
                line.append(value.strftime("%d.%m.%Y") if value else "")
            elif key == "added_time":
                line.append(value.strftime("%d.%m.%Y %I:%M %p") if value else "")
            elif key == "received_date":
                line.append(value.strftime("%d.%m.%Y") if value else "")
            else:
                line.append(value or "")
        ws.append(line)

    t = ctx["totals"]
    total_row = []
    for key, _label in PURCHASE_REPORT_COLUMNS:
        if key in t and key not in ("rate",):
            total_row.append(float(t[key]))
        elif key == "item":
            total_row.append("Total")
        else:
            total_row.append("")
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = bold

    widths = [11, 16, 14, 26, 11, 12, 22, 17, 19, 10, 10, 8, 12, 13, 8,
              14, 22, 18, 14, 14, 14, 24, 14, 12, 20]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="purchase_report.xlsx"'
    wb.save(response)
    return response
