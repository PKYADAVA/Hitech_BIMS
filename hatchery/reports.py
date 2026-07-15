"""Shared engine for the simple filterable hatchery reports (Egg Purchase,
Incubation, Dispatch & Sales). Mirrors the Hatch Register Report behavior:
on-screen table + totals footer, print header, and CSV/XLSX export driven by
the same GET filters.

Each report supplies: ``columns`` — a list of ``(key, label, kind)`` tuples
(kind in text/num/money/pct) — plus plain row dicts keyed by column key.
"""
import csv
from datetime import date
from decimal import Decimal

from django.http import HttpResponse
from django.shortcuts import render


def _export_value(kind, value):
    if kind == "money":
        return float(value or 0)
    if kind in ("num", "pct"):
        return value if value not in ("", None) else 0
    return value if value not in ("", None) else ""


def _format_cell(kind, value):
    if kind == "pct":
        return f"{value}%", "text-end"
    if kind == "money":
        return f"{Decimal(value or 0):,.2f}", "text-end"
    if kind == "num":
        return (value if value not in ("", None) else "-"), "text-end"
    return (value if value not in ("", None) else "-"), ""


def report_totals(columns, rows):
    sum_keys = [key for key, _, kind in columns if kind in ("num", "money")]
    totals = {key: Decimal("0") for key in sum_keys}
    for row in rows:
        for key in sum_keys:
            totals[key] += Decimal(str(row.get(key) or 0))
    return totals


def _display_rows(columns, rows):
    display = []
    for row in rows:
        cells = []
        for key, _, kind in columns:
            value, align = _format_cell(kind, row.get(key))
            cells.append({"value": value, "align": align})
        display.append(cells)
    return display


def _display_totals(columns, totals):
    first_key = columns[0][0]
    cells = []
    for key, _, kind in columns:
        if key == first_key:
            cells.append({"value": "TOTAL", "align": ""})
        elif key in totals:
            value, align = _format_cell(kind, totals[key])
            cells.append({"value": value, "align": align})
        else:
            cells.append({"value": "", "align": ""})
    return cells


def _csv_response(columns, rows, slug):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{slug}_{date.today().isoformat()}.csv"'
    writer = csv.writer(response)
    writer.writerow([label for _, label, _ in columns])
    for row in rows:
        writer.writerow([_export_value(kind, row.get(key)) for key, _, kind in columns])
    totals = report_totals(columns, rows)
    first_key = columns[0][0]
    writer.writerow([
        "TOTAL" if key == first_key
        else (_export_value(kind, totals[key]) if key in totals else "")
        for key, _, kind in columns
    ])
    return response


def _xlsx_response(columns, rows, title, slug, filter_line):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append([f"Hi Tech Farms — {title}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([filter_line or "All records"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append([label for _, label, _ in columns])
    for cell in ws[header_row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F3B73", end_color="1F3B73", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([_export_value(kind, row.get(key)) for key, _, kind in columns])

    totals = report_totals(columns, rows)
    first_key = columns[0][0]
    total_row = []
    for key, _, kind in columns:
        if key == first_key:
            total_row.append("TOTAL")
        elif key in totals:
            total_row.append(_export_value(kind, totals[key]))
        else:
            total_row.append("")
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col_idx, (_, label, _kind) in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=header_row_idx, column=col_idx).column_letter].width = max(12, len(label) + 3)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{slug}_{date.today().isoformat()}.xlsx"'
    wb.save(response)
    return response


def respond_report(request, *, title, icon, subtitle, active_tab, slug, columns, rows,
                   filters, extra_filters=None):
    """Render the shared report page, or return an export response when the
    ``export`` GET param is csv/xlsx. ``filters`` must contain from_date/to_date;
    ``extra_filters`` is a list of {name, label, options, selected} select boxes."""
    export = request.GET.get("export", "").strip().lower()
    filter_bits = []
    if filters.get("from_date") or filters.get("to_date"):
        filter_bits.append(f"Date: {filters.get('from_date') or '…'} to {filters.get('to_date') or '…'}")
    for f in (extra_filters or []):
        if f["selected"]:
            filter_bits.append(f"{f['label']}: {f['selected']}")
    filter_line = " | ".join(filter_bits)

    if export == "csv":
        return _csv_response(columns, rows, slug)
    if export == "xlsx":
        return _xlsx_response(columns, rows, title, slug, filter_line)

    totals = report_totals(columns, rows)
    return render(request, "hatchery_generic_report.html", {
        "title": title, "icon": icon, "subtitle": subtitle, "active_tab": active_tab,
        "columns": columns,
        "display_rows": _display_rows(columns, rows),
        "display_totals": _display_totals(columns, totals),
        "row_count": len(rows),
        "filters": filters,
        "extra_filters": extra_filters or [],
        "filter_line": filter_line,
    })
