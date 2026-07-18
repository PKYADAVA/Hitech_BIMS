import csv
from datetime import date
from decimal import Decimal
from typing import Optional
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.http import Http404, HttpResponse, JsonResponse
from django.core.exceptions import ValidationError
from django.db import transaction
import json
import logging

from account.models import ChartOfAccount, CompanyProfile, TermsConditions
from inventory.models import Item, Warehouse
from purchase.models import Supplier
from hatchery_master.models import Hatchery, Setter, Hatcher, STATES_AND_TERRITORIES
from hr.models import Employee
from sales.models import Customer, CustomerShippingAddress

from .models import (
    HatchSetting, HatchEggIntake, HatchHatcherOutput, HatchSalesLine,
    EggPurchase, EggPurchaseItem, EggGrading, EggGradingHatchItem,
    DeliveryChallan, DeliveryChallanItem, TraySetting, TraySettingLine,
    HatchEntry, HatchEntryVaccine, ChickSale, ChickSaleItem, ChangeRequest,
)

logger = logging.getLogger(__name__)


class BaseAPIView(View):
    """Base class for API views with common error handling."""

    def handle_exception(self, e: Exception) -> JsonResponse:
        logger.error(f"Error in {self.__class__.__name__}: {str(e)}", exc_info=True)
        if isinstance(e, Http404):
            return JsonResponse({"error": str(e)}, status=404)
        if isinstance(e, ValidationError):
            return JsonResponse({"error": str(e)}, status=400)
        return JsonResponse({"error": "Internal server error"}, status=500)


@method_decorator(login_required, name="dispatch")
class HatchSettingListTemplateView(View):
    """Renders the hatch register list page."""

    def get(self, request):
        return render(request, "hatchery_list.html")


@method_decorator(login_required, name="dispatch")
class HatchSettingFormTemplateView(View):
    """Renders the full hatch register form page for add/edit."""

    def get(self, request, id: Optional[int] = None, request_mode: bool = False):
        context = {
            "hatch_setting_id": id,
            "request_mode": request_mode,
            "suppliers": Supplier.objects.all(),
            "setter_nos": (
                Setter.objects.filter(is_active=True)
                .values_list("setter_no", flat=True).distinct().order_by("setter_no")
            ),
            "hatcher_nos": (
                Hatcher.objects.filter(is_active=True)
                .values_list("hatcher_no", flat=True).distinct().order_by("hatcher_no")
            ),
            "next_batch_flock_no": HatchSetting._next_batch_flock_no() if id is None else None,
        }
        return render(request, "hatchery_form.html", context)


@login_required
def next_batch_flock_no(request):
    return JsonResponse({"batch_flock_no": HatchSetting._next_batch_flock_no()})


def _egg_intake_to_dict(row: HatchEggIntake) -> dict:
    return {
        "id": row.id,
        "sub_lot_flock": row.sub_lot_flock,
        "setter_no": row.setter_no,
        "no_trays": row.no_trays,
        "tray_size": row.tray_size,
        "total_eggs": row.total_eggs,
    }


def _hatcher_output_to_dict(row: HatchHatcherOutput) -> dict:
    return {
        "id": row.id,
        "hatcher_no": row.hatcher_no,
        "infertile_qty": row.infertile_qty,
        "early_dead_qty": row.early_dead_qty,
        "blasting_qty": row.blasting_qty,
        "transfer_qty": row.transfer_qty,
        "dead_in_shell_qty": row.dead_in_shell_qty,
        "culls_malf_qty": row.culls_malf_qty,
        "saleable_chicks": row.saleable_chicks,
    }


def _sales_line_to_dict(row: HatchSalesLine) -> dict:
    return {
        "id": row.id,
        "trader_customer_name": row.trader_customer_name,
        "chicks_sold": row.chicks_sold,
        "discount_percent": str(row.discount_percent),
        "free_chicks": row.free_chicks,
        "billed_chicks": row.billed_chicks,
        "rate": str(row.rate),
        "total_amount": str(row.total_amount),
        "payment_status": row.payment_status,
        "delivery_notes": row.delivery_notes,
    }


def _hatch_setting_to_dict(hs: HatchSetting) -> dict:
    return {
        "id": hs.id,
        "setting_no": hs.setting_no,
        "batch_flock_no": hs.batch_flock_no,
        "supplier_name": hs.supplier_name,
        "primary_machine_nos": hs.primary_machine_nos,
        "avg_egg_weight": hs.avg_egg_weight,
        "received_date": hs.received_date.isoformat() if hs.received_date else None,
        "received_time": hs.received_time.strftime("%H:%M") if hs.received_time else None,
        "setting_date": hs.setting_date.isoformat() if hs.setting_date else None,
        "transfer_date": hs.transfer_date.isoformat() if hs.transfer_date else None,
        "hatch_date": hs.hatch_date.isoformat() if hs.hatch_date else None,
        "push_time": hs.push_time.strftime("%H:%M") if hs.push_time else None,
        "received_qty": hs.received_qty,
        "breakage_qty": hs.breakage_qty,
        "crack_qty": hs.crack_qty,
        "setting_qty": hs.setting_qty,
        "breakage_percent": hs.breakage_percent(),
        "crack_percent": hs.crack_percent(),
        "setter_temperature": hs.setter_temperature,
        "setter_humidity": hs.setter_humidity,
        "hatcher_temperature": hs.hatcher_temperature,
        "hatcher_humidity": hs.hatcher_humidity,
        "avg_chick_weight": hs.avg_chick_weight,
        "medicine_vaccine": hs.medicine_vaccine,
        "packing_boxes_used": hs.packing_boxes_used,
        "remarks": hs.remarks,
        "prepared_by": hs.prepared_by,
        "verified_by": hs.verified_by,
        "egg_intakes": [_egg_intake_to_dict(r) for r in hs.egg_intakes.all()],
        "hatcher_outputs": [_hatcher_output_to_dict(r) for r in hs.hatcher_outputs.all()],
        "sales_lines": [_sales_line_to_dict(r) for r in hs.sales_lines.all()],
        "total_setting_qty": hs.total_setting_qty(),
        "total_saleable_chicks": hs.total_saleable_chicks(),
        "hatch_percent": hs.hatch_percent(),
        "total_chicks_sold": hs.total_chicks_sold(),
        "unsold_chicks": hs.unsold_chicks(),
        "total_sales_amount": str(hs.total_sales_amount()),
    }


@method_decorator(login_required, name="dispatch")
class HatchSettingAPI(BaseAPIView):
    """API endpoints for HatchSetting (hatch register) records, including nested rows."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                hs = get_object_or_404(
                    HatchSetting.objects.prefetch_related(
                        "egg_intakes", "hatcher_outputs", "sales_lines"
                    ),
                    id=id,
                )
                return JsonResponse(_hatch_setting_to_dict(hs))

            hatch_settings = HatchSetting.objects.all()
            results = []
            for hs in hatch_settings:
                results.append({
                    "id": hs.id,
                    "setting_no": hs.setting_no,
                    "batch_flock_no": hs.batch_flock_no,
                    "supplier_name": hs.supplier_name,
                    "setting_date": hs.setting_date.isoformat() if hs.setting_date else None,
                    "transfer_date": hs.transfer_date.isoformat() if hs.transfer_date else None,
                    "hatch_date": hs.hatch_date.isoformat() if hs.hatch_date else None,
                    "total_setting_qty": hs.total_setting_qty(),
                    "total_saleable_chicks": hs.total_saleable_chicks(),
                    "hatch_percent": hs.hatch_percent(),
                    "total_chicks_sold": hs.total_chicks_sold(),
                    "unsold_chicks": hs.unsold_chicks(),
                })
            return JsonResponse(results, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body.decode("utf-8"))
            hs = self._save_hatch_setting(data)
            return JsonResponse({"message": "Hatch setting created", "id": hs.id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id: int) -> JsonResponse:
        try:
            data = json.loads(request.body.decode("utf-8"))
            hs = self._save_hatch_setting(data, hatch_setting_id=id)
            return JsonResponse({"message": "Hatch setting updated", "id": hs.id})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            hs = get_object_or_404(HatchSetting, id=id)
            hs.delete()
            return JsonResponse({"message": "Hatch setting deleted"})
        except Exception as e:
            return self.handle_exception(e)

    def _save_hatch_setting(self, data: dict, hatch_setting_id: Optional[int] = None) -> HatchSetting:
        header_fields = [
            "setting_no", "supplier_name", "primary_machine_nos", "avg_egg_weight",
            "received_date", "received_time", "setting_date", "transfer_date",
            "hatch_date", "push_time", "received_qty", "breakage_qty",
            "crack_qty", "setting_qty", "setter_temperature", "setter_humidity",
            "hatcher_temperature", "hatcher_humidity", "avg_chick_weight",
            "medicine_vaccine", "packing_boxes_used", "remarks",
            "prepared_by", "verified_by",
        ]
        header_data = {f: data[f] for f in header_fields if f in data and data[f] not in ("", None)}

        if hatch_setting_id:
            hs = get_object_or_404(HatchSetting, id=hatch_setting_id)
            for field, value in header_data.items():
                setattr(hs, field, value)
            hs.full_clean(exclude=["batch_flock_no"])
            hs.save()
            hs.egg_intakes.all().delete()
            hs.hatcher_outputs.all().delete()
            hs.sales_lines.all().delete()
        else:
            hs = HatchSetting(**header_data)
            hs.full_clean(exclude=["batch_flock_no"])
            hs.save()

        for row in data.get("egg_intakes", []):
            HatchEggIntake.objects.create(
                hatch_setting=hs,
                sub_lot_flock=row.get("sub_lot_flock", ""),
                setter_no=row["setter_no"],
                no_trays=row.get("no_trays") or None,
                tray_size=row.get("tray_size") or None,
                total_eggs=row.get("total_eggs") or None,
            )

        for row in data.get("hatcher_outputs", []):
            HatchHatcherOutput.objects.create(
                hatch_setting=hs,
                hatcher_no=row["hatcher_no"],
                infertile_qty=row.get("infertile_qty") or 0,
                early_dead_qty=row.get("early_dead_qty") or 0,
                blasting_qty=row.get("blasting_qty") or 0,
                transfer_qty=row.get("transfer_qty") or 0,
                dead_in_shell_qty=row.get("dead_in_shell_qty") or 0,
                culls_malf_qty=row.get("culls_malf_qty") or 0,
                saleable_chicks=row.get("saleable_chicks") or 0,
            )

        for row in data.get("sales_lines", []):
            HatchSalesLine.objects.create(
                hatch_setting=hs,
                trader_customer_name=row["trader_customer_name"],
                chicks_sold=row.get("chicks_sold") or 0,
                discount_percent=row.get("discount_percent") or 0,
                free_chicks=row.get("free_chicks") or 0,
                billed_chicks=row.get("billed_chicks") or 0,
                rate=row.get("rate") or 0,
                total_amount=row.get("total_amount") or 0,
                payment_status=row.get("payment_status") or "unpaid",
                delivery_notes=row.get("delivery_notes", ""),
            )

        return hs


@method_decorator(login_required, name="dispatch")
class EggPurchaseListTemplateView(View):
    """Renders the egg purchase list page."""

    def get(self, request):
        return render(request, "egg_purchase_list.html")


@method_decorator(login_required, name="dispatch")
class EggPurchaseFormTemplateView(View):
    """Renders the egg purchase add/edit page."""

    def get(self, request, id: Optional[int] = None, request_mode: bool = False):
        context = {
            "egg_purchase_id": id,
            "request_mode": request_mode,
            "suppliers": Supplier.objects.all().order_by("name"),
            "warehouses": Warehouse.objects.all().order_by("name"),
            "items": Item.objects.all().order_by("item_code"),
            "pay_accounts": ChartOfAccount.objects.filter(status="Active").order_by("code"),
            "next_transaction_no": EggPurchase._next_transaction_no() if id is None else None,
        }
        return render(request, "egg_purchase_form.html", context)


@login_required
def egg_purchase_next_number(request):
    return JsonResponse({"transaction_no": EggPurchase._next_transaction_no()})


def _egg_purchase_item_to_dict(row: EggPurchaseItem) -> dict:
    return {
        "id": row.id,
        "item": row.item_id,
        "item_code": row.item.item_code,
        "item_description": row.item.description,
        "sent_qty": str(row.sent_qty),
        "rcv_qty": str(row.rcv_qty),
        "free_qty": str(row.free_qty),
        "no_of_boxes": str(row.no_of_boxes),
        "amount": str(row.amount),
        "discount_percent": str(row.discount_percent),
        "discount_amount": str(row.discount_amount),
        "total_amount": str(row.total_amount),
        "rate": str(row.rate),
    }


def _egg_purchase_to_dict(ep: EggPurchase) -> dict:
    return {
        "id": ep.id,
        "transaction_no": ep.transaction_no,
        "date": ep.date.isoformat() if ep.date else None,
        "supplier": ep.supplier_id,
        "supplier_name": ep.supplier.name,
        "warehouse": ep.warehouse_id,
        "warehouse_name": ep.warehouse.name,
        "dc_no": ep.dc_no,
        "vehicle": ep.vehicle,
        "driver": ep.driver,
        "freight_type": ep.freight_type,
        "payment_mode": ep.payment_mode,
        "pay_account": ep.pay_account_id,
        "freight_account": ep.freight_account_id,
        "freight_amount": str(ep.freight_amount),
        "tcs_applicable": ep.tcs_applicable,
        "tcs_percent": str(ep.tcs_percent),
        "tcs_amount": str(ep.tcs_amount()),
        "remarks": ep.remarks,
        "items": [_egg_purchase_item_to_dict(row) for row in ep.items.all()],
        "gross_amount": str(ep.gross_amount()),
        "net_quantity": str(ep.net_quantity()),
        "net_rate": str(ep.net_rate()),
        "net_amount": str(ep.net_amount()),
        "supplier_amount": str(ep.supplier_amount()),
    }


@method_decorator(login_required, name="dispatch")
class EggPurchaseAPI(BaseAPIView):
    """API endpoints for EggPurchase records, including nested item rows."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                ep = get_object_or_404(
                    EggPurchase.objects.select_related("supplier", "warehouse").prefetch_related("items__item"),
                    id=id,
                )
                return JsonResponse(_egg_purchase_to_dict(ep))

            queryset = EggPurchase.objects.select_related("supplier", "warehouse").prefetch_related("items__item")
            supplier_id = request.GET.get("supplier")
            if supplier_id:
                queryset = queryset.filter(supplier_id=supplier_id)

            results = []
            for ep in queryset:
                item_names = ", ".join(row.item.item_code for row in ep.items.all())
                results.append({
                    "id": ep.id,
                    "transaction_no": ep.transaction_no,
                    "date": ep.date.isoformat() if ep.date else None,
                    "dc_no": ep.dc_no,
                    "supplier": ep.supplier_id,
                    "supplier_name": ep.supplier.name,
                    "item_names": item_names,
                    "items": [{"item": row.item_id, "item_code": row.item.item_code, "rcv_qty": str(row.rcv_qty)} for row in ep.items.all()],
                    "net_quantity": str(ep.net_quantity()),
                    "net_rate": str(ep.net_rate()),
                    "net_amount": str(ep.net_amount()),
                    "warehouse_name": ep.warehouse.name,
                })
            return JsonResponse(results, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body.decode("utf-8"))
            ep = self._save_egg_purchase(data)
            from account.services.auto_posting import post_document
            post_document(ep, user=request.user)
            return JsonResponse({"message": "Egg purchase created", "id": ep.id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id: int) -> JsonResponse:
        try:
            data = json.loads(request.body.decode("utf-8"))
            ep = self._save_egg_purchase(data, egg_purchase_id=id)
            from account.services.auto_posting import post_document
            post_document(ep, user=request.user)
            return JsonResponse({"message": "Egg purchase updated", "id": ep.id})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            ep = get_object_or_404(EggPurchase, id=id)
            ep.delete()
            return JsonResponse({"message": "Egg purchase deleted"})
        except Exception as e:
            return self.handle_exception(e)

    def _save_egg_purchase(self, data: dict, egg_purchase_id: Optional[int] = None) -> EggPurchase:
        header_fields = [
            "date", "dc_no", "vehicle", "driver", "freight_type", "payment_mode",
            "freight_amount", "tcs_applicable", "tcs_percent", "remarks",
        ]
        header_data = {f: data[f] for f in header_fields if f in data and data[f] not in ("", None)}
        header_data["tcs_applicable"] = bool(data.get("tcs_applicable"))

        header_data["supplier"] = get_object_or_404(Supplier, id=data["supplier"])
        header_data["warehouse"] = get_object_or_404(Warehouse, id=data["warehouse"])
        header_data["pay_account"] = get_object_or_404(ChartOfAccount, id=data["pay_account"])
        if data.get("freight_account"):
            header_data["freight_account"] = get_object_or_404(ChartOfAccount, id=data["freight_account"])
        else:
            header_data["freight_account"] = None

        if egg_purchase_id:
            ep = get_object_or_404(EggPurchase, id=egg_purchase_id)
            for field, value in header_data.items():
                setattr(ep, field, value)
            ep.full_clean(exclude=["transaction_no"])
            ep.save()
            ep.items.all().delete()
        else:
            ep = EggPurchase(**header_data)
            ep.full_clean(exclude=["transaction_no"])
            ep.save()

        for row in data.get("items", []):
            EggPurchaseItem.objects.create(
                egg_purchase=ep,
                item=get_object_or_404(Item, id=row["item"]),
                sent_qty=Decimal(str(row.get("sent_qty") or 0)),
                rcv_qty=Decimal(str(row.get("rcv_qty") or 0)),
                free_qty=Decimal(str(row.get("free_qty") or 0)),
                no_of_boxes=Decimal(str(row.get("no_of_boxes") or 0)),
                amount=Decimal(str(row.get("amount") or 0)),
                discount_percent=Decimal(str(row.get("discount_percent") or 0)),
                discount_amount=Decimal(str(row.get("discount_amount") or 0)),
            )

        return ep


@method_decorator(login_required, name="dispatch")
class EggGradingListTemplateView(View):
    """Renders the egg grading list page."""

    def get(self, request):
        return render(request, "egg_grading_list.html")


@method_decorator(login_required, name="dispatch")
class EggGradingFormTemplateView(View):
    """Renders the egg grading add/edit page."""

    def get(self, request, id: Optional[int] = None, request_mode: bool = False):
        context = {
            "egg_grading_id": id,
            "request_mode": request_mode,
            "suppliers": Supplier.objects.all().order_by("name"),
            "storage_locations": Warehouse.objects.all().order_by("name"),
            "items": Item.objects.all().order_by("item_code"),
            "next_transaction_no": EggGrading._next_transaction_no() if id is None else None,
        }
        return render(request, "egg_grading_form.html", context)


@login_required
def egg_grading_next_number(request):
    return JsonResponse({"transaction_no": EggGrading._next_transaction_no()})


@login_required
def egg_grading_stock_check(request):
    purchase_invoice = request.GET.get("purchase_invoice")
    item = request.GET.get("item")
    exclude_id = request.GET.get("exclude_id")
    if not purchase_invoice or not item:
        return JsonResponse({"error": "purchase_invoice and item are required"}, status=400)
    stock = EggGrading.available_stock(purchase_invoice, item, exclude_id=exclude_id)
    return JsonResponse({"stock": str(stock)})


def _egg_grading_hatch_item_to_dict(row: EggGradingHatchItem) -> dict:
    return {
        "id": row.id,
        "hatch_item": row.hatch_item_id,
        "hatch_item_code": row.hatch_item.item_code,
        "quantity": str(row.quantity),
    }


def _egg_grading_to_dict(eg: EggGrading) -> dict:
    return {
        "id": eg.id,
        "transaction_no": eg.transaction_no,
        "date": eg.date.isoformat() if eg.date else None,
        "storage_location": eg.storage_location_id,
        "storage_location_name": eg.storage_location.name,
        "supplier": eg.supplier_id,
        "supplier_name": eg.supplier.name,
        "purchase_invoice": eg.purchase_invoice_id,
        "purchase_invoice_no": eg.purchase_invoice.transaction_no,
        "item": eg.item_id,
        "item_code": eg.item.item_code,
        "quantity": str(eg.quantity),
        "broken_eggs": str(eg.broken_eggs),
        "damage_eggs": str(eg.damage_eggs),
        "misshapped_eggs": str(eg.misshapped_eggs),
        "dirty_eggs": str(eg.dirty_eggs),
        "total_rejections": str(eg.total_rejections()),
        "eggs_to_stock": str(eg.eggs_to_stock()),
        "hatch_items": [_egg_grading_hatch_item_to_dict(row) for row in eg.hatch_items.all()],
        "total_hatch_qty": str(eg.total_hatch_qty()),
    }


@method_decorator(login_required, name="dispatch")
class EggGradingAPI(BaseAPIView):
    """API endpoints for EggGrading records, including nested hatch item rows."""

    def get(self, request, id: Optional[int] = None) -> JsonResponse:
        try:
            if id:
                eg = get_object_or_404(
                    EggGrading.objects.select_related(
                        "storage_location", "supplier", "purchase_invoice", "item"
                    ).prefetch_related("hatch_items__hatch_item"),
                    id=id,
                )
                return JsonResponse(_egg_grading_to_dict(eg))

            results = []
            for eg in EggGrading.objects.select_related(
                "storage_location", "supplier", "purchase_invoice", "item"
            ).prefetch_related("hatch_items__hatch_item"):
                hatch_egg_summary = ", ".join(
                    f"{row.hatch_item.item_code} ({row.quantity})" for row in eg.hatch_items.all()
                )
                results.append({
                    "id": eg.id,
                    "transaction_no": eg.transaction_no,
                    "date": eg.date.isoformat() if eg.date else None,
                    "item_code": eg.item.item_code,
                    "quantity": str(eg.quantity),
                    "hatch_egg_summary": hatch_egg_summary,
                    "storage_location_name": eg.storage_location.name,
                })
            return JsonResponse(results, safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request) -> JsonResponse:
        try:
            data = json.loads(request.body.decode("utf-8"))
            eg = self._save_egg_grading(data)
            return JsonResponse({"message": "Egg grading created", "id": eg.id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id: int) -> JsonResponse:
        try:
            data = json.loads(request.body.decode("utf-8"))
            eg = self._save_egg_grading(data, egg_grading_id=id)
            return JsonResponse({"message": "Egg grading updated", "id": eg.id})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id: int) -> JsonResponse:
        try:
            eg = get_object_or_404(EggGrading, id=id)
            eg.delete()
            return JsonResponse({"message": "Egg grading deleted"})
        except Exception as e:
            return self.handle_exception(e)

    def _save_egg_grading(self, data: dict, egg_grading_id: Optional[int] = None) -> EggGrading:
        header_fields = ["date", "quantity", "broken_eggs", "damage_eggs", "misshapped_eggs", "dirty_eggs"]
        header_data = {f: data[f] for f in header_fields if f in data and data[f] not in ("", None)}

        header_data["storage_location"] = get_object_or_404(Warehouse, id=data["storage_location"])
        header_data["supplier"] = get_object_or_404(Supplier, id=data["supplier"])
        header_data["purchase_invoice"] = get_object_or_404(EggPurchase, id=data["purchase_invoice"])
        header_data["item"] = get_object_or_404(Item, id=data["item"])

        if egg_grading_id:
            eg = get_object_or_404(EggGrading, id=egg_grading_id)
            for field, value in header_data.items():
                setattr(eg, field, value)
            eg.full_clean(exclude=["transaction_no"])
            eg.save()
            eg.hatch_items.all().delete()
        else:
            eg = EggGrading(**header_data)
            eg.full_clean(exclude=["transaction_no"])
            eg.save()

        for row in data.get("hatch_items", []):
            if not row.get("hatch_item"):
                continue
            EggGradingHatchItem.objects.create(
                egg_grading=eg,
                hatch_item=get_object_or_404(Item, id=row["hatch_item"]),
                quantity=Decimal(str(row.get("quantity") or 0)),
            )

        return eg


# ---------------------------------------------------------------------------
# Delivery Challans
@method_decorator(login_required, name="dispatch")
class DeliveryChallanListTemplateView(View):
    def get(self, request):
        return render(request, "delivery_challan_list.html")


@method_decorator(login_required, name="dispatch")
class DeliveryChallanFormTemplateView(View):
    def get(self, request, id: Optional[int] = None, request_mode: bool = False):
        return render(request, "delivery_challan_form.html", {
            "delivery_challan_id": id,
            "request_mode": request_mode,
            "customers": Customer.objects.filter(contact_type__in=["Customer", "Supplier & Customer"]).order_by("name"),
            "items": Item.objects.all().order_by("item_code"),
            "next_challan_no": DeliveryChallan.next_challan_no() if id is None else None,
            "states_and_union_territories": STATES_AND_TERRITORIES,
            "company": CompanyProfile.get_solo(),
            "terms_conditions": TermsConditions.objects.filter(party_type=TermsConditions.PartyType.CUSTOMER).exclude(condition__isnull=True).exclude(condition__exact="").order_by("type"),
        })


def _delivery_challan_to_dict(challan):
    company = CompanyProfile.get_solo()
    return {
        "id": challan.id, "challan_no": challan.challan_no, "date": challan.date.isoformat(),
        "place_of_supply": challan.place_of_supply, "overall_discount": str(challan.overall_discount or 0),
        "customer": challan.customer_id, "customer_name": challan.customer.name,
        "customer_address": challan.customer.address, "customer_mobile": challan.customer.mobile,
        "customer_gstin": challan.customer.gstin or "", "customer_email": challan.customer.email or "",
        "shipping_address": challan.shipping_address, "transport_mode": challan.transport_mode,
        "vehicle_no": challan.vehicle_no, "driver_name": challan.driver_name,
        "driver_mobile": challan.driver_mobile, "transporter_name": challan.transporter_name,
        "transport_document_no": challan.transport_document_no,
        "transport_document_date": challan.transport_document_date.isoformat() if challan.transport_document_date else "",
        "eway_bill_no": challan.eway_bill_no,
        "eway_bill_date": challan.eway_bill_date.isoformat() if challan.eway_bill_date else "",
        "print_price_details": challan.print_price_details, "terms": challan.terms,
        "total_quantity": str(challan.total_quantity()), "total_units": str(challan.total_units()),
        "total_amount": str(challan.total_amount()), "total_tax": str(challan.total_tax()),
        "grand_total": str(challan.grand_total()),
        "company": {
            "name": company.name, "address": company.address, "state": company.state,
            "mobile": company.mobile, "email": company.email, "gstin": company.gstin, "pan": company.pan,
            "bank_name": company.bank_name, "bank_account_no": company.bank_account_no,
            "ifsc_code": company.ifsc_code, "bank_branch": company.bank_branch,
        },
        "items": [{
            "id": row.id, "item": row.item_id, "item_code": row.item.item_code,
            "item_description": row.item.description, "hsn_code": row.item.hsn_code or "",
            "packing_size": str(row.packing_size), "units": str(row.units),
            "quantity": str(row.quantity), "unit": row.unit, "price": str(row.price),
            "discount_percent": str(row.discount_percent), "tax_percent": str(row.tax_percent),
            "amount": str(row.amount),
        } for row in challan.items.select_related("item")],
    }


@method_decorator(login_required, name="dispatch")
class DeliveryChallanAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None):
        try:
            if id:
                return JsonResponse(_delivery_challan_to_dict(get_object_or_404(DeliveryChallan.objects.select_related("customer"), id=id)))
            challans = DeliveryChallan.objects.select_related("customer").prefetch_related("items", "chick_sales").all()
            return JsonResponse([{
                "id": c.id, "challan_no": c.challan_no, "date": c.date.isoformat(),
                "customer_name": c.customer.name, "vehicle_no": c.vehicle_no,
                "total_quantity": str(c.total_quantity()), "total_amount": str(c.total_amount()),
                "converted_bill_no": next((cs.bill_no for cs in c.chick_sales.all()), ""),
            } for c in challans], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request):
        try:
            return JsonResponse({"id": self._save(json.loads(request.body)).id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id):
        try:
            return JsonResponse({"id": self._save(json.loads(request.body), id).id})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id):
        try:
            get_object_or_404(DeliveryChallan, id=id).delete()
            return JsonResponse({"message": "Delivery challan deleted"})
        except Exception as e:
            return self.handle_exception(e)

    def _save(self, data, challan_id=None):
        fields = ["date", "place_of_supply", "shipping_address", "transport_mode",
                  "vehicle_no", "driver_name", "driver_mobile", "transporter_name", "transport_document_no",
                  "transport_document_date", "eway_bill_no", "eway_bill_date", "terms"]
        values = {field: data.get(field) or (None if field.endswith("_date") else "") for field in fields}
        values["customer"] = get_object_or_404(Customer, id=data["customer"])
        values["print_price_details"] = bool(data.get("print_price_details", True))
        values["overall_discount"] = Decimal(str(data.get("overall_discount") or 0))
        if challan_id:
            challan = get_object_or_404(DeliveryChallan, id=challan_id)
            for field, value in values.items(): setattr(challan, field, value)
            challan.full_clean(exclude=["challan_no"])
            challan.save(); challan.items.all().delete()
        else:
            challan = DeliveryChallan(**values)
            challan.full_clean(exclude=["challan_no"])
            challan.save()
        for row in data.get("items", []):
            DeliveryChallanItem.objects.create(
                challan=challan, item=get_object_or_404(Item, id=row["item"]),
                packing_size=Decimal(str(row.get("packing_size") or 1)), units=Decimal(str(row.get("units") or 1)),
                quantity=Decimal(str(row.get("quantity") or 0)), unit=row.get("unit", ""),
                price=Decimal(str(row.get("price") or 0)), discount_percent=Decimal(str(row.get("discount_percent") or 0)),
                tax_percent=Decimal(str(row.get("tax_percent") or 0)),
            )
        return challan


# Hatchery Sheet Report
# ---------------------------------------------------------------------------

# Report columns as (row-key, display label, kind). `kind` drives formatting,
# alignment and totals behaviour, and is shared by the on-screen table, the
# Excel sheet and the CSV so all three stay in sync.
#   text  - left aligned, no total
#   date  - left aligned, no total
#   num   - right aligned integer, summed in the totals row
#   pct   - right aligned percentage (value + "%"), recomputed in totals
#   money - right aligned 2dp amount, summed in the totals row
HATCH_REPORT_COLUMNS = [
    # Identity
    ("setting_no", "Setting No", "text"),
    ("batch_flock_no", "Batch/Flock No", "text"),
    ("supplier_name", "Supplier", "text"),
    ("primary_machine_nos", "Machine Nos", "text"),
    ("avg_egg_weight", "Avg Egg Wt", "text"),
    ("received_date", "Received Date", "date"),
    ("received_time", "Received Time", "text"),
    ("setting_date", "Setting Date", "date"),
    ("transfer_date", "Transfer Date", "date"),
    ("hatch_date", "Hatch Date", "date"),
    ("push_time", "Push Time", "text"),
    # Egg intake
    ("received_qty", "Received Qty", "num"),
    ("breakage_qty", "Damage", "num"),
    ("breakage_pct", "Damage %", "pct"),
    ("crack_qty", "Broken", "num"),
    ("crack_pct", "Broken %", "pct"),
    ("setting_qty", "Setting Qty", "num"),
    # Candling & hatch output
    ("infertile_qty", "Infertile", "num"),
    ("infertile_pct", "Infertile %", "pct"),
    ("early_dead_qty", "Early Dead", "num"),
    ("early_dead_pct", "Early Dead %", "pct"),
    ("blasting_qty", "Blasting", "num"),
    ("blasting_pct", "Blasting %", "pct"),
    ("transfer_qty", "Transfer Qty", "num"),
    ("transfer_pct", "Transfer %", "pct"),
    ("dead_in_shell_qty", "Dead-In-Shell", "num"),
    ("culls_malf_qty", "Culls / Malf.", "num"),
    ("saleable_chicks", "Saleable Chicks", "num"),
    ("hatch_pct", "Hatch %", "pct"),
    # Environmental & consumables
    ("avg_chick_weight", "Avg Chick Wt", "text"),
    ("medicine_vaccine", "Medicine / Vaccine", "text"),
    ("packing_boxes_used", "Packing Boxes", "num"),
    # Sales
    ("chicks_sold", "Chicks Sold", "num"),
    ("unsold_chicks", "Unsold", "num"),
    ("sales_amount", "Sales Amount", "money"),
]

# Percentage columns and the (numerator, denominator) qty keys they derive
# from. Used for both per-row and totals-row percentages.
HATCH_REPORT_PCT_SOURCES = {
    "breakage_pct": ("breakage_qty", "received_qty"),
    "crack_pct": ("crack_qty", "received_qty"),
    "infertile_pct": ("infertile_qty", "setting_qty"),
    "early_dead_pct": ("early_dead_qty", "setting_qty"),
    "blasting_pct": ("blasting_qty", "setting_qty"),
    "transfer_pct": ("transfer_qty", "setting_qty"),
    "hatch_pct": ("saleable_chicks", "setting_qty"),
}

# Column keys summed into the totals footer.
HATCH_REPORT_SUM_KEYS = [key for key, _, kind in HATCH_REPORT_COLUMNS if kind in ("num", "money")]


def _pct(numerator, denominator):
    return round(numerator / denominator * 100, 2) if denominator else 0


def _build_hatch_report_rows(queryset):
    """Turn prefetched HatchSetting records into flat report row dicts.

    Aggregates are computed in Python from the prefetched child rows to avoid
    a per-record DB round trip (unlike the model's own helper methods).
    """
    rows = []
    for hs in queryset:
        outputs = list(hs.hatcher_outputs.all())
        sales = list(hs.sales_lines.all())

        infertile = sum(o.infertile_qty for o in outputs)
        early_dead = sum(o.early_dead_qty for o in outputs)
        blasting = sum(o.blasting_qty for o in outputs)
        transfer = sum(o.transfer_qty for o in outputs)
        dead_in_shell = sum(o.dead_in_shell_qty for o in outputs)
        culls = sum(o.culls_malf_qty for o in outputs)
        saleable = sum(o.saleable_chicks for o in outputs)
        sold = sum(s.chicks_sold for s in sales)
        sales_amount = sum((s.total_amount for s in sales), Decimal("0"))

        row = {
            "id": hs.id,
            "setting_no": hs.setting_no,
            "batch_flock_no": hs.batch_flock_no or "",
            "supplier_name": hs.supplier_name,
            "primary_machine_nos": hs.primary_machine_nos or "",
            "avg_egg_weight": hs.avg_egg_weight or "",
            "received_date": hs.received_date.isoformat() if hs.received_date else "",
            "received_time": hs.received_time.strftime("%H:%M") if hs.received_time else "",
            "setting_date": hs.setting_date.isoformat() if hs.setting_date else "",
            "transfer_date": hs.transfer_date.isoformat() if hs.transfer_date else "",
            "hatch_date": hs.hatch_date.isoformat() if hs.hatch_date else "",
            "push_time": hs.push_time.strftime("%H:%M") if hs.push_time else "",
            "received_qty": hs.received_qty,
            "breakage_qty": hs.breakage_qty,
            "crack_qty": hs.crack_qty,
            "setting_qty": hs.setting_qty,
            "infertile_qty": infertile,
            "early_dead_qty": early_dead,
            "blasting_qty": blasting,
            "transfer_qty": transfer,
            "dead_in_shell_qty": dead_in_shell,
            "culls_malf_qty": culls,
            "saleable_chicks": saleable,
            "setter_temperature": hs.setter_temperature or "",
            "setter_humidity": hs.setter_humidity or "",
            "hatcher_temperature": hs.hatcher_temperature or "",
            "hatcher_humidity": hs.hatcher_humidity or "",
            "avg_chick_weight": hs.avg_chick_weight or "",
            "medicine_vaccine": hs.medicine_vaccine or "",
            "packing_boxes_used": hs.packing_boxes_used or 0,
            "chicks_sold": sold,
            "unsold_chicks": saleable - sold,
            "sales_amount": sales_amount,
        }
        for pct_key, (num_key, den_key) in HATCH_REPORT_PCT_SOURCES.items():
            row[pct_key] = _pct(row[num_key], row[den_key])
        rows.append(row)
    return rows


def _hatch_report_totals(rows):
    totals = {key: 0 for key in HATCH_REPORT_SUM_KEYS}
    totals["sales_amount"] = Decimal("0")
    for row in rows:
        for key in HATCH_REPORT_SUM_KEYS:
            totals[key] += row[key] or 0
    for pct_key, (num_key, den_key) in HATCH_REPORT_PCT_SOURCES.items():
        totals[pct_key] = _pct(totals.get(num_key, 0), totals.get(den_key, 0))
    return totals


def _filter_hatch_report_queryset(request):
    """Applies the report bar filters (date range on setting_date, supplier,
    payment status) and returns (queryset, applied_filters_dict)."""
    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()
    supplier = request.GET.get("supplier", "").strip()
    payment_status = request.GET.get("payment_status", "").strip()

    qs = HatchSetting.objects.prefetch_related("hatcher_outputs", "sales_lines")
    if from_date:
        qs = qs.filter(setting_date__gte=from_date)
    if to_date:
        qs = qs.filter(setting_date__lte=to_date)
    if supplier:
        qs = qs.filter(supplier_name=supplier)
    if payment_status:
        qs = qs.filter(sales_lines__payment_status=payment_status).distinct()
    qs = qs.order_by("setting_date", "id")

    return qs, {
        "from_date": from_date,
        "to_date": to_date,
        "supplier": supplier,
        "payment_status": payment_status,
    }


def _report_export_value(kind, value):
    """Coerce a cell to a spreadsheet/CSV-friendly value (numbers stay numeric)."""
    if kind == "money":
        return float(value or 0)
    if kind in ("num", "pct"):
        return value if value not in ("", None) else 0
    return value if value not in ("", None) else ""


def _hatch_report_csv_response(rows):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="hatchery_report_{date.today().isoformat()}.csv"'
    writer = csv.writer(response)
    writer.writerow([label for _, label, _ in HATCH_REPORT_COLUMNS])
    for row in rows:
        writer.writerow([_report_export_value(kind, row[key]) for key, _, kind in HATCH_REPORT_COLUMNS])
    totals = _hatch_report_totals(rows)
    writer.writerow([
        "TOTAL" if key == "setting_no"
        else (_report_export_value(kind, totals[key]) if key in totals else "")
        for key, _, kind in HATCH_REPORT_COLUMNS
    ])
    return response


def _hatch_report_xlsx_response(rows, filters):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Hatchery Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F3B73", end_color="1F3B73", fill_type="solid")
    title_font = Font(bold=True, size=14)

    ws.append(["Hi Tech Farms — Hatch Register Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HATCH_REPORT_COLUMNS))
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    range_bits = []
    if filters["from_date"] or filters["to_date"]:
        range_bits.append(f"Setting Date: {filters['from_date'] or '…'} to {filters['to_date'] or '…'}")
    if filters["supplier"]:
        range_bits.append(f"Supplier: {filters['supplier']}")
    if filters["payment_status"]:
        range_bits.append(f"Payment: {filters['payment_status'].title()}")
    ws.append([" | ".join(range_bits) if range_bits else "All records"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HATCH_REPORT_COLUMNS))
    ws.append([])

    header_row_idx = ws.max_row + 1
    ws.append([label for _, label, _ in HATCH_REPORT_COLUMNS])
    for cell in ws[header_row_idx]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([_report_export_value(kind, row[key]) for key, _, kind in HATCH_REPORT_COLUMNS])

    totals = _hatch_report_totals(rows)
    total_row = []
    for key, _, kind in HATCH_REPORT_COLUMNS:
        if key == "setting_no":
            total_row.append("TOTAL")
        elif key in totals:
            total_row.append(_report_export_value(kind, totals[key]))
        else:
            total_row.append("")
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for col_idx, (_, label, _kind) in enumerate(HATCH_REPORT_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=header_row_idx, column=col_idx).column_letter].width = max(12, len(label) + 3)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="hatchery_report_{date.today().isoformat()}.xlsx"'
    wb.save(response)
    return response


def _format_hatch_report_cell(kind, value):
    """Returns (display_value, css_align_class) for one report cell so the
    template can render rows without needing variable-key dict lookups."""
    if kind == "pct":
        return f"{value}%", "text-end"
    if kind == "money":
        return f"{Decimal(value or 0):,.2f}", "text-end"
    if kind == "num":
        return (value if value not in ("", None) else "-"), "text-end"
    return (value if value not in ("", None) else "-"), ""


def _hatch_report_display_rows(rows):
    """Each row -> ordered list of {value, align} cells matching HATCH_REPORT_COLUMNS."""
    display = []
    for row in rows:
        cells = []
        for key, _, kind in HATCH_REPORT_COLUMNS:
            value, align = _format_hatch_report_cell(kind, row[key])
            cells.append({"value": value, "align": align})
        display.append(cells)
    return display


def _hatch_report_display_totals(totals):
    cells = []
    for key, _, kind in HATCH_REPORT_COLUMNS:
        if key == "setting_no":
            cells.append({"value": "TOTAL", "align": ""})
        elif key in totals:
            value, align = _format_hatch_report_cell(kind, totals[key])
            cells.append({"value": value, "align": align})
        else:
            cells.append({"value": "", "align": ""})
    return cells


@method_decorator(login_required, name="dispatch")
class HatchReportView(View):
    """Filterable hatchery sheet report with on-screen table, print and
    Excel/CSV export (all driven by the same GET filters)."""

    def get(self, request):
        qs, filters = _filter_hatch_report_queryset(request)
        rows = _build_hatch_report_rows(qs)

        export = request.GET.get("export", "").strip().lower()
        if export == "csv":
            return _hatch_report_csv_response(rows)
        if export == "xlsx":
            return _hatch_report_xlsx_response(rows, filters)

        totals = _hatch_report_totals(rows)
        context = {
            "columns": HATCH_REPORT_COLUMNS,
            "display_rows": _hatch_report_display_rows(rows),
            "display_totals": _hatch_report_display_totals(totals),
            "row_count": len(rows),
            "filters": filters,
            "suppliers": (
                HatchSetting.objects.exclude(supplier_name="")
                .values_list("supplier_name", flat=True).distinct().order_by("supplier_name")
            ),
            "payment_statuses": HatchSetting.PAYMENT_STATUS_CHOICES,
            "has_filters": any(filters.values()),
        }
        return render(request, "hatchery_report.html", context)


@method_decorator(login_required, name="dispatch")
class TraySetListTemplateView(View):
    def get(self, request):
        return render(request, "tray_set_list.html")


@method_decorator(login_required, name="dispatch")
class TraySetFormTemplateView(View):
    def get(self, request, id: Optional[int] = None, request_mode: bool = False):
        return render(request, "tray_set_form.html", {
            "tray_setting_id": id,
            "request_mode": request_mode,
            "hatcheries": Hatchery.objects.filter(is_active=True).order_by("hatchery_name"),
            "employees": Employee.objects.all().order_by("full_name"),
            "gradings": EggGrading.objects.all().order_by("-date", "-id"),
            "suppliers": Supplier.objects.all().order_by("name"),
            "setters": Setter.objects.filter(is_active=True).select_related("hatchery").order_by("setter_no"),
            "items": Item.objects.all().order_by("item_code"),
            "next_setting_no": TraySetting.next_setting_no() if id is None else None,
        })


def _tray_setting_to_dict(ts):
    return {
        "id": ts.id, "setting_no": ts.setting_no,
        "hatchery": ts.hatchery_id, "hatchery_name": ts.hatchery.hatchery_name,
        "setting_date": ts.setting_date.isoformat(),
        "transfer_date": ts.transfer_date.isoformat() if ts.transfer_date else "",
        "hatch_date": ts.hatch_date.isoformat(),
        "setting_time": ts.setting_time.strftime("%H:%M") if ts.setting_time else "",
        "loaded_by": ts.loaded_by,
        "loaded_by_name": ts.loaded_by,
        "grading": ts.grading_id, "grading_no": ts.grading.transaction_no,
        "total_eggs": str(ts.total_eggs()), "total_broken": str(ts.total_broken()),
        "total_damaged": str(ts.total_damaged()), "total_eggs_set": str(ts.total_eggs_set()),
        "total_expected_chicks": str(ts.total_expected_chicks()),
        "supplier_names": ts.supplier_names(),
        "lines": [{
            "id": line.id, "supplier": line.supplier_id,
            "supplier_name": line.supplier.name if line.supplier else "",
            "setter": line.setter_id, "setter_no": line.setter.setter_no,
            "item": line.item_id,
            "total_eggs": str(line.total_eggs), "broken": str(line.broken),
            "damaged": str(line.damaged), "eggs_set": str(line.eggs_set),
            "avg_weight": str(line.avg_weight), "expected_chicks": line.expected_chicks,
        } for line in ts.lines.select_related("supplier", "setter")],
    }


@method_decorator(login_required, name="dispatch")
class TraySettingAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None):
        try:
            if id:
                ts = get_object_or_404(
                    TraySetting.objects.select_related("hatchery", "grading"), id=id)
                return JsonResponse(_tray_setting_to_dict(ts))
            settings = TraySetting.objects.select_related(
                "hatchery", "grading").prefetch_related("lines__supplier").all()
            return JsonResponse([{
                "id": ts.id, "setting_no": ts.setting_no,
                "setting_date": ts.setting_date.isoformat(), "hatch_date": ts.hatch_date.isoformat(),
                "hatchery_name": ts.hatchery.hatchery_name, "supplier_names": ts.supplier_names(),
                "grading_no": ts.grading.transaction_no,
                "total_eggs_set": str(ts.total_eggs_set()), "total_broken": str(ts.total_broken()),
                "total_damaged": str(ts.total_damaged()), "total_eggs": str(ts.total_eggs()),
                "setting_time": ts.setting_time.strftime("%H:%M") if ts.setting_time else "",
                "loaded_by_name": ts.loaded_by,
                "total_expected_chicks": str(ts.total_expected_chicks()),
            } for ts in settings], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request):
        try:
            return JsonResponse({"id": self._save(json.loads(request.body)).id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id):
        try:
            return JsonResponse({"id": self._save(json.loads(request.body), id).id})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id):
        try:
            get_object_or_404(TraySetting, id=id).delete()
            return JsonResponse({"message": "Tray setting deleted"})
        except Exception as e:
            return self.handle_exception(e)

    def _save(self, data, tray_setting_id=None):
        values = {
            "hatchery": get_object_or_404(Hatchery, id=data["hatchery"]),
            "grading": get_object_or_404(EggGrading, id=data["grading"]),
            "setting_date": data["setting_date"],
            "transfer_date": data.get("transfer_date") or None,
            "hatch_date": data.get("hatch_date") or None,
            "setting_time": data.get("setting_time") or None,
            "loaded_by": (data.get("loaded_by") or "").strip(),
        }
        from datetime import timedelta
        if not values["transfer_date"]:
            values["transfer_date"] = date.fromisoformat(values["setting_date"]) + timedelta(days=18)
        if not values["hatch_date"]:
            values["hatch_date"] = date.fromisoformat(values["setting_date"]) + timedelta(days=21)
        if tray_setting_id:
            ts = get_object_or_404(TraySetting, id=tray_setting_id)
            for field, value in values.items():
                setattr(ts, field, value)
            ts.full_clean(exclude=["setting_no"])
            ts.save()
            ts.lines.all().delete()
        else:
            ts = TraySetting(**values)
            ts.full_clean(exclude=["setting_no"])
            ts.save()
        for row in data.get("lines", []):
            if not row.get("setter"):
                continue
            TraySettingLine.objects.create(
                tray_setting=ts,
                supplier=Supplier.objects.filter(id=row.get("supplier") or 0).first(),
                setter=get_object_or_404(Setter, id=row["setter"]),
                item=Item.objects.filter(id=row.get("item") or 0).first(),
                total_eggs=Decimal(str(row.get("total_eggs") or 0)),
                broken=Decimal(str(row.get("broken") or 0)),
                damaged=Decimal(str(row.get("damaged") or 0)),
                eggs_set=Decimal(str(row.get("eggs_set") or 0)),
                avg_weight=Decimal(str(row.get("avg_weight") or 0)),
                expected_chicks=int(float(row.get("expected_chicks") or 0)),
            )
        if not ts.lines.exists():
            raise ValidationError("Add at least one setter line.")
        return ts


@method_decorator(login_required, name="dispatch")
class HatchEntryListTemplateView(View):
    def get(self, request):
        return render(request, "hatch_entry_list.html")


@method_decorator(login_required, name="dispatch")
class HatchEntryFormTemplateView(View):
    def get(self, request, id: Optional[int] = None, request_mode: bool = False):
        used = HatchEntry.objects.exclude(id=id or 0).values_list("tray_setting_id", flat=True)
        tray_settings = TraySetting.objects.exclude(id__in=used).select_related("hatchery", "grading__purchase_invoice")
        return render(request, "hatch_entry_form.html", {
            "hatch_entry_id": id,
            "request_mode": request_mode,
            "hatcheries": Hatchery.objects.filter(is_active=True).order_by("hatchery_name"),
            "tray_settings": json.dumps([{
                "id": ts.id, "setting_no": ts.setting_no, "hatchery": ts.hatchery_id,
                "setting_date": ts.setting_date.isoformat(), "hatch_date": ts.hatch_date.isoformat(),
                "eggs_set": str(ts.total_eggs_set()), "expected_chicks": str(ts.total_expected_chicks()),
                "purchase_qty": str(ts.grading.purchase_invoice.net_quantity()),
                "purchase_rate": str(ts.grading.purchase_invoice.net_rate()),
                "purchase_amount": str(ts.grading.purchase_invoice.net_amount()),
            } for ts in tray_settings]),
            "items": Item.objects.all().order_by("item_code"),
            "next_transaction_no": HatchEntry.next_transaction_no() if id is None else None,
        })


def _hatch_entry_to_dict(he):
    ts = he.tray_setting
    return {
        "id": he.id, "transaction_no": he.transaction_no,
        "tray_setting": ts.id, "setting_no": ts.setting_no,
        "hatchery": ts.hatchery_id, "hatchery_name": ts.hatchery.hatchery_name,
        "setting_date": ts.setting_date.isoformat(),
        "hatch_date": (he.hatch_date or ts.hatch_date).isoformat(),
        "eggs_total": str(he.eggs_total), "egg_rate": str(he.egg_rate), "eggs_amount": str(he.eggs_amount),
        "chicks_total": str(he.chicks_total), "chick_rate": str(he.chick_rate), "chicks_amount": str(he.chicks_amount),
        "net_rate": str(he.net_rate), "net_amount": str(he.net_amount),
        "remarks": he.remarks,
        "vaccines": [{
            "id": v.id, "item": v.item_id, "item_label": f"{v.item.item_code} - {v.item.description}",
            "quantity": str(v.quantity), "rate": str(v.rate), "amount": str(v.amount),
        } for v in he.vaccines.select_related("item")],
    }


@method_decorator(login_required, name="dispatch")
class HatchEntryAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None):
        try:
            if id:
                he = get_object_or_404(
                    HatchEntry.objects.select_related("tray_setting__hatchery"), id=id)
                return JsonResponse(_hatch_entry_to_dict(he))
            entries = HatchEntry.objects.select_related("tray_setting__hatchery").all()
            return JsonResponse([{
                "id": he.id, "transaction_no": he.transaction_no,
                "setting_date": he.tray_setting.setting_date.isoformat(),
                "hatch_date": (he.hatch_date or he.tray_setting.hatch_date).isoformat(),
                "tse_no": he.tray_setting_id, "setting_no": he.tray_setting.setting_no,
                "hatchery_name": he.tray_setting.hatchery.hatchery_name,
                "description": he.remarks, "eggs_total": str(he.eggs_total),
                "chicks_total": str(he.chicks_total), "net_amount": str(he.net_amount),
            } for he in entries], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request):
        try:
            return JsonResponse({"id": self._save(json.loads(request.body)).id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id):
        try:
            return JsonResponse({"id": self._save(json.loads(request.body), id).id})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id):
        try:
            get_object_or_404(HatchEntry, id=id).delete()
            return JsonResponse({"message": "Hatch entry deleted"})
        except Exception as e:
            return self.handle_exception(e)

    def _save(self, data, hatch_entry_id=None):
        ts = get_object_or_404(TraySetting, id=data["tray_setting"])
        clash = HatchEntry.objects.filter(tray_setting=ts).exclude(id=hatch_entry_id or 0)
        if clash.exists():
            raise ValidationError("A hatch entry already exists for this setting number.")
        values = {
            "tray_setting": ts,
            "hatch_date": data.get("hatch_date") or ts.hatch_date,
            "chicks_total": Decimal(str(data.get("chicks_total") or 0)),
            "remarks": data.get("remarks") or "",
        }
        if hatch_entry_id:
            he = get_object_or_404(HatchEntry, id=hatch_entry_id)
            for field, value in values.items():
                setattr(he, field, value)
        else:
            he = HatchEntry(**values)
        he.apply_purchase_snapshot()
        he.full_clean(exclude=["transaction_no"])
        he.save()
        if hatch_entry_id:
            he.vaccines.all().delete()
        for row in data.get("vaccines", []):
            if not row.get("item"):
                continue
            HatchEntryVaccine.objects.create(
                hatch_entry=he, item=get_object_or_404(Item, id=row["item"]),
                quantity=Decimal(str(row.get("quantity") or 0)),
                rate=Decimal(str(row.get("rate") or 0)),
            )
        he.recalculate_net()
        return he


@method_decorator(login_required, name="dispatch")
class ChickSaleListTemplateView(View):
    def get(self, request):
        return render(request, "chick_sale_list.html")


@method_decorator(login_required, name="dispatch")
class ChickSaleFormTemplateView(View):
    def get(self, request, id: Optional[int] = None, request_mode: bool = False):
        from_challan = None
        challan_id = request.GET.get("from_challan")
        if id is None and challan_id:
            challan = DeliveryChallan.objects.filter(id=challan_id).select_related("customer").first()
            if challan and challan.chick_sales.exists():
                challan = None  # already converted; conversion re-opens only if that sale is deleted
            if challan:
                from_challan = {
                    "id": challan.id, "challan_no": challan.challan_no,
                    "customer": challan.customer_id, "shipping_address": challan.shipping_address,
                    "vehicle": challan.vehicle_no, "driver": challan.driver_name,
                    "terms": challan.terms,
                    "items": [{
                        "item": row.item_id, "total_qty": str(row.quantity),
                        "sale_rate": str(row.price),
                    } for row in challan.items.all()],
                }
        return render(request, "chick_sale_form.html", {
            "chick_sale_id": id,
            "request_mode": request_mode,
            "customers": Customer.objects.filter(contact_type__in=["Customer", "Supplier & Customer"]).order_by("name"),
            "warehouses": Warehouse.objects.all().order_by("name"),
            "items": Item.objects.all().order_by("item_code"),
            "accounts": ChartOfAccount.objects.all().order_by("code"),
            "states_and_union_territories": STATES_AND_TERRITORIES,
            "next_bill_no": ChickSale.next_bill_no() if id is None else None,
            "from_challan": json.dumps(from_challan),
            "terms_conditions": TermsConditions.objects.filter(party_type=TermsConditions.PartyType.CUSTOMER).exclude(condition__isnull=True).exclude(condition__exact="").order_by("type"),
        })


def _chick_sale_to_dict(cs):
    company = CompanyProfile.get_solo()
    return {
        "id": cs.id, "bill_no": cs.bill_no, "date": cs.date.isoformat(),
        "customer": cs.customer_id, "customer_name": cs.customer.name,
        "customer_address": cs.customer.address, "customer_mobile": cs.customer.mobile,
        "customer_gstin": cs.customer.gstin or "", "customer_email": cs.customer.email or "",
        "warehouse": cs.warehouse_id, "warehouse_name": cs.warehouse.name,
        "delivery_challan": cs.delivery_challan_id,
        "dc_no": cs.delivery_challan.challan_no if cs.delivery_challan else "",
        "shipping_address": cs.shipping_address, "vehicle": cs.vehicle, "driver": cs.driver,
        "freight_type": cs.freight_type, "payment_mode": cs.payment_mode,
        "pay_account": cs.pay_account_id, "freight_account": cs.freight_account_id,
        "freight_amount": str(cs.freight_amount), "final_amount": str(cs.final_amount),
        "avg_amount": str(cs.avg_amount), "profit_amount": str(cs.profit_amount),
        "terms": cs.terms, "remarks": cs.remarks,
        "company": {
            "name": company.name, "address": company.address, "state": company.state,
            "mobile": company.mobile, "email": company.email, "gstin": company.gstin, "pan": company.pan,
            "bank_name": company.bank_name, "bank_account_no": company.bank_account_no,
            "ifsc_code": company.ifsc_code, "bank_branch": company.bank_branch,
        },
        "items": [{
            "id": row.id, "item": row.item_id, "item_code": row.item.item_code,
            "item_description": row.item.description, "farm": row.farm,
            "total_qty": str(row.total_qty), "mortality": str(row.mortality),
            "culls": str(row.culls), "sale_qty": str(row.sale_qty),
            "discount_percent": str(row.discount_percent), "discount_amount": str(row.discount_amount),
            "free_qty": str(row.free_qty),
            "net_qty": str(row.net_qty), "sale_rate": str(row.sale_rate), "amount": str(row.amount),
        } for row in cs.items.select_related("item")],
    }


@method_decorator(login_required, name="dispatch")
class ChickSaleAPI(BaseAPIView):
    def get(self, request, id: Optional[int] = None):
        try:
            if id:
                cs = get_object_or_404(
                    ChickSale.objects.select_related("customer", "warehouse", "delivery_challan"), id=id)
                return JsonResponse(_chick_sale_to_dict(cs))
            sales = ChickSale.objects.select_related(
                "customer", "warehouse", "delivery_challan").prefetch_related("items__item").all()
            return JsonResponse([{
                "id": cs.id, "bill_no": cs.bill_no, "date": cs.date.isoformat(),
                "dc_no": cs.delivery_challan.challan_no if cs.delivery_challan else "",
                "customer_name": cs.customer.name, "item_names": cs.item_names(),
                "total_birds": str(cs.total_birds()), "total_net_qty": str(cs.total_net_qty()),
                "total_free_qty": str(cs.total_free_qty()), "avg_amount": str(cs.avg_amount),
                "final_amount": str(cs.final_amount), "warehouse_name": cs.warehouse.name,
            } for cs in sales], safe=False)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def post(self, request):
        try:
            cs = self._save(json.loads(request.body))
            from account.services.auto_posting import post_document
            post_document(cs, user=request.user)
            return JsonResponse({"id": cs.id}, status=201)
        except Exception as e:
            return self.handle_exception(e)

    @transaction.atomic
    def put(self, request, id):
        try:
            cs = self._save(json.loads(request.body), id)
            from account.services.auto_posting import post_document
            post_document(cs, user=request.user)
            return JsonResponse({"id": cs.id})
        except Exception as e:
            return self.handle_exception(e)

    def delete(self, request, id):
        try:
            get_object_or_404(ChickSale, id=id).delete()
            return JsonResponse({"message": "Chick sale deleted"})
        except Exception as e:
            return self.handle_exception(e)

    def _save(self, data, chick_sale_id=None):
        challan = DeliveryChallan.objects.filter(id=data.get("delivery_challan") or 0).first()
        if challan and challan.chick_sales.exclude(id=chick_sale_id or 0).exists():
            raise ValidationError(f"Challan {challan.challan_no} has already been converted to a chick sale.")
        values = {
            "date": data["date"],
            "customer": get_object_or_404(Customer, id=data["customer"]),
            "warehouse": get_object_or_404(Warehouse, id=data["warehouse"]),
            "delivery_challan": challan,
            "shipping_address": data.get("shipping_address") or "",
            "vehicle": data.get("vehicle") or "",
            "driver": data.get("driver") or "",
            "freight_type": data.get("freight_type") or "Paid by Customer",
            "payment_mode": data.get("payment_mode") or "pay_later",
            "pay_account": ChartOfAccount.objects.filter(id=data.get("pay_account") or 0).first(),
            "freight_account": ChartOfAccount.objects.filter(id=data.get("freight_account") or 0).first(),
            "freight_amount": Decimal(str(data.get("freight_amount") or 0)),
            "terms": data.get("terms") or "",
            "remarks": data.get("remarks") or "",
        }
        if chick_sale_id:
            cs = get_object_or_404(ChickSale, id=chick_sale_id)
            for field, value in values.items():
                setattr(cs, field, value)
            cs.full_clean(exclude=["bill_no"])
            cs.save()
            cs.items.all().delete()
        else:
            cs = ChickSale(**values)
            cs.full_clean(exclude=["bill_no"])
            cs.save()
        for row in data.get("items", []):
            if not row.get("item"):
                continue
            ChickSaleItem.objects.create(
                sale=cs, item=get_object_or_404(Item, id=row["item"]),
                farm=row.get("farm") or "",
                total_qty=Decimal(str(row.get("total_qty") or 0)),
                mortality=Decimal(str(row.get("mortality") or 0)),
                culls=Decimal(str(row.get("culls") or 0)),
                sale_qty=Decimal(str(row.get("sale_qty") or 0)),
                discount_percent=Decimal(str(row.get("discount_percent") or 0)),
                discount_amount=Decimal(str(row.get("discount_amount") or 0)),
                free_qty=Decimal(str(row.get("free_qty") or 0)),
                net_qty=Decimal(str(row.get("net_qty") or 0)),
                sale_rate=Decimal(str(row.get("sale_rate") or 0)),
            )
        if not cs.items.exists():
            raise ValidationError("Add at least one item line.")
        cs.recalculate()
        return cs


# --------------------------------------------------------------------------
# Hatchery reports (shared engine in hatchery/reports.py)
# --------------------------------------------------------------------------
from .reports import respond_report  # noqa: E402

EGG_PURCHASE_REPORT_COLUMNS = [
    ("transaction_no", "Trnum", "text"),
    ("date", "Date", "text"),
    ("dc_no", "DC No", "text"),
    ("supplier", "Supplier", "text"),
    ("warehouse", "Farm/Warehouse", "text"),
    ("items", "Items", "text"),
    ("sent_qty", "Sent Qty", "num"),
    ("rcv_qty", "Rcv Qty", "num"),
    ("net_rate", "Rate", "money"),
    ("gross_amount", "Gross Amount", "money"),
    ("freight_amount", "Freight", "money"),
    ("tcs_amount", "TCS", "money"),
    ("net_amount", "Net Amount", "money"),
    ("payment_mode", "Payment", "text"),
]


@method_decorator(login_required, name="dispatch")
class EggPurchaseReportView(View):
    def get(self, request):
        from_date = request.GET.get("from_date", "").strip()
        to_date = request.GET.get("to_date", "").strip()
        supplier = request.GET.get("supplier", "").strip()

        qs = EggPurchase.objects.select_related("supplier", "warehouse").prefetch_related("items__item")
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
        if supplier:
            qs = qs.filter(supplier_id=supplier)
        qs = qs.order_by("date", "id")

        rows = []
        for ep in qs:
            items = list(ep.items.all())
            rows.append({
                "transaction_no": ep.transaction_no,
                "date": ep.date.isoformat(),
                "dc_no": ep.dc_no or "",
                "supplier": ep.supplier.name,
                "warehouse": ep.warehouse.name,
                "items": ", ".join(sorted({r.item.item_code for r in items})),
                "sent_qty": sum((r.sent_qty for r in items), Decimal("0")),
                "rcv_qty": sum((r.rcv_qty for r in items), Decimal("0")),
                "net_rate": ep.net_rate(),
                "gross_amount": ep.gross_amount(),
                "freight_amount": ep.freight_amount,
                "tcs_amount": ep.tcs_amount(),
                "net_amount": ep.net_amount(),
                "payment_mode": ep.get_payment_mode_display(),
            })

        return respond_report(
            request, title="Egg Purchase Report", icon="fas fa-shopping-basket",
            subtitle="Egg purchases with quantities, freight, TCS and net amounts",
            active_tab="egg_purchase_report", slug="egg_purchase_report",
            columns=EGG_PURCHASE_REPORT_COLUMNS, rows=rows,
            filters={"from_date": from_date, "to_date": to_date},
            extra_filters=[{
                "name": "supplier", "label": "Supplier", "selected": supplier,
                "options": [(s.id, s.name) for s in Supplier.objects.order_by("name")],
            }],
        )


INCUBATION_REPORT_COLUMNS = [
    ("grading_no", "Grading No", "text"),
    ("grading_date", "Grading Date", "text"),
    ("supplier", "Supplier", "text"),
    ("item", "Item", "text"),
    ("graded_qty", "Graded Qty", "num"),
    ("rejected_qty", "Rejected", "num"),
    ("setting_no", "Setting No", "text"),
    ("setting_date", "Setting Date", "text"),
    ("hatchery", "Hatchery", "text"),
    ("eggs_set", "Eggs Set", "num"),
    ("transfer_date", "Transfer Date", "text"),
    ("hatch_trnum", "Hatch Trnum", "text"),
    ("hatch_date", "Hatch Date", "text"),
    ("saleable_chicks", "Saleable Chicks", "num"),
    ("hatch_pct", "Hatch %", "pct"),
    ("egg_cost", "Egg Cost", "money"),
    ("vaccine_cost", "Vaccine Cost", "money"),
    ("net_cost", "Net Cost", "money"),
    ("net_rate", "Net Rate/Chick", "money"),
]


@method_decorator(login_required, name="dispatch")
class IncubationReportView(View):
    """Egg Grading + Tray Set + Hatch Entry in one report, one row per tray
    setting, shared columns (supplier, item, dates) shown once."""

    def get(self, request):
        from_date = request.GET.get("from_date", "").strip()
        to_date = request.GET.get("to_date", "").strip()
        hatchery = request.GET.get("hatchery", "").strip()

        qs = TraySetting.objects.select_related(
            "hatchery", "grading__supplier", "grading__item").prefetch_related("lines", "hatch_entry__vaccines")
        if from_date:
            qs = qs.filter(setting_date__gte=from_date)
        if to_date:
            qs = qs.filter(setting_date__lte=to_date)
        if hatchery:
            qs = qs.filter(hatchery_id=hatchery)
        qs = qs.order_by("setting_date", "id")

        rows = []
        for ts in qs:
            g = ts.grading
            try:
                he = ts.hatch_entry
            except HatchEntry.DoesNotExist:
                he = None
            eggs_set = ts.total_eggs_set()
            saleable = he.chicks_total if he else 0
            rejected = (g.broken_eggs or 0) + (g.damage_eggs or 0) + (g.misshapped_eggs or 0) + (g.dirty_eggs or 0)
            rows.append({
                "grading_no": g.transaction_no,
                "grading_date": g.date.isoformat(),
                "supplier": g.supplier.name,
                "item": g.item.item_code,
                "graded_qty": g.quantity,
                "rejected_qty": rejected,
                "setting_no": ts.setting_no,
                "setting_date": ts.setting_date.isoformat(),
                "hatchery": ts.hatchery.hatchery_name,
                "eggs_set": eggs_set,
                "transfer_date": ts.transfer_date.isoformat() if ts.transfer_date else "",
                "hatch_trnum": he.transaction_no if he else "",
                "hatch_date": (he.hatch_date or ts.hatch_date).isoformat() if he else ts.hatch_date.isoformat(),
                "saleable_chicks": saleable,
                "hatch_pct": _pct(saleable, eggs_set),
                "egg_cost": he.eggs_amount if he else 0,
                "vaccine_cost": he.vaccine_total() if he else 0,
                "net_cost": he.net_amount if he else 0,
                "net_rate": he.net_rate if he else 0,
            })

        return respond_report(
            request, title="Incubation Report", icon="fas fa-egg",
            subtitle="Egg grading, tray settings and hatch outcome in one sheet",
            active_tab="incubation_report", slug="incubation_report",
            columns=INCUBATION_REPORT_COLUMNS, rows=rows,
            filters={"from_date": from_date, "to_date": to_date},
            extra_filters=[{
                "name": "hatchery", "label": "Hatchery", "selected": hatchery,
                "options": [(h.id, h.hatchery_name) for h in Hatchery.objects.order_by("hatchery_name")],
            }],
        )


DELIVERY_CHALLAN_REPORT_COLUMNS = [
    ("challan_no", "Challan No", "text"),
    ("date", "Date", "text"),
    ("customer", "Customer", "text"),
    ("shipped_to", "Shipped To", "text"),
    ("place_of_supply", "Place of Supply", "text"),
    ("transporter", "Transporter", "text"),
    ("vehicle_no", "Vehicle", "text"),
    ("items", "Items", "text"),
    ("packing_size", "Packing Size", "text"),
    ("units", "No. of Units", "num"),
    ("quantity", "Quantity", "num"),
    ("tax", "Tax", "money"),
    ("discount", "Discount", "money"),
    ("amount", "Amount", "money"),
    ("status", "Status", "text"),
]


@method_decorator(login_required, name="dispatch")
class DeliveryChallanReportView(View):
    @staticmethod
    def _shipped_to_name(dc):
        """Ship-to party: the contact person on the customer's saved shipping
        address matching this challan, falling back to the customer name."""
        match = CustomerShippingAddress.objects.filter(
            customer_id=dc.customer_id, address=dc.shipping_address).first()
        if match and match.contact_person:
            return match.contact_person
        return dc.customer.name

    def get(self, request):
        from_date = request.GET.get("from_date", "").strip()
        to_date = request.GET.get("to_date", "").strip()
        customer = request.GET.get("customer", "").strip()

        qs = DeliveryChallan.objects.select_related("customer").prefetch_related("items__item", "chick_sales")
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
        if customer:
            qs = qs.filter(customer_id=customer)
        qs = qs.order_by("date", "id")

        rows = []
        for dc in qs:
            items = list(dc.items.all())
            sale = dc.chick_sales.first()
            rows.append({
                "challan_no": dc.challan_no,
                "date": dc.date.isoformat(),
                "customer": dc.customer.name,
                "shipped_to": self._shipped_to_name(dc),
                "place_of_supply": dc.place_of_supply or "",
                "transporter": dc.transporter_name or "",
                "vehicle_no": dc.vehicle_no or "",
                "items": ", ".join(sorted({r.item.item_code for r in items})),
                "packing_size": ", ".join(sorted({"%g" % r.packing_size for r in items}, key=float)),
                "units": sum((r.units for r in items), Decimal("0")),
                "quantity": dc.total_quantity(),
                "tax": dc.total_tax(),
                "discount": dc.overall_discount or 0,
                "amount": dc.grand_total(),
                "status": f"Converted ({sale.bill_no})" if sale else "Pending",
            })

        return respond_report(
            request, title="Delivery Challan Report", icon="fas fa-truck",
            subtitle="Dispatch challans with quantities, amounts and conversion status",
            active_tab="delivery_challan_report", slug="delivery_challan_report",
            columns=DELIVERY_CHALLAN_REPORT_COLUMNS, rows=rows,
            filters={"from_date": from_date, "to_date": to_date},
            extra_filters=[{
                "name": "customer", "label": "Customer", "selected": customer,
                "options": [(c.id, c.name) for c in Customer.objects.order_by("name")],
            }],
        )


CHICK_SALE_REPORT_COLUMNS = [
    ("bill_no", "Bill No", "text"),
    ("date", "Date", "text"),
    ("dc_no", "DC No", "text"),
    ("customer", "Customer", "text"),
    ("warehouse", "Farm/Warehouse", "text"),
    ("items", "Items", "text"),
    ("birds", "Sale Qty", "num"),
    ("free_qty", "Free Qty", "num"),
    ("billed_qty", "Billed Qty", "num"),
    ("avg_rate", "Avg Rate", "money"),
    ("freight", "Freight", "money"),
    ("final_amount", "Final Amount", "money"),
]


@method_decorator(login_required, name="dispatch")
class ChickSaleReportView(View):
    def get(self, request):
        from_date = request.GET.get("from_date", "").strip()
        to_date = request.GET.get("to_date", "").strip()
        customer = request.GET.get("customer", "").strip()

        qs = ChickSale.objects.select_related("customer", "warehouse", "delivery_challan").prefetch_related("items__item")
        if from_date:
            qs = qs.filter(date__gte=from_date)
        if to_date:
            qs = qs.filter(date__lte=to_date)
        if customer:
            qs = qs.filter(customer_id=customer)
        qs = qs.order_by("date", "id")

        rows = []
        for cs in qs:
            items = list(cs.items.all())
            rows.append({
                "bill_no": cs.bill_no,
                "date": cs.date.isoformat(),
                "dc_no": cs.delivery_challan.challan_no if cs.delivery_challan else "",
                "customer": cs.customer.name,
                "warehouse": cs.warehouse.name,
                "items": ", ".join(sorted({r.item.item_code for r in items})),
                "birds": sum((r.sale_qty for r in items), Decimal("0")),
                "free_qty": cs.total_free_qty(),
                "billed_qty": cs.total_net_qty(),
                "avg_rate": cs.avg_amount,
                "freight": cs.freight_amount,
                "final_amount": cs.final_amount,
            })

        return respond_report(
            request, title="Chick Sale Report", icon="fas fa-cash-register",
            subtitle="Chick sale invoices with quantities, rates, freight and profit",
            active_tab="chick_sale_report", slug="chick_sale_report",
            columns=CHICK_SALE_REPORT_COLUMNS, rows=rows,
            filters={"from_date": from_date, "to_date": to_date},
            extra_filters=[{
                "name": "customer", "label": "Customer", "selected": customer,
                "options": [(c.id, c.name) for c in Customer.objects.order_by("name")],
            }],
        )


# --------------------------------------------------------------------------
# Change requests: users without edit/delete rights submit proposed changes;
# users holding the right review and apply them.
# --------------------------------------------------------------------------
from django.utils import timezone as _tz
from user.access import user_can

CHANGE_REQUEST_HANDLERS = {
    "egg_purchase": {
        "api": "/egg_purchase_api/",
        "label": "Egg Purchase", "tab": "egg_purchase_list", "model": EggPurchase,
        "save": lambda data, oid: EggPurchaseAPI()._save_egg_purchase(data, oid),
        "number": lambda obj: obj.transaction_no,
    },
    "egg_grading": {
        "api": "/egg_grading_api/",
        "label": "Egg Grading", "tab": "egg_grading_list", "model": EggGrading,
        "save": lambda data, oid: EggGradingAPI()._save_egg_grading(data, oid),
        "number": lambda obj: obj.transaction_no,
    },
    "tray_set": {
        "api": "/tray_set_api/",
        "label": "Tray Set", "tab": "tray_set_list", "model": TraySetting,
        "save": lambda data, oid: TraySettingAPI()._save(data, oid),
        "number": lambda obj: obj.setting_no,
    },
    "hatch_entry": {
        "api": "/hatch_entry_api/",
        "label": "Hatch Entry", "tab": "hatch_entry_list", "model": HatchEntry,
        "save": lambda data, oid: HatchEntryAPI()._save(data, oid),
        "number": lambda obj: obj.transaction_no,
    },
    "hatchery": {
        "api": "/hatchery_api/",
        "label": "Hatch Register", "tab": "hatchery_list", "model": HatchSetting,
        "save": lambda data, oid: HatchSettingAPI()._save_hatch_setting(data, oid),
        "number": lambda obj: obj.setting_no,
    },
    "delivery_challan": {
        "api": "/delivery_challan_api/",
        "label": "Delivery Challan", "tab": "delivery_challan_list", "model": DeliveryChallan,
        "save": lambda data, oid: DeliveryChallanAPI()._save(data, oid),
        "number": lambda obj: obj.challan_no,
    },
    "chick_sale": {
        "api": "/chick_sale_api/",
        "label": "Chick Sale", "tab": "chick_sale_list", "model": ChickSale,
        "save": lambda data, oid: ChickSaleAPI()._save(data, oid),
        "number": lambda obj: obj.bill_no,
    },
}


@method_decorator(login_required, name="dispatch")
class ChangeRequestListTemplateView(View):
    def get(self, request):
        return render(request, "change_request_list.html")


def _change_request_to_dict(cr, user):
    handler = CHANGE_REQUEST_HANDLERS.get(cr.module, {})
    return {
        "id": cr.id, "module": cr.module, "module_label": handler.get("label", cr.module),
        "object_id": cr.object_id, "object_label": cr.object_label,
        "action": cr.action, "action_label": cr.get_action_display(),
        "status": cr.status, "note": cr.note,
        "payload": cr.payload,
        "requested_by": cr.requested_by.username,
        "requested_at": _tz.localtime(cr.requested_at).strftime("%Y-%m-%d %H:%M"),
        "reviewed_by": cr.reviewed_by.username if cr.reviewed_by else "",
        "reviewed_at": _tz.localtime(cr.reviewed_at).strftime("%Y-%m-%d %H:%M") if cr.reviewed_at else "",
        "review_note": cr.review_note,
        "can_review": user_can(user, handler.get("tab", ""), cr.action) if handler else False,
        "detail_api": handler.get("api", ""),
    }


@method_decorator(login_required, name="dispatch")
class ChangeRequestAPI(BaseAPIView):
    @staticmethod
    def _fk_labels():
        """id -> human label per payload field name, so the diff viewer can show
        names instead of raw foreign-key ids (both current and proposed sides)."""
        items = {str(i.id): f"{i.item_code} - {i.description}" for i in Item.objects.all()}
        warehouses = {str(w.id): w.name for w in Warehouse.objects.all()}
        accounts = {str(a.id): f"{a.code} - {a.description}" for a in ChartOfAccount.objects.all()}
        return {
            "supplier": {str(s.id): s.name for s in Supplier.objects.all()},
            "customer": {str(c.id): c.name for c in Customer.objects.all()},
            "warehouse": warehouses,
            "storage_location": warehouses,
            "item": items,
            "hatch_item": items,
            "pay_account": accounts,
            "freight_account": accounts,
            "hatchery": {str(h.id): h.hatchery_name for h in Hatchery.objects.all()},
            "setter": {str(s.id): s.setter_no for s in Setter.objects.all()},
            "grading": {str(g.id): g.transaction_no for g in EggGrading.objects.all()},
            "purchase_invoice": {str(p.id): p.transaction_no for p in EggPurchase.objects.all()},
            "tray_setting": {str(t.id): t.setting_no for t in TraySetting.objects.all()},
            "delivery_challan": {str(d.id): d.challan_no for d in DeliveryChallan.objects.all()},
        }

    def get(self, request):
        try:
            requests_qs = ChangeRequest.objects.select_related("requested_by", "reviewed_by").all()[:300]
            return JsonResponse({
                "requests": [_change_request_to_dict(cr, request.user) for cr in requests_qs],
                "labels": self._fk_labels(),
            })
        except Exception as e:
            return self.handle_exception(e)

    def post(self, request):
        """Create a change request: {module, object_id, action, payload?, note?}."""
        try:
            data = json.loads(request.body)
            handler = CHANGE_REQUEST_HANDLERS.get(data.get("module"))
            if not handler:
                raise ValidationError("Unknown module for change request.")
            if not user_can(request.user, handler["tab"], "view"):
                raise ValidationError("You do not have access to this module.")
            action = data.get("action")
            if action not in ("edit", "delete"):
                raise ValidationError("Invalid action.")
            obj = get_object_or_404(handler["model"], id=data.get("object_id"))
            if action == "edit" and not data.get("payload"):
                raise ValidationError("No proposed changes supplied.")
            cr = ChangeRequest.objects.create(
                module=data["module"], object_id=obj.id,
                object_label=handler["number"](obj),
                action=action,
                payload=data.get("payload") if action == "edit" else None,
                note=data.get("note") or "",
                requested_by=request.user,
            )
            return JsonResponse({"id": cr.id, "message": "Change request submitted for approval."}, status=201)
        except Exception as e:
            return self.handle_exception(e)


@method_decorator(login_required, name="dispatch")
class ChangeRequestReviewAPI(BaseAPIView):
    """POST /change_request_api/<id>/approve|reject/ — reviewer must hold the
    matching edit/delete right on the target module's tab."""

    @transaction.atomic
    def post(self, request, id, decision):
        try:
            cr = get_object_or_404(ChangeRequest, id=id)
            if cr.status != "pending":
                raise ValidationError("This request has already been reviewed.")
            handler = CHANGE_REQUEST_HANDLERS.get(cr.module)
            if not handler:
                raise ValidationError("Unknown module for change request.")
            if not user_can(request.user, handler["tab"], cr.action):
                return JsonResponse(
                    {"error": "You do not have permission to review this request."}, status=403)
            data = json.loads(request.body) if request.body else {}
            if decision == "approve":
                obj = handler["model"].objects.filter(id=cr.object_id).first()
                if obj is None:
                    raise ValidationError("The record no longer exists.")
                if cr.action == "delete":
                    obj.delete()
                else:
                    handler["save"](cr.payload, cr.object_id)
                cr.status = "approved"
            elif decision == "reject":
                cr.status = "rejected"
            else:
                raise ValidationError("Invalid decision.")
            cr.reviewed_by = request.user
            cr.reviewed_at = _tz.now()
            cr.review_note = data.get("review_note") or ""
            cr.save()
            return JsonResponse({"message": f"Request {cr.status}."})
        except Exception as e:
            return self.handle_exception(e)
